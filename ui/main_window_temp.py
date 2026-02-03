"""
主界�?
"""
import sqlite3
from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QLineEdit, QPushButton, QTableWidget,
                             QTableWidgetItem, QTabWidget, QMessageBox,
                             QFileDialog, QComboBox, QCheckBox, QDateEdit, QHeaderView)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont
from datetime import datetime


class MainWindow(QMainWindow):
    def __init__(self, db_manager, import_service, export_service, user):
        super().__init__()
        self.db_manager = db_manager
        self.import_service = import_service
        self.export_service = export_service
        self.user = user
        self.current_youth_id = None
        self.init_ui()
    
    def setup_table_style(self, table):
        """设置表格的统一样式，包括字体大小和行高"""
        # 设置表格字体
        table_font = QFont()
        table_font.setPointSize(11)  # 设置字体大小�?1pt
        table.setFont(table_font)
        
        # 设置表头字体
        header_font = QFont()
        header_font.setPointSize(11)  # 表头字体大小
        header_font.setBold(True)  # 表头加粗
        table.horizontalHeader().setFont(header_font)
        
        # 设置默认行高（根据字体自适应�?
        table.verticalHeader().setDefaultSectionSize(40)  # 设置行高�?0像素
        
        # 设置表格样式
        table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
                background-color: white;
                selection-background-color: #e3f2fd;
                selection-color: black;
                font-size: 11pt;
            }
            QTableWidget::item {
                padding: 5px;
                color: black;
            }
            QTableWidget::item:selected {
                color: black;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 8px;
                border: 1px solid #d0d0d0;
                font-weight: bold;
                font-size: 11pt;
            }
        """)
    
    def setup_button_style(self, button, button_type='normal'):
        """设置按钮的统一样式
        
        Args:
            button: QPushButton对象
            button_type: 按钮类型 ('normal', 'primary', 'danger', 'success')
        """
        # 设置按钮字体
        button_font = QFont()
        button_font.setPointSize(10)  # 按钮字体大小10pt
        button.setFont(button_font)
        
        # 设置按钮最小尺�?
        button.setMinimumHeight(32)  # 最小高�?2像素
        button.setMinimumWidth(80)   # 最小宽�?0像素
        
        # 根据按钮类型设置不同的样�?
        if button_type == 'primary':
            # 主要按钮（蓝色）
            button.setStyleSheet("""
                QPushButton {
                    background-color: #3498DB;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 16px;
                    font-size: 10pt;
                    min-height: 32px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #2980B9;
                }
                QPushButton:pressed {
                    background-color: #21618C;
                }
                QPushButton:disabled {
                    background-color: #BDC3C7;
                    color: #7F8C8D;
                }
            """)
        elif button_type == 'danger':
            # 危险按钮（红色）
            button.setStyleSheet("""
                QPushButton {
                    background-color: #E74C3C;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 16px;
                    font-size: 10pt;
                    min-height: 32px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #C0392B;
                }
                QPushButton:pressed {
                    background-color: #A93226;
                }
                QPushButton:disabled {
                    background-color: #BDC3C7;
                    color: #7F8C8D;
                }
            """)
        elif button_type == 'success':
            # 成功按钮（绿色）
            button.setStyleSheet("""
                QPushButton {
                    background-color: #27AE60;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 16px;
                    font-size: 10pt;
                    min-height: 32px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
                QPushButton:pressed {
                    background-color: #1E8449;
                }
                QPushButton:disabled {
                    background-color: #BDC3C7;
                    color: #7F8C8D;
                }
            """)
        else:
            # 普通按钮（灰色边框�?
            button.setStyleSheet("""
                QPushButton {
                    background-color: white;
                    color: #2C3E50;
                    border: 1px solid #BDC3C7;
                    border-radius: 4px;
                    padding: 6px 16px;
                    font-size: 10pt;
                    min-height: 32px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #ECF0F1;
                    border-color: #95A5A6;
                }
                QPushButton:pressed {
                    background-color: #D5DBDB;
                }
                QPushButton:disabled {
                    background-color: #F8F9F9;
                    color: #BDC3C7;
                    border-color: #D5DBDB;
                }
            """)
    
    def init_ui(self):
        self.setWindowTitle(f'一人一策记录本 - {self.user.unit} - {self.user.username}')
        
        # 默认全屏显示
        self.showMaximized()
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局 - 水平布局
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # 左侧导航�?
        nav_widget = self.create_navigation()
        main_layout.addWidget(nav_widget)
        
        # 右侧内容区域
        content_widget = QWidget()
        content_layout = QVBoxLayout()
        content_layout.setContentsMargins(0, 0, 0, 0)
        
        # 标签页内容（不显示标签）
        self.tabs = QTabWidget()
        self.tabs.tabBar().hide()  # 隐藏标签�?
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background: white;
            }
        """)
        
        self.create_tabs()
        content_layout.addWidget(self.tabs)
        
        # 底部操作按钮
        button_layout = self.create_bottom_buttons()
        content_layout.addLayout(button_layout)
        
        content_widget.setLayout(content_layout)
        main_layout.addWidget(content_widget)
        
        central_widget.setLayout(main_layout)
        
        # 初始化每日情况统计数�?
        self.load_daily_stats_data()
        
        # 初始化镇街谈心谈话数�?
        self.load_town_interview_data()
        
        # 初始化领导谈心谈话数�?
        self.load_leader_interview_data()
        
        # 初始化异常情况统计数�?
        self.load_exception_statistics_data()
        self.load_leader_interview_data()
    
    def create_navigation(self):
        """创建左侧导航�?""
        nav_widget = QWidget()
        nav_widget.setFixedWidth(240)
        nav_widget.setStyleSheet("""
            QWidget {
                background: #34495e;
            }
        """)
        
        nav_layout = QVBoxLayout()
        nav_layout.setContentsMargins(0, 0, 0, 0)
        nav_layout.setSpacing(0)
        
        # 顶部标题
        title_label = QLabel('一人一策系�?)
        title_label.setStyleSheet("""
            QLabel {
                background: #2c3e50;
                color: white;
                font-size: 16px;
                font-weight: bold;
                padding: 20px;
                border-bottom: 2px solid #1abc9c;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        nav_layout.addWidget(title_label)
        
        # 导航按钮
        self.nav_buttons = []
        nav_items = [
            '基本信息',
            '病史筛查情况',
            '镇街谈心谈话情况',
            '领导谈心谈话情况',
            '每日情况统计',
            '异常情况统计',
            '隐性疾病及心理问题筛查情况'
        ]
        
        for index, item in enumerate(nav_items):
            btn = QPushButton(item)
            # 为最后一个按钮设置稍小的字体
            if index == len(nav_items) - 1:
                btn.setStyleSheet("""
                    QPushButton {
                        background: #34495e;
                        color: #ecf0f1;
                        border: none;
                        border-bottom: 1px solid #2c3e50;
                        padding: 18px 10px;
                        text-align: left;
                        font-size: 14px;
                    }
                    QPushButton:hover {
                        background: #3d566e;
                        color: white;
                    }
                    QPushButton:checked {
                        background: #3498db;
                        color: white;
                        font-weight: bold;
                        border-left: 4px solid #e74c3c;
                    }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton {
                        background: #34495e;
                        color: #ecf0f1;
                        border: none;
                        border-bottom: 1px solid #2c3e50;
                        padding: 18px 15px;
                        text-align: left;
                        font-size: 15px;
                    }
                    QPushButton:hover {
                        background: #3d566e;
                        color: white;
                    }
                    QPushButton:checked {
                        background: #3498db;
                        color: white;
                        font-weight: bold;
                        border-left: 4px solid #e74c3c;
                    }
                """)
            btn.setCheckable(True)
            btn.clicked.connect(lambda checked, idx=index: self.switch_tab(idx))
            self.nav_buttons.append(btn)
            nav_layout.addWidget(btn)
        
        # 默认选中第一�?
        self.nav_buttons[0].setChecked(True)
        
        # 添加弹性空�?
        nav_layout.addStretch()
        
        nav_widget.setLayout(nav_layout)
        return nav_widget
    
    def switch_tab(self, index):
        """切换标签�?""
        # 取消其他按钮的选中状�?
        for i, btn in enumerate(self.nav_buttons):
            btn.setChecked(i == index)
        
        # 切换到对应的标签�?
        self.tabs.setCurrentIndex(index)
        
        # 如果切换到异常情况统计标签页，自动刷新数�?
        if index == 5:  # 异常情况统计是第6个标签页（索引为5�?
            self.load_exception_statistics_data()
    
    def create_search_bar(self):
        layout = QHBoxLayout()
        
        self.name_input = QLineEdit()
        self.name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.name_input.setPlaceholderText('姓名')
        self.name_input.returnPressed.connect(self.search_youth)
        
        self.id_card_input = QLineEdit()
        self.id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.id_card_input.setPlaceholderText('身份证号')
        self.id_card_input.returnPressed.connect(self.search_youth)
        
        self.district_input = QLineEdit()
        self.district_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.district_input.setPlaceholderText('区划')
        self.district_input.returnPressed.connect(self.search_youth)
        
        self.street_school_input = QLineEdit()
        self.street_school_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.street_school_input.setPlaceholderText('镇街/毕业院校')
        self.street_school_input.returnPressed.connect(self.search_youth)
        
        # 新增三个搜索框：连、排、班
        self.company_input = QLineEdit()
        self.company_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.company_input.setPlaceholderText('�?)
        self.company_input.returnPressed.connect(self.search_youth)
        
        self.platoon_input = QLineEdit()
        self.platoon_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.platoon_input.setPlaceholderText('�?)
        self.platoon_input.returnPressed.connect(self.search_youth)
        
        self.squad_input = QLineEdit()
        self.squad_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.squad_input.setPlaceholderText('�?)
        self.squad_input.returnPressed.connect(self.search_youth)
        
        search_btn = QPushButton('搜索')
        search_btn.clicked.connect(self.search_youth)
        self.setup_button_style(search_btn, 'primary')
        
        reset_btn = QPushButton('重置')
        reset_btn.clicked.connect(self.reset_search)
        self.setup_button_style(reset_btn, 'normal')
        
        layout.addWidget(QLabel('搜索:'))
        layout.addWidget(self.name_input)
        layout.addWidget(self.id_card_input)
        layout.addWidget(self.district_input)
        layout.addWidget(self.street_school_input)
        layout.addWidget(self.company_input)
        layout.addWidget(self.platoon_input)
        layout.addWidget(self.squad_input)
        layout.addWidget(search_btn)
        layout.addWidget(reset_btn)
        
        return layout
    
    def create_tabs(self):
        """创建7个板块标签页"""
        tab_names = [
            '基本信息',
            '病史筛查情况',
            '镇街谈心谈话情况',
            '领导谈心谈话情况',
            '每日情况统计',
            '异常情况统计',
            '隐性疾病及心理问题筛查情况'
        ]
        
        self.tab_tables = {}
        
        for tab_name in tab_names:
            if tab_name == '基本信息':
                # 基本信息标签页特殊处理，包含搜索和列�?
                tab_widget = self.create_basic_info_tab()
            elif tab_name == '病史筛查情况':
                # 病史筛查情况标签页特殊处�?
                tab_widget = self.create_medical_screening_tab()
            elif tab_name == '镇街谈心谈话情况':
                # 镇街谈心谈话情况标签页特殊处�?
                tab_widget = self.create_town_interview_tab()
            elif tab_name == '领导谈心谈话情况':
                # 领导谈心谈话情况标签页特殊处�?
                tab_widget = self.create_leader_interview_tab()
            elif tab_name == '异常情况统计':
                # 异常情况统计标签页特殊处�?
                tab_widget = self.create_exception_statistics_tab()
            else:
                tab_widget = QWidget()
                tab_layout = QVBoxLayout()
                
                # 特殊功能按钮 - 搜索功能放在最前面（表格之前）
                if tab_name == '每日情况统计':
                    # 添加搜索功能（放在顶部）
                    search_layout = QHBoxLayout()
                    search_layout.addWidget(QLabel('搜索:'))
                    
                    self.daily_name_input = QLineEdit()
                    self.daily_name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
                    self.daily_name_input.setPlaceholderText('姓名')
                    self.daily_name_input.returnPressed.connect(self.search_daily_stats)
                    search_layout.addWidget(self.daily_name_input)
                    
                    self.daily_id_card_input = QLineEdit()
                    self.daily_id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
                    self.daily_id_card_input.setPlaceholderText('公民身份号码')
                    self.daily_id_card_input.returnPressed.connect(self.search_daily_stats)
                    search_layout.addWidget(self.daily_id_card_input)
                    
                    search_btn = QPushButton('搜索')
                    search_btn.clicked.connect(self.search_daily_stats)
                    self.setup_button_style(search_btn, 'primary')
                    search_layout.addWidget(search_btn)
                    
                    reset_search_btn = QPushButton('重置')
                    reset_search_btn.clicked.connect(self.reset_daily_stats_search)
                    self.setup_button_style(reset_search_btn, 'normal')
                    search_layout.addWidget(reset_search_btn)
                    
                    tab_layout.addLayout(search_layout)
                    
                    # 添加全选复选框�?
                    select_all_layout = QHBoxLayout()
                    self.daily_select_all_checkbox = QCheckBox('全�?)
                    self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
                    self.daily_select_all_checkbox.setStyleSheet("""
                        QCheckBox {
                            font-size: 14px;
                            font-weight: bold;
                            padding: 5px;
                        }
                        QCheckBox::indicator {
                            width: 20px;
                            height: 20px;
                            border: 2px solid #7F8C8D;
                            border-radius: 2px;
                            background-color: white;
                        }
                        QCheckBox::indicator:hover {
                            border-color: #3498DB;
                            background-color: #ECF0F1;
                        }
                        QCheckBox::indicator:checked {
                            background-color: #3498DB;
                            border-color: #2980B9;
                        }
                    """)
                    select_all_layout.addWidget(self.daily_select_all_checkbox)
                    select_all_layout.addStretch()
                    tab_layout.addLayout(select_all_layout)
                
                # 创建表格
                table = QTableWidget()
                table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # 禁止直接编辑
                self.tab_tables[tab_name] = table
                
                # 应用统一的表格样�?
                self.setup_table_style(table)
                
                tab_layout.addWidget(table)
                
                # 添加操作按钮
                button_layout = QHBoxLayout()
                
                # 导入按钮 - 每日情况统计改为添加信息
                if tab_name == '每日情况统计':
                    import_btn = QPushButton(f'添加信息')
                    import_btn.clicked.connect(lambda checked, t=tab_name: self.add_daily_info(t))
                    self.setup_button_style(import_btn, 'normal')
                    button_layout.addWidget(import_btn)
                    
                    # 批量添加信息按钮
                    batch_add_btn = QPushButton('批量添加信息')
                    self.setup_button_style(batch_add_btn, 'success')
                    batch_add_btn.clicked.connect(self.batch_add_daily_info)
                    button_layout.addWidget(batch_add_btn)
                    
                    # 批量删除按钮
                    batch_delete_btn = QPushButton('批量删除')
                    self.setup_button_style(batch_delete_btn, 'danger')
                    batch_delete_btn.clicked.connect(self.batch_delete_daily_stats)
                    button_layout.addWidget(batch_delete_btn)
                else:
                    import_btn = QPushButton(f'导入数据')
                    import_btn.clicked.connect(lambda checked, t=tab_name: self.import_data(t))
                    self.setup_button_style(import_btn, 'normal')
                    button_layout.addWidget(import_btn)
                
                # 导出按钮
                export_btn = QPushButton(f'导出数据')
                export_btn.clicked.connect(lambda checked, t=tab_name: self.export_data(t))
                self.setup_button_style(export_btn, 'normal')
                button_layout.addWidget(export_btn)
                
                # 特殊功能按钮
                if tab_name == '每日情况统计':
                    # 添加日期筛选功�?
                    filter_layout = QHBoxLayout()
                    
                    filter_layout.addWidget(QLabel('日期筛�?'))
                    
                    self.start_date_input = QDateEdit()
                    self.start_date_input.setDate(QDate.currentDate().addDays(-30))  # 默认30天前
                    self.start_date_input.setCalendarPopup(True)
                    filter_layout.addWidget(QLabel('�?'))
                    filter_layout.addWidget(self.start_date_input)
                    
                    self.end_date_input = QDateEdit()
                    self.end_date_input.setDate(QDate.currentDate())  # 默认今天
                    self.end_date_input.setCalendarPopup(True)
                    filter_layout.addWidget(QLabel('�?'))
                    filter_layout.addWidget(self.end_date_input)
                    
                    filter_btn = QPushButton('筛�?)
                    filter_btn.clicked.connect(self.filter_daily_stats)
                    self.setup_button_style(filter_btn, 'primary')
                    filter_layout.addWidget(filter_btn)
                    
                    reset_filter_btn = QPushButton('重置')
                    reset_filter_btn.clicked.connect(self.reset_daily_stats_filter)
                    self.setup_button_style(reset_filter_btn, 'normal')
                    filter_layout.addWidget(reset_filter_btn)
                    
                    filter_layout.addStretch()
                    tab_layout.addLayout(filter_layout)
                
                button_layout.addStretch()
                tab_layout.addLayout(button_layout)
                
                tab_widget.setLayout(tab_layout)
            
            self.tabs.addTab(tab_widget, tab_name)
    
    def create_basic_info_tab(self):
        """创建基本信息标签�?""
        tab_widget = QWidget()
        layout = QVBoxLayout()
        
        # 搜索�?
        search_layout = self.create_search_bar()
        layout.addLayout(search_layout)
        
        # 全选按钮行
        select_all_layout = QHBoxLayout()
        self.select_all_checkbox = QCheckBox('全�?)
        self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
        self.select_all_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #7F8C8D;
                border-radius: 2px;
                background-color: white;
            }
            QCheckBox::indicator:hover {
                border-color: #3498DB;
                background-color: #ECF0F1;
            }
            QCheckBox::indicator:checked {
                background-color: #3498DB;
                border-color: #2980B9;
            }
        """)
        select_all_layout.addWidget(self.select_all_checkbox)
        select_all_layout.addStretch()
        layout.addLayout(select_all_layout)
        
        # 成员详细信息表格
        self.search_table = QTableWidget()
        # 设置所有字段列（旧结构：只保留一列编号）
        self.basic_info_headers = [
            '选择', '姓名', '性别', '公民身份号码', '民族', '政治面貌',
            '区划', '镇街', '连', '排', '班', '带训班长信息', '在营状态', '离营时间', '离营情况说明',
            '毕业或所在院校', '文化程度', '所学专业', '学业情况', '学习类型',
            '电话', '户籍地', '常住地', '家庭情况',
            '父母电话', '个人经历', '证明人', '电话', '操作', '删除'
        ]
        self.search_table.setColumnCount(len(self.basic_info_headers))
        self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
        self.search_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.search_table.setAlternatingRowColors(True)
        self.search_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # 禁止直接编辑
        self.search_table.horizontalHeader().setStretchLastSection(False)
        self.search_table.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        
        # 应用统一的表格样�?
        self.setup_table_style(self.search_table)
        
        # 添加双击事件处理
        self.search_table.cellDoubleClicked.connect(self.on_basic_info_double_click)
        
        layout.addWidget(self.search_table)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        import_btn = QPushButton('导入数据')
        import_btn.clicked.connect(lambda: self.import_data('基本信息'))
        self.setup_button_style(import_btn, 'normal')
        button_layout.addWidget(import_btn)
        
        export_btn = QPushButton('导出数据')
        export_btn.clicked.connect(lambda: self.export_data('基本信息'))
        self.setup_button_style(export_btn, 'normal')
        button_layout.addWidget(export_btn)
        
        add_btn = QPushButton('添加信息')
        add_btn.clicked.connect(self.add_youth)
        self.setup_button_style(add_btn, 'normal')
        button_layout.addWidget(add_btn)
        
        batch_delete_btn = QPushButton('批量删除')
        self.setup_button_style(batch_delete_btn, 'danger')
        batch_delete_btn.clicked.connect(self.batch_delete_youth)
        button_layout.addWidget(batch_delete_btn)
        
        import_camp_btn = QPushButton('导入入营点验')
        self.setup_button_style(import_camp_btn, 'primary')
        import_camp_btn.clicked.connect(self.import_camp_verification_data)
        button_layout.addWidget(import_camp_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        tab_widget.setLayout(layout)
        
        # 加载所有青年详细信�?
        self.load_all_youth_detailed()
        
        return tab_widget
    
    def create_town_interview_tab(self):
        """创建镇街谈心谈话情况标签�?""
        tab_widget = QWidget()
        layout = QVBoxLayout()
        
        # 搜索�?
        search_layout = QHBoxLayout()
        
        search_layout.addWidget(QLabel('搜索:'))
        
        self.town_name_input = QLineEdit()
        self.town_name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.town_name_input.setPlaceholderText('姓名')
        self.town_name_input.returnPressed.connect(self.search_town_interview)
        search_layout.addWidget(self.town_name_input)
        
        self.town_id_card_input = QLineEdit()
        self.town_id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.town_id_card_input.setPlaceholderText('身份证号')
        self.town_id_card_input.returnPressed.connect(self.search_town_interview)
        search_layout.addWidget(self.town_id_card_input)
        
        search_btn = QPushButton('搜索')
        search_btn.clicked.connect(self.search_town_interview)
        self.setup_button_style(search_btn, 'primary')
        search_layout.addWidget(search_btn)
        
        reset_btn = QPushButton('重置')
        reset_btn.clicked.connect(self.reset_town_interview_search)
        self.setup_button_style(reset_btn, 'normal')
        search_layout.addWidget(reset_btn)
        
        layout.addLayout(search_layout)
        
        # 全选按钮行
        select_all_layout = QHBoxLayout()
        self.town_select_all_checkbox = QCheckBox('全�?)
        self.town_select_all_checkbox.stateChanged.connect(self.toggle_town_select_all)
        self.town_select_all_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #7F8C8D;
                border-radius: 2px;
                background-color: white;
            }
            QCheckBox::indicator:hover {
                border-color: #3498DB;
                background-color: #ECF0F1;
            }
            QCheckBox::indicator:checked {
                background-color: #3498DB;
                border-color: #2980B9;
            }
        """)
        select_all_layout.addWidget(self.town_select_all_checkbox)
        select_all_layout.addStretch()
        layout.addLayout(select_all_layout)
        
        # 镇街谈心谈话表格
        self.town_interview_table = QTableWidget()
        self.town_interview_headers = [
            '选择', '日期', '姓名', '性别', '公民身份号码', '应征地', '连', '排', '班', '带训班长信息', '思想', '精神', '走访调查情况', '操作', '删除'
        ]
        self.town_interview_table.setColumnCount(len(self.town_interview_headers))
        self.town_interview_table.setHorizontalHeaderLabels(self.town_interview_headers)
        self.town_interview_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.town_interview_table.setAlternatingRowColors(True)
        self.town_interview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # 禁止编辑
        
        # 应用统一的表格样�?
        self.setup_table_style(self.town_interview_table)
        
        # 设置表格自适应屏幕
        header = self.town_interview_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)  # 允许手动调整
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 选择�?
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 日期�?
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 姓名�?
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 性别�?
        header.setSectionResizeMode(4, QHeaderView.Stretch)  # 公民身份号码列自动拉�?
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 思想�?
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 精神�?
        header.setSectionResizeMode(7, QHeaderView.Stretch)  # 走访调查情况列自动拉�?
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # 操作�?
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # 删除�?
        header.setStretchLastSection(False)
        header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        
        layout.addWidget(self.town_interview_table)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        add_btn = QPushButton('添加信息')
        add_btn.clicked.connect(self.add_town_interview)
        self.setup_button_style(add_btn, 'normal')
        button_layout.addWidget(add_btn)
        
        import_btn = QPushButton('导入文件')
        import_btn.clicked.connect(self.import_town_interview_files)
        self.setup_button_style(import_btn, 'normal')
        button_layout.addWidget(import_btn)
        
        export_btn = QPushButton('导出数据')
        export_btn.clicked.connect(lambda: self.export_data('镇街谈心谈话情况'))
        self.setup_button_style(export_btn, 'normal')
        button_layout.addWidget(export_btn)
        
        batch_delete_btn = QPushButton('批量删除')
        self.setup_button_style(batch_delete_btn, 'danger')
        batch_delete_btn.clicked.connect(self.batch_delete_town_interview)
        button_layout.addWidget(batch_delete_btn)
        
        refresh_btn = QPushButton('刷新')
        refresh_btn.clicked.connect(self.load_town_interview_data)
        self.setup_button_style(refresh_btn, 'normal')
        button_layout.addWidget(refresh_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        tab_widget.setLayout(layout)
        
        return tab_widget
    
    def create_leader_interview_tab(self):
        """创建领导谈心谈话情况标签�?""
        tab_widget = QWidget()
        layout = QVBoxLayout()
        
        # 搜索�?
        search_layout = QHBoxLayout()
        
        search_layout.addWidget(QLabel('搜索:'))
        
        self.leader_name_input = QLineEdit()
        self.leader_name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.leader_name_input.setPlaceholderText('姓名')
        self.leader_name_input.returnPressed.connect(self.search_leader_interview)
        search_layout.addWidget(self.leader_name_input)
        
        self.leader_id_card_input = QLineEdit()
        self.leader_id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.leader_id_card_input.setPlaceholderText('身份证号')
        self.leader_id_card_input.returnPressed.connect(self.search_leader_interview)
        search_layout.addWidget(self.leader_id_card_input)
        
        search_btn = QPushButton('搜索')
        search_btn.clicked.connect(self.search_leader_interview)
        self.setup_button_style(search_btn, 'primary')
        search_layout.addWidget(search_btn)
        
        reset_btn = QPushButton('重置')
        reset_btn.clicked.connect(self.reset_leader_interview_search)
        self.setup_button_style(reset_btn, 'normal')
        search_layout.addWidget(reset_btn)
        
        layout.addLayout(search_layout)
        
        # 全选按钮行
        select_all_layout = QHBoxLayout()
        self.leader_select_all_checkbox = QCheckBox('全�?)
        self.leader_select_all_checkbox.stateChanged.connect(self.toggle_leader_select_all)
        self.leader_select_all_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #7F8C8D;
                border-radius: 2px;
                background-color: white;
            }
            QCheckBox::indicator:hover {
                border-color: #3498DB;
                background-color: #ECF0F1;
            }
            QCheckBox::indicator:checked {
                background-color: #3498DB;
                border-color: #2980B9;
            }
        """)
        select_all_layout.addWidget(self.leader_select_all_checkbox)
        select_all_layout.addStretch()
        layout.addLayout(select_all_layout)
        
        # 领导谈心谈话表格
        self.leader_interview_table = QTableWidget()
        self.leader_interview_headers = [
            '选择', '日期', '姓名', '性别', '公民身份号码', '应征地', '连', '排', '班', '带训班长信息', '思想', '精神', '走访调查情况', '操作', '删除'
        ]
        self.leader_interview_table.setColumnCount(len(self.leader_interview_headers))
        self.leader_interview_table.setHorizontalHeaderLabels(self.leader_interview_headers)
        self.leader_interview_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.leader_interview_table.setAlternatingRowColors(True)
        self.leader_interview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # 禁止编辑
        
        # 应用统一的表格样�?
        self.setup_table_style(self.leader_interview_table)
        
        # 设置表格自适应屏幕
        header = self.leader_interview_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)  # 允许手动调整
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 选择�?
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 日期�?
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 姓名�?
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 性别�?
        header.setSectionResizeMode(4, QHeaderView.Stretch)  # 公民身份号码列自动拉�?
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 思想�?
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 精神�?
        header.setSectionResizeMode(7, QHeaderView.Stretch)  # 走访调查情况列自动拉�?
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # 操作�?
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # 删除�?
        header.setStretchLastSection(False)
        header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        
        layout.addWidget(self.leader_interview_table)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        add_btn = QPushButton('添加信息')
        add_btn.clicked.connect(self.add_leader_interview)
        self.setup_button_style(add_btn, 'normal')
        button_layout.addWidget(add_btn)
        
        import_btn = QPushButton('导入文件')
        import_btn.clicked.connect(self.import_leader_interview_files)
        self.setup_button_style(import_btn, 'normal')
        button_layout.addWidget(import_btn)
        
        export_btn = QPushButton('导出数据')
        export_btn.clicked.connect(lambda: self.export_data('领导谈心谈话情况'))
        self.setup_button_style(export_btn, 'normal')
        button_layout.addWidget(export_btn)
        
        batch_delete_btn = QPushButton('批量删除')
        self.setup_button_style(batch_delete_btn, 'danger')
        batch_delete_btn.clicked.connect(self.batch_delete_leader_interview)
        button_layout.addWidget(batch_delete_btn)
        
        refresh_btn = QPushButton('刷新')
        refresh_btn.clicked.connect(self.load_leader_interview_data)
        self.setup_button_style(refresh_btn, 'normal')
        button_layout.addWidget(refresh_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        tab_widget.setLayout(layout)
        
        return tab_widget
    
    def create_exception_statistics_tab(self):
        """创建异常情况统计标签�?""
        tab_widget = QWidget()
        layout = QVBoxLayout()
        
        # 搜索�?
        search_layout = QHBoxLayout()
        
        search_layout.addWidget(QLabel('搜索:'))
        
        self.exception_name_input = QLineEdit()
        self.exception_name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.exception_name_input.setPlaceholderText('姓名')
        self.exception_name_input.returnPressed.connect(self.search_exception_statistics)
        search_layout.addWidget(self.exception_name_input, 1)  # 添加拉伸因子
        
        self.exception_id_card_input = QLineEdit()
        self.exception_id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.exception_id_card_input.setPlaceholderText('身份证号')
        self.exception_id_card_input.returnPressed.connect(self.search_exception_statistics)
        search_layout.addWidget(self.exception_id_card_input, 1)  # 添加拉伸因子
        
        search_btn = QPushButton('搜索')
        search_btn.clicked.connect(self.search_exception_statistics)
        self.setup_button_style(search_btn, 'primary')
        search_layout.addWidget(search_btn)
        
        reset_btn = QPushButton('重置')
        reset_btn.clicked.connect(self.reset_exception_statistics_search)
        self.setup_button_style(reset_btn, 'normal')
        search_layout.addWidget(reset_btn)
        
        # 移除addStretch()，让输入框自动填充空�?
        layout.addLayout(search_layout)
        
        # 表格上方的控制栏（包含全选和时间范围�?
        table_control_layout = QHBoxLayout()
        
        # 左侧：全选复选框
        self.exception_select_all_checkbox = QCheckBox('全�?)
        self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
        self.exception_select_all_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #7F8C8D;
                border-radius: 2px;
                background-color: white;
            }
            QCheckBox::indicator:hover {
                border-color: #3498DB;
                background-color: #ECF0F1;
            }
            QCheckBox::indicator:checked {
                background-color: #3498DB;
                border-color: #2980B9;
            }
        """)
        table_control_layout.addWidget(self.exception_select_all_checkbox)
        
        # 中间：弹性空�?
        table_control_layout.addStretch()
        
        # 右侧：时间筛选器
        # 日期筛选区�?
        date_filter_layout = QHBoxLayout()
        
        # 日期筛选标�?
        date_filter_layout.addWidget(QLabel('日期筛�? �?'))
        
        # 开始日期选择�?
        from PyQt5.QtWidgets import QDateEdit
        from PyQt5.QtCore import QDate
        self.exception_start_date = QDateEdit()
        self.exception_start_date.setDate(QDate.currentDate().addDays(-30))  # 默认30天前
        self.exception_start_date.setCalendarPopup(True)
        self.exception_start_date.setStyleSheet("""
            QDateEdit {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
                min-width: 120px;
                color: #2c3e50;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        date_filter_layout.addWidget(self.exception_start_date)
        
        date_filter_layout.addWidget(QLabel('�?'))
        
        # 结束日期选择�?
        self.exception_end_date = QDateEdit()
        self.exception_end_date.setDate(QDate.currentDate())  # 默认今天
        self.exception_end_date.setCalendarPopup(True)
        self.exception_end_date.setStyleSheet("""
            QDateEdit {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
                min-width: 120px;
                color: #2c3e50;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        date_filter_layout.addWidget(self.exception_end_date)
        
        # 筛选按�?
        filter_btn = QPushButton('筛�?)
        filter_btn.clicked.connect(self.filter_exception_by_date_range)
        filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 60px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        date_filter_layout.addWidget(filter_btn)
        
        table_control_layout.addLayout(date_filter_layout)
        
        # 时间范围筛选下拉框
        table_control_layout.addWidget(QLabel('时间范围:'))
        self.exception_time_range_combo = QComboBox()
        self.exception_time_range_combo.addItems(['当天', '三天�?, '七天�?, '半个月内', '全部', '�?])
        self.exception_time_range_combo.setCurrentText('全部')  # 默认选择全部
        self.exception_time_range_combo.currentTextChanged.connect(self.on_exception_time_range_changed)
        self.exception_time_range_combo.setStyleSheet("""
            QComboBox {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
                min-width: 100px;
                color: #2c3e50;
            }
            QComboBox:hover {
                border-color: #3498db;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #7f8c8d;
                margin-right: 5px;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #bdc3c7;
                background-color: white;
                selection-background-color: #3498db;
                selection-color: white;
                color: #2c3e50;
            }
            QComboBox QAbstractItemView::item {
                padding: 8px;
                border: none;
                color: #2c3e50;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #ecf0f1;
                color: #2c3e50;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)
        table_control_layout.addWidget(self.exception_time_range_combo)
        
        layout.addLayout(table_control_layout)
        
        # 异常情况统计表格
        self.exception_statistics_table = QTableWidget()
        self.exception_statistics_headers = [
            '选择', '日期', '姓名', '性别', '公民身份号码', '思想', '身体', '精神', '其他', '操作'
        ]
        self.exception_statistics_table.setColumnCount(len(self.exception_statistics_headers))
        self.exception_statistics_table.setHorizontalHeaderLabels(self.exception_statistics_headers)
        self.exception_statistics_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.exception_statistics_table.setAlternatingRowColors(True)
        self.exception_statistics_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # 应用统一的表格样�?
        self.setup_table_style(self.exception_statistics_table)  # 禁止编辑
        
        # 添加双击事件处理
        self.exception_statistics_table.cellDoubleClicked.connect(self.on_exception_statistics_double_click)
        
        # 设置表格占满屏幕
        self.exception_statistics_table.setSizeAdjustPolicy(QTableWidget.SizeAdjustPolicy.AdjustToContentsOnFirstShow)
        
        # 设置表格列宽
        header = self.exception_statistics_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)  # 所有列平均分配宽度
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 选择�?
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 日期�?
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 姓名�?
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 性别�?
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 思想�?
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 身体�?
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # 精神�?
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # 操作�?
        # 让身份证号码列和异常来源列自动拉伸填充剩余空�?
        header.setSectionResizeMode(4, QHeaderView.Stretch)  # 公民身份号码�?
        header.setSectionResizeMode(8, QHeaderView.Stretch)  # 异常来源�?
        header.setStretchLastSection(False)  # 不自动拉伸最后一�?
        header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        
        layout.addWidget(self.exception_statistics_table)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        export_btn = QPushButton('导出数据')
        export_btn.clicked.connect(lambda: self.export_data('异常情况统计'))
        self.setup_button_style(export_btn, 'normal')
        button_layout.addWidget(export_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        tab_widget.setLayout(layout)
        
        return tab_widget
    
    def create_bottom_buttons(self):
        layout = QHBoxLayout()
        layout.addStretch()
        return layout
    
    def load_all_youth(self):
        """加载所有青年信息（已废弃，改用详细显示�?""
        self.load_all_youth_detailed()
    
    def load_all_youth_detailed(self):
        """加载所有青年详细信息（完整版，默认显示）（旧结构）"""
        try:
            conn = None
            try:
                conn = self.db_manager.get_connection()
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id_card, name, gender, nation, political_status,
                           district, street, company, platoon, squad,
                           squad_leader, camp_status, leave_time, situation_note,
                           school, education_level, major, study_status, study_type,
                           phone, household_address, residence_address, family_info,
                           parent_phone, personal_experience,
                           reference_person, reference_phone, id
                    FROM youth ORDER BY id
                """)
                results = cursor.fetchall()
                self.display_detailed_results(results)
            finally:
                if conn:
                    conn.close()
        except Exception as e:
            QMessageBox.warning(self, "加载错误", f"加载青年信息时发生错误：{str(e)}")
            self.search_table.setRowCount(0)
    
    def search_youth(self):
        """搜索青年（旧结构�?""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.select_all_checkbox.stateChanged.disconnect()
                self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            
            name = self.name_input.text().strip()
            id_card = self.id_card_input.text().strip()
            district = self.district_input.text().strip()
            street_school = self.street_school_input.text().strip()
            company = self.company_input.text().strip()
            platoon = self.platoon_input.text().strip()
            squad = self.squad_input.text().strip()

            conn = None
            try:
                conn = self.db_manager.get_connection()
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                sql = """
                    SELECT id_card, name, gender, nation, political_status,
                           district, street, company, platoon, squad,
                           squad_leader, camp_status, leave_time, situation_note,
                           school, education_level, major, study_status, study_type,
                           phone, household_address, residence_address, family_info,
                           parent_phone, personal_experience,
                           reference_person, reference_phone, id
                    FROM youth WHERE 1=1
                """
                params = []

                if name:
                    sql += " AND name LIKE ?"
                    params.append(f"%{name}%")

                if id_card:
                    sql += " AND id_card LIKE ?"
                    params.append(f"%{id_card}%")

                if district:
                    sql += " AND district LIKE ?"
                    params.append(f"%{district}%")

                if street_school:
                    sql += " AND (street LIKE ? OR school LIKE ?)"
                    params.append(f"%{street_school}%")
                    params.append(f"%{street_school}%")

                if company:
                    sql += " AND company LIKE ?"
                    params.append(f"%{company}%")

                if platoon:
                    sql += " AND platoon LIKE ?"
                    params.append(f"%{platoon}%")

                if squad:
                    sql += " AND squad LIKE ?"
                    params.append(f"%{squad}%")

                cursor.execute(sql, params)
                results = cursor.fetchall()
                
                self.search_table.setColumnCount(len(self.basic_info_headers))
                self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
                self.display_detailed_results(results)
                    
            finally:
                if conn:
                    conn.close()
                
        except Exception as e:
            QMessageBox.critical(self, "搜索错误", f"搜索时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'select_all_checkbox'):
                    self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            except:
                pass
            self.reset_search()
    
    def reset_search(self):
        """重置搜索"""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.select_all_checkbox.stateChanged.disconnect()
                self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            
            self.name_input.clear()
            self.id_card_input.clear()
            self.district_input.clear()
            self.street_school_input.clear()
            self.company_input.clear()
            self.platoon_input.clear()
            self.squad_input.clear()

            self.search_table.clearContents()
            self.search_table.setRowCount(0)
            
            for row in range(self.search_table.rowCount()):
                for col in range(self.search_table.columnCount()):
                    widget = self.search_table.cellWidget(row, col)
                    if widget:
                        self.search_table.removeCellWidget(row, col)
            
            self.search_table.setColumnCount(len(self.basic_info_headers))
            self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
            
            self.load_all_youth_detailed()
            
        except Exception as e:
            QMessageBox.warning(self, "重置错误", f"重置搜索时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'select_all_checkbox'):
                    self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            except:
                pass

    
    def display_detailed_results(self, results):
        """显示详细结果（完整版�? 包含选择框、操作按钮、删除按钮（旧结构）"""
        self.search_table.setRowCount(len(results))
        
        # 保存当前结果数据供双击事件使�?
        self.current_results = results

        for row, data in enumerate(results):
            # 第一列：添加复选框
            checkbox = QCheckBox()
            checkbox.setStyleSheet("""
                    QCheckBox {
                        spacing: 0px;
                    }
                    QCheckBox::indicator {
                        width: 20px;
                        height: 20px;
                        border: 2px solid #7F8C8D;
                        border-radius: 2px;
                        background-color: white;
                    }
                    QCheckBox::indicator:hover {
                        border-color: #3498DB;
                        background-color: #ECF0F1;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #3498DB;
                        border-color: #2980B9;
                    }
                """)
            checkbox.setMinimumSize(24, 24)  # 确保复选框有足够空间显�?
            
            # 添加状态变化监听，当单个复选框状态改变时更新全选按钮状�?
            checkbox.stateChanged.connect(self.update_select_all_state)
            
            checkbox_layout = QHBoxLayout()
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.setContentsMargins(5, 5, 5, 5)
            checkbox_layout.addWidget(checkbox)
            checkbox_container = QWidget()
            checkbox_container.setLayout(checkbox_layout)
            self.search_table.setCellWidget(row, 0, checkbox_container)
            
            # 数据库返回的字段顺序�?
            # 0:id_card, 1:name, 2:gender, 3:nation, 4:political_status, 5:district, 6:street, 7:company, 8:platoon, 9:squad,
            # 10:squad_leader, 11:camp_status, 12:leave_time, 13:situation_note,
            # 14:school, 15:education_level, 16:major, 17:study_status, 18:study_type, 19:phone, 20:household_address, 21:residence_address,
            # 22:family_info, 23:parent_phone, 24:personal_experience, 25:reference_person, 26:reference_phone, 27:id
            
            # 表头顺序（从�?列开始）�?
            # 姓名 性别 公民身份号码 民族 政治面貌 区划 镇街 连 排 班 带训班长信息 在营状态 离营时间 情况说明 毕业或所在院校 文化程度 所学专业 学业情况 学习类型
            # 电话 户籍地 常住地 家庭情况 父母电话 个人经历 证明人 电话
            
            # 映射关系：表头列 -> 数据索引
            column_mapping = [
                1,   # 姓名 -> name
                2,   # 性别 -> gender
                0,   # 公民身份号码 -> id_card
                3,   # 民族 -> nation
                4,   # 政治面貌 -> political_status
                5,   # 区划 -> district
                6,   # 镇街 -> street
                7,   # 连 -> company
                8,   # 排 -> platoon
                9,   # 班 -> squad
                10,  # 带训班长信息 -> squad_leader
                11,  # 在营状态 -> camp_status
                12,  # 离营时间 -> leave_time
                13,  # 情况说明 -> situation_note
                14,  # 毕业或所在院�?-> school
                15,  # 文化程度 -> education_level
                16,  # 所学专�?-> major
                17,  # 学业情况 -> study_status
                18,  # 学习类型 -> study_type
                19,  # 电话 -> phone
                20,  # 户籍�?-> household_address
                21,  # 常住�?-> residence_address
                22,  # 家庭情况 -> family_info
                23,  # 父母电话 -> parent_phone
                24,  # 个人经历 -> personal_experience
                25,  # 证明�?-> reference_person
                26   # 电话(证明人电�? -> reference_phone
            ]
            
            # 显示数据字段（从�?列到�?1列）
            for col_idx, data_idx in enumerate(column_mapping, start=1):
                if hasattr(data, '__getitem__') and not hasattr(data, 'keys'):
                    value = data[data_idx] if data_idx < len(data) else ''
                else:
                    field_names = ['id_card', 'name', 'gender', 'nation', 'political_status',
                                  'district', 'street', 'company', 'platoon', 'squad',
                                  'squad_leader', 'camp_status', 'leave_time', 'situation_note',
                                  'school', 'education_level', 'major', 'study_status', 'study_type',
                                  'phone', 'household_address', 'residence_address', 'family_info',
                                  'parent_phone', 'personal_experience',
                                  'reference_person', 'reference_phone', 'id']
                    value = data[field_names[data_idx]] if data_idx < len(field_names) else ''
                
                display_text = str(value or '').replace('\n', ' ').replace('\r', ' ')
                if len(display_text) > 100:
                    display_text = display_text[:100] + '...'
                
                item = QTableWidgetItem(display_text)
                item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
                self.search_table.setItem(row, col_idx, item)

            # 倒数第二列：查看详情按钮
            if hasattr(data, '__getitem__') and not hasattr(data, 'keys'):
                id_card = data[0]  # 第一个字段是id_card
            else:
                id_card = data['id_card']
            
            view_btn = QPushButton('查看详情')
            view_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝�?
            view_btn.clicked.connect(lambda checked, idc=id_card: self.view_details(idc))
            view_btn_layout = QHBoxLayout()
            view_btn_layout.setAlignment(Qt.AlignCenter)
            view_btn_layout.setContentsMargins(5, 5, 5, 5)
            view_btn_layout.addWidget(view_btn)
            view_btn_container = QWidget()
            view_btn_container.setLayout(view_btn_layout)
            self.search_table.setCellWidget(row, len(self.basic_info_headers) - 2, view_btn_container)

            # 最后一列：删除按钮
            delete_btn = QPushButton('删除')
            delete_btn.setStyleSheet('background-color: #FFB6C1;')  # 淡红�?
            delete_btn.clicked.connect(lambda checked, idc=id_card: self.delete_single_youth(idc))
            delete_btn_layout = QHBoxLayout()
            delete_btn_layout.setAlignment(Qt.AlignCenter)
            delete_btn_layout.setContentsMargins(5, 5, 5, 5)
            delete_btn_layout.addWidget(delete_btn)
            delete_btn_container = QWidget()
            delete_btn_container.setLayout(delete_btn_layout)
            self.search_table.setCellWidget(row, len(self.basic_info_headers) - 1, delete_btn_container)

        # 根据内容调整列宽，让每列宽度适配该列中最长的内容
        # 使用 QTimer.singleShot 确保在表格完全渲染后再调整列�?
        from PyQt5.QtCore import QTimer
        QTimer.singleShot(0, self.search_table.resizeColumnsToContents)
    
    def on_basic_info_double_click(self, row, column):
        """基本信息表格双击事件处理"""
        try:
            # 从原始数据中获取身份证号，而不是从表格单元�?
            # 需要先获取当前显示的数�?
            if hasattr(self, 'current_results') and row < len(self.current_results):
                data = self.current_results[row]
                if hasattr(data, '__getitem__') and not hasattr(data, 'keys'):
                    id_card = data[0]  # 第一个字段是id_card
                else:
                    id_card = data['id_card']
                # 调用查看详情功能
                self.view_details(id_card)
            else:
                QMessageBox.warning(self, '错误', '无法获取行数�?)
        except Exception as e:
            QMessageBox.warning(self, '错误', f'双击查看详情时发生错�? {str(e)}')
    
    def view_details(self, id_card):
        """查看青年详情"""
        try:
            from ui.youth_detail_dialog import YouthDetailDialog
            dialog = YouthDetailDialog(self.db_manager, self.export_service, id_card, self)
            dialog.data_updated.connect(self.refresh_table_data)
            dialog.exec_()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开详情窗口时发生错�? {str(e)}')
    
    def refresh_table_data(self):
        """刷新表格数据"""
        try:
            self.search_table.clearContents()
            self.search_table.setRowCount(0)
            
            for row in range(self.search_table.rowCount()):
                for col in range(self.search_table.columnCount()):
                    widget = self.search_table.cellWidget(row, col)
                    if widget:
                        self.search_table.removeCellWidget(row, col)
            
            self.search_table.setColumnCount(len(self.basic_info_headers))
            self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
            
            self.load_all_youth_detailed()
            
        except Exception as e:
            print(f"刷新表格数据时发生错�? {str(e)}")
    
    def add_youth(self):
        """添加青年信息"""
        try:
            from ui.add_youth_dialog import AddYouthDialog
            dialog = AddYouthDialog(self.db_manager, self)
            
            # 连接信号，添加成功后刷新表格
            dialog.data_updated.connect(self.load_all_youth_detailed)
            
            if dialog.exec_():
                self.load_all_youth_detailed()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开添加青年信息窗口时发生错�? {str(e)}')
    
    def delete_single_youth(self, id_card):
        """删除单个青年信息"""
        try:
            reply = QMessageBox.question(
                self, 
                '确认删除', 
                '确定要删除这条青年信息吗？\n删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                conn = None
                try:
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM youth WHERE id_card = ?", (id_card,))
                    conn.commit()
                    
                    QMessageBox.information(self, '删除成功', '青年信息已删�?)
                    self.load_all_youth_detailed()
                    
                finally:
                    if conn:
                        conn.close()
                        
        except Exception as e:
            QMessageBox.critical(self, '删除失败', f'删除青年信息时发生错误：{str(e)}')
    
    def toggle_select_all(self, state):
        """全�?取消全�?""
        is_checked = (state == Qt.CheckState.Checked.value)
        
        # 临时标记，防止在全选操作时触发单个复选框的状态更�?
        self._updating_select_all = True
        
        for row in range(self.search_table.rowCount()):
            checkbox_widget = self.search_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(is_checked)
        
        # 重置标记
        self._updating_select_all = False
    
    def update_select_all_state(self):
        """更新基本信息全选按钮状�?""
        # 如果正在执行全选操作，则不更新全选按钮状�?
        if getattr(self, '_updating_select_all', False):
            return
        
        try:
            total_rows = self.search_table.rowCount()
            if total_rows == 0:
                return
            
            checked_count = 0
            for row in range(total_rows):
                checkbox_widget = self.search_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        checked_count += 1
            
            # 临时断开全选按钮的信号连接，避免触发toggle_select_all
            self.select_all_checkbox.stateChanged.disconnect()
            
            # 更新全选按钮状�?
            if checked_count == 0:
                # 没有选中任何�?
                self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            elif checked_count == total_rows:
                # 全部选中
                self.select_all_checkbox.setCheckState(Qt.CheckState.Checked)
            else:
                # 部分选中
                self.select_all_checkbox.setCheckState(Qt.CheckState.PartiallyChecked)
            
            # 重新连接信号
            self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            
        except Exception as e:
            print(f"更新基本信息全选按钮状态时出错: {e}")
            # 确保信号重新连接
            try:
                self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            except:
                pass
    
    def batch_delete_youth(self):
        """批量删除青年信息（旧结构�?""
        try:
            selected_id_cards = []
            selected_names = []
            selected_rows_data = []  # 保存原始数据用于调试
            
            # 先从当前显示的数据中获取
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id_card, name, gender, nation, political_status,
                       school, education_level, major, study_status, study_type,
                       phone, household_address, residence_address, family_info,
                       district, street, parent_phone, personal_experience,
                       reference_person, reference_phone, id
                FROM youth ORDER BY id
            """)
            all_results = cursor.fetchall()
            conn.close()
            
            for row in range(self.search_table.rowCount()):
                checkbox_widget = self.search_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        # 从原始数据中获取id_card（第一个字段）
                        if row < len(all_results):
                            data = all_results[row]
                            id_card = data['id_card'] if hasattr(data, 'keys') else data[0]
                            name = data['name'] if hasattr(data, 'keys') else data[1]
                            
                            if id_card:
                                selected_id_cards.append(id_card)
                                selected_names.append(name)
            
            if not selected_id_cards:
                QMessageBox.information(self, '提示', '请先选择要删除的青年信息')
                return
            
            names_text = '�?.join(selected_names[:5])
            if len(selected_names) > 5:
                names_text += f' 等{len(selected_names)}�?
            
            reply = QMessageBox.question(
                self,
                '确认批量删除',
                f'确定要删除以下青年信息吗？\n{names_text}\n\n共{len(selected_id_cards)}条记录，删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                conn = None
                try:
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    
                    # 逐个删除以确保成�?
                    deleted_count = 0
                    for id_card in selected_id_cards:
                        cursor.execute("DELETE FROM youth WHERE id_card = ?", (id_card,))
                        if cursor.rowcount > 0:
                            deleted_count += 1
                    
                    conn.commit()
                    
                    QMessageBox.information(self, '删除成功', f'成功删除 {deleted_count} 条青年信�?)
                    
                    # 取消全�?
                    if hasattr(self, 'select_all_checkbox'):
                        self.select_all_checkbox.setChecked(False)
                    
                    self.load_all_youth_detailed()
                    
                finally:
                    if conn:
                        conn.close()
                        
        except Exception as e:
            QMessageBox.critical(self, '批量删除失败', f'批量删除青年信息时发生错误：{str(e)}')
    
    def load_module_data(self, module_name, youth_data):
        """加载模块数据"""
        table = self.tab_tables[module_name]
        
        if module_name == '基本信息' and youth_data:
            table.setColumnCount(2)
            table.setHorizontalHeaderLabels(['字段', '�?])
            fields = [
                ('姓名', youth_data[1]),
                ('身份证号', youth_data[2]),
                ('区划', youth_data[3]),
                ('镇街', youth_data[4]),
                ('性别', youth_data[5]),
                ('出生日期', youth_data[6]),
                ('学历', youth_data[7]),
                ('政治面貌', youth_data[8]),
                ('电话', youth_data[9]),
                ('地址', youth_data[10])
            ]
            table.setRowCount(len(fields))
            for row, (field, value) in enumerate(fields):
                table.setItem(row, 0, QTableWidgetItem(field))
                table.setItem(row, 1, QTableWidgetItem(str(value or '')))
        else:
            # 加载其他模块数据
            pass
    
    def import_data(self, tab_name):
        """导入数据"""
        file_path, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '', 'Excel Files (*.xlsx *.xls)')
        if not file_path:
            return
        
        if tab_name == '基本信息':
            count, error = self.import_service.import_youth_from_excel(file_path)
            if error:
                # 如果有部分成功导入，显示详细信息
                if count > 0:
                    QMessageBox.warning(self, '导入完成（有错误�?, error)
                    # 刷新显示
                    self.search_table.setColumnCount(len(self.basic_info_headers))
                    self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
                    self.load_all_youth_detailed()
                else:
                    QMessageBox.warning(self, '导入失败', error)
            else:
                QMessageBox.information(self, '导入成功', f'成功导入 {count} 条记�?)
                # 导入成功后刷新显示详细信�?
                self.search_table.setColumnCount(len(self.basic_info_headers))
                self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
                self.load_all_youth_detailed()
        elif tab_name == '异常情况统计':
            count, error = self.import_service.import_abnormal_stats(file_path)
            if error:
                QMessageBox.warning(self, '导入失败', error)
            else:
                QMessageBox.information(self, '导入成功', f'成功导入 {count} 条记�?)
        elif tab_name == '隐性疾病及心理问题筛查情况':
            count, error = self.import_service.import_health_screening(file_path)
            if error:
                QMessageBox.warning(self, '导入失败', error)
            else:
                QMessageBox.information(self, '导入成功', f'成功导入 {count} 条记�?)
    
    def upload_document(self, tab_name):
        """上传扫描文档"""
        if not self.current_youth_id:
            QMessageBox.warning(self, '提示', '请先选择一个青�?)
            return
        
        file_path, _ = QFileDialog.getOpenFileName(self, '选择文档', '', 'All Files (*.*)')
        if file_path:
            QMessageBox.information(self, '成功', '文档上传成功')
    
    def add_daily_info(self, tab_name):
        """添加每日情况统计信息"""
        try:
            from ui.daily_record_dialog import DailyRecordDialog
            dialog = DailyRecordDialog(self.db_manager, self)
            dialog.data_updated.connect(self.load_daily_stats_data)
            if dialog.exec_():
                # 刷新每日情况统计表格数据
                self.load_daily_stats_data()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开添加每日记录窗口时发生错�? {str(e)}')
    
    def batch_add_daily_info(self):
        """批量添加每日情况统计信息"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QDateEdit, QComboBox, QTextEdit
            from PyQt5.QtCore import QDate
            
            # 创建对话�?
            dialog = QDialog(self)
            dialog.setWindowTitle('批量添加每日情况记录')
            dialog.setFixedSize(500, 350)
            dialog.setModal(True)
            
            layout = QVBoxLayout()
            layout.setSpacing(15)
            
            # 标题
            title = QLabel('为所有基本信息中的人员批量添加每日情况记�?)
            title.setStyleSheet('font-size: 14px; font-weight: bold; padding: 5px;')
            layout.addWidget(title)
            
            # 表单
            from PyQt5.QtWidgets import QFormLayout
            form_layout = QFormLayout()
            form_layout.setSpacing(10)
            
            # 日期
            date_input = QDateEdit()
            date_input.setDate(QDate.currentDate())
            date_input.setCalendarPopup(True)
            form_layout.addRow('日期:', date_input)
            
            # 思想
            mood_input = QComboBox()
            mood_input.addItems(['正常', '异常'])
            form_layout.addRow('思想:', mood_input)
            
            # 身体
            physical_input = QComboBox()
            physical_input.addItems(['正常', '异常'])
            form_layout.addRow('身体:', physical_input)
            
            # 精神
            mental_input = QComboBox()
            mental_input.addItems(['正常', '异常'])
            form_layout.addRow('精神:', mental_input)
            
            # 其他
            notes_input = QLineEdit()
            notes_input.setPlaceholderText('请输入其他备注信息（可选）')
            form_layout.addRow('其他:', notes_input)
            
            layout.addLayout(form_layout)
            
            # 提示信息
            info_label = QLabel('将为基本信息表中的所有人员添加记录。\n如果某人在该日期已有记录，将询问是否覆盖�?)
            info_label.setStyleSheet('color: #7F8C8D; font-size: 12px; padding: 5px;')
            layout.addWidget(info_label)
            
            # 按钮
            button_layout = QHBoxLayout()
            
            ok_btn = QPushButton('确定')
            ok_btn.setStyleSheet("""
                QPushButton {
                    background-color: #27AE60;
                    color: white;
                    border: none;
                    padding: 8px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
            """)
            
            def on_ok_clicked():
                try:
                    date = date_input.date().toString('yyyy-MM-dd')
                    mood = mood_input.currentText()
                    physical = physical_input.currentText()
                    mental = mental_input.currentText()
                    notes = notes_input.text().strip()
                    
                    # 获取所有基本信息中的人�?
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    cursor.execute("SELECT id, name, id_card FROM youth ORDER BY id")
                    all_youth = cursor.fetchall()
                    
                    if not all_youth:
                        conn.close()
                        QMessageBox.warning(dialog, '提示', '基本信息表中没有人员数据')
                        return
                    
                    # 检查每个人员是否已有该日期的记�?
                    duplicates = []
                    new_records = []
                    
                    for youth in all_youth:
                        youth_id = youth[0]
                        youth_name = youth[1]
                        youth_id_card = youth[2]
                        
                        cursor.execute("""
                            SELECT id FROM daily_stat 
                            WHERE youth_id = ? AND record_date = ?
                        """, (youth_id, date))
                        existing = cursor.fetchone()
                        
                        if existing:
                            duplicates.append((existing[0], youth_id, youth_name, youth_id_card))
                        else:
                            new_records.append((youth_id, youth_name, youth_id_card))
                    
                    conn.close()
                    
                    # 如果有重复记录，询问是否覆盖
                    if duplicates:
                        duplicate_names = [f"{d[2]}({d[3]})" for d in duplicates[:5]]
                        names_text = '�?.join(duplicate_names)
                        if len(duplicates) > 5:
                            names_text += f' 等{len(duplicates)}�?
                        
                        reply = QMessageBox.question(
                            dialog,
                            '发现重复记录',
                            f'以下人员�?{date} 已有记录：\n{names_text}\n\n共{len(duplicates)}条重复记录。\n\n是否覆盖这些记录？\n（选择"�?将覆盖重复记录并添加新记录，选择"�?将只添加新记录）',
                            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
                            QMessageBox.StandardButton.No
                        )
                        
                        if reply == QMessageBox.StandardButton.Cancel:
                            return
                        
                        should_update = (reply == QMessageBox.StandardButton.Yes)
                    else:
                        should_update = False
                    
                    # 执行批量添加/更新
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    
                    added_count = 0
                    updated_count = 0
                    
                    # 添加新记�?
                    for youth_id, youth_name, youth_id_card in new_records:
                        # 使用新的插入方法
                        self.db_manager.insert_daily_stat(
                            youth_id, date, mood, physical, mental, notes
                        )
                        added_count += 1
                    
                    # 更新重复记录（如果用户选择覆盖�?
                    if should_update and duplicates:
                        for existing_id, youth_id, youth_name, youth_id_card in duplicates:
                            # 使用新的更新方法
                            self.db_manager.update_daily_stat(
                                existing_id, date, mood, physical, mental, notes
                            )
                            updated_count += 1
                    
                    conn.commit()
                    conn.close()
                    
                    # 显示结果
                    result_msg = f'批量添加完成！\n\n'
                    if added_count > 0:
                        result_msg += f'新增记录：{added_count} 条\n'
                    if updated_count > 0:
                        result_msg += f'更新记录：{updated_count} 条\n'
                    if duplicates and not should_update:
                        result_msg += f'跳过重复：{len(duplicates)} �?
                    
                    QMessageBox.information(dialog, '成功', result_msg)
                    
                    # 刷新数据
                    self.load_daily_stats_data()
                    
                    # 关闭对话�?
                    dialog.accept()
                    
                except Exception as e:
                    QMessageBox.warning(dialog, '错误', f'批量添加失败: {str(e)}')
            
            ok_btn.clicked.connect(on_ok_clicked)
            
            cancel_btn = QPushButton('取消')
            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #95A5A6;
                    color: white;
                    border: none;
                    padding: 8px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #7F8C8D;
                }
            """)
            cancel_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(ok_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开批量添加窗口时发生错�? {str(e)}')
    
    def filter_daily_stats(self):
        """按日期筛选每日情况统�?""
        try:
            if '每日情况统计' not in self.tab_tables:
                return
                
            table = self.tab_tables['每日情况统计']
            
            # 获取日期范围
            start_date = self.start_date_input.date().toString('yyyy-MM-dd')
            end_date = self.end_date_input.date().toString('yyyy-MM-dd')
            
            # 设置表头 - 与load_daily_stats_data保持一�?
            headers = ['公民身份证号�?, '姓名', '日期', '思想', '身体', '精神', '训练', '管理', '其他', '修改', '删除']
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            
            # 获取筛选后的数�?- 按添加顺序排序（先添加的在前面），只获取需要显示的字段
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.id, d.record_date, y.name, y.id_card, 
                       d.mood, d.physical_condition, d.mental_state, d.training, d.management, d.notes, d.youth_id
                FROM daily_stat d
                JOIN youth y ON d.youth_id = y.id
                WHERE d.record_date >= ? AND d.record_date <= ?
                ORDER BY d.id ASC
            ''', (start_date, end_date))
            results = cursor.fetchall()
            conn.close()
            
            # 显示数据
            table.setRowCount(len(results))
            for row, data in enumerate(results):
                # 数据库返回字段顺序：
                # 0:d.id, 1:d.record_date, 2:y.name, 3:y.id_card, 
                # 4:d.mood, 5:d.physical_condition, 6:d.mental_state, 7:d.training, 8:d.management, 9:d.notes, 10:d.youth_id
                
                # 第一列：公民身份证号�?
                table.setItem(row, 0, QTableWidgetItem(str(data[3] or '')))
                
                # 第二列：姓名
                table.setItem(row, 1, QTableWidgetItem(str(data[2] or '')))
                
                # 第三列：日期
                date_value = data[1]
                if date_value:
                    try:
                        from datetime import datetime
                        if isinstance(date_value, str):
                            date_obj = datetime.strptime(date_value, '%Y-%m-%d')
                            formatted_date = date_obj.strftime('%Y-%m-%d')
                        else:
                            formatted_date = str(date_value)
                        table.setItem(row, 2, QTableWidgetItem(formatted_date))
                    except:
                        table.setItem(row, 2, QTableWidgetItem(str(date_value or '')))
                else:
                    table.setItem(row, 2, QTableWidgetItem(''))
                
                # 第四到九列：状态字段，添加红色高亮显示异常状�?
                status_fields = [data[4], data[5], data[6], data[7], data[8], data[9]]  # mood, physical, mental, training, management, notes
                for col, value in enumerate(status_fields, 3):
                    item = QTableWidgetItem(str(value or ''))
                    # 如果状态为"异常"，设置红色文�?
                    if col < 8 and str(value) == '异常':  # �?列是状态字段（思想、身体、精神、训练、管理）
                        from PyQt5.QtGui import QColor
                        item.setForeground(QColor('red'))
                        font = item.font()
                        font.setBold(True)
                        item.setFont(font)
                    table.setItem(row, col, item)
                
                # 第十列：修改按钮
                edit_btn = QPushButton('修改')
                edit_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝�?
                edit_btn.clicked.connect(lambda checked, record_id=data[0], youth_id=data[10]: self.edit_daily_record(record_id, youth_id))
                edit_btn_layout = QHBoxLayout()
                edit_btn_layout.setAlignment(Qt.AlignCenter)
                edit_btn_layout.setContentsMargins(5, 5, 5, 5)
                edit_btn_layout.addWidget(edit_btn)
                edit_btn_container = QWidget()
                edit_btn_container.setLayout(edit_btn_layout)
                table.setCellWidget(row, 9, edit_btn_container)
                
                # 第十一列：删除按钮
                del_btn = QPushButton('删除')
                del_btn.setStyleSheet('background-color: #FFB6C1;')  # 淡红�?
                del_btn.clicked.connect(lambda checked, record_id=data[0]: self.delete_daily_record(record_id))
                del_btn_layout = QHBoxLayout()
                del_btn_layout.setAlignment(Qt.AlignCenter)
                del_btn_layout.setContentsMargins(5, 5, 5, 5)
                del_btn_layout.addWidget(del_btn)
                del_btn_container = QWidget()
                del_btn_container.setLayout(del_btn_layout)
                table.setCellWidget(row, 10, del_btn_container)
            
            # 设置表格自适应屏幕 - 与load_daily_stats_data保持一�?
            header = table.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Interactive)  # 允许手动调整
            header.setSectionResizeMode(0, QHeaderView.Stretch)  # 公民身份证号码列
            header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 姓名�?
            header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 日期�?
            header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 思想�?
            header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 身体�?
            header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 精神�?
            header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 训练�?
            header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # 管理�?
            header.setSectionResizeMode(8, QHeaderView.Stretch)  # 其他�?
            header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # 修改�?
            header.setSectionResizeMode(10, QHeaderView.ResizeToContents)  # 删除�?
            header.setStretchLastSection(False)
            
            QMessageBox.information(self, '筛选完�?, f'找到 {len(results)} 条记�?)
            
        except Exception as e:
            QMessageBox.warning(self, "筛选错�?, f"筛选每日情况统计数据时发生错误：{str(e)}")
    
    def reset_daily_stats_filter(self):
        """重置每日情况统计筛�?""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'daily_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.daily_select_all_checkbox.stateChanged.disconnect()
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            
            # 重置日期范围
            self.start_date_input.setDate(QDate.currentDate().addDays(-30))
            self.end_date_input.setDate(QDate.currentDate())
            
            # 重新加载所有数�?
            self.load_daily_stats_data()
            
        except Exception as e:
            QMessageBox.warning(self, "重置错误", f"重置筛选时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'daily_select_all_checkbox'):
                    self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            except:
                pass
    
    def search_daily_stats(self):
        """搜索每日情况统计"""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'daily_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.daily_select_all_checkbox.stateChanged.disconnect()
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            
            name = self.daily_name_input.text().strip()
            id_card = self.daily_id_card_input.text().strip()
            
            if '每日情况统计' not in self.tab_tables:
                return
            
            table = self.tab_tables['每日情况统计']
            
            # 设置表头 - 与load_daily_stats_data保持一�?
            headers = ['公民身份证号�?, '姓名', '日期', '思想', '身体', '精神', '训练', '管理', '其他', '修改', '删除']
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            
            sql = '''
                SELECT d.id, d.record_date, y.name, y.id_card, 
                       d.mood, d.physical_condition, d.mental_state, d.training, d.management, d.notes, d.youth_id
                FROM daily_stat d
                JOIN youth y ON d.youth_id = y.id
                WHERE 1=1
            '''
            params = []
            
            if name:
                sql += " AND y.name LIKE ?"
                params.append(f"%{name}%")
            
            if id_card:
                sql += " AND y.id_card LIKE ?"
                params.append(f"%{id_card}%")
            
            sql += " ORDER BY d.id ASC"
            
            cursor.execute(sql, params)
            results = cursor.fetchall()
            conn.close()
            
            # 显示搜索结果
            table.setRowCount(len(results))
            for row, data in enumerate(results):
                # 数据库返回字段顺序：
                # 0:d.id, 1:d.record_date, 2:y.name, 3:y.id_card, 
                # 4:d.mood, 5:d.physical_condition, 6:d.mental_state, 7:d.training, 8:d.management, 9:d.notes, 10:d.youth_id
                
                # 第一列：公民身份证号�?
                table.setItem(row, 0, QTableWidgetItem(str(data[3] or '')))
                
                # 第二列：姓名
                table.setItem(row, 1, QTableWidgetItem(str(data[2] or '')))
                
                # 第三列：日期
                date_value = data[1]
                if date_value:
                    try:
                        from datetime import datetime
                        if isinstance(date_value, str):
                            date_obj = datetime.strptime(date_value, '%Y-%m-%d')
                            formatted_date = date_obj.strftime('%Y-%m-%d')
                        else:
                            formatted_date = str(date_value)
                        table.setItem(row, 2, QTableWidgetItem(formatted_date))
                    except:
                        table.setItem(row, 2, QTableWidgetItem(str(date_value or '')))
                else:
                    table.setItem(row, 2, QTableWidgetItem(''))
                
                # 第四到九列：状态字段，添加红色高亮显示异常状�?
                status_fields = [data[4], data[5], data[6], data[7], data[8], data[9]]  # mood, physical, mental, training, management, notes
                for col, value in enumerate(status_fields, 3):
                    item = QTableWidgetItem(str(value or ''))
                    # 如果状态为"异常"，设置红色文�?
                    if col < 8 and str(value) == '异常':  # �?列是状态字段（思想、身体、精神、训练、管理）
                        from PyQt5.QtGui import QColor
                        item.setForeground(QColor('red'))
                        font = item.font()
                        font.setBold(True)
                        item.setFont(font)
                    table.setItem(row, col, item)
                
                # 第十列：修改按钮
                edit_btn = QPushButton('修改')
                edit_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝�?
                edit_btn.clicked.connect(lambda checked, record_id=data[0], youth_id=data[10]: self.edit_daily_record(record_id, youth_id))
                edit_btn_layout = QHBoxLayout()
                edit_btn_layout.setAlignment(Qt.AlignCenter)
                edit_btn_layout.setContentsMargins(5, 5, 5, 5)
                edit_btn_layout.addWidget(edit_btn)
                edit_btn_container = QWidget()
                edit_btn_container.setLayout(edit_btn_layout)
                table.setCellWidget(row, 9, edit_btn_container)
                
                # 第十一列：删除按钮
                del_btn = QPushButton('删除')
                del_btn.setStyleSheet('background-color: #FFB6C1;')  # 淡红�?
                del_btn.clicked.connect(lambda checked, record_id=data[0]: self.delete_daily_record(record_id))
                del_btn_layout = QHBoxLayout()
                del_btn_layout.setAlignment(Qt.AlignCenter)
                del_btn_layout.setContentsMargins(5, 5, 5, 5)
                del_btn_layout.addWidget(del_btn)
                del_btn_container = QWidget()
                del_btn_container.setLayout(del_btn_layout)
                table.setCellWidget(row, 10, del_btn_container)
            
            # 设置表格自适应屏幕 - 与load_daily_stats_data保持一�?
            header = table.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Interactive)  # 允许手动调整
            header.setSectionResizeMode(0, QHeaderView.Stretch)  # 公民身份证号码列
            header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 姓名�?
            header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 日期�?
            header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 思想�?
            header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 身体�?
            header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 精神�?
            header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 训练�?
            header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # 管理�?
            header.setSectionResizeMode(8, QHeaderView.Stretch)  # 其他�?
            header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # 修改�?
            header.setSectionResizeMode(10, QHeaderView.ResizeToContents)  # 删除�?
            header.setStretchLastSection(False)
            
        except Exception as e:
            QMessageBox.warning(self, "搜索错误", f"搜索每日情况统计时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'daily_select_all_checkbox'):
                    self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            except:
                pass
    
    def reset_daily_stats_search(self):
        """重置每日情况统计搜索"""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'daily_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.daily_select_all_checkbox.stateChanged.disconnect()
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            
            self.daily_name_input.clear()
            self.daily_id_card_input.clear()
            self.load_daily_stats_data()
            
        except Exception as e:
            QMessageBox.warning(self, "重置错误", f"重置搜索时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'daily_select_all_checkbox'):
                    self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            except:
                pass
    
    def edit_daily_record(self, record_id, youth_id):
        """修改每日记录"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QDateEdit, QComboBox, QFormLayout
            from PyQt5.QtCore import QDate
            
            # 获取记录数据
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.id, d.record_date, d.mood, d.physical_condition, d.mental_state, d.notes,
                       y.name, y.id_card
                FROM daily_stat d
                JOIN youth y ON d.youth_id = y.id
                WHERE d.id=?
            ''', (record_id,))
            record = cursor.fetchone()
            conn.close()
            
            if not record:
                QMessageBox.warning(self, '错误', '记录不存�?)
                return
            
            dialog = QDialog(self)
            dialog.setWindowTitle('修改每日记录')
            dialog.setMinimumSize(500, 300)  # 设置最小尺�?
            dialog.resize(500, 300)  # 设置初始尺寸，但允许调整
            dialog.setModal(True)
            
            # 允许窗口调整大小，只显示最大化和关闭按�?
            dialog.setWindowFlags(Qt.Dialog | 
                                 Qt.WindowMaximizeButtonHint | 
                                 Qt.WindowCloseButtonHint)
            
            layout = QVBoxLayout()
            layout.setSpacing(10)
            
            # 标题
            title = QLabel(f'修改每日记录 - {record[6]} ({record[7]})')
            title.setStyleSheet('font-size: 14px; font-weight: bold;')
            layout.addWidget(title)
            
            # 表单
            form_layout = QFormLayout()
            form_layout.setSpacing(10)
            
            # 日期
            date_input = QDateEdit()
            date_input.setCalendarPopup(True)
            if record[1]:
                date_parts = record[1].split('-')
                if len(date_parts) == 3:
                    date_input.setDate(QDate(int(date_parts[0]), int(date_parts[1]), int(date_parts[2])))
            form_layout.addRow('日期:', date_input)
            
            # 思想
            mood_input = QComboBox()
            mood_input.addItems(['正常', '异常'])
            mood_input.setCurrentText(record[2] or '正常')
            form_layout.addRow('思想:', mood_input)
            
            # 身体
            physical_input = QComboBox()
            physical_input.addItems(['正常', '异常'])
            physical_input.setCurrentText(record[3] or '正常')
            form_layout.addRow('身体:', physical_input)
            
            # 精神
            mental_input = QComboBox()
            mental_input.addItems(['正常', '异常'])
            mental_input.setCurrentText(record[4] or '正常')
            form_layout.addRow('精神:', mental_input)
            
            # 其他
            notes_input = QLineEdit()
            notes_input.setText(record[5] or '')
            form_layout.addRow('其他:', notes_input)
            
            layout.addLayout(form_layout)
            
            # 按钮
            button_layout = QHBoxLayout()
            
            save_btn = QPushButton('保存')
            save_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3498DB;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #2980B9;
                }
            """)
            
            def save_changes():
                try:
                    # 使用新的更新方法
                    self.db_manager.update_daily_stat(
                        record_id,
                        date_input.date().toString('yyyy-MM-dd'), 
                        mood_input.currentText(),
                        physical_input.currentText(),
                        mental_input.currentText(),
                        notes_input.text().strip()
                    )
                    
                    QMessageBox.information(dialog, '成功', '记录已更�?)
                    self.load_daily_stats_data()
                    dialog.accept()
                except Exception as e:
                    QMessageBox.warning(dialog, '错误', f'更新失败: {str(e)}')
            
            save_btn.clicked.connect(save_changes)
            
            cancel_btn = QPushButton('取消')
            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #95A5A6;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #7F8C8D;
                }
            """)
            cancel_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(save_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.warning(self, '错误', f'修改记录时发生错�? {str(e)}')
    
    def delete_daily_record(self, record_id):
        """删除每日记录"""
        try:
            reply = QMessageBox.question(self, '确认删除', 
                                       '确定要删除这条记录吗�?,
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            
            if reply == QMessageBox.StandardButton.Yes:
                # 使用新的删除方法
                success = self.db_manager.delete_daily_stat(record_id)
                
                if success:
                    QMessageBox.information(self, '成功', '记录已删�?)
                    self.load_daily_stats_data()
                else:
                    QMessageBox.warning(self, '错误', '删除失败')
        except Exception as e:
            QMessageBox.warning(self, '错误', f'删除记录时发生错�? {str(e)}')
    
    def toggle_daily_select_all(self, state):
        """每日情况统计全�?取消全�?""
        if '每日情况统计' not in self.tab_tables:
            return
        
        table = self.tab_tables['每日情况统计']
        is_checked = (state == Qt.CheckState.Checked.value)
        
        # 临时标记，防止在全选操作时触发单个复选框的状态更�?
        self._updating_daily_select_all = True
        
        for row in range(table.rowCount()):
            checkbox_widget = table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(is_checked)
        
        # 重置标记
        self._updating_daily_select_all = False
    
    def update_daily_select_all_state(self):
        """更新每日情况统计全选按钮状�?""
        # 如果正在执行全选操作，则不更新全选按钮状�?
        if getattr(self, '_updating_daily_select_all', False):
            return
        
        try:
            if '每日情况统计' not in self.tab_tables:
                return
            
            table = self.tab_tables['每日情况统计']
            total_rows = table.rowCount()
            if total_rows == 0:
                return
            
            checked_count = 0
            for row in range(total_rows):
                checkbox_widget = table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        checked_count += 1
            
            # 临时断开全选按钮的信号连接，避免触发toggle_daily_select_all
            self.daily_select_all_checkbox.stateChanged.disconnect()
            
            # 更新全选按钮状�?
            if checked_count == 0:
                # 没有选中任何�?
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            elif checked_count == total_rows:
                # 全部选中
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.Checked)
            else:
                # 部分选中
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.PartiallyChecked)
            
            # 重新连接信号
            self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            
        except Exception as e:
            print(f"更新每日统计全选按钮状态时出错: {e}")
            # 确保信号重新连接
            try:
                self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            except:
                pass
    
    def get_selected_daily_stats_data(self):
        """获取每日情况统计中选中的数�?""
        try:
            if '每日情况统计' not in self.tab_tables:
                return []
            
            table = self.tab_tables['每日情况统计']
            selected_data = []
            
            # 获取所有数据用于对�?
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT y.id_card, y.name, d.record_date, d.mood, d.physical_condition, 
                       d.mental_state, d.notes
                FROM daily_stat d
                JOIN youth y ON d.youth_id = y.id
                ORDER BY d.id ASC
            ''')
            all_records = cursor.fetchall()
            conn.close()
            
            # 检查选中的行
            for row in range(table.rowCount()):
                checkbox_widget = table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_records):
                            selected_data.append(all_records[row])
            
            return selected_data
            
        except Exception as e:
            print(f"获取选中每日情况统计数据时出�? {e}")
            return []
    
    def batch_delete_daily_stats(self):
        """批量删除每日情况统计"""
        try:
            if '每日情况统计' not in self.tab_tables:
                return
            
            table = self.tab_tables['每日情况统计']
            
            # 先获取所有记录数�?
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.id, y.name, d.record_date
                FROM daily_stat d
                JOIN youth y ON d.youth_id = y.id
                ORDER BY d.id ASC
            ''')
            all_records = cursor.fetchall()
            conn.close()
            
            # 获取选中的记录ID和姓�?
            selected_ids = []
            selected_names = []
            
            for row in range(table.rowCount()):
                checkbox_widget = table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_records):
                            record = all_records[row]
                            selected_ids.append(record[0])  # record[0] �?id
                            selected_names.append(f"{record[1]}({record[2]})")  # record[1] �?name, record[2] �?date
            
            if not selected_ids:
                QMessageBox.warning(self, "提示", "请选择要删除的记录")
                return
            
            # 确认删除
            names_text = '�?.join(selected_names[:5])
            if len(selected_names) > 5:
                names_text += f' �?
            
            reply = QMessageBox.question(
                self,
                '确认批量删除',
                f'确定要删除以下每日情况记录吗？\n{names_text}\n\n共{len(selected_ids)}条记录，删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # 删除记录
                conn = self.db_manager.get_connection()
                cursor = conn.cursor()
                
                # 使用新的批量删除方法
                deleted_count = self.db_manager.delete_daily_stats(selected_ids)
                
                conn.commit()
                conn.close()
                
                # 先刷新数据（这会重新创建表格和checkbox�?
                self.load_daily_stats_data()
                
                # 显示成功消息
                QMessageBox.information(self, '删除成功', f'成功删除 {deleted_count} 条记�?)
                
        except Exception as e:
            QMessageBox.critical(self, '批量删除失败', f'批量删除每日情况统计时发生错误：{str(e)}')

    def load_daily_stats_data(self):
        """加载每日情况统计数据"""
        try:
            if '每日情况统计' not in self.tab_tables:
                return
                
            table = self.tab_tables['每日情况统计']
            
            # 设置表头 - 简化显示，按照用户要求的格式，添加操作按钮
            headers = ['公民身份证号�?, '姓名', '日期', '思想', '身体', '精神', '训练', '管理', '其他', '修改', '删除']
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            
            # 获取数据 - 按添加顺序排序（先添加的在前面），只获取需要显示的字段
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.id, d.record_date, y.name, y.id_card, 
                       d.mood, d.physical_condition, d.mental_state, d.training, d.management, d.notes, d.youth_id
                FROM daily_stat d
                JOIN youth y ON d.youth_id = y.id
                ORDER BY d.id ASC
            ''')
            results = cursor.fetchall()
            conn.close()
            
            # 显示数据
            table.setRowCount(len(results))
            for row, data in enumerate(results):
                # 数据库返回字段顺序：
                # 0:d.id, 1:d.record_date, 2:y.name, 3:y.id_card, 
                # 4:d.mood, 5:d.physical_condition, 6:d.mental_state, 7:d.training, 8:d.management, 9:d.notes, 10:d.youth_id
                
                # 第一列：公民身份证号�?
                table.setItem(row, 0, QTableWidgetItem(str(data[3] or '')))
                
                # 第二列：姓名
                table.setItem(row, 1, QTableWidgetItem(str(data[2] or '')))
                
                # 第三列：日期
                date_value = data[1]
                if date_value:
                    try:
                        from datetime import datetime
                        if isinstance(date_value, str):
                            date_obj = datetime.strptime(date_value, '%Y-%m-%d')
                            formatted_date = date_obj.strftime('%Y-%m-%d')
                        else:
                            formatted_date = str(date_value)
                        table.setItem(row, 2, QTableWidgetItem(formatted_date))
                    except:
                        table.setItem(row, 2, QTableWidgetItem(str(date_value or '')))
                else:
                    table.setItem(row, 2, QTableWidgetItem(''))
                
                # 第四到九列：状态字段，添加红色高亮显示异常状�?
                status_fields = [data[4], data[5], data[6], data[7], data[8], data[9]]  # mood, physical, mental, training, management, notes
                for col, value in enumerate(status_fields, 3):
                    item = QTableWidgetItem(str(value or ''))
                    # 如果状态为"异常"，设置红色文�?
                    if col < 8 and str(value) == '异常':  # �?列是状态字段（思想、身体、精神、训练、管理）
                        from PyQt5.QtGui import QColor
                        item.setForeground(QColor('red'))
                        font = item.font()
                        font.setBold(True)
                        item.setFont(font)
                    table.setItem(row, col, item)
                
                # 第十列：修改按钮
                edit_btn = QPushButton('修改')
                edit_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝�?
                edit_btn.clicked.connect(lambda checked, record_id=data[0], youth_id=data[10]: self.edit_daily_record(record_id, youth_id))
                edit_btn_layout = QHBoxLayout()
                edit_btn_layout.setAlignment(Qt.AlignCenter)
                edit_btn_layout.setContentsMargins(5, 5, 5, 5)
                edit_btn_layout.addWidget(edit_btn)
                edit_btn_container = QWidget()
                edit_btn_container.setLayout(edit_btn_layout)
                table.setCellWidget(row, 9, edit_btn_container)
                
                # 第十一列：删除按钮
                del_btn = QPushButton('删除')
                del_btn.setStyleSheet('background-color: #FFB6C1;')  # 淡红�?
                del_btn.clicked.connect(lambda checked, record_id=data[0]: self.delete_daily_record(record_id))
                del_btn_layout = QHBoxLayout()
                del_btn_layout.setAlignment(Qt.AlignCenter)
                del_btn_layout.setContentsMargins(5, 5, 5, 5)
                del_btn_layout.addWidget(del_btn)
                del_btn_container = QWidget()
                del_btn_container.setLayout(del_btn_layout)
                table.setCellWidget(row, 10, del_btn_container)
            
            # 设置表格自适应屏幕
            header = table.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Interactive)  # 允许手动调整
            header.setSectionResizeMode(0, QHeaderView.Stretch)  # 公民身份证号码列
            header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 姓名�?
            header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 日期�?
            header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 思想�?
            header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 身体�?
            header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 精神�?
            header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 训练�?
            header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # 管理�?
            header.setSectionResizeMode(8, QHeaderView.Stretch)  # 其他�?
            header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # 修改�?
            header.setSectionResizeMode(10, QHeaderView.ResizeToContents)  # 删除�?
            header.setStretchLastSection(False)
            header.setStretchLastSection(False)
            
        except Exception as e:
            QMessageBox.warning(self, "加载错误", f"加载每日情况统计数据时发生错误：{str(e)}")

    def add_daily_record(self):
        """添加每日记录"""
        QMessageBox.information(self, '提示', '请先在搜索结果中选择一个青年，然后点击"查看详情"进入详情页面添加每日记录')

    def export_data(self, tab_name):
        """导出数据"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            f'导出{tab_name}', 
            f'{tab_name}_{datetime.now().strftime("%Y%m%d")}.xlsx',
            'Excel Files (*.xlsx)'
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = tab_name
            
            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            if tab_name == '基本信息':
                # 导出基本信息（不包含序号�?
                headers = [
                    '姓名', '性别', '公民身份号码', '民族', '政治面貌',
                    '区划', '镇街', '连', '排', '班', '带训班长信息', '在营状态', '离营时间', '离营情况说明',
                    '毕业或所在院校', '文化程度', '所学专业', '学业情况', '学习类型',
                    '电话', '户籍地', '常住地', '家庭情况',
                    '父母电话', '个人经历', '证明人', '电话'
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # 获取选中的数据或全部数据
                selected_data = self.get_selected_basic_info_data()
                if not selected_data:
                    # 没有选择时导出全部数�?
                    results = self.db_manager.search_youth()
                else:
                    # 有选择时导出选中数据
                    results = selected_data
                
                for row_idx, youth in enumerate(results, 2):
                    # 数据库返回字段顺序：
                    # 0:id_card, 1:name, 2:gender, 3:nation, 4:political_status, 5:district, 6:street, 7:company, 8:platoon, 9:squad,
                    # 10:squad_leader, 11:camp_status, 12:leave_time, 13:situation_note,
                    # 14:school, 15:education_level, 16:major, 17:study_status, 18:study_type, 19:phone, 20:household_address, 21:residence_address,
                    # 22:family_info, 23:parent_phone, 24:personal_experience, 25:reference_person, 26:reference_phone, 27:id
                    
                    # 表头顺序（不含序号）：姓名, 性别, 公民身份号码, 民族, 政治面貌, 区划, 镇街, 连, 排, 班, 带训班长信息, 在营状态, 离营时间, 情况说明, 毕业或所在院校, 文化程度, 所学专业, 学业情况, 学习类型, 电话, 户籍地, 常住地, 家庭情况, 父母电话, 个人经历, 证明人, 电话
                    
                    # 映射关系：表头列 -> 数据索引
                    column_mapping = [
                        1,   # 姓名 -> name
                        2,   # 性别 -> gender  
                        0,   # 公民身份号码 -> id_card
                        3,   # 民族 -> nation
                        4,   # 政治面貌 -> political_status
                        5,   # 区划 -> district
                        6,   # 镇街 -> street
                        7,   # 连 -> company
                        8,   # 排 -> platoon
                        9,   # 班 -> squad
                        10,  # 带训班长信息 -> squad_leader
                        11,  # 在营状态 -> camp_status
                        12,  # 离营时间 -> leave_time
                        13,  # 情况说明 -> situation_note
                        14,  # 毕业或所在院�?-> school
                        15,  # 文化程度 -> education_level
                        16,  # 所学专�?-> major
                        17,  # 学业情况 -> study_status
                        18,  # 学习类型 -> study_type
                        19,  # 电话 -> phone
                        20,  # 户籍�?-> household_address
                        21,  # 常住�?-> residence_address
                        22,  # 家庭情况 -> family_info
                        23,  # 父母电话 -> parent_phone
                        24,  # 个人经历 -> personal_experience
                        25,  # 证明�?-> reference_person
                        26   # 电话(证明人电�? -> reference_phone
                    ]
                    
                    # 填充数据（从�?列开始，不再有序号列�?
                    for col_idx, data_idx in enumerate(column_mapping, start=1):
                        value = youth[data_idx] if data_idx < len(youth) else ''
                        ws.cell(row=row_idx, column=col_idx, value=str(value or ''))
                
            elif tab_name == '异常情况统计':
                headers = ['日期', '姓名', '性别', '公民身份号码', '连', '排', '班', '应征地', '带训班长', '思想', '身体', '精神', '训练', '管理', '其他']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # 获取选中的数据或全部数据
                selected_records = self.get_selected_exception_statistics_data()
                
                if not selected_records:
                    # 没有选择时导出全部数据 - 使用异常统计视图
                    data = self.db_manager.get_exception_statistics_view_data()
                    results = []
                    for record in data:
                        # 转换ViewRecord对象为元组格式，按照导出表头顺序
                        result_tuple = (
                            record[14],  # 日期
                            record[1],   # 姓名
                            record[2],   # 性别
                            record[0],   # 公民身份号码
                            record[3],   # 连
                            record[4],   # 排
                            record[5],   # 班
                            record[7],   # 应征地
                            record[6],   # 带训班长
                            record[8],   # 思想是否异常
                            record[9],   # 身体是否异常
                            record[10],  # 精神是否异常
                            record[11],  # 训练是否异常
                            record[12],  # 管理是否异常
                            f"来源:{record[13]}" if record[13] else ""  # 其他（异常来源）
                        )
                        results.append(result_tuple)
                else:
                    # 有选择时导出选中数据
                    results = selected_records
                
                for row_idx, record in enumerate(results, 2):
                    # 直接使用记录数据，因为已经是正确的格式
                    for col_idx, value in enumerate(record, 1):
                        ws.cell(row=row_idx, column=col_idx, value=str(value or ''))
            
            elif tab_name == '隐性疾病及心理问题筛查情况':
                headers = ['公民身份号码', '筛查类型', '结果', '筛查日期', '后续跟进']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                
                conn = self.db_manager.get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT y.id_card, h.screening_type, h.result, h.screening_date, h.follow_up
                    FROM health_screening h
                    JOIN youth y ON h.youth_id = y.id
                    ORDER BY h.screening_date DESC
                ''')
                results = cursor.fetchall()
                conn.close()
                
                for row_idx, record in enumerate(results, 2):
                    for col_idx, value in enumerate(record, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            elif tab_name == '每日情况统计':
                headers = ['公民身份号码', '姓名', '日期', '思想', '身体', '精神', '其他']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # 获取选中的数据或全部数据
                selected_records = self.get_selected_daily_stats_data()
                
                if not selected_records:
                    # 没有选择时导出全部数�?
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT y.id_card, y.name, d.record_date, d.mood, d.physical_condition, 
                               d.mental_state, d.notes
                        FROM daily_stat d
                        JOIN youth y ON d.youth_id = y.id
                        ORDER BY d.id ASC
                    ''')
                    results = cursor.fetchall()
                    conn.close()
                else:
                    # 有选择时导出选中数据
                    results = selected_records
                
                for row_idx, record in enumerate(results, 2):
                    for col_idx, value in enumerate(record, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            elif tab_name in ['病史调查情况', '镇街谈心谈话情况', '领导谈心谈话情况']:
                if tab_name == '镇街谈心谈话情况':
                    # 镇街谈心谈话情况特殊处理 - xlsx + 图片文件夹格�?
                    import os
                    import shutil
                    
                    # 创建导出文件�?
                    export_folder = file_path.replace('.xlsx', '_导出')
                    if os.path.exists(export_folder):
                        shutil.rmtree(export_folder)
                    os.makedirs(export_folder)
                    
                    # 创建图片文件�?
                    images_folder = os.path.join(export_folder, '走访调查图片')
                    os.makedirs(images_folder)
                    
                    # 设置表头
                    headers = ['日期', '姓名', '性别', '公民身份号码', '思想', '精神', '走访调查情况', '导出时间']
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    # 获取选中的数据或全部数据
                    selected_records = self.get_selected_town_interview_data()
                    if not selected_records:
                        # 没有选择时导出全部数�?
                        records = self.db_manager.search_town_interviews()
                    else:
                        # 有选择时导出选中数据
                        records = selected_records
                    
                    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    for row_idx, record in enumerate(records, 2):
                        # record格式: (id, youth_id_card, youth_name, gender, interview_date, thoughts, spirit, created_at)
                        ws.cell(row=row_idx, column=1, value=record[4] or '')  # 日期
                        ws.cell(row=row_idx, column=2, value=record[2] or '')  # 姓名
                        ws.cell(row=row_idx, column=3, value=record[3] or '')  # 性别
                        ws.cell(row=row_idx, column=4, value=record[1] or '')  # 公民身份号码
                        ws.cell(row=row_idx, column=5, value=record[5] or '')  # 思想
                        ws.cell(row=row_idx, column=6, value=record[6] or '')  # 精神
                        ws.cell(row=row_idx, column=8, value=export_time)  # 导出时间
                        
                        # 处理图片
                        try:
                            image_data = self.db_manager.get_town_interview_image(record[0])
                            if image_data:
                                # 保存图片到文件夹
                                image_filename = f"{record[2]}_{record[1]}_走访调查.jpg"
                                image_path = os.path.join(images_folder, image_filename)
                                
                                with open(image_path, 'wb') as img_file:
                                    img_file.write(image_data)
                                
                                # 在Excel中创建超链接
                                relative_path = f"走访调查图片/{image_filename}"
                                cell = ws.cell(row=row_idx, column=7, value="查看图片")
                                cell.hyperlink = relative_path
                                cell.style = "Hyperlink"
                            else:
                                ws.cell(row=row_idx, column=7, value='无图�?)
                        except Exception as e:
                            print(f"处理图片时出�? {e}")
                            ws.cell(row=row_idx, column=7, value='图片处理失败')
                    
                    # 保存Excel文件到导出文件夹
                    excel_path = os.path.join(export_folder, '镇街谈心谈话数据.xlsx')
                    wb.save(excel_path)
                    
                    # 显示成功消息
                    QMessageBox.information(self, '成功', f'数据已导出到:\n{export_folder}\n\n包含:\n- Excel文件: 镇街谈心谈话数据.xlsx\n- 图片文件�? 走访调查图片')
                    return  # 提前返回，不执行后面的通用保存逻辑
                
                elif tab_name == '领导谈心谈话情况':
                    # 领导谈心谈话情况特殊处理 - xlsx + 图片文件夹格�?
                    import os
                    import shutil
                    
                    # 创建导出文件�?
                    export_folder = file_path.replace('.xlsx', '_导出')
                    if os.path.exists(export_folder):
                        shutil.rmtree(export_folder)
                    os.makedirs(export_folder)
                    
                    # 创建图片文件�?
                    images_folder = os.path.join(export_folder, '走访调查图片')
                    os.makedirs(images_folder)
                    
                    # 设置表头
                    headers = ['日期', '姓名', '性别', '公民身份号码', '思想', '精神', '走访调查情况', '导出时间']
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    # 获取选中的数据或全部数据
                    selected_records = self.get_selected_leader_interview_data()
                    if not selected_records:
                        # 没有选择时导出全部数�?
                        records = self.db_manager.search_leader_interviews()
                    else:
                        # 有选择时导出选中数据
                        records = selected_records
                    
                    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    for row_idx, record in enumerate(records, 2):
                        # record格式: (id, youth_id_card, youth_name, gender, interview_date, thoughts, spirit, created_at)
                        ws.cell(row=row_idx, column=1, value=record[4] or '')  # 日期
                        ws.cell(row=row_idx, column=2, value=record[2] or '')  # 姓名
                        ws.cell(row=row_idx, column=3, value=record[3] or '')  # 性别
                        ws.cell(row=row_idx, column=4, value=record[1] or '')  # 公民身份号码
                        ws.cell(row=row_idx, column=5, value=record[5] or '')  # 思想
                        ws.cell(row=row_idx, column=6, value=record[6] or '')  # 精神
                        ws.cell(row=row_idx, column=8, value=export_time)  # 导出时间
                        
                        # 处理图片
                        try:
                            image_data = self.db_manager.get_leader_interview_image(record[0])
                            if image_data:
                                # 保存图片到文件夹
                                image_filename = f"{record[2]}_{record[1]}_走访调查.jpg"
                                image_path = os.path.join(images_folder, image_filename)
                                
                                with open(image_path, 'wb') as img_file:
                                    img_file.write(image_data)
                                
                                # 在Excel中创建超链接
                                relative_path = f"走访调查图片/{image_filename}"
                                cell = ws.cell(row=row_idx, column=7, value="查看图片")
                                cell.hyperlink = relative_path
                                cell.style = "Hyperlink"
                            else:
                                ws.cell(row=row_idx, column=7, value='无图�?)
                        except Exception as e:
                            print(f"处理图片时出�? {e}")
                            ws.cell(row=row_idx, column=7, value='图片处理失败')
                    
                    # 保存Excel文件到导出文件夹
                    excel_path = os.path.join(export_folder, '领导谈心谈话数据.xlsx')
                    wb.save(excel_path)
                    
                    # 显示成功消息
                    QMessageBox.information(self, '成功', f'数据已导出到:\n{export_folder}\n\n包含:\n- Excel文件: 领导谈心谈话数据.xlsx\n- 图片文件�? 走访调查图片')
                    return  # 提前返回，不执行后面的通用保存逻辑
                    
                else:
                    # 其他情况的原有处�?
                    headers = ['公民身份号码', '姓名', '文件路径', '日期', '备注']
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    table_map = {
                        '病史调查情况': 'medical_history',
                        '领导谈心谈话情况': 'leader_interview'
                    }
                    table_name = table_map[tab_name]
                    date_field = 'upload_date' if tab_name == '病史调查情况' else 'interview_date'
                    
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    cursor.execute(f'''
                        SELECT y.id_card, y.name, t.file_path, t.{date_field}, t.notes
                        FROM {table_name} t
                        JOIN youth y ON t.youth_id = y.id
                        ORDER BY t.{date_field} DESC
                    ''')
                    results = cursor.fetchall()
                    conn.close()
                    
                    for row_idx, record in enumerate(results, 2):
                        for col_idx, value in enumerate(record, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 调整列宽（仅对非镇街谈心谈话情况�?
            if tab_name != '镇街谈心谈话情况':
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
                
                wb.save(file_path)
                QMessageBox.information(self, '成功', f'数据已导出到:\n{file_path}')
            
        except Exception as e:
            error_msg = str(e)
            if "Permission denied" in error_msg or "PermissionError" in error_msg:
                QMessageBox.warning(self, '导出失败', 
                    f'导出失败：文件访问权限被拒绝\n\n可能的原因：\n'
                    f'1. 目标文件正在被Excel或其他程序打开，请关闭后重试\n'
                    f'2. 没有写入权限，请选择其他保存位置（如文档文件夹）\n'
                    f'3. 文件被系统锁定\n\n'
                    f'建议解决方案：\n'
                    f'�?关闭可能打开该文件的Excel程序\n'
                    f'�?选择桌面或文档文件夹作为保存位置\n'
                    f'�?以管理员身份运行程序\n\n'
                    f'详细错误信息：{error_msg}')
            else:
                QMessageBox.warning(self, '导出失败', f'导出失败: {error_msg}')

    def print_all_info(self):
        """打印全部信息"""
        QMessageBox.information(self, '提示', '请在各个模块中使用导出功能导出数据后打印')
    
    def open_town_interview_manager(self):
        """打开镇街谈心谈话情况管理�?""
        try:
            from ui.town_interview_manager import TownInterviewManager
            dialog = TownInterviewManager(self.db_manager, self)
            dialog.exec_()
            # 刷新当前标签页数�?
            self.load_town_interview_data()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开镇街谈心谈话管理器时发生错误: {str(e)}')
    
    def load_town_interview_data(self):
        """加载镇街谈心谈话数据到表�?""
        try:
            # 获取数据
            records = self.db_manager.search_town_interviews()
            self.display_town_interview_records(records)
            
        except Exception as e:
            QMessageBox.warning(self, "加载错误", f"加载镇街谈心谈话数据时发生错误：{str(e)}")
    
    def search_town_interview(self):
        """搜索镇街谈心谈话记录"""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'town_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.town_select_all_checkbox.stateChanged.disconnect()
                self.town_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.town_select_all_checkbox.stateChanged.connect(self.toggle_town_select_all)
            
            name = self.town_name_input.text().strip()
            id_card = self.town_id_card_input.text().strip()
            
            records = self.db_manager.search_town_interviews(name=name, id_card=id_card)
            self.display_town_interview_records(records)
            
        except Exception as e:
            QMessageBox.warning(self, "搜索错误", f"搜索镇街谈心谈话记录时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'town_select_all_checkbox'):
                    self.town_select_all_checkbox.stateChanged.connect(self.toggle_town_select_all)
            except:
                pass
    
    def reset_town_interview_search(self):
        """重置镇街谈心谈话搜索"""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'town_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.town_select_all_checkbox.stateChanged.disconnect()
                self.town_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.town_select_all_checkbox.stateChanged.connect(self.toggle_town_select_all)
            
            self.town_name_input.clear()
            self.town_id_card_input.clear()
            self.load_town_interview_data()
            
        except Exception as e:
            QMessageBox.warning(self, "重置错误", f"重置搜索时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'town_select_all_checkbox'):
                    self.town_select_all_checkbox.stateChanged.connect(self.toggle_town_select_all)
            except:
                pass
    
    def display_town_interview_records(self, records):
        """显示镇街谈心谈话记录"""
        try:
            self.town_interview_table.setRowCount(len(records))
            
            for row, record in enumerate(records):
                # record格式: (id, youth_id_card, youth_name, gender, interview_date, thoughts, spirit, created_at)
                
                # 选择�?
                checkbox = QCheckBox()
                checkbox.setStyleSheet("""
                    QCheckBox {
                        spacing: 0px;
                    }
                    QCheckBox::indicator {
                        width: 20px;
                        height: 20px;
                        border: 2px solid #7F8C8D;
                        border-radius: 2px;
                        background-color: white;
                    }
                    QCheckBox::indicator:hover {
                        border-color: #3498DB;
                        background-color: #ECF0F1;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #3498DB;
                        border-color: #2980B9;
                    }
                """)
                checkbox.setMinimumSize(24, 24)  # 确保复选框有足够空间显�?
                
                # 添加状态变化监听，当单个复选框状态改变时更新全选按钮状�?
                checkbox.stateChanged.connect(self.update_town_select_all_state)
                
                checkbox_layout = QHBoxLayout()
                checkbox_layout.setAlignment(Qt.AlignCenter)
                checkbox_layout.setContentsMargins(5, 5, 5, 5)
                checkbox_layout.addWidget(checkbox)
                checkbox_container = QWidget()
                checkbox_container.setLayout(checkbox_layout)
                self.town_interview_table.setCellWidget(row, 0, checkbox_container)
                
                # 日期
                self.town_interview_table.setItem(row, 1, QTableWidgetItem(str(record[4] or '')))
                
                # 姓名
                self.town_interview_table.setItem(row, 2, QTableWidgetItem(str(record[2] or '')))
                
                # 性别
                self.town_interview_table.setItem(row, 3, QTableWidgetItem(str(record[3] or '')))
                
                # 公民身份号码
                self.town_interview_table.setItem(row, 4, QTableWidgetItem(str(record[1] or '')))
                
                # 思想（截断长文本�?
                thoughts = record[5][:50] + "..." if record[5] and len(record[5]) > 50 else (record[5] or "")
                self.town_interview_table.setItem(row, 5, QTableWidgetItem(thoughts))
                
                # 精神（截断长文本�?
                spirit = record[6][:50] + "..." if record[6] and len(record[6]) > 50 else (record[6] or "")
                self.town_interview_table.setItem(row, 6, QTableWidgetItem(spirit))
                
                # 走访调查情况（查看图片按钮）
                try:
                    image_data = self.db_manager.get_town_interview_image(record[0])
                    if image_data:
                        view_image_btn = QPushButton('查看图片')
                        view_image_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝�?
                        view_image_btn.clicked.connect(lambda checked, rid=record[0]: self.view_town_interview_image(rid))
                    else:
                        view_image_btn = QPushButton('无图�?)
                        view_image_btn.setEnabled(False)
                        view_image_btn.setStyleSheet("color: gray;")
                except:
                    view_image_btn = QPushButton('无图�?)
                    view_image_btn.setEnabled(False)
                    view_image_btn.setStyleSheet("color: gray;")
                
                view_image_btn_layout = QHBoxLayout()
                view_image_btn_layout.setAlignment(Qt.AlignCenter)
                view_image_btn_layout.setContentsMargins(2, 2, 2, 2)
                view_image_btn_layout.addWidget(view_image_btn, 1)  # 添加拉伸因子
                view_image_btn_container = QWidget()
                view_image_btn_container.setLayout(view_image_btn_layout)
                self.town_interview_table.setCellWidget(row, 7, view_image_btn_container)
                
                # 操作按钮（编辑信息）
                edit_btn = QPushButton('编辑信息')
                edit_btn.setStyleSheet("background-color: #ADD8E6;")  # 淡蓝�?
                edit_btn.clicked.connect(lambda checked, record_data=record: self.edit_town_interview_by_record(record_data))
                edit_btn_layout = QHBoxLayout()
                edit_btn_layout.setAlignment(Qt.AlignCenter)
                edit_btn_layout.setContentsMargins(5, 5, 5, 5)
                edit_btn_layout.addWidget(edit_btn)
                edit_btn_container = QWidget()
                edit_btn_container.setLayout(edit_btn_layout)
                self.town_interview_table.setCellWidget(row, 8, edit_btn_container)
                
                # 删除按钮
                delete_btn = QPushButton('删除')
                delete_btn.setStyleSheet("background-color: #FFB6C1;")  # 淡红�?
                delete_btn.clicked.connect(lambda checked, record_data=record: self.delete_single_town_interview(record_data))
                delete_btn_layout = QHBoxLayout()
                delete_btn_layout.setAlignment(Qt.AlignCenter)
                delete_btn_layout.setContentsMargins(5, 5, 5, 5)
                delete_btn_layout.addWidget(delete_btn)
                delete_btn_container = QWidget()
                delete_btn_container.setLayout(delete_btn_layout)
                self.town_interview_table.setCellWidget(row, 9, delete_btn_container)
            
            # 不调�?resizeColumnsToContents()，保持创建表格时设置的自适应模式
            
        except Exception as e:
            QMessageBox.warning(self, "显示错误", f"显示镇街谈心谈话记录时发生错误：{str(e)}")
    
    def toggle_town_select_all(self, state):
        """镇街谈心谈话全�?取消全�?""
        is_checked = (state == Qt.CheckState.Checked.value)
        
        # 临时标记，防止在全选操作时触发单个复选框的状态更�?
        self._updating_town_select_all = True
        
        for row in range(self.town_interview_table.rowCount()):
            checkbox_widget = self.town_interview_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(is_checked)
        
        # 重置标记
        self._updating_town_select_all = False
    
    def update_town_select_all_state(self):
        """更新镇街谈心谈话全选按钮状�?""
        # 如果正在执行全选操作，则不更新全选按钮状�?
        if getattr(self, '_updating_town_select_all', False):
            return
        
        try:
            total_rows = self.town_interview_table.rowCount()
            if total_rows == 0:
                return
            
            checked_count = 0
            for row in range(total_rows):
                checkbox_widget = self.town_interview_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        checked_count += 1
            
            # 临时断开全选按钮的信号连接，避免触发toggle_town_select_all
            self.town_select_all_checkbox.stateChanged.disconnect()
            
            # 更新全选按钮状�?
            if checked_count == 0:
                # 没有选中任何�?
                self.town_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            elif checked_count == total_rows:
                # 全部选中
                self.town_select_all_checkbox.setCheckState(Qt.CheckState.Checked)
            else:
                # 部分选中
                self.town_select_all_checkbox.setCheckState(Qt.CheckState.PartiallyChecked)
            
            # 重新连接信号
            self.town_select_all_checkbox.stateChanged.connect(self.toggle_town_select_all)
            
        except Exception as e:
            print(f"更新镇街谈话全选按钮状态时出错: {e}")
            # 确保信号重新连接
            try:
                self.town_select_all_checkbox.stateChanged.connect(self.toggle_town_select_all)
            except:
                pass
    
    def add_town_interview(self):
        """添加镇街谈心谈话记录"""
        try:
            from ui.town_interview_dialog import TownInterviewDialog
            dialog = TownInterviewDialog(self.db_manager, self)
            if dialog.exec_():
                self.load_town_interview_data()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开添加镇街谈心谈话对话框时发生错误: {str(e)}')
    
    def edit_town_interview(self):
        """编辑镇街谈心谈话记录（旧方法，保留兼容性）"""
        current_row = self.town_interview_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "提示", "请选择要编辑的记录")
            return
        
        try:
            # 获取所有记录，根据界面行号找到对应记录
            records = self.db_manager.search_town_interviews()
            if current_row < len(records):
                record_data = records[current_row]
                self.edit_town_interview_by_record(record_data)
            else:
                QMessageBox.warning(self, "错误", "找不到对应的记录")
                
        except Exception as e:
            QMessageBox.warning(self, '错误', f'编辑镇街谈心谈话记录时发生错�? {str(e)}')
    
    def edit_town_interview_by_record(self, record_data):
        """根据记录数据编辑镇街谈心谈话记录"""
        try:
            from ui.town_interview_dialog import TownInterviewDialog
            dialog = TownInterviewDialog(self.db_manager, self, record_data)
            if dialog.exec_():
                self.load_town_interview_data()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'编辑镇街谈心谈话记录时发生错�? {str(e)}')
    
    def batch_delete_town_interview(self):
        """批量删除镇街谈心谈话记录"""
        try:
            # 获取选中的记�?
            selected_rows = []
            for row in range(self.town_interview_table.rowCount()):
                checkbox_widget = self.town_interview_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        selected_rows.append(row)
            
            if not selected_rows:
                QMessageBox.warning(self, "提示", "请选择要删除的记录")
                return
            
            # 获取所有记录，根据选中行号找到对应的记录ID
            records = self.db_manager.search_town_interviews()
            selected_ids = []
            selected_names = []
            
            for row in selected_rows:
                if row < len(records):
                    record = records[row]
                    selected_ids.append(record[0])  # record[0] �?id
                    selected_names.append(record[2])  # record[2] �?youth_name
            
            if not selected_ids:
                QMessageBox.warning(self, "错误", "找不到要删除的记�?)
                return
            
            # 确认删除
            names_text = '�?.join(selected_names[:5])
            if len(selected_names) > 5:
                names_text += f' 等{len(selected_names)}�?
            
            reply = QMessageBox.question(
                self,
                '确认批量删除',
                f'确定要删除以下镇街谈心谈话记录吗？\n{names_text}\n\n共{len(selected_ids)}条记录，删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                deleted_count = self.db_manager.delete_town_interviews(selected_ids)
                QMessageBox.information(self, '删除成功', f'成功删除 {deleted_count} 条记�?)
                
                # 取消全�?
                if hasattr(self, 'town_select_all_checkbox'):
                    self.town_select_all_checkbox.setChecked(False)
                
                self.load_town_interview_data()
                
        except Exception as e:
            QMessageBox.critical(self, '批量删除失败', f'批量删除镇街谈心谈话记录时发生错误：{str(e)}')
    
    def delete_single_town_interview(self, record_data):
        """删除单个镇街谈心谈话记录"""
        try:
            # record_data格式: (id, youth_id_card, youth_name, gender, interview_date, thoughts, spirit, created_at)
            record_id = record_data[0]
            youth_name = record_data[2]
            
            reply = QMessageBox.question(
                self,
                '确认删除',
                f'确定要删�?{youth_name} 的镇街谈心谈话记录吗？\n删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                deleted_count = self.db_manager.delete_town_interviews([record_id])
                if deleted_count > 0:
                    QMessageBox.information(self, '删除成功', '记录已删�?)
                    self.load_town_interview_data()
                else:
                    QMessageBox.warning(self, '删除失败', '未能删除记录')
                    
        except Exception as e:
            QMessageBox.critical(self, '删除失败', f'删除镇街谈心谈话记录时发生错误：{str(e)}')
    
    def view_town_interview_image(self, record_id):
        """查看镇街谈心谈话图片"""
        try:
            image_data = self.db_manager.get_town_interview_image(record_id)
            if not image_data:
                QMessageBox.information(self, "提示", "该记录没有图�?)
                return
            
            from ui.image_viewer_dialog import ImageViewerDialog
            dialog = ImageViewerDialog(image_data, "走访调查图片", self)
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"查看图片失败: {str(e)}")
    
    def open_town_interview_manager(self):
        """打开镇街谈心谈话情况管理器（已废弃）"""
        # 这个方法已经不再使用，功能已集成到主窗口
        pass
    
    def get_selected_basic_info_data(self):
        """获取基本信息中选中的数�?""
        try:
            selected_data = []
            
            # 获取所有数据用于对�?
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id_card, name, gender, nation, political_status,
                       school, education_level, major, study_status, study_type,
                       phone, household_address, residence_address, family_info,
                       district, street, parent_phone, personal_experience,
                       reference_person, reference_phone, id
                FROM youth ORDER BY id
            """)
            all_results = cursor.fetchall()
            conn.close()
            
            # 检查选中的行
            for row in range(self.search_table.rowCount()):
                checkbox_widget = self.search_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_results):
                            selected_data.append(all_results[row])
            
            return selected_data
            
        except Exception as e:
            print(f"获取选中基本信息数据时出�? {e}")
            return []
    
    def get_selected_town_interview_data(self):
        """获取镇街谈心谈话中选中的数�?""
        try:
            selected_data = []
            
            # 获取所有数据用于对�?
            all_records = self.db_manager.search_town_interviews()
            
            # 检查选中的行
            for row in range(self.town_interview_table.rowCount()):
                checkbox_widget = self.town_interview_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_records):
                            selected_data.append(all_records[row])
            
            return selected_data
            
        except Exception as e:
            print(f"获取选中镇街谈心谈话数据时出�? {e}")
            return []
    
    def load_leader_interview_data(self):
        """加载领导谈心谈话数据到表�?""
        try:
            # 获取数据
            records = self.db_manager.search_leader_interviews()
            self.display_leader_interview_records(records)
            
        except Exception as e:
            QMessageBox.warning(self, "加载错误", f"加载领导谈心谈话数据时发生错误：{str(e)}")
    
    def search_leader_interview(self):
        """搜索领导谈心谈话记录"""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'leader_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.leader_select_all_checkbox.stateChanged.disconnect()
                self.leader_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.leader_select_all_checkbox.stateChanged.connect(self.toggle_leader_select_all)
            
            name = self.leader_name_input.text().strip()
            id_card = self.leader_id_card_input.text().strip()
            
            records = self.db_manager.search_leader_interviews(name=name, id_card=id_card)
            self.display_leader_interview_records(records)
            
        except Exception as e:
            QMessageBox.warning(self, "搜索错误", f"搜索领导谈心谈话记录时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'leader_select_all_checkbox'):
                    self.leader_select_all_checkbox.stateChanged.connect(self.toggle_leader_select_all)
            except:
                pass
    
    def reset_leader_interview_search(self):
        """重置领导谈心谈话搜索"""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'leader_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.leader_select_all_checkbox.stateChanged.disconnect()
                self.leader_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.leader_select_all_checkbox.stateChanged.connect(self.toggle_leader_select_all)
            
            self.leader_name_input.clear()
            self.leader_id_card_input.clear()
            self.load_leader_interview_data()
            
        except Exception as e:
            QMessageBox.warning(self, "重置错误", f"重置搜索时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'leader_select_all_checkbox'):
                    self.leader_select_all_checkbox.stateChanged.connect(self.toggle_leader_select_all)
            except:
                pass
    
    def display_leader_interview_records(self, records):
        """显示领导谈心谈话记录"""
        try:
            self.leader_interview_table.setRowCount(len(records))
            
            for row, record in enumerate(records):
                # record格式: (id, youth_id_card, youth_name, gender, interview_date, thoughts, spirit, created_at)
                
                # 选择�?
                checkbox = QCheckBox()
                checkbox.setStyleSheet("""
                    QCheckBox {
                        spacing: 0px;
                    }
                    QCheckBox::indicator {
                        width: 20px;
                        height: 20px;
                        border: 2px solid #7F8C8D;
                        border-radius: 2px;
                        background-color: white;
                    }
                    QCheckBox::indicator:hover {
                        border-color: #3498DB;
                        background-color: #ECF0F1;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #3498DB;
                        border-color: #2980B9;
                    }
                """)
                checkbox.setMinimumSize(24, 24)  # 确保复选框有足够空间显�?
                
                # 添加状态变化监听，当单个复选框状态改变时更新全选按钮状�?
                checkbox.stateChanged.connect(self.update_leader_select_all_state)
                
                checkbox_layout = QHBoxLayout()
                checkbox_layout.setAlignment(Qt.AlignCenter)
                checkbox_layout.setContentsMargins(5, 5, 5, 5)
                checkbox_layout.addWidget(checkbox)
                checkbox_container = QWidget()
                checkbox_container.setLayout(checkbox_layout)
                self.leader_interview_table.setCellWidget(row, 0, checkbox_container)
                
                # 日期
                self.leader_interview_table.setItem(row, 1, QTableWidgetItem(str(record[4] or '')))
                
                # 姓名
                self.leader_interview_table.setItem(row, 2, QTableWidgetItem(str(record[2] or '')))
                
                # 性别
                self.leader_interview_table.setItem(row, 3, QTableWidgetItem(str(record[3] or '')))
                
                # 公民身份号码
                self.leader_interview_table.setItem(row, 4, QTableWidgetItem(str(record[1] or '')))
                
                # 思想（截断长文本�?
                thoughts = record[5][:50] + "..." if record[5] and len(record[5]) > 50 else (record[5] or "")
                self.leader_interview_table.setItem(row, 5, QTableWidgetItem(thoughts))
                
                # 精神（截断长文本�?
                spirit = record[6][:50] + "..." if record[6] and len(record[6]) > 50 else (record[6] or "")
                self.leader_interview_table.setItem(row, 6, QTableWidgetItem(spirit))
                
                # 走访调查情况（查看图片按钮）
                try:
                    image_data = self.db_manager.get_leader_interview_image(record[0])
                    if image_data:
                        view_image_btn = QPushButton('查看图片')
                        view_image_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝�?
                        view_image_btn.clicked.connect(lambda checked, rid=record[0]: self.view_leader_interview_image(rid))
                    else:
                        view_image_btn = QPushButton('无图�?)
                        view_image_btn.setEnabled(False)
                        view_image_btn.setStyleSheet("color: gray;")
                except:
                    view_image_btn = QPushButton('无图�?)
                    view_image_btn.setEnabled(False)
                    view_image_btn.setStyleSheet("color: gray;")
                
                view_image_btn_layout = QHBoxLayout()
                view_image_btn_layout.setAlignment(Qt.AlignCenter)
                view_image_btn_layout.setContentsMargins(2, 2, 2, 2)
                view_image_btn_layout.addWidget(view_image_btn, 1)  # 添加拉伸因子
                view_image_btn_container = QWidget()
                view_image_btn_container.setLayout(view_image_btn_layout)
                self.leader_interview_table.setCellWidget(row, 7, view_image_btn_container)
                
                # 操作按钮（编辑信息）
                edit_btn = QPushButton('编辑信息')
                edit_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝�?
                edit_btn.clicked.connect(lambda checked, record_data=record: self.edit_leader_interview_by_record(record_data))
                edit_btn_layout = QHBoxLayout()
                edit_btn_layout.setAlignment(Qt.AlignCenter)
                edit_btn_layout.setContentsMargins(5, 5, 5, 5)
                edit_btn_layout.addWidget(edit_btn)
                edit_btn_container = QWidget()
                edit_btn_container.setLayout(edit_btn_layout)
                self.leader_interview_table.setCellWidget(row, 8, edit_btn_container)
                
                # 删除按钮
                delete_btn = QPushButton('删除')
                delete_btn.setStyleSheet('background-color: #FFB6C1;')  # 淡红�?
                delete_btn.clicked.connect(lambda checked, record_data=record: self.delete_single_leader_interview(record_data))
                delete_btn_layout = QHBoxLayout()
                delete_btn_layout.setAlignment(Qt.AlignCenter)
                delete_btn_layout.setContentsMargins(5, 5, 5, 5)
                delete_btn_layout.addWidget(delete_btn)
                delete_btn_container = QWidget()
                delete_btn_container.setLayout(delete_btn_layout)
                self.leader_interview_table.setCellWidget(row, 9, delete_btn_container)
            
            # 不调�?resizeColumnsToContents()，保持创建表格时设置的自适应模式
            
        except Exception as e:
            QMessageBox.warning(self, "显示错误", f"显示领导谈心谈话记录时发生错误：{str(e)}")
    
    def toggle_leader_select_all(self, state):
        """领导谈心谈话全�?取消全�?""
        is_checked = (state == Qt.CheckState.Checked.value)
        
        # 临时标记，防止在全选操作时触发单个复选框的状态更�?
        self._updating_leader_select_all = True
        
        for row in range(self.leader_interview_table.rowCount()):
            checkbox_widget = self.leader_interview_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(is_checked)
        
        # 重置标记
        self._updating_leader_select_all = False
    
    def update_leader_select_all_state(self):
        """更新领导谈心谈话全选按钮状�?""
        # 如果正在执行全选操作，则不更新全选按钮状�?
        if getattr(self, '_updating_leader_select_all', False):
            return
        
        try:
            total_rows = self.leader_interview_table.rowCount()
            if total_rows == 0:
                return
            
            checked_count = 0
            for row in range(total_rows):
                checkbox_widget = self.leader_interview_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        checked_count += 1
            
            # 临时断开全选按钮的信号连接，避免触发toggle_leader_select_all
            self.leader_select_all_checkbox.stateChanged.disconnect()
            
            # 更新全选按钮状�?
            if checked_count == 0:
                # 没有选中任何�?
                self.leader_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            elif checked_count == total_rows:
                # 全部选中
                self.leader_select_all_checkbox.setCheckState(Qt.CheckState.Checked)
            else:
                # 部分选中
                self.leader_select_all_checkbox.setCheckState(Qt.CheckState.PartiallyChecked)
            
            # 重新连接信号
            self.leader_select_all_checkbox.stateChanged.connect(self.toggle_leader_select_all)
            
        except Exception as e:
            print(f"更新领导谈话全选按钮状态时出错: {e}")
            # 确保信号重新连接
            try:
                self.leader_select_all_checkbox.stateChanged.connect(self.toggle_leader_select_all)
            except:
                pass
    
    def add_leader_interview(self):
        """添加领导谈心谈话记录"""
        try:
            from ui.leader_interview_dialog import LeaderInterviewDialog
            dialog = LeaderInterviewDialog(self.db_manager, self)
            if dialog.exec_():
                self.load_leader_interview_data()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开添加领导谈心谈话对话框时发生错误: {str(e)}')
    
    def edit_leader_interview_by_record(self, record_data):
        """根据记录数据编辑领导谈心谈话记录"""
        try:
            from ui.leader_interview_dialog import LeaderInterviewDialog
            dialog = LeaderInterviewDialog(self.db_manager, self, record_data)
            if dialog.exec_():
                self.load_leader_interview_data()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'编辑领导谈心谈话记录时发生错�? {str(e)}')
    
    def delete_single_leader_interview(self, record_data):
        """删除单个领导谈心谈话记录"""
        try:
            # record_data格式: (id, youth_id_card, youth_name, gender, interview_date, thoughts, spirit, created_at)
            record_id = record_data[0]
            youth_name = record_data[2]
            
            reply = QMessageBox.question(
                self,
                '确认删除',
                f'确定要删�?{youth_name} 的领导谈心谈话记录吗？\n删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                deleted_count = self.db_manager.delete_leader_interviews([record_id])
                if deleted_count > 0:
                    QMessageBox.information(self, '删除成功', '记录已删�?)
                    self.load_leader_interview_data()
                else:
                    QMessageBox.warning(self, '删除失败', '未能删除记录')
                    
        except Exception as e:
            QMessageBox.critical(self, '删除失败', f'删除领导谈心谈话记录时发生错误：{str(e)}')
    
    def batch_delete_leader_interview(self):
        """批量删除领导谈心谈话记录"""
        try:
            # 获取选中的记�?
            selected_rows = []
            for row in range(self.leader_interview_table.rowCount()):
                checkbox_widget = self.leader_interview_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        selected_rows.append(row)
            
            if not selected_rows:
                QMessageBox.warning(self, "提示", "请选择要删除的记录")
                return
            
            # 获取所有记录，根据选中行号找到对应的记录ID
            records = self.db_manager.search_leader_interviews()
            selected_ids = []
            selected_names = []
            
            for row in selected_rows:
                if row < len(records):
                    record = records[row]
                    selected_ids.append(record[0])  # record[0] �?id
                    selected_names.append(record[2])  # record[2] �?youth_name
            
            if not selected_ids:
                QMessageBox.warning(self, "错误", "找不到要删除的记�?)
                return
            
            # 确认删除
            names_text = '�?.join(selected_names[:5])
            if len(selected_names) > 5:
                names_text += f' 等{len(selected_names)}�?
            
            reply = QMessageBox.question(
                self,
                '确认批量删除',
                f'确定要删除以下领导谈心谈话记录吗？\n{names_text}\n\n共{len(selected_ids)}条记录，删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                deleted_count = self.db_manager.delete_leader_interviews(selected_ids)
                QMessageBox.information(self, '删除成功', f'成功删除 {deleted_count} 条记�?)
                
                # 取消全�?
                if hasattr(self, 'leader_select_all_checkbox'):
                    self.leader_select_all_checkbox.setChecked(False)
                
                self.load_leader_interview_data()
                
        except Exception as e:
            QMessageBox.critical(self, '批量删除失败', f'批量删除领导谈心谈话记录时发生错误：{str(e)}')
    
    def view_leader_interview_image(self, record_id):
        """查看领导谈心谈话图片"""
        try:
            image_data = self.db_manager.get_leader_interview_image(record_id)
            if not image_data:
                QMessageBox.information(self, "提示", "该记录没有图�?)
                return
            
            from ui.image_viewer_dialog import ImageViewerDialog
            dialog = ImageViewerDialog(image_data, "走访调查图片", self)
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"查看图片失败: {str(e)}")
    
    def get_selected_leader_interview_data(self):
        """获取领导谈心谈话中选中的数�?""
        try:
            selected_data = []
            
            # 获取所有数据用于对�?
            all_records = self.db_manager.search_leader_interviews()
            
            # 检查选中的行
            for row in range(self.leader_interview_table.rowCount()):
                checkbox_widget = self.leader_interview_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_records):
                            selected_data.append(all_records[row])
            
            return selected_data
            
        except Exception as e:
            print(f"获取选中领导谈心谈话数据时出�? {e}")
            return []

    def get_selected_exception_statistics_data(self):
        """获取异常情况统计中选中的数据"""
        try:
            selected_data = []
            
            # 获取所有异常统计视图数据用于对照
            all_records = self.db_manager.get_exception_statistics_view_data()
            
            # 检查选中的行
            for row in range(self.exception_statistics_table.rowCount()):
                checkbox_widget = self.exception_statistics_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_records):
                            record = all_records[row]
                            # 转换为导出格式：日期, 姓名, 性别, 公民身份号码, 连, 排, 班, 应征地, 带训班长, 思想, 身体, 精神, 训练, 管理, 其他
                            selected_tuple = (
                                record[14],  # 日期
                                record[1],   # 姓名
                                record[2],   # 性别
                                record[0],   # 公民身份号码
                                record[3],   # 连
                                record[4],   # 排
                                record[5],   # 班
                                record[7],   # 应征地
                                record[6],   # 带训班长
                                record[8],   # 思想是否异常
                                record[9],   # 身体是否异常
                                record[10],  # 精神是否异常
                                record[11],  # 训练是否异常
                                record[12],  # 管理是否异常
                                f"来源:{record[13]}" if record[13] else ""  # 其他（异常来源）
                            )
                            selected_data.append(selected_tuple)
            
            return selected_data
            
        except Exception as e:
            print(f"获取选中异常情况统计数据时出错: {e}")
            return []

    
    # ==================== 病史筛查情况模块 ====================
    
    def create_medical_screening_tab(self):
        """创建病史筛查标签�?""
        tab_widget = QWidget()
        layout = QVBoxLayout()
        
        # 搜索�?
        search_layout = QHBoxLayout()
        
        search_layout.addWidget(QLabel('搜索:'))
        
        self.medical_name_input = QLineEdit()
        self.medical_name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.medical_name_input.setPlaceholderText('姓名')
        self.medical_name_input.returnPressed.connect(self.search_medical_screening)
        search_layout.addWidget(self.medical_name_input)
        
        self.medical_id_card_input = QLineEdit()
        self.medical_id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.medical_id_card_input.setPlaceholderText('身份证号')
        self.medical_id_card_input.returnPressed.connect(self.search_medical_screening)
        search_layout.addWidget(self.medical_id_card_input)
        
        search_btn = QPushButton('搜索')
        search_btn.clicked.connect(self.search_medical_screening)
        self.setup_button_style(search_btn, 'primary')
        search_layout.addWidget(search_btn)
        
        reset_btn = QPushButton('重置')
        reset_btn.clicked.connect(self.reset_medical_screening_search)
        self.setup_button_style(reset_btn, 'normal')
        search_layout.addWidget(reset_btn)
        
        layout.addLayout(search_layout)
        
        # 全选复选框�?
        select_all_layout = QHBoxLayout()
        self.medical_select_all_checkbox = QCheckBox('全�?)
        self.medical_select_all_checkbox.stateChanged.connect(self.toggle_medical_select_all)
        self.medical_select_all_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #7F8C8D;
                border-radius: 2px;
                background-color: white;
            }
            QCheckBox::indicator:hover {
                border-color: #3498DB;
                background-color: #ECF0F1;
            }
            QCheckBox::indicator:checked {
                background-color: #3498DB;
                border-color: #2980B9;
            }
        """)
        select_all_layout.addWidget(self.medical_select_all_checkbox)
        select_all_layout.addStretch()
        layout.addLayout(select_all_layout)
        
        # 病史筛查表格
        self.medical_screening_table = QTableWidget()
        medical_headers = ['选择', '姓名', '性别', '公民身份号码', '筛查情况', '筛查日期', '身体状况', '精神状况', '操作', '删除']
        self.medical_screening_table.setColumnCount(len(medical_headers))
        self.medical_screening_table.setHorizontalHeaderLabels(medical_headers)
        self.medical_screening_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.medical_screening_table.setAlternatingRowColors(True)
        self.medical_screening_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # 应用统一的表格样�?
        self.setup_table_style(self.medical_screening_table)  # 禁止直接编辑
        # 设置文本省略模式为末尾省�?
        self.medical_screening_table.setTextElideMode(Qt.TextElideMode.ElideRight)
        
        # 确保行高设置生效
        self.medical_screening_table.verticalHeader().setDefaultSectionSize(40)
        
        # 连接双击事件
        self.medical_screening_table.cellDoubleClicked.connect(self.show_medical_screening_detail)
        
        # 设置表格自适应屏幕
        from PyQt5.QtWidgets import QHeaderView
        header = self.medical_screening_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)  # 允许手动调整
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 选择�?
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 姓名�?
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 性别�?
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 身份证列
        header.setSectionResizeMode(4, QHeaderView.Stretch)  # 筛查情况列自动拉�?
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 筛查日期�?
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 身体状况�?
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # 精神状况�?
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # 操作�?
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # 删除�?
        header.setStretchLastSection(False)
        header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        
        layout.addWidget(self.medical_screening_table)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        import_btn = QPushButton('导入数据')
        import_btn.clicked.connect(lambda: self.import_medical_screening_data())
        self.setup_button_style(import_btn, 'normal')
        button_layout.addWidget(import_btn)
        
        export_btn = QPushButton('导出数据')
        export_btn.clicked.connect(lambda: self.export_medical_screening_data())
        self.setup_button_style(export_btn, 'normal')
        button_layout.addWidget(export_btn)
        
        add_btn = QPushButton('添加信息')
        add_btn.clicked.connect(self.add_medical_screening_record)
        self.setup_button_style(add_btn, 'normal')
        button_layout.addWidget(add_btn)
        
        batch_delete_btn = QPushButton('批量删除')
        self.setup_button_style(batch_delete_btn, 'danger')
        batch_delete_btn.clicked.connect(self.batch_delete_medical_screening)
        button_layout.addWidget(batch_delete_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        tab_widget.setLayout(layout)
        
        # 加载病史筛查数据
        self.load_medical_screening_data()
        
        return tab_widget
    
    def load_medical_screening_data(self):
        """加载病史筛查数据"""
        try:
            conn = None
            try:
                conn = self.db_manager.get_connection()
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT ms.id, ms.name, ms.gender, ms.id_card, ms.screening_result, ms.screening_date, 
                           ms.physical_status, ms.mental_status
                    FROM medical_screening ms
                    ORDER BY ms.id
                """)
                results = cursor.fetchall()
                self.display_medical_screening_results(results)
            finally:
                if conn:
                    conn.close()
        except Exception as e:
            QMessageBox.warning(self, "加载错误", f"加载病史筛查数据时发生错误：{str(e)}")
            self.medical_screening_table.setRowCount(0)
    
    def search_medical_screening(self):
        """搜索病史筛查记录"""
        try:
            name = self.medical_name_input.text().strip()
            id_card = self.medical_id_card_input.text().strip()
            
            conn = None
            try:
                conn = self.db_manager.get_connection()
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                sql = """
                    SELECT ms.id, ms.name, ms.gender, ms.id_card, ms.screening_result, ms.screening_date, 
                           ms.physical_status, ms.mental_status
                    FROM medical_screening ms
                    WHERE 1=1
                """
                params = []
                
                if name:
                    sql += " AND ms.name LIKE ?"
                    params.append(f"%{name}%")
                
                if id_card:
                    sql += " AND ms.id_card LIKE ?"
                    params.append(f"%{id_card}%")
                
                sql += " ORDER BY ms.id"
                
                cursor.execute(sql, params)
                results = cursor.fetchall()
                self.display_medical_screening_results(results)
                
            finally:
                if conn:
                    conn.close()
                    
        except Exception as e:
            QMessageBox.warning(self, "搜索错误", f"搜索病史筛查记录时发生错误：{str(e)}")
    
    def reset_medical_screening_search(self):
        """重置病史筛查搜索"""
        self.medical_name_input.clear()
        self.medical_id_card_input.clear()
        self.load_medical_screening_data()
    
    def display_medical_screening_results(self, results):
        """显示病史筛查结果"""
        self.medical_screening_table.setRowCount(len(results))

        for row, data in enumerate(results):
            # 第一列：复选框
            checkbox = QCheckBox()
            checkbox.setStyleSheet("""
                    QCheckBox {
                        spacing: 0px;
                    }
                    QCheckBox::indicator {
                        width: 20px;
                        height: 20px;
                        border: 2px solid #7F8C8D;
                        border-radius: 2px;
                        background-color: white;
                    }
                    QCheckBox::indicator:hover {
                        border-color: #3498DB;
                        background-color: #ECF0F1;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #3498DB;
                        border-color: #2980B9;
                    }
                """)
            checkbox.setMinimumSize(24, 24)  # 确保复选框有足够空间显�?
            
            checkbox_layout = QHBoxLayout()
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.setContentsMargins(5, 5, 5, 5)
            checkbox_layout.addWidget(checkbox)
            checkbox_container = QWidget()
            checkbox_container.setLayout(checkbox_layout)
            self.medical_screening_table.setCellWidget(row, 0, checkbox_container)

            # 第二列：姓名
            name_text = str(data['name'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(name_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            self.medical_screening_table.setItem(row, 1, item)

            # 第三列：性别
            gender_text = str(data['gender'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(gender_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            self.medical_screening_table.setItem(row, 2, item)

            # 第四列：公民身份号码
            id_card_text = str(data['id_card'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(id_card_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            self.medical_screening_table.setItem(row, 3, item)

            # 第五列：筛查情况
            screening_text = str(data['screening_result'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(screening_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            # 设置工具提示显示完整内容
            item.setToolTip(screening_text)
            self.medical_screening_table.setItem(row, 4, item)

            # 第六列：筛查日期
            date_text = str(data['screening_date'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(date_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            self.medical_screening_table.setItem(row, 5, item)

            # 第七列：身体状况
            physical_text = str(data['physical_status'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(physical_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            self.medical_screening_table.setItem(row, 6, item)

            # 第八列：精神状况
            mental_text = str(data['mental_status'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(mental_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            self.medical_screening_table.setItem(row, 7, item)

            # 第九列：查看详情按钮
            medical_id = data['id']
            view_btn = QPushButton('查看详情')
            view_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝�?
            view_btn.clicked.connect(lambda checked, mid=medical_id: self.view_medical_screening_details(mid))
            view_btn_layout = QHBoxLayout()
            view_btn_layout.setAlignment(Qt.AlignCenter)
            view_btn_layout.setContentsMargins(5, 5, 5, 5)
            view_btn_layout.addWidget(view_btn)
            view_btn_container = QWidget()
            view_btn_container.setLayout(view_btn_layout)
            self.medical_screening_table.setCellWidget(row, 8, view_btn_container)

            # 第十列：删除按钮
            delete_btn = QPushButton('删除')
            delete_btn.setStyleSheet('background-color: #FFB6C1;')  # 淡红�?
            delete_btn.clicked.connect(lambda checked, mid=medical_id: self.delete_medical_screening(mid))
            delete_btn_layout = QHBoxLayout()
            delete_btn_layout.setAlignment(Qt.AlignCenter)
            delete_btn_layout.setContentsMargins(5, 5, 5, 5)
            delete_btn_layout.addWidget(delete_btn)
            delete_btn_container = QWidget()
            delete_btn_container.setLayout(delete_btn_layout)
            self.medical_screening_table.setCellWidget(row, 9, delete_btn_container)
        
        # 确保所有行的行高都�?0像素
        for row in range(self.medical_screening_table.rowCount()):
            self.medical_screening_table.setRowHeight(row, 40)
    
    def add_medical_screening_record(self):
        """添加病史筛查记录"""
        from ui.add_medical_screening_dialog import AddMedicalScreeningDialog
        
        dialog = AddMedicalScreeningDialog(self.db_manager, self)
        dialog.data_updated.connect(self.load_medical_screening_data)
        dialog.exec_()
    
    def show_medical_screening_detail(self, row, column):
        """显示病史筛查详情（双击事件）"""
        try:
            # 获取该行的记录ID（存储在�?列的按钮数据中）
            # 我们需要从数据库重新查询该行的完整数据
            
            # 获取身份证号和日期来查询
            id_card_item = self.medical_screening_table.item(row, 3)
            date_item = self.medical_screening_table.item(row, 5)
            
            if not id_card_item or not date_item:
                return
            
            id_card = id_card_item.text()
            screening_date = date_item.text()
            
            # 从数据库查询完整记录
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, youth_id_card, name, gender, id_card, 
                       screening_result, screening_date, physical_status, mental_status
                FROM medical_screening 
                WHERE id_card = ? AND screening_date = ?
            ''', (id_card, screening_date))
            
            record = cursor.fetchone()
            conn.close()
            
            if record:
                from ui.medical_screening_detail_dialog import MedicalScreeningDetailDialog
                dialog = MedicalScreeningDetailDialog(self.db_manager, record, self)
                dialog.data_updated.connect(self.load_medical_screening_data)
                dialog.exec_()
            else:
                QMessageBox.warning(self, '错误', '未找到该记录')
                
        except Exception as e:
            QMessageBox.warning(self, '错误', f'查看详情时发生错�? {str(e)}')
    
    def delete_medical_screening(self, medical_id):
        """删除单条病史筛查记录"""
        try:
            reply = QMessageBox.question(
                self, 
                '确认删除', 
                '确定要删除这条病史筛查记录吗？\n删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                conn = None
                try:
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    
                    cursor.execute("DELETE FROM medical_screening WHERE id = ?", (medical_id,))
                    conn.commit()
                    
                    QMessageBox.information(self, '删除成功', '病史筛查记录已删除')
                    self.load_medical_screening_data()
                    
                finally:
                    if conn:
                        conn.close()
                        
        except Exception as e:
            QMessageBox.critical(self, '删除失败', f'删除病史筛查记录时发生错误：{str(e)}')
    
    def batch_delete_medical_screening(self):
        """批量删除病史筛查记录"""
        try:
            selected_ids = []
            selected_names = []
            
            # 先从当前显示的数据中获取
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, name FROM medical_screening ORDER BY id
            """)
            all_results = cursor.fetchall()
            conn.close()
            
            for row in range(self.medical_screening_table.rowCount()):
                checkbox_widget = self.medical_screening_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_results):
                            data = all_results[row]
                            medical_id = data['id']
                            name = data['name']
                            
                            if medical_id:
                                selected_ids.append(medical_id)
                                selected_names.append(name)
            
            if not selected_ids:
                QMessageBox.information(self, '提示', '请先选择要删除的病史筛查记录')
                return
            
            names_text = '�?.join(selected_names[:5])
            if len(selected_names) > 5:
                names_text += f' 等{len(selected_names)}�?
            
            reply = QMessageBox.question(
                self,
                '确认批量删除',
                f'确定要删除以下病史筛查记录吗？\n{names_text}\n\n共{len(selected_ids)}条记录，删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                conn = None
                try:
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    
                    # 在删除前收集需要同步的记录信息
                    deleted_count = 0
                    for medical_id in selected_ids:
                        cursor.execute("DELETE FROM medical_screening WHERE id = ?", (medical_id,))
                        if cursor.rowcount > 0:
                            deleted_count += 1
                    
                    conn.commit()
                    
                    QMessageBox.information(self, '删除成功', f'成功删除 {deleted_count} 条病史筛查记录')
                    
                    # 取消全选
                    if hasattr(self, 'medical_select_all_checkbox'):
                        self.medical_select_all_checkbox.setChecked(False)
                    
                    self.load_medical_screening_data()
                    
                finally:
                    if conn:
                        conn.close()
                        
        except Exception as e:
            QMessageBox.critical(self, '批量删除失败', f'批量删除病史筛查记录时发生错误：{str(e)}')
    
    def toggle_medical_select_all(self, state):
        """全�?取消全选病史筛�?""
        is_checked = (state == Qt.CheckState.Checked.value)
        for row in range(self.medical_screening_table.rowCount()):
            checkbox_widget = self.medical_screening_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(is_checked)
    
    def view_medical_screening_details(self, medical_id):
        """查看病史筛查详情"""
        try:
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, youth_id_card, name, gender, id_card, 
                       screening_result, screening_date, physical_status, mental_status
                FROM medical_screening 
                WHERE id = ?
            ''', (medical_id,))
            
            record = cursor.fetchone()
            conn.close()
            
            if record:
                from ui.medical_screening_detail_dialog import MedicalScreeningDetailDialog
                dialog = MedicalScreeningDetailDialog(self.db_manager, record, self)
                dialog.data_updated.connect(self.load_medical_screening_data)
                dialog.exec_()
            else:
                QMessageBox.warning(self, '错误', '未找到该记录')
                
        except Exception as e:
            QMessageBox.warning(self, '错误', f'查看详情时发生错�? {str(e)}')
    
    def import_medical_screening_data(self):
        """导入病史筛查数据"""
        file_path, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '', 'Excel Files (*.xlsx *.xls)')
        if not file_path:
            return
        
        count, error = self.import_service.import_medical_screening(file_path)
        if error:
            if count > 0:
                QMessageBox.warning(self, '导入完成（有错误�?, error)
                self.load_medical_screening_data()
            else:
                QMessageBox.warning(self, '导入失败', error)
        else:
            QMessageBox.information(self, '导入成功', f'成功导入 {count} 条病史筛查记�?)
            self.load_medical_screening_data()
            # 额外刷新所有已打开的青年详情对话框（如果有的话），确保详情页也能看到新导入的数�?
            try:
                from PyQt5.QtWidgets import QApplication
                from ui.youth_detail_dialog import YouthDetailDialog

                for widget in QApplication.topLevelWidgets():
                    if isinstance(widget, YouthDetailDialog):
                        if hasattr(widget, 'load_medical_screening_data'):
                            try:
                                widget.load_medical_screening_data()
                            except Exception:
                                # 忽略单个对话框的刷新错误
                                pass
            except Exception:
                pass
    
    def export_medical_screening_data(self):
        """导出病史筛查数据 - 勾选记录时导出选中的，不勾选时导出全部"""
        try:
            # 检查是否有选中的记�?
            selected_ids = []
            selected_names = []
            
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, name FROM medical_screening ORDER BY id
            """)
            all_results = cursor.fetchall()
            conn.close()
            
            for row in range(self.medical_screening_table.rowCount()):
                checkbox_widget = self.medical_screening_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_results):
                            data = all_results[row]
                            medical_id = data['id']
                            name = data['name']
                            
                            if medical_id:
                                selected_ids.append(medical_id)
                                selected_names.append(name)
            
            # 根据是否有选中记录决定导出方式
            if selected_ids:
                # 有选中记录，导出选中�?
                names_text = '�?.join(selected_names[:5])
                if len(selected_names) > 5:
                    names_text += f' 等{len(selected_names)}�?
                
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    '导出选中的病史筛查数�?,
                    f'病史筛查数据_选中_{datetime.now().strftime("%Y%m%d")}.xlsx',
                    'Excel Files (*.xlsx)'
                )
                
                if not file_path:
                    return
                
                # 调用导出服务导出选中的记�?
                success, message = self.export_service.export_selected_medical_screening_to_excel(file_path, selected_ids)
                
                if success:
                    QMessageBox.information(self, '导出成功', f'{message}\n\n导出记录：{names_text}')
                    
                    # 取消全�?
                    if hasattr(self, 'medical_select_all_checkbox'):
                        self.medical_select_all_checkbox.setChecked(False)
                else:
                    QMessageBox.warning(self, '导出失败', message)
            else:
                # 没有选中记录，导出全�?
                file_path, _ = QFileDialog.getSaveFileName(
                    self, 
                    '导出全部病史筛查数据', 
                    f'病史筛查数据_全部_{datetime.now().strftime("%Y%m%d")}.xlsx',
                    'Excel Files (*.xlsx)'
                )
                
                if not file_path:
                    return
                
                success, message = self.export_service.export_medical_screening_to_excel(file_path)
                if success:
                    QMessageBox.information(self, '导出成功', message)
                else:
                    QMessageBox.warning(self, '导出失败', message)
                    
        except Exception as e:
            QMessageBox.critical(self, '导出失败', f'导出病史筛查记录时发生错误：{str(e)}')
    
    def import_camp_verification_data(self):
        """导入入营点验情况数据"""
        file_path, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '', 'Excel Files (*.xlsx *.xls)')
        if not file_path:
            return
        
        try:
            from services.camp_verification_import_service import CampVerificationImportService
            
            # 创建导入服务
            import_service = CampVerificationImportService(self.db_manager)
            
            # 执行导入
            success_count, fail_count, error_messages, skipped_records = import_service.import_from_excel(file_path)
            
            # 生成消息
            total = success_count + fail_count + len(skipped_records)
            
            if total == 0:
                # 检查是否是表头问题
                if error_messages:
                    error_detail = '\n'.join(error_messages)
                    QMessageBox.critical(self, '导入失败', f'导入失败！\n\n{error_detail}')
                else:
                    QMessageBox.warning(self, '导入失败', '没有找到有效的数据行')
                return
            
            # 构建导入结果消息
            message = f'导入完成！\n\n成功：{success_count} 条\n失败：{fail_count} 条\n重复：{len(skipped_records)} �?
            
            # 如果有被跳过（重复）的数据，添加提示和详细列�?
            if len(skipped_records) > 0:
                message += '\n\n以下数据已经存在（身份证号重复）：\n'
                for idx, record in enumerate(skipped_records[:10], 1):
                    message += f'{idx}. 姓名: {record["username"]}，身份证: {record["user_id"]}\n'
                
                if len(skipped_records) > 10:
                    message += f'... 还有 {len(skipped_records) - 10} 条重复数�?
                
                QMessageBox.warning(self, '导入完成（有数据重复�?, message)
            elif fail_count == 0:
                # 全部成功，没有重�?
                QMessageBox.information(self, '导入完成', message)
            else:
                # 有失败数�?
                error_text = '\n'.join(error_messages[:10])
                if len(error_messages) > 10:
                    error_text += f'\n... 还有 {len(error_messages) - 10} 条错�?
                
                message += f'\n\n错误详情：\n{error_text}'
                QMessageBox.warning(self, '导入完成（有错误�?, message)
            
        except Exception as e:
            QMessageBox.critical(self, '导入错误', f'导入入营点验情况数据时发生错误：\n{str(e)}')
    
    def load_exception_statistics_data(self):
        """加载异常情况统计数据到表�?""
        try:
            # 重置全选按钮状态（除非是从时间范围变化触发的）
            if hasattr(self, 'exception_select_all_checkbox') and not getattr(self, '_loading_from_time_range', False):
                # 临时断开信号连接，避免触发全选操�?
                self.exception_select_all_checkbox.stateChanged.disconnect()
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
            # 获取当前选择的时间范�?
            time_range = getattr(self, 'exception_time_range_combo', None)
            if time_range:
                selected_range = time_range.currentText()
            else:
                selected_range = '全部'  # 默认�?
            
            # 如果选择"�?，则不加载任何数据，保持当前显示
            if selected_range == '�?:
                return
            
            # 根据时间范围获取异常情况统计数据
            all_records = self.db_manager.get_exception_statistics_by_time_range(selected_range)
            self.display_exception_statistics_records(all_records)
            
        except Exception as e:
            QMessageBox.warning(self, "加载错误", f"加载异常情况统计数据时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'exception_select_all_checkbox'):
                    self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            except:
                pass
    
    def filter_exception_by_date_range(self):
        """根据日期范围筛选异常情况统计数�?""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'exception_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.exception_select_all_checkbox.stateChanged.disconnect()
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
            # 将下拉框选项设置�?�?
            if hasattr(self, 'exception_time_range_combo'):
                # 临时断开信号连接，避免触发下拉框变化事件
                self.exception_time_range_combo.currentTextChanged.disconnect()
                self.exception_time_range_combo.setCurrentText('�?)
                # 重新连接信号
                self.exception_time_range_combo.currentTextChanged.connect(self.on_exception_time_range_changed)
            
            # 获取日期范围
            start_date = self.exception_start_date.date().toString('yyyy-MM-dd')
            end_date = self.exception_end_date.date().toString('yyyy-MM-dd')
            
            # 验证日期范围
            if start_date > end_date:
                QMessageBox.warning(self, "日期错误", "开始日期不能晚于结束日�?)
                return
            
            # 根据日期范围获取数据
            all_records = self.db_manager.get_exception_statistics_by_date_range(start_date, end_date)
            
            # 如果有搜索条件，进行过滤
            name = self.exception_name_input.text().strip()
            id_card = self.exception_id_card_input.text().strip()
            
            if name or id_card:
                filtered_records = []
                for record in all_records:
                    if name and name not in record.name:
                        continue
                    if id_card and id_card not in record.user_id:
                        continue
                    filtered_records.append(record)
                all_records = filtered_records
            
            self.display_exception_statistics_records(all_records)
            
        except Exception as e:
            QMessageBox.warning(self, "筛选错�?, f"按日期范围筛选数据时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'exception_select_all_checkbox'):
                    self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
                if hasattr(self, 'exception_time_range_combo'):
                    self.exception_time_range_combo.currentTextChanged.connect(self.on_exception_time_range_changed)
            except:
                pass

    def on_exception_time_range_changed(self, selected_range):
        """时间范围下拉框变化处�?""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'exception_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.exception_select_all_checkbox.stateChanged.disconnect()
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
            # 如果选择的不�?�?，则清空日期筛选器
            if selected_range != '�? and hasattr(self, 'exception_start_date') and hasattr(self, 'exception_end_date'):
                from PyQt5.QtCore import QDate
                # 重置日期选择器为默认�?
                self.exception_start_date.setDate(QDate.currentDate().addDays(-30))
                self.exception_end_date.setDate(QDate.currentDate())
            
            # 如果选择"�?，不进行任何筛选操作，保持当前显示
            if selected_range == '�?:
                return
            
            # 检查是否有搜索条件
            name = self.exception_name_input.text().strip()
            id_card = self.exception_id_card_input.text().strip()
            
            if name or id_card:
                # 如果有搜索条件，执行搜索（会自动应用新的时间范围�?
                self.search_exception_statistics()
            else:
                # 如果没有搜索条件，直接加载数�?
                # 设置标记，表示是从时间范围变化触发的加载
                self._loading_from_time_range = True
                
                # 重新加载数据
                self.load_exception_statistics_data()
                
                # 重置标记
                self._loading_from_time_range = False
            
        except Exception as e:
            QMessageBox.warning(self, "时间范围切换错误", f"切换时间范围时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'exception_select_all_checkbox'):
                    self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            except:
                pass
            
        except Exception as e:
            QMessageBox.warning(self, "筛选错�?, f"按时间范围筛选数据时发生错误：{str(e)}")
            # 重置标记和确保信号重新连�?
            self._loading_from_time_range = False
            try:
                if hasattr(self, 'exception_select_all_checkbox'):
                    self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            except:
                pass
    
    def search_exception_statistics(self):
        """搜索异常情况统计记录"""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'exception_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.exception_select_all_checkbox.stateChanged.disconnect()
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
            name = self.exception_name_input.text().strip()
            id_card = self.exception_id_card_input.text().strip()
            
            # 获取当前选择的时间范�?
            time_range = getattr(self, 'exception_time_range_combo', None)
            if time_range:
                selected_range = time_range.currentText()
            else:
                selected_range = '全部'  # 默认�?
            
            # 如果选择"�?，则获取所有数�?
            if selected_range == '�?:
                all_records = self.db_manager.get_all_exception_statistics()
            else:
                # 根据时间范围获取数据
                all_records = self.db_manager.get_exception_statistics_by_time_range(selected_range)
            
            # 过滤记录
            filtered_records = []
            for record in all_records:
                if name and name not in record.name:
                    continue
                if id_card and id_card not in record.user_id:
                    continue
                filtered_records.append(record)
            
            self.display_exception_statistics_records(filtered_records)
            
        except Exception as e:
            QMessageBox.warning(self, "搜索错误", f"搜索异常情况统计记录时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'exception_select_all_checkbox'):
                    self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            except:
                pass
    
    def reset_exception_statistics_search(self):
        """重置异常情况统计搜索"""
        try:
            # 重置全选按钮状�?
            if hasattr(self, 'exception_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操�?
                self.exception_select_all_checkbox.stateChanged.disconnect()
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
            self.exception_name_input.clear()
            self.exception_id_card_input.clear()
            # 重置时间范围为默认�?
            if hasattr(self, 'exception_time_range_combo'):
                self.exception_time_range_combo.setCurrentText('全部')
            self.load_exception_statistics_data()
            
        except Exception as e:
            QMessageBox.warning(self, "重置错误", f"重置搜索时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'exception_select_all_checkbox'):
                    self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            except:
                pass
    
    def display_exception_statistics_records(self, records):
        """显示异常情况统计记录"""
        try:
            self.exception_statistics_table.setRowCount(len(records))
            
            for row, record in enumerate(records):
                # 选择�?
                checkbox = QCheckBox()
                checkbox.setStyleSheet("""
                    QCheckBox {
                        spacing: 0px;
                    }
                    QCheckBox::indicator {
                        width: 20px;
                        height: 20px;
                        border: 2px solid #7F8C8D;
                        border-radius: 2px;
                        background-color: white;
                    }
                    QCheckBox::indicator:hover {
                        border-color: #3498DB;
                        background-color: #ECF0F1;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #3498DB;
                        border-color: #2980B9;
                    }
                """)
                checkbox.setMinimumSize(24, 24)  # 确保复选框有足够空间显�?
                
                # 添加状态变化监听，当单个复选框状态改变时更新全选按钮状�?
                checkbox.stateChanged.connect(self.update_exception_select_all_state)
                
                checkbox_layout = QHBoxLayout()
                checkbox_layout.setAlignment(Qt.AlignCenter)
                checkbox_layout.setContentsMargins(5, 5, 5, 5)
                checkbox_layout.addWidget(checkbox)
                checkbox_container = QWidget()
                checkbox_container.setLayout(checkbox_layout)
                self.exception_statistics_table.setCellWidget(row, 0, checkbox_container)
                
                # 日期
                self.exception_statistics_table.setItem(row, 1, QTableWidgetItem(str(record.data or '')))
                
                # 姓名
                self.exception_statistics_table.setItem(row, 2, QTableWidgetItem(str(record.name or '')))
                
                # 性别
                self.exception_statistics_table.setItem(row, 3, QTableWidgetItem(str(record.gender or '')))
                
                # 公民身份号码
                self.exception_statistics_table.setItem(row, 4, QTableWidgetItem(str(record.user_id or '')))
                
                # 思想 (0=正常, 1=异常)
                thought_text = '异常' if record.thought else '正常'
                self.exception_statistics_table.setItem(row, 5, QTableWidgetItem(thought_text))
                
                # 身体 (0=正常, 1=异常)
                body_text = '异常' if record.body else '正常'
                self.exception_statistics_table.setItem(row, 6, QTableWidgetItem(body_text))
                
                # 精神 (0=正常, 1=异常)
                spirit_text = '异常' if record.Spirit else '正常'
                self.exception_statistics_table.setItem(row, 7, QTableWidgetItem(spirit_text))
                
                # 异常来源
                source_text = str(record.other or '')  # other字段现在存储异常来源文本
                self.exception_statistics_table.setItem(row, 8, QTableWidgetItem(source_text))
                
                # 查看详情按钮
                view_btn = QPushButton('查看详情')
                view_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝�?
                view_btn.clicked.connect(lambda checked, user_id=record.user_id, name=record.name: 
                                       self.show_exception_statistics_detail(user_id, name))
                self.exception_statistics_table.setCellWidget(row, 9, view_btn)
                
        except Exception as e:
            QMessageBox.warning(self, "显示错误", f"显示异常情况统计记录时发生错误：{str(e)}")
    
    def on_exception_statistics_double_click(self, row, column):
        """异常统计表格双击事件处理"""
        try:
            # 获取当前行的用户ID和姓�?
            user_id_item = self.exception_statistics_table.item(row, 4)  # 公民身份号码�?
            name_item = self.exception_statistics_table.item(row, 2)     # 姓名�?
            
            if user_id_item and name_item:
                user_id = user_id_item.text()
                name = name_item.text()
                self.show_exception_statistics_detail(user_id, name)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"打开详情时发生错误：{str(e)}")
    
    def show_exception_statistics_detail(self, user_id, name):
        """显示异常统计详情弹窗"""
        try:
            from .exception_statistics_detail_dialog import ExceptionStatisticsDetailDialog
            dialog = ExceptionStatisticsDetailDialog(self.db_manager, self, user_id, name)
            dialog.exec_()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"打开异常统计详情时发生错误：{str(e)}")
    
    def toggle_exception_select_all(self, state):
        """异常情况统计全�?取消全�?""
        is_checked = (state == Qt.CheckState.Checked.value)
        
        # 临时标记，防止在全选操作时触发单个复选框的状态更�?
        self._updating_exception_select_all = True
        
        for row in range(self.exception_statistics_table.rowCount()):
            checkbox_widget = self.exception_statistics_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(is_checked)
        
        # 重置标记
        self._updating_exception_select_all = False
    
    def update_exception_select_all_state(self):
        """更新异常情况统计全选按钮状�?""
        # 如果正在执行全选操作，则不更新全选按钮状�?
        if getattr(self, '_updating_exception_select_all', False):
            return
        
        try:
            total_rows = self.exception_statistics_table.rowCount()
            if total_rows == 0:
                return
            
            checked_count = 0
            for row in range(total_rows):
                checkbox_widget = self.exception_statistics_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        checked_count += 1
            
            # 临时断开全选按钮的信号连接，避免触发toggle_exception_select_all
            self.exception_select_all_checkbox.stateChanged.disconnect()
            
            # 更新全选按钮状�?
            if checked_count == 0:
                # 没有选中任何�?
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            elif checked_count == total_rows:
                # 全部选中
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Checked)
            else:
                # 部分选中
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.PartiallyChecked)
            
            # 重新连接信号
            self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
        except Exception as e:
            print(f"更新异常统计全选按钮状态时出错: {e}")
            # 确保信号重新连接
            try:
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            except:
                pass
    
    def view_exception_statistics(self, record_id):
        """查看异常情况统计记录详情"""
        try:
            # 获取记录详情
            all_records = self.db_manager.get_all_exception_statistics()
            record = None
            for r in all_records:
                if r.id == record_id:
                    record = r
                    break
            
            if not record:
                QMessageBox.warning(self, '错误', '找不到该记录')
                return
            
            # 使用新的详情弹窗
            self.show_exception_statistics_detail(record.user_id, record.name)
            
        except Exception as e:
            QMessageBox.critical(self, '查看失败', f'查看异常情况统计记录时发生错误：{str(e)}')
    
    def delete_single_exception_statistics(self, record_id, name):
        """删除单个异常情况统计记录"""
        try:
            reply = QMessageBox.question(
                self,
                '确认删除',
                f'确定要删�?{name} 的异常情况统计记录吗？\n删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # 根据record_id找到对应的user_id
                all_records = self.db_manager.get_all_exception_statistics()
                for record in all_records:
                    if record.id == record_id:
                        # 删除该用户的异常情况统计
                        self.db_manager.delete_exception_statistics(record.user_id)
                        QMessageBox.information(self, '删除成功', '记录已删�?)
                        self.load_exception_statistics_data()
                        return
                
                QMessageBox.warning(self, '删除失败', '未能找到记录')
                    
        except Exception as e:
            QMessageBox.critical(self, '删除失败', f'删除异常情况统计记录时发生错误：{str(e)}')
    


    def import_town_interview_files(self):
        """导入镇街谈心谈话文件"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QDateEdit, QPushButton, QFileDialog, QMessageBox, QProgressBar
            from PyQt5.QtCore import QDate, QThread, pyqtSignal
            import os
            from datetime import datetime
            
            # 创建导入对话框
            dialog = QDialog(self)
            dialog.setWindowTitle('导入镇街谈心谈话文件')
            dialog.setModal(True)
            dialog.resize(400, 200)
            
            layout = QVBoxLayout()
            
            # 日期选择
            date_layout = QHBoxLayout()
            date_layout.addWidget(QLabel('选择日期:'))
            date_edit = QDateEdit()
            date_edit.setDate(QDate.currentDate())
            date_edit.setCalendarPopup(True)
            date_layout.addWidget(date_edit)
            layout.addLayout(date_layout)
            
            # 文件选择
            file_layout = QHBoxLayout()
            file_layout.addWidget(QLabel('选择文件:'))
            file_btn = QPushButton('选择文件')
            selected_files = []
            
            def select_files():
                files, _ = QFileDialog.getOpenFileNames(
                    dialog, '选择文件', '',
                    '支持的文件 (*.pdf *.jpg *.jpeg *.png *.bmp *.gif *.tiff);;PDF文件 (*.pdf);;图片文件 (*.jpg *.jpeg *.png *.bmp *.gif *.tiff);;所有文件 (*.*)'
                )
                selected_files.clear()
                selected_files.extend(files)
                file_btn.setText(f'已选择 {len(files)} 个文件')
            
            file_btn.clicked.connect(select_files)
            file_layout.addWidget(file_btn)
            layout.addLayout(file_layout)
            
            # 按钮
            button_layout = QHBoxLayout()
            import_btn = QPushButton('开始导入')
            cancel_btn = QPushButton('取消')
            
            def start_import():
                if not selected_files:
                    QMessageBox.warning(dialog, '警告', '请先选择文件')
                    return
                
                interview_date = date_edit.date().toString('yyyy-MM-dd')
                self.process_interview_files(selected_files, interview_date, 'town')
                dialog.accept()
            
            import_btn.clicked.connect(start_import)
            cancel_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(import_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导入文件时发生错误: {str(e)}')
    
    def import_leader_interview_files(self):
        """导入领导谈心谈话文件"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QDateEdit, QPushButton, QFileDialog, QMessageBox, QProgressBar
            from PyQt5.QtCore import QDate, QThread, pyqtSignal
            import os
            from datetime import datetime
            
            # 创建导入对话框
            dialog = QDialog(self)
            dialog.setWindowTitle('导入领导谈心谈话文件')
            dialog.setModal(True)
            dialog.resize(400, 200)
            
            layout = QVBoxLayout()
            
            # 日期选择
            date_layout = QHBoxLayout()
            date_layout.addWidget(QLabel('选择日期:'))
            date_edit = QDateEdit()
            date_edit.setDate(QDate.currentDate())
            date_edit.setCalendarPopup(True)
            date_layout.addWidget(date_edit)
            layout.addLayout(date_layout)
            
            # 文件选择
            file_layout = QHBoxLayout()
            file_layout.addWidget(QLabel('选择文件:'))
            file_btn = QPushButton('选择文件')
            selected_files = []
            
            def select_files():
                files, _ = QFileDialog.getOpenFileNames(
                    dialog, '选择文件', '',
                    '支持的文件 (*.pdf *.jpg *.jpeg *.png *.bmp *.gif *.tiff);;PDF文件 (*.pdf);;图片文件 (*.jpg *.jpeg *.png *.bmp *.gif *.tiff);;所有文件 (*.*)'
                )
                selected_files.clear()
                selected_files.extend(files)
                file_btn.setText(f'已选择 {len(files)} 个文件')
            
            file_btn.clicked.connect(select_files)
            file_layout.addWidget(file_btn)
            layout.addLayout(file_layout)
            
            # 按钮
            button_layout = QHBoxLayout()
            import_btn = QPushButton('开始导入')
            cancel_btn = QPushButton('取消')
            
            def start_import():
                if not selected_files:
                    QMessageBox.warning(dialog, '警告', '请先选择文件')
                    return
                
                interview_date = date_edit.date().toString('yyyy-MM-dd')
                self.process_interview_files(selected_files, interview_date, 'leader')
                dialog.accept()
            
            import_btn.clicked.connect(start_import)
            cancel_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(import_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导入文件时发生错误: {str(e)}')
    
    def process_interview_files(self, file_paths, interview_date, interview_type):
        """处理导入的谈心谈话文件"""
        try:
            import os
            from PyQt5.QtWidgets import QProgressDialog, QMessageBox
            from PyQt5.QtCore import Qt
            
            # 创建进度对话框
            progress = QProgressDialog('正在处理文件...', '取消', 0, len(file_paths), self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.show()
            
            success_count = 0
            failed_files = []
            
            for i, file_path in enumerate(file_paths):
                if progress.wasCanceled():
                    break
                
                progress.setValue(i)
                progress.setLabelText(f'正在处理: {os.path.basename(file_path)}')
                
                try:
                    # 从文件名提取身份证号
                    filename = os.path.basename(file_path)
                    name_without_ext = os.path.splitext(filename)[0]
                    
                    if '-' not in name_without_ext:
                        failed_files.append(f'{filename}: 文件名格式错误（应为：姓名-身份证号）')
                        continue
                    
                    parts = name_without_ext.split('-')
                    if len(parts) < 2:
                        failed_files.append(f'{filename}: 文件名格式错误（应为：姓名-身份证号）')
                        continue
                    
                    youth_name = parts[0]
                    id_card = parts[-1]  # 取最后一部分作为身份证号
                    
                    # 验证身份证号是否存在于基本信息表
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    cursor.execute('SELECT name, gender FROM youth WHERE id_card = ?', (id_card,))
                    youth_info = cursor.fetchone()
                    conn.close()
                    
                    if not youth_info:
                        failed_files.append(f'{filename}: 身份证号 {id_card} 在基本信息表中不存在')
                        continue
                    
                    db_name, gender = youth_info
                    
                    # 读取文件数据
                    file_ext = os.path.splitext(filename)[1].lower()
                    if file_ext == '.pdf':
                        # 转换PDF为JPG
                        image_data = self.convert_pdf_to_jpg(file_path)
                        if not image_data:
                            failed_files.append(f'{filename}: PDF转换失败')
                            continue
                    else:
                        # 读取图片文件
                        with open(file_path, 'rb') as f:
                            image_data = f.read()
                    
                    # 插入数据库
                    if interview_type == 'town':
                        self.db_manager.insert_town_interview(
                            id_card, db_name, gender, interview_date, image_data, '正常', '正常'
                        )
                    else:  # leader
                        self.db_manager.insert_leader_interview(
                            id_card, db_name, gender, interview_date, image_data, '正常', '正常'
                        )
                    
                    success_count += 1
                    
                except Exception as e:
                    failed_files.append(f'{filename}: {str(e)}')
            
            progress.setValue(len(file_paths))
            progress.close()
            
            # 刷新数据
            if interview_type == 'town':
                self.load_town_interview_data()
            else:
                self.load_leader_interview_data()
            
            # 显示结果
            result_msg = f'导入完成！\n成功导入: {success_count} 个文件'
            if failed_files:
                result_msg += f'\n失败: {len(failed_files)} 个文件\n\n失败详情:\n' + '\n'.join(failed_files[:10])
                if len(failed_files) > 10:
                    result_msg += f'\n... 还有 {len(failed_files) - 10} 个失败文件'
            
            QMessageBox.information(self, '导入结果', result_msg)
            
        except Exception as e:
            QMessageBox.critical(self, '错误', f'处理文件时发生错误: {str(e)}')
    
    def convert_pdf_to_jpg(self, pdf_path):
        """将PDF文件转换为JPG图片"""
        try:
            # 尝试导入pdf2image库
            try:
                from pdf2image import convert_from_path
            except ImportError:
                return None
            
            # 转换PDF的第一页为图片
            images = convert_from_path(pdf_path, first_page=1, last_page=1, dpi=200)
            
            if not images:
                return None
            
            # 将PIL图片转换为字节数据
            import io
            img_byte_arr = io.BytesIO()
            images[0].save(img_byte_arr, format='JPEG', quality=85)
            img_byte_arr = img_byte_arr.getvalue()
            
            return img_byte_arr
            
        except Exception as e:
            print(f"PDF转换错误: {e}")
            return None