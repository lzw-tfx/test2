"""
主界面
"""
import sqlite3
from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QLineEdit, QPushButton, QTableWidget,
                             QTableWidgetItem, QTabWidget, QMessageBox,
                             QFileDialog, QComboBox, QCheckBox, QDateEdit, QHeaderView)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor
from datetime import datetime


class MainWindow(QMainWindow):
    def __init__(self, db_manager, import_service, export_service, user):
        super().__init__()
        self.db_manager = db_manager
        self.import_service = import_service
        self.export_service = export_service
        self.user = user
        self.current_youth_id = None
        
        # 创建谈心谈话基类实例
        from ui.interview_base import InterviewBase
        self.town_interview_base = InterviewBase(self, 'town')
        self.leader_interview_base = InterviewBase(self, 'leader')
        
        self.init_ui()
    
    def setup_table_style(self, table):
        """设置表格的统一样式，包括字体大小和行高"""
        # 设置表格字体
        table_font = QFont()
        table_font.setPointSize(11)  # 设置字体大小为11pt
        table.setFont(table_font)
        
        # 设置表头字体
        header_font = QFont()
        header_font.setPointSize(11)  # 表头字体大小
        header_font.setBold(True)  # 表头加粗
        table.horizontalHeader().setFont(header_font)
        
        # 设置默认行高（根据字体自适应）
        table.verticalHeader().setDefaultSectionSize(40)  # 设置行高为40像素
        
        # 设置表格样式
        table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
                background-color: white;
                selection-background-color: #e3f2fd;
                selection-color: black;
                font-size: 11pt;
            }
            QTableWidget::item {
                padding: 5px;
                color: black;
            }
            QTableWidget::item:selected {
                color: black;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 8px;
                border: 1px solid #d0d0d0;
                font-weight: bold;
                font-size: 11pt;
            }
        """)
    
    def setup_button_style(self, button, button_type='normal'):
        """设置按钮的统一样式
        
        Args:
            button: QPushButton对象
            button_type: 按钮类型 ('normal', 'primary', 'danger', 'success')
        """
        # 设置按钮字体
        button_font = QFont()
        button_font.setPointSize(10)  # 按钮字体大小10pt
        button.setFont(button_font)
        
        # 设置按钮最小尺寸
        button.setMinimumHeight(32)  # 最小高度32像素
        button.setMinimumWidth(80)   # 最小宽度80像素
        
        # 根据按钮类型设置不同的样式
        if button_type == 'primary':
            # 主要按钮（蓝色）
            button.setStyleSheet("""
                QPushButton {
                    background-color: #3498DB;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 16px;
                    font-size: 10pt;
                    min-height: 32px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #2980B9;
                }
                QPushButton:pressed {
                    background-color: #21618C;
                }
                QPushButton:disabled {
                    background-color: #BDC3C7;
                    color: #7F8C8D;
                }
            """)
        elif button_type == 'danger':
            # 危险按钮（红色）
            button.setStyleSheet("""
                QPushButton {
                    background-color: #E74C3C;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 16px;
                    font-size: 10pt;
                    min-height: 32px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #C0392B;
                }
                QPushButton:pressed {
                    background-color: #A93226;
                }
                QPushButton:disabled {
                    background-color: #BDC3C7;
                    color: #7F8C8D;
                }
            """)
        elif button_type == 'success':
            # 成功按钮（绿色）
            button.setStyleSheet("""
                QPushButton {
                    background-color: #27AE60;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 16px;
                    font-size: 10pt;
                    min-height: 32px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
                QPushButton:pressed {
                    background-color: #1E8449;
                }
                QPushButton:disabled {
                    background-color: #BDC3C7;
                    color: #7F8C8D;
                }
            """)
        else:
            # 普通按钮（灰色边框）
            button.setStyleSheet("""
                QPushButton {
                    background-color: white;
                    color: #2C3E50;
                    border: 1px solid #BDC3C7;
                    border-radius: 4px;
                    padding: 6px 16px;
                    font-size: 10pt;
                    min-height: 32px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #ECF0F1;
                    border-color: #95A5A6;
                }
                QPushButton:pressed {
                    background-color: #D5DBDB;
                }
                QPushButton:disabled {
                    background-color: #F8F9F9;
                    color: #BDC3C7;
                    border-color: #D5DBDB;
                }
            """)
    
    def init_ui(self):
        self.setWindowTitle(f'一人一策记录本 - {self.user.unit} - {self.user.username}')
        
        # 默认全屏显示
        self.showMaximized()
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局 - 水平布局
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # 左侧导航栏
        nav_widget = self.create_navigation()
        main_layout.addWidget(nav_widget)
        
        # 右侧内容区域
        content_widget = QWidget()
        content_layout = QVBoxLayout()
        content_layout.setContentsMargins(0, 0, 0, 0)
        
        # 标签页内容（不显示标签）
        self.tabs = QTabWidget()
        self.tabs.tabBar().hide()  # 隐藏标签栏
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background: white;
            }
        """)
        
        self.create_tabs()
        content_layout.addWidget(self.tabs)
        
        # 底部操作按钮
        button_layout = self.create_bottom_buttons()
        content_layout.addLayout(button_layout)
        
        content_widget.setLayout(content_layout)
        main_layout.addWidget(content_widget)
        
        central_widget.setLayout(main_layout)
        
        # 延迟加载数据，只在切换到对应标签页时才加载
        # 初始化时不加载数据，提高启动速度
        
        # 但是需要立即加载基本信息，因为这是默认显示的标签页
        self._loaded_tabs = set()
        self.load_all_youth_detailed()
        self._loaded_tabs.add('基本信息')
    
    def create_navigation(self):
        """创建左侧导航栏"""
        nav_widget = QWidget()
        nav_widget.setFixedWidth(240)
        nav_widget.setStyleSheet("""
            QWidget {
                background: #34495e;
            }
        """)
        
        nav_layout = QVBoxLayout()
        nav_layout.setContentsMargins(0, 0, 0, 0)
        nav_layout.setSpacing(0)
        
        # 顶部标题
        title_label = QLabel('一人一策系统')
        title_label.setStyleSheet("""
            QLabel {
                background: #2c3e50;
                color: white;
                font-size: 16px;
                font-weight: bold;
                padding: 20px;
                border-bottom: 2px solid #1abc9c;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        nav_layout.addWidget(title_label)
        
        # 时间筛选区域
        filter_widget = QWidget()
        filter_widget.setStyleSheet("""
            QWidget {
                background: #2c3e50;
                padding: 10px;
            }
            QLabel {
                color: white;
                font-size: 14px;
                font-weight: bold;
                margin: 5px 0;
            }
            QComboBox {
                background-color: white;
                color: black;
                font-size: 13px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 4px 8px;
                min-height: 20px;
            }
            QComboBox:hover {
                border-color: #3498db;
            }
            QComboBox:focus {
                border-color: #3498db;
                outline: none;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                color: black;
                font-size: 13px;
                selection-background-color: #3498db;
                selection-color: white;
                border: 1px solid #bdc3c7;
            }
            QComboBox QAbstractItemView::item {
                padding: 6px;
                color: black;
                background-color: white;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #3498db;
                color: white;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #ecf0f1;
                color: black;
            }
        """)
        
        filter_layout = QVBoxLayout()
        filter_layout.setContentsMargins(10, 10, 10, 10)
        filter_layout.setSpacing(8)
        
        # 年份和半年选择 - 分两行显示，每行包含标签和下拉框
        
        # 年份选择行
        year_row_layout = QHBoxLayout()
        year_row_layout.setSpacing(12)  # 增加间距
        year_row_layout.setContentsMargins(0, 0, 0, 0)
        
        year_label = QLabel('年份：')
        year_label.setFixedWidth(60)  # 增加标签宽度
        year_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        year_row_layout.addWidget(year_label)
        
        self.year_combo = QComboBox()
        # 添加"全部"选项，然后是今年到之前的年份，总共10年，按时间倒序排列
        current_year = datetime.now().year
        years = ['全部'] + [str(year) for year in range(current_year, current_year - 10, -1)]
        self.year_combo.addItems(years)
        self.year_combo.setCurrentText(str(current_year))
        self.year_combo.setFixedWidth(100)  # 固定宽度确保显示完整
        self.year_combo.currentTextChanged.connect(self.on_time_filter_changed)
        year_row_layout.addWidget(self.year_combo)
        
        year_row_layout.addStretch()  # 添加弹性空间
        filter_layout.addLayout(year_row_layout)
        
        # 半年选择行
        half_year_row_layout = QHBoxLayout()
        half_year_row_layout.setSpacing(12)  # 增加间距
        half_year_row_layout.setContentsMargins(0, 0, 0, 0)
        
        half_year_label = QLabel('半年：')
        half_year_label.setFixedWidth(60)  # 增加标签宽度
        half_year_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        half_year_row_layout.addWidget(half_year_label)
        
        self.half_year_combo = QComboBox()
        self.half_year_combo.addItems(['上半年', '下半年'])
        # 设置默认值为当前半年
        current_month = datetime.now().month
        if current_month <= 6:
            self.half_year_combo.setCurrentText('上半年')
        else:
            self.half_year_combo.setCurrentText('下半年')
        self.half_year_combo.setFixedWidth(100)  # 固定宽度确保显示完整
        self.half_year_combo.currentTextChanged.connect(self.on_time_filter_changed)
        half_year_row_layout.addWidget(self.half_year_combo)
        
        half_year_row_layout.addStretch()  # 添加弹性空间
        filter_layout.addLayout(half_year_row_layout)
        
        filter_widget.setLayout(filter_layout)
        nav_layout.addWidget(filter_widget)
        
        # 导航按钮
        self.nav_buttons = []
        nav_items = [
            '基本信息',
            '病史筛查情况',
            '镇街谈心谈话情况',
            '领导谈心谈话情况',
            '每日情况统计',
            '异常情况统计',
            '体检情况统计表'
        ]
        
        for index, item in enumerate(nav_items):
            btn = QPushButton(item)
            # 所有按钮使用统一样式
            btn.setStyleSheet("""
                QPushButton {
                    background: #34495e;
                    color: #ecf0f1;
                    border: none;
                    border-bottom: 1px solid #2c3e50;
                    padding: 18px 15px;
                    text-align: left;
                    font-size: 15px;
                }
                QPushButton:hover {
                    background: #3d566e;
                    color: white;
                }
                QPushButton:checked {
                    background: #3498db;
                    color: white;
                    font-weight: bold;
                    border-left: 4px solid #e74c3c;
                }
            """)
            btn.setCheckable(True)
            btn.clicked.connect(lambda checked, idx=index: self.switch_tab(idx))
            self.nav_buttons.append(btn)
            nav_layout.addWidget(btn)
        
        # 默认选中第一个
        self.nav_buttons[0].setChecked(True)
        
        # 添加弹性空间
        nav_layout.addStretch()
        
        nav_widget.setLayout(nav_layout)
        return nav_widget
    
    def switch_tab(self, index):
        """切换标签页"""
        # 取消其他按钮的选中状态
        for i, btn in enumerate(self.nav_buttons):
            btn.setChecked(i == index)
        
        # 切换到对应的标签页
        self.tabs.setCurrentIndex(index)
        
        # 延迟加载数据 - 只在首次切换到标签页时加载数据
        tab_names = [
            '基本信息',
            '病史筛查情况', 
            '镇街谈心谈话情况',
            '领导谈心谈话情况',
            '每日情况统计',
            '异常情况统计',
            '体检情况统计表'
        ]
        
        if index < len(tab_names):
            tab_name = tab_names[index]
            
            # 检查是否已经加载过数据
            if not hasattr(self, '_loaded_tabs'):
                self._loaded_tabs = set()
            
            if tab_name not in self._loaded_tabs:
                # 首次加载该标签页的数据
                if index == 0:  # 基本信息
                    self.load_all_youth_detailed()
                elif index == 2:  # 镇街谈心谈话情况
                    self.load_town_interview_data()
                elif index == 3:  # 领导谈心谈话情况
                    self.load_leader_interview_data()
                elif index == 4:  # 每日情况统计
                    self.load_daily_stats_data()
                elif index == 5:  # 异常情况统计
                    self.load_exception_statistics_data()
                elif index == 6:  # 体检情况统计表
                    self.load_physical_examination_data()
                
                # 标记为已加载
                self._loaded_tabs.add(tab_name)
            elif index == 5:  # 异常情况统计每次都刷新
                self.load_exception_statistics_data()
    
    def on_time_filter_changed(self):
        """时间筛选条件改变时的处理"""
        # 重新加载当前标签页的数据
        self.refresh_current_tab_data()
    
    def refresh_current_tab_data(self):
        """刷新当前标签页的数据"""
        current_index = self.tabs.currentIndex()
        if current_index == 0:  # 基本信息
            self.load_all_youth_detailed()
        elif current_index == 1:  # 病史筛查情况
            self.load_medical_screening_data()
        elif current_index == 2:  # 镇街谈心谈话情况
            self.load_town_interview_data()
        elif current_index == 3:  # 领导谈心谈话情况
            self.load_leader_interview_data()
        elif current_index == 4:  # 每日情况统计
            self.load_daily_stats_data()
        elif current_index == 5:  # 异常情况统计
            self.load_exception_statistics_data()
        elif current_index == 6:  # 体检情况统计表
            self.load_physical_examination_data()
    
    def get_time_filter_condition(self, date_field):
        """获取时间筛选条件的SQL语句片段
        
        Args:
            date_field: 要筛选的日期字段名
            
        Returns:
            tuple: (where_condition, params) SQL条件和参数
        """
        year = self.year_combo.currentText()
        
        # 如果选择了"全部"，则不添加时间筛选条件
        if year == '全部':
            return "", []
        
        half_year = self.half_year_combo.currentText()
        
        if half_year == '上半年':
            start_date = f"{year}-01-01"
            end_date = f"{year}-06-30"
        else:  # 下半年
            start_date = f"{year}-07-01"
            end_date = f"{year}-12-31"
        
        condition = f"({date_field} >= ? AND {date_field} <= ?)"
        params = [start_date, end_date]
        
        return condition, params
    
    def create_search_bar(self):
        layout = QHBoxLayout()
        
        self.name_input = QLineEdit()
        self.name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.name_input.setPlaceholderText('姓名')
        self.name_input.returnPressed.connect(self.search_youth)
        
        self.id_card_input = QLineEdit()
        self.id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.id_card_input.setPlaceholderText('身份证号')
        self.id_card_input.returnPressed.connect(self.search_youth)
        
        # 合并为应征地搜索框
        self.recruitment_place_input = QLineEdit()
        self.recruitment_place_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.recruitment_place_input.setPlaceholderText('应征地')
        self.recruitment_place_input.returnPressed.connect(self.search_youth)
        
        # 新增三个搜索框：连、排、班
        self.company_input = QLineEdit()
        self.company_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.company_input.setPlaceholderText('连')
        self.company_input.returnPressed.connect(self.search_youth)
        
        self.platoon_input = QLineEdit()
        self.platoon_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.platoon_input.setPlaceholderText('排')
        self.platoon_input.returnPressed.connect(self.search_youth)
        
        self.squad_input = QLineEdit()
        self.squad_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.squad_input.setPlaceholderText('班')
        self.squad_input.returnPressed.connect(self.search_youth)
        
        search_btn = QPushButton('搜索')
        search_btn.clicked.connect(self.search_youth)
        self.setup_button_style(search_btn, 'primary')
        
        reset_btn = QPushButton('重置')
        reset_btn.clicked.connect(self.reset_search)
        self.setup_button_style(reset_btn, 'normal')
        
        layout.addWidget(QLabel('搜索:'))
        layout.addWidget(self.name_input)
        layout.addWidget(self.id_card_input)
        layout.addWidget(self.recruitment_place_input)
        layout.addWidget(self.company_input)
        layout.addWidget(self.platoon_input)
        layout.addWidget(self.squad_input)
        layout.addWidget(search_btn)
        layout.addWidget(reset_btn)
        
        return layout
    
    def create_tabs(self):
        """创建7个板块标签页"""
        tab_names = [
            '基本信息',
            '病史筛查情况',
            '镇街谈心谈话情况',
            '领导谈心谈话情况',
            '每日情况统计',
            '异常情况统计',
            '体检情况统计表'
        ]
        
        self.tab_tables = {}
        
        for tab_name in tab_names:
            if tab_name == '基本信息':
                # 基本信息标签页特殊处理，包含搜索和列表
                tab_widget = self.create_basic_info_tab()
            elif tab_name == '病史筛查情况':
                # 病史筛查情况标签页特殊处理
                tab_widget = self.create_medical_screening_tab()
            elif tab_name == '镇街谈心谈话情况':
                # 镇街谈心谈话情况标签页特殊处理
                tab_widget = self.create_town_interview_tab()
            elif tab_name == '领导谈心谈话情况':
                # 领导谈心谈话情况标签页特殊处理
                tab_widget = self.create_leader_interview_tab()
            elif tab_name == '异常情况统计':
                # 异常情况统计标签页特殊处理
                tab_widget = self.create_exception_statistics_tab()
            elif tab_name == '每日情况统计':
                # 每日情况统计标签页特殊处理
                tab_widget = self.create_daily_stats_tab()
            elif tab_name == '体检情况统计表':
                # 体检情况统计表标签页特殊处理
                tab_widget = self.create_physical_examination_tab()
            else:
                tab_widget = QWidget()
                tab_layout = QVBoxLayout()
                
                # 创建表格
                table = QTableWidget()
                table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # 禁止直接编辑
                self.tab_tables[tab_name] = table
                
                # 应用统一的表格样式
                self.setup_table_style(table)
                
                tab_layout.addWidget(table)
                
                # 添加操作按钮
                button_layout = QHBoxLayout()
                
                import_btn = QPushButton(f'导入数据')
                import_btn.clicked.connect(lambda checked, t=tab_name: self.import_data(t))
                self.setup_button_style(import_btn, 'normal')
                button_layout.addWidget(import_btn)
                
                # 导出按钮
                export_btn = QPushButton(f'导出数据')
                export_btn.clicked.connect(lambda checked, t=tab_name: self.export_data(t))
                self.setup_button_style(export_btn, 'normal')
                button_layout.addWidget(export_btn)
                
                button_layout.addStretch()
                tab_layout.addLayout(button_layout)
                
                tab_widget.setLayout(tab_layout)
            
            self.tabs.addTab(tab_widget, tab_name)
    
    def create_basic_info_tab(self):
        """创建基本信息标签页"""
        tab_widget = QWidget()
        layout = QVBoxLayout()
        
        # 搜索栏
        search_layout = self.create_search_bar()
        layout.addLayout(search_layout)
        
        # 全选按钮行
        select_all_layout = QHBoxLayout()
        self.select_all_checkbox = QCheckBox('全选')
        self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
        self.select_all_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #7F8C8D;
                border-radius: 2px;
                background-color: white;
            }
            QCheckBox::indicator:hover {
                border-color: #3498DB;
                background-color: #ECF0F1;
            }
            QCheckBox::indicator:checked {
                background-color: #3498DB;
                border-color: #2980B9;
            }
        """)
        select_all_layout.addWidget(self.select_all_checkbox)
        select_all_layout.addStretch()
        layout.addLayout(select_all_layout)
        
        # 成员详细信息表格
        self.search_table = QTableWidget()
        # 设置完整的青年基本情况统计表字段（用于导出）
        self.basic_info_headers_full = [
            '选择', '序号', '姓名', '公民身份号码', '性别', '出生日期', '民族', '政治面貌', '宗教信仰', '籍贯',
            '文化程度', '学业情况', '学习类型', '入营时间', '应征地', '经常居住地址', '户籍所在地', '邮编',
            '本人电话', '家庭电话', '毕业(就读)学校', '所学专业', '入学时间',
            '初检医院', '初检结论', '初检时间', '体检结论', '体检时间', '体检不合格原因', '主检医师意见', '毕业时间',
            '连', '排', '班', '带训班长信息', '在营状态', '离营时间', '离营原因', '操作', '删除'
        ]
        
        # 设置显示用的表头（不包含序号）
        self.basic_info_headers = [
            '选择', '姓名', '公民身份号码', '性别', '出生日期', '民族', '政治面貌', '宗教信仰', '籍贯',
            '文化程度', '学业情况', '学习类型', '入营时间', '应征地', '经常居住地址', '户籍所在地', '邮编',
            '本人电话', '家庭电话', '毕业(就读)学校', '所学专业', '入学时间',
            '初检医院', '初检结论', '初检时间', '体检结论', '体检时间', '体检不合格原因', '主检医师意见', '毕业时间',
            '连', '排', '班', '带训班长信息', '在营状态', '离营时间', '离营原因', '操作', '删除'
        ]
        self.search_table.setColumnCount(len(self.basic_info_headers))
        self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
        self.search_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.search_table.setAlternatingRowColors(True)
        self.search_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # 禁止直接编辑
        self.search_table.horizontalHeader().setStretchLastSection(False)
        self.search_table.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        
        # 应用统一的表格样式
        self.setup_table_style(self.search_table)
        
        # 添加双击事件处理
        self.search_table.cellDoubleClicked.connect(self.on_basic_info_double_click)
        
        layout.addWidget(self.search_table)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        import_btn = QPushButton('导入数据')
        import_btn.clicked.connect(lambda: self.import_data('基本信息'))
        self.setup_button_style(import_btn, 'normal')
        button_layout.addWidget(import_btn)
        
        export_btn = QPushButton('导出数据')
        export_btn.clicked.connect(lambda: self.export_data('基本信息'))
        self.setup_button_style(export_btn, 'normal')
        button_layout.addWidget(export_btn)
        
        add_btn = QPushButton('添加信息')
        add_btn.clicked.connect(self.add_youth)
        self.setup_button_style(add_btn, 'normal')
        button_layout.addWidget(add_btn)
        
        batch_delete_btn = QPushButton('批量删除')
        self.setup_button_style(batch_delete_btn, 'danger')
        batch_delete_btn.clicked.connect(self.batch_delete_youth)
        button_layout.addWidget(batch_delete_btn)
        
        import_camp_btn = QPushButton('导入入营点验')
        self.setup_button_style(import_camp_btn, 'primary')
        import_camp_btn.clicked.connect(self.import_camp_verification_data)
        button_layout.addWidget(import_camp_btn)
        
        import_political_btn = QPushButton('导入政治考核情况')
        self.setup_button_style(import_political_btn, 'primary')
        import_political_btn.clicked.connect(self.import_political_assessment_data)
        button_layout.addWidget(import_political_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        tab_widget.setLayout(layout)
        
        # 不在初始化时加载数据，改为延迟加载
        
        return tab_widget
    
    def create_town_interview_tab(self):
        """创建镇街谈心谈话情况标签页"""
        return self.town_interview_base.create_interview_tab()
    
    def create_leader_interview_tab(self):
        """创建领导谈心谈话情况标签页"""
        return self.leader_interview_base.create_interview_tab()
    
    def create_exception_statistics_tab(self):
        """创建异常情况统计标签页"""
        tab_widget = QWidget()
        layout = QVBoxLayout()
        
        # 搜索栏 - 单行布局
        search_container = QVBoxLayout()
        
        # 搜索框行
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel('搜索:'))
        
        self.exception_name_input = QLineEdit()
        self.exception_name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.exception_name_input.setPlaceholderText('姓名')
        self.exception_name_input.returnPressed.connect(self.search_exception_statistics)
        search_layout.addWidget(self.exception_name_input, 1)
        
        self.exception_id_card_input = QLineEdit()
        self.exception_id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.exception_id_card_input.setPlaceholderText('身份证号')
        self.exception_id_card_input.returnPressed.connect(self.search_exception_statistics)
        search_layout.addWidget(self.exception_id_card_input, 1)
        
        self.exception_recruitment_place_input = QLineEdit()
        self.exception_recruitment_place_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.exception_recruitment_place_input.setPlaceholderText('应征地')
        self.exception_recruitment_place_input.returnPressed.connect(self.search_exception_statistics)
        search_layout.addWidget(self.exception_recruitment_place_input, 1)
        
        self.exception_company_input = QLineEdit()
        self.exception_company_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.exception_company_input.setPlaceholderText('连')
        self.exception_company_input.returnPressed.connect(self.search_exception_statistics)
        search_layout.addWidget(self.exception_company_input, 1)
        
        self.exception_platoon_input = QLineEdit()
        self.exception_platoon_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.exception_platoon_input.setPlaceholderText('排')
        self.exception_platoon_input.returnPressed.connect(self.search_exception_statistics)
        search_layout.addWidget(self.exception_platoon_input, 1)
        
        self.exception_squad_input = QLineEdit()
        self.exception_squad_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.exception_squad_input.setPlaceholderText('班')
        self.exception_squad_input.returnPressed.connect(self.search_exception_statistics)
        search_layout.addWidget(self.exception_squad_input, 1)
        
        # 搜索和重置按钮
        search_btn = QPushButton('搜索')
        search_btn.clicked.connect(self.search_exception_statistics)
        self.setup_button_style(search_btn, 'primary')
        search_layout.addWidget(search_btn)
        
        reset_btn = QPushButton('重置')
        reset_btn.clicked.connect(self.reset_exception_statistics_search)
        self.setup_button_style(reset_btn, 'normal')
        search_layout.addWidget(reset_btn)
        
        search_container.addLayout(search_layout)
        layout.addLayout(search_container)
        
        # 表格上方的控制栏（包含全选和时间范围）
        table_control_layout = QHBoxLayout()
        
        # 左侧：全选复选框
        self.exception_select_all_checkbox = QCheckBox('全选')
        self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
        self.exception_select_all_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #7F8C8D;
                border-radius: 2px;
                background-color: white;
            }
            QCheckBox::indicator:hover {
                border-color: #3498DB;
                background-color: #ECF0F1;
            }
            QCheckBox::indicator:checked {
                background-color: #3498DB;
                border-color: #2980B9;
            }
        """)
        table_control_layout.addWidget(self.exception_select_all_checkbox)
        
        # 中间：弹性空间
        table_control_layout.addStretch()
        
        # 右侧：时间筛选器
        # 日期筛选区域
        date_filter_layout = QHBoxLayout()
        
        # 日期筛选标签
        date_filter_layout.addWidget(QLabel('日期筛选: 从:'))
        
        # 开始日期选择器
        from PyQt5.QtWidgets import QDateEdit
        from PyQt5.QtCore import QDate
        self.exception_start_date = QDateEdit()
        self.exception_start_date.setDate(QDate.currentDate().addDays(-30))  # 默认30天前
        self.exception_start_date.setCalendarPopup(True)
        date_filter_layout.addWidget(self.exception_start_date)
        
        date_filter_layout.addWidget(QLabel('到:'))
        
        # 结束日期选择器
        self.exception_end_date = QDateEdit()
        self.exception_end_date.setDate(QDate.currentDate())  # 默认今天
        self.exception_end_date.setCalendarPopup(True)
        date_filter_layout.addWidget(self.exception_end_date)
        
        # 筛选按钮
        filter_btn = QPushButton('筛选')
        filter_btn.clicked.connect(self.filter_exception_by_date_range)
        filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 60px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        date_filter_layout.addWidget(filter_btn)
        
        table_control_layout.addLayout(date_filter_layout)
        
        # 时间范围筛选下拉框
        table_control_layout.addWidget(QLabel('时间范围:'))
        self.exception_time_range_combo = QComboBox()
        self.exception_time_range_combo.addItems(['当天', '三天内', '七天内', '半个月内', '全部', '无'])
        self.exception_time_range_combo.setCurrentText('全部')  # 默认选择全部
        self.exception_time_range_combo.currentTextChanged.connect(self.on_exception_time_range_changed)
        self.exception_time_range_combo.setStyleSheet("""
            QComboBox {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
                min-width: 100px;
                color: #2c3e50;
            }
            QComboBox:hover {
                border-color: #3498db;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #7f8c8d;
                margin-right: 5px;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #bdc3c7;
                background-color: white;
                selection-background-color: #3498db;
                selection-color: white;
                color: #2c3e50;
            }
            QComboBox QAbstractItemView::item {
                padding: 8px;
                border: none;
                color: #2c3e50;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #ecf0f1;
                color: #2c3e50;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)
        table_control_layout.addWidget(self.exception_time_range_combo)
        
        layout.addLayout(table_control_layout)
        
        # 异常情况统计表格
        self.exception_statistics_table = QTableWidget()
        self.exception_statistics_headers = [
            '选择', '姓名', '日期', '公民身份号码', '性别', '应征地', '连', '排', '班', '带训班长信息', 
            '思想', '身体', '精神', '训练', '管理', '其他', '操作'
        ]
        self.exception_statistics_table.setColumnCount(len(self.exception_statistics_headers))
        self.exception_statistics_table.setHorizontalHeaderLabels(self.exception_statistics_headers)
        self.exception_statistics_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.exception_statistics_table.setAlternatingRowColors(True)
        self.exception_statistics_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # 应用统一的表格样式
        self.setup_table_style(self.exception_statistics_table)  # 禁止编辑
        
        # 添加双击事件处理
        self.exception_statistics_table.cellDoubleClicked.connect(self.on_exception_statistics_double_click)
        
        # 设置表格占满屏幕
        self.exception_statistics_table.setSizeAdjustPolicy(QTableWidget.SizeAdjustPolicy.AdjustToContentsOnFirstShow)
        
        # 设置表格列宽
        header = self.exception_statistics_table.horizontalHeader()
        
        # 设置各列的调整模式，确保内容完整显示
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 选择列
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 姓名列
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 日期列
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 公民身份号码列
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 性别列
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 应征地列
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 连列
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # 排列
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # 班列
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # 带训班长信息列
        header.setSectionResizeMode(10, QHeaderView.ResizeToContents)  # 思想列
        header.setSectionResizeMode(11, QHeaderView.ResizeToContents)  # 身体列
        header.setSectionResizeMode(12, QHeaderView.ResizeToContents)  # 精神列
        header.setSectionResizeMode(13, QHeaderView.ResizeToContents)  # 训练列
        header.setSectionResizeMode(14, QHeaderView.ResizeToContents)  # 管理列
        header.setSectionResizeMode(15, QHeaderView.ResizeToContents)  # 其他列（异常来源）
        header.setSectionResizeMode(16, QHeaderView.ResizeToContents)  # 操作列
        
        # 设置表格可以水平滚动
        self.exception_statistics_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        # 设置最小列宽，确保数据完整显示
        header.setMinimumSectionSize(80)  # 设置最小列宽
        
        # 为重要列设置合适的最小宽度
        self.exception_statistics_table.setColumnWidth(1, max(120, header.sectionSizeHint(1)))  # 姓名列
        self.exception_statistics_table.setColumnWidth(3, max(180, header.sectionSizeHint(3)))  # 公民身份号码列
        self.exception_statistics_table.setColumnWidth(15, max(200, header.sectionSizeHint(15)))  # 异常来源列
        
        # 允许用户手动调整列宽
        header.setSectionsMovable(False)  # 不允许移动列
        header.setStretchLastSection(False)  # 不自动拉伸最后一列
        header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        
        layout.addWidget(self.exception_statistics_table)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        export_btn = QPushButton('导出数据')
        export_btn.clicked.connect(lambda: self.export_data('异常情况统计'))
        self.setup_button_style(export_btn, 'normal')
        button_layout.addWidget(export_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        tab_widget.setLayout(layout)
        
        return tab_widget
    
    def create_bottom_buttons(self):
        layout = QHBoxLayout()
        layout.addStretch()
        return layout
    
    def load_all_youth(self):
        """加载所有青年信息（已废弃，改用详细显示）"""
        self.load_all_youth_detailed()
    
    def load_all_youth_detailed(self):
        """加载所有青年详细信息（完整版，默认显示）（新结构）"""
        try:
            conn = None
            try:
                conn = self.db_manager.get_connection()
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                # 获取时间筛选条件
                time_condition, time_params = self.get_time_filter_condition("camp_entry_time")
                
                base_query = """
                    SELECT id_card, name, gender, birth_date, nation, political_status,
                           religion, native_place, education_level, study_status, study_type,
                           camp_entry_time, recruitment_place, residence_address, household_address,
                           postal_code, personal_phone, family_phone, school, major,
                           enrollment_time, initial_hospital, initial_conclusion, initial_time,
                           physical_conclusion, physical_time, physical_disqualification,
                           chief_doctor_opinion, graduation_time, company, platoon, squad,
                           squad_leader, camp_status, leave_time, leave_reason, id
                    FROM youth
                """
                
                if time_condition:
                    query = base_query + f" WHERE {time_condition} ORDER BY id"
                    cursor.execute(query, time_params)
                else:
                    query = base_query + " ORDER BY id"
                    cursor.execute(query)
                
                results = cursor.fetchall()
                self.display_detailed_results(results)
            finally:
                if conn:
                    conn.close()
        except Exception as e:
            QMessageBox.warning(self, "加载错误", f"加载青年信息时发生错误：{str(e)}")
            self.search_table.setRowCount(0)
    
    def search_youth(self):
        """搜索青年（新结构）"""
        try:
            # 重置全选按钮状态
            if hasattr(self, 'select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操作
                self.select_all_checkbox.stateChanged.disconnect()
                self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            
            name = self.name_input.text().strip()
            id_card = self.id_card_input.text().strip()
            recruitment_place = self.recruitment_place_input.text().strip()  # 应征地搜索
            company = self.company_input.text().strip()
            platoon = self.platoon_input.text().strip()
            squad = self.squad_input.text().strip()

            conn = None
            try:
                conn = self.db_manager.get_connection()
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                sql = """
                    SELECT id_card, name, gender, birth_date, nation, political_status,
                           religion, native_place, education_level, study_status, study_type,
                           camp_entry_time, recruitment_place, residence_address, household_address,
                           postal_code, personal_phone, family_phone, school, major,
                           enrollment_time, initial_hospital, initial_conclusion, initial_time,
                           physical_conclusion, physical_time, physical_disqualification,
                           chief_doctor_opinion, graduation_time, company, platoon, squad,
                           squad_leader, camp_status, leave_time, leave_reason, id
                    FROM youth WHERE 1=1
                """
                params = []

                if name:
                    sql += " AND name LIKE ?"
                    params.append(f"%{name}%")

                if id_card:
                    sql += " AND id_card LIKE ?"
                    params.append(f"%{id_card}%")

                if recruitment_place:
                    sql += " AND recruitment_place LIKE ?"
                    params.append(f"%{recruitment_place}%")

                if company:
                    sql += " AND company LIKE ?"
                    params.append(f"%{company}%")

                if platoon:
                    sql += " AND platoon LIKE ?"
                    params.append(f"%{platoon}%")

                if squad:
                    sql += " AND squad LIKE ?"
                    params.append(f"%{squad}%")

                cursor.execute(sql, params)
                results = cursor.fetchall()
                
                self.search_table.setColumnCount(len(self.basic_info_headers))
                self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
                self.display_detailed_results(results)
                    
            finally:
                if conn:
                    conn.close()
                
        except Exception as e:
            QMessageBox.critical(self, "搜索错误", f"搜索时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'select_all_checkbox'):
                    self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            except:
                pass
            self.reset_search()
    
    def reset_search(self):
        """重置搜索"""
        try:
            # 重置全选按钮状态
            if hasattr(self, 'select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操作
                self.select_all_checkbox.stateChanged.disconnect()
                self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            
            self.name_input.clear()
            self.id_card_input.clear()
            self.recruitment_place_input.clear()  # 应征地输入框
            self.company_input.clear()
            self.platoon_input.clear()
            self.squad_input.clear()

            self.search_table.clearContents()
            self.search_table.setRowCount(0)
            
            for row in range(self.search_table.rowCount()):
                for col in range(self.search_table.columnCount()):
                    widget = self.search_table.cellWidget(row, col)
                    if widget:
                        self.search_table.removeCellWidget(row, col)
            
            self.search_table.setColumnCount(len(self.basic_info_headers))
            self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
            
            self.load_all_youth_detailed()
            
        except Exception as e:
            QMessageBox.warning(self, "重置错误", f"重置搜索时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'select_all_checkbox'):
                    self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            except:
                pass

    
    def display_detailed_results(self, results):
        """显示详细结果（完整版）- 包含选择框、操作按钮、删除按钮（新结构）"""
        self.search_table.setRowCount(len(results))
        
        # 保存当前结果数据供双击事件使用
        self.current_results = results

        for row, data in enumerate(results):
            # 第一列：添加复选框
            checkbox = QCheckBox()
            checkbox.setStyleSheet("""
                    QCheckBox {
                        spacing: 0px;
                    }
                    QCheckBox::indicator {
                        width: 20px;
                        height: 20px;
                        border: 2px solid #7F8C8D;
                        border-radius: 2px;
                        background-color: white;
                    }
                    QCheckBox::indicator:hover {
                        border-color: #3498DB;
                        background-color: #ECF0F1;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #3498DB;
                        border-color: #2980B9;
                    }
                """)
            checkbox.setMinimumSize(24, 24)  # 确保复选框有足够空间显示
            
            # 添加状态变化监听，当单个复选框状态改变时更新全选按钮状态
            checkbox.stateChanged.connect(self.update_select_all_state)
            
            checkbox_layout = QHBoxLayout()
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.setContentsMargins(5, 5, 5, 5)
            checkbox_layout.addWidget(checkbox)
            checkbox_container = QWidget()
            checkbox_container.setLayout(checkbox_layout)
            self.search_table.setCellWidget(row, 0, checkbox_container)
            
            # 数据库返回的字段顺序（新结构）：
            # 0:id_card, 1:name, 2:gender, 3:birth_date, 4:nation, 5:political_status, 6:religion, 7:native_place,
            # 8:education_level, 9:study_status, 10:study_type, 11:camp_entry_time, 12:recruitment_place,
            # 13:residence_address, 14:household_address, 15:postal_code, 16:personal_phone, 17:family_phone,
            # 18:school, 19:major, 20:enrollment_time, 21:initial_hospital, 22:initial_conclusion, 23:initial_time,
            # 24:physical_conclusion, 25:physical_time, 26:physical_disqualification, 27:chief_doctor_opinion,
            # 28:graduation_time, 29:company, 30:platoon, 31:squad, 32:squad_leader, 33:camp_status,
            # 34:leave_time, 35:leave_reason, 36:id
            
            # 表头顺序（从第1列开始，不包含序号）：
            # 姓名, 公民身份号码, 性别, 出生日期, 民族, 政治面貌, 宗教信仰, 籍贯,
            # 文化程度, 学业情况, 学习类型, 入营时间, 应征地, 经常居住地址, 户籍所在地, 邮编,
            # 本人电话, 家庭电话, 毕业(就读)学校, 所学专业, 入学时间,
            # 初检医院, 初检结论, 初检时间, 体检结论, 体检时间, 体检不合格原因, 主检医师意见, 毕业时间,
            # 连, 排, 班, 带训班长信息, 在营状态, 离营时间, 离营原因
            
            # 映射关系：表头列 -> 数据索引（不包含序号列）
            column_mapping = [
                1,   # 姓名 -> name
                0,   # 公民身份号码 -> id_card
                2,   # 性别 -> gender
                3,   # 出生日期 -> birth_date
                4,   # 民族 -> nation
                5,   # 政治面貌 -> political_status
                6,   # 宗教信仰 -> religion
                7,   # 籍贯 -> native_place
                8,   # 文化程度 -> education_level
                9,   # 学业情况 -> study_status
                10,  # 学习类型 -> study_type
                11,  # 入营时间 -> camp_entry_time
                12,  # 应征地 -> recruitment_place
                13,  # 经常居住地址 -> residence_address
                14,  # 户籍所在地 -> household_address
                15,  # 邮编 -> postal_code
                16,  # 本人电话 -> personal_phone
                17,  # 家庭电话 -> family_phone
                18,  # 毕业(就读)学校 -> school
                19,  # 所学专业 -> major
                20,  # 入学时间 -> enrollment_time
                21,  # 初检医院 -> initial_hospital
                22,  # 初检结论 -> initial_conclusion
                23,  # 初检时间 -> initial_time
                24,  # 体检结论 -> physical_conclusion
                25,  # 体检时间 -> physical_time
                26,  # 体检不合格原因 -> physical_disqualification
                27,  # 主检医师意见 -> chief_doctor_opinion
                28,  # 毕业时间 -> graduation_time
                29,  # 连 -> company
                30,  # 排 -> platoon
                31,  # 班 -> squad
                32,  # 带训班长信息 -> squad_leader
                33,  # 在营状态 -> camp_status
                34,  # 离营时间 -> leave_time
                35   # 离营原因 -> leave_reason
            ]
            
            # 显示数据字段（从第1列开始，不包含序号）
            for col_idx, data_idx in enumerate(column_mapping, start=1):
                if hasattr(data, '__getitem__') and not hasattr(data, 'keys'):
                    value = data[data_idx] if data_idx < len(data) else ''
                else:
                    field_names = ['id_card', 'name', 'gender', 'birth_date', 'nation', 'political_status',
                                  'religion', 'native_place', 'education_level', 'study_status', 'study_type',
                                  'camp_entry_time', 'recruitment_place', 'residence_address', 'household_address',
                                  'postal_code', 'personal_phone', 'family_phone', 'school', 'major',
                                  'enrollment_time', 'initial_hospital', 'initial_conclusion', 'initial_time',
                                  'physical_conclusion', 'physical_time', 'physical_disqualification',
                                  'chief_doctor_opinion', 'graduation_time', 'company', 'platoon', 'squad',
                                  'squad_leader', 'camp_status', 'leave_time', 'leave_reason', 'id']
                    value = data[field_names[data_idx]] if data_idx < len(field_names) else ''
                
                display_text = str(value or '').replace('\n', ' ').replace('\r', ' ')
                if len(display_text) > 100:
                    display_text = display_text[:100] + '...'
                
                item = QTableWidgetItem(display_text)
                item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
                self.search_table.setItem(row, col_idx, item)

            # 倒数第二列：查看详情按钮
            if hasattr(data, '__getitem__') and not hasattr(data, 'keys'):
                id_card = data[0]  # 第一个字段是id_card
            else:
                id_card = data['id_card']
            
            view_btn = QPushButton('查看详情')
            view_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝色
            view_btn.clicked.connect(lambda checked, idc=id_card: self.view_details(idc))
            view_btn_layout = QHBoxLayout()
            view_btn_layout.setAlignment(Qt.AlignCenter)
            view_btn_layout.setContentsMargins(5, 5, 5, 5)
            view_btn_layout.addWidget(view_btn)
            view_btn_container = QWidget()
            view_btn_container.setLayout(view_btn_layout)
            self.search_table.setCellWidget(row, len(self.basic_info_headers) - 2, view_btn_container)

            # 最后一列：删除按钮
            delete_btn = QPushButton('删除')
            delete_btn.setStyleSheet('background-color: #FFB6C1;')  # 淡红色
            delete_btn.clicked.connect(lambda checked, idc=id_card: self.delete_single_youth(idc))
            delete_btn_layout = QHBoxLayout()
            delete_btn_layout.setAlignment(Qt.AlignCenter)
            delete_btn_layout.setContentsMargins(5, 5, 5, 5)
            delete_btn_layout.addWidget(delete_btn)
            delete_btn_container = QWidget()
            delete_btn_container.setLayout(delete_btn_layout)
            self.search_table.setCellWidget(row, len(self.basic_info_headers) - 1, delete_btn_container)

        # 根据内容调整列宽，让每列宽度适配该列中最长的内容
        # 使用 QTimer.singleShot 确保在表格完全渲染后再调整列宽
        from PyQt5.QtCore import QTimer
        QTimer.singleShot(0, self.search_table.resizeColumnsToContents)
    
    def on_basic_info_double_click(self, row, column):
        """基本信息表格双击事件处理"""
        try:
            # 从原始数据中获取身份证号，而不是从表格单元格
            # 需要先获取当前显示的数据
            if hasattr(self, 'current_results') and row < len(self.current_results):
                data = self.current_results[row]
                if hasattr(data, '__getitem__') and not hasattr(data, 'keys'):
                    id_card = data[0]  # 第一个字段是id_card
                else:
                    id_card = data['id_card']
                # 调用查看详情功能
                self.view_details(id_card)
            else:
                QMessageBox.warning(self, '错误', '无法获取行数据')
        except Exception as e:
            QMessageBox.warning(self, '错误', f'双击查看详情时发生错误: {str(e)}')
    
    def view_details(self, id_card):
        """查看青年详情"""
        try:
            from ui.youth_detail_dialog import YouthDetailDialog
            dialog = YouthDetailDialog(self.db_manager, self.export_service, id_card, self)
            dialog.data_updated.connect(self.refresh_table_data)
            dialog.exec_()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开详情窗口时发生错误: {str(e)}')
    
    def refresh_table_data(self):
        """刷新表格数据"""
        try:
            self.search_table.clearContents()
            self.search_table.setRowCount(0)
            
            for row in range(self.search_table.rowCount()):
                for col in range(self.search_table.columnCount()):
                    widget = self.search_table.cellWidget(row, col)
                    if widget:
                        self.search_table.removeCellWidget(row, col)
            
            self.search_table.setColumnCount(len(self.basic_info_headers))
            self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
            
            self.load_all_youth_detailed()
            
        except Exception as e:
            print(f"刷新表格数据时发生错误: {str(e)}")
    
    def add_youth(self):
        """添加青年信息"""
        try:
            from ui.add_youth_dialog_simple import AddYouthDialog
            dialog = AddYouthDialog(self.db_manager, self)
            
            # 连接信号，添加成功后刷新表格
            dialog.data_updated.connect(self.load_all_youth_detailed)
            
            if dialog.exec_():
                self.load_all_youth_detailed()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开添加青年信息窗口时发生错误: {str(e)}')
    
    def delete_single_youth(self, id_card):
        """删除单个青年信息"""
        try:
            reply = QMessageBox.question(
                self, 
                '确认删除', 
                '确定要删除这条青年信息吗？\n删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                conn = None
                try:
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM youth WHERE id_card = ?", (id_card,))
                    conn.commit()
                    
                    QMessageBox.information(self, '删除成功', '青年信息已删除')
                    self.load_all_youth_detailed()
                    
                finally:
                    if conn:
                        conn.close()
                        
        except Exception as e:
            QMessageBox.critical(self, '删除失败', f'删除青年信息时发生错误：{str(e)}')
    
    def toggle_select_all(self, state):
        """全选/取消全选"""
        is_checked = (state == Qt.CheckState.Checked.value)
        
        # 临时标记，防止在全选操作时触发单个复选框的状态更新
        self._updating_select_all = True
        
        for row in range(self.search_table.rowCount()):
            checkbox_widget = self.search_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(is_checked)
        
        # 重置标记
        self._updating_select_all = False
    
    def update_select_all_state(self):
        """更新基本信息全选按钮状态"""
        # 如果正在执行全选操作，则不更新全选按钮状态
        if getattr(self, '_updating_select_all', False):
            return
        
        try:
            total_rows = self.search_table.rowCount()
            if total_rows == 0:
                return
            
            checked_count = 0
            for row in range(total_rows):
                checkbox_widget = self.search_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        checked_count += 1
            
            # 临时断开全选按钮的信号连接，避免触发toggle_select_all
            self.select_all_checkbox.stateChanged.disconnect()
            
            # 更新全选按钮状态
            if checked_count == 0:
                # 没有选中任何项
                self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            elif checked_count == total_rows:
                # 全部选中
                self.select_all_checkbox.setCheckState(Qt.CheckState.Checked)
            else:
                # 部分选中
                self.select_all_checkbox.setCheckState(Qt.CheckState.PartiallyChecked)
            
            # 重新连接信号
            self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            
        except Exception as e:
            print(f"更新基本信息全选按钮状态时出错: {e}")
            # 确保信号重新连接
            try:
                self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
            except:
                pass
    
    def batch_delete_youth(self):
        """批量删除青年信息（旧结构）"""
        try:
            selected_id_cards = []
            selected_names = []
            selected_rows_data = []  # 保存原始数据用于调试
            
            # 先从当前显示的数据中获取
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id_card, name, gender, nation, political_status,
                       school, education_level, major, study_status, study_type,
                       phone, household_address, residence_address, family_info,
                       district, street, parent_phone, personal_experience,
                       reference_person, reference_phone, id
                FROM youth ORDER BY id
            """)
            all_results = cursor.fetchall()
            conn.close()
            
            for row in range(self.search_table.rowCount()):
                checkbox_widget = self.search_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        # 从原始数据中获取id_card（第一个字段）
                        if row < len(all_results):
                            data = all_results[row]
                            id_card = data['id_card'] if hasattr(data, 'keys') else data[0]
                            name = data['name'] if hasattr(data, 'keys') else data[1]
                            
                            if id_card:
                                selected_id_cards.append(id_card)
                                selected_names.append(name)
            
            if not selected_id_cards:
                QMessageBox.information(self, '提示', '请先选择要删除的青年信息')
                return
            
            names_text = '、'.join(selected_names[:5])
            if len(selected_names) > 5:
                names_text += f' 等{len(selected_names)}人'
            
            reply = QMessageBox.question(
                self,
                '确认批量删除',
                f'确定要删除以下青年信息吗？\n{names_text}\n\n共{len(selected_id_cards)}条记录，删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                conn = None
                try:
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    
                    # 逐个删除以确保成功
                    deleted_count = 0
                    for id_card in selected_id_cards:
                        cursor.execute("DELETE FROM youth WHERE id_card = ?", (id_card,))
                        if cursor.rowcount > 0:
                            deleted_count += 1
                    
                    conn.commit()
                    
                    QMessageBox.information(self, '删除成功', f'成功删除 {deleted_count} 条青年信息')
                    
                    # 取消全选
                    if hasattr(self, 'select_all_checkbox'):
                        self.select_all_checkbox.setChecked(False)
                    
                    self.load_all_youth_detailed()
                    
                finally:
                    if conn:
                        conn.close()
                        
        except Exception as e:
            QMessageBox.critical(self, '批量删除失败', f'批量删除青年信息时发生错误：{str(e)}')
    
    def load_module_data(self, module_name, youth_data):
        """加载模块数据"""
        table = self.tab_tables[module_name]
        
        if module_name == '基本信息' and youth_data:
            table.setColumnCount(2)
            table.setHorizontalHeaderLabels(['字段', '值'])
            fields = [
                ('姓名', youth_data[1]),
                ('身份证号', youth_data[2]),
                ('区划', youth_data[3]),
                ('镇街', youth_data[4]),
                ('性别', youth_data[5]),
                ('出生日期', youth_data[6]),
                ('学历', youth_data[7]),
                ('政治面貌', youth_data[8]),
                ('电话', youth_data[9]),
                ('地址', youth_data[10])
            ]
            table.setRowCount(len(fields))
            for row, (field, value) in enumerate(fields):
                table.setItem(row, 0, QTableWidgetItem(field))
                table.setItem(row, 1, QTableWidgetItem(str(value or '')))
        else:
            # 加载其他模块数据
            pass
    
    def import_data(self, tab_name):
        """导入数据"""
        file_path, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '', 'Excel Files (*.xlsx *.xls)')
        if not file_path:
            return
        
        if tab_name == '基本信息':
            count, error = self.import_service.import_youth_from_excel(file_path)
            if error:
                # 如果有部分成功导入，显示详细信息
                if count > 0:
                    QMessageBox.warning(self, '导入完成（有错误）', error)
                    # 刷新显示
                    self.search_table.setColumnCount(len(self.basic_info_headers))
                    self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
                    self.load_all_youth_detailed()
                else:
                    QMessageBox.warning(self, '导入失败', error)
            else:
                QMessageBox.information(self, '导入成功', f'成功导入 {count} 条记录')
                # 导入成功后刷新显示详细信息
                self.search_table.setColumnCount(len(self.basic_info_headers))
                self.search_table.setHorizontalHeaderLabels(self.basic_info_headers)
                self.load_all_youth_detailed()
        elif tab_name == '每日情况统计':
            count, error = self.import_service.import_daily_stats(file_path)
            if error:
                if count > 0:
                    QMessageBox.warning(self, '导入完成（有错误）', error)
                    # 刷新每日情况统计数据
                    self.load_daily_stats_data()
                else:
                    QMessageBox.warning(self, '导入失败', error)
            else:
                QMessageBox.information(self, '导入成功', f'成功导入 {count} 条记录')
                # 导入成功后刷新数据
                self.load_daily_stats_data()
        elif tab_name == '异常情况统计':
            count, error = self.import_service.import_abnormal_stats(file_path)
            if error:
                QMessageBox.warning(self, '导入失败', error)
            else:
                QMessageBox.information(self, '导入成功', f'成功导入 {count} 条记录')
    
    def upload_document(self, tab_name):
        """上传扫描文档"""
        if not self.current_youth_id:
            QMessageBox.warning(self, '提示', '请先选择一个青年')
            return
        
        file_path, _ = QFileDialog.getOpenFileName(self, '选择文档', '', 'All Files (*.*)')
        if file_path:
            QMessageBox.information(self, '成功', '文档上传成功')
    
    def add_daily_info(self, tab_name):
        """添加每日情况统计信息"""
        try:
            from ui.daily_record_dialog import DailyRecordDialog
            dialog = DailyRecordDialog(self.db_manager, self)
            dialog.data_updated.connect(self.load_daily_stats_data)
            if dialog.exec_():
                # 刷新每日情况统计表格数据
                self.load_daily_stats_data()
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开添加每日记录窗口时发生错误: {str(e)}')
    
    def batch_add_daily_info(self):
        """批量添加每日情况统计信息"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QDateEdit, QComboBox, QTextEdit
            from PyQt5.QtCore import QDate
            
            # 创建对话框
            dialog = QDialog(self)
            dialog.setWindowTitle('批量添加每日情况记录')
            dialog.setFixedSize(500, 450)
            dialog.setModal(True)
            
            layout = QVBoxLayout()
            layout.setSpacing(15)
            
            # 标题
            title = QLabel('为所有基本信息中的人员批量添加每日情况记录')
            title.setStyleSheet('font-size: 14px; font-weight: bold; padding: 5px;')
            layout.addWidget(title)
            
            # 表单
            from PyQt5.QtWidgets import QFormLayout
            form_layout = QFormLayout()
            form_layout.setSpacing(10)
            
            # 日期
            date_input = QDateEdit()
            date_input.setDate(QDate.currentDate())
            date_input.setCalendarPopup(True)
            form_layout.addRow('日期:', date_input)
            
            # 思想
            mood_input = QComboBox()
            mood_input.addItems(['正常', '异常'])
            form_layout.addRow('思想:', mood_input)
            
            # 身体
            physical_input = QComboBox()
            physical_input.addItems(['正常', '异常'])
            form_layout.addRow('身体:', physical_input)
            
            # 精神
            mental_input = QComboBox()
            mental_input.addItems(['正常', '异常'])
            form_layout.addRow('精神:', mental_input)
            
            # 训练
            training_input = QComboBox()
            training_input.addItems(['正常', '异常'])
            form_layout.addRow('训练:', training_input)
            
            # 管理
            management_input = QComboBox()
            management_input.addItems(['正常', '异常'])
            form_layout.addRow('管理:', management_input)
            
            # 其他
            notes_input = QLineEdit()
            notes_input.setPlaceholderText('请输入其他备注信息（可选）')
            form_layout.addRow('其他:', notes_input)
            
            layout.addLayout(form_layout)
            
            # 提示信息
            info_label = QLabel('将为基本信息表中的所有人员添加记录。\n如果某人在该日期已有记录，将询问是否覆盖。')
            info_label.setStyleSheet('color: #7F8C8D; font-size: 12px; padding: 5px;')
            layout.addWidget(info_label)
            
            # 按钮
            button_layout = QHBoxLayout()
            
            ok_btn = QPushButton('确定')
            ok_btn.setStyleSheet("""
                QPushButton {
                    background-color: #27AE60;
                    color: white;
                    border: none;
                    padding: 8px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
            """)
            
            def on_ok_clicked():
                try:
                    date = date_input.date().toString('yyyy-MM-dd')
                    mood = mood_input.currentText()
                    physical = physical_input.currentText()
                    mental = mental_input.currentText()
                    training = training_input.currentText()
                    management = management_input.currentText()
                    notes = notes_input.text().strip()
                    
                    # 获取所有基本信息中的人员
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    cursor.execute("SELECT id, name, id_card FROM youth ORDER BY id")
                    all_youth = cursor.fetchall()
                    
                    if not all_youth:
                        conn.close()
                        QMessageBox.warning(dialog, '提示', '基本信息表中没有人员数据')
                        return
                    
                    # 检查每个人员是否已有该日期的记录
                    duplicates = []
                    new_records = []
                    
                    for youth in all_youth:
                        youth_id = youth[0]
                        youth_name = youth[1]
                        youth_id_card = youth[2]
                        
                        cursor.execute("""
                            SELECT id FROM daily_stat 
                            WHERE youth_id = ? AND record_date = ?
                        """, (youth_id, date))
                        existing = cursor.fetchone()
                        
                        if existing:
                            duplicates.append((existing[0], youth_id, youth_name, youth_id_card))
                        else:
                            new_records.append((youth_id, youth_name, youth_id_card))
                    
                    conn.close()
                    
                    # 如果有重复记录，询问是否覆盖
                    if duplicates:
                        duplicate_names = [f"{d[2]}({d[3]})" for d in duplicates[:5]]
                        names_text = '、'.join(duplicate_names)
                        if len(duplicates) > 5:
                            names_text += f' 等{len(duplicates)}人'
                        
                        reply = QMessageBox.question(
                            dialog,
                            '发现重复记录',
                            f'以下人员在 {date} 已有记录：\n{names_text}\n\n共{len(duplicates)}条重复记录。\n\n是否覆盖这些记录？\n（选择"是"将覆盖重复记录并添加新记录，选择"否"将只添加新记录）',
                            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
                            QMessageBox.StandardButton.No
                        )
                        
                        if reply == QMessageBox.StandardButton.Cancel:
                            return
                        
                        should_update = (reply == QMessageBox.StandardButton.Yes)
                    else:
                        should_update = False
                    
                    # 执行批量添加/更新
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    
                    added_count = 0
                    updated_count = 0
                    
                    # 添加新记录
                    for youth_id, youth_name, youth_id_card in new_records:
                        # 使用正确的参数调用插入方法
                        self.db_manager.insert_daily_stat(
                            youth_id, date, mood, physical, mental, training, management, notes
                        )
                        added_count += 1
                    
                    # 更新重复记录（如果用户选择覆盖）
                    if should_update and duplicates:
                        for existing_id, youth_id, youth_name, youth_id_card in duplicates:
                            # 使用正确的参数调用更新方法
                            self.db_manager.update_daily_stat(
                                existing_id, date, mood, physical, mental, training, management, notes
                            )
                            updated_count += 1
                    
                    conn.commit()
                    conn.close()
                    
                    # 显示结果
                    result_msg = f'批量添加完成！\n\n'
                    if added_count > 0:
                        result_msg += f'新增记录：{added_count} 条\n'
                    if updated_count > 0:
                        result_msg += f'更新记录：{updated_count} 条\n'
                    if duplicates and not should_update:
                        result_msg += f'跳过重复：{len(duplicates)} 条'
                    
                    QMessageBox.information(dialog, '成功', result_msg)
                    
                    # 刷新数据
                    self.load_daily_stats_data()
                    
                    # 关闭对话框
                    dialog.accept()
                    
                except Exception as e:
                    QMessageBox.warning(dialog, '错误', f'批量添加失败: {str(e)}')
            
            ok_btn.clicked.connect(on_ok_clicked)
            
            cancel_btn = QPushButton('取消')
            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #95A5A6;
                    color: white;
                    border: none;
                    padding: 8px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #7F8C8D;
                }
            """)
            cancel_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(ok_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开批量添加窗口时发生错误: {str(e)}')
    
    def edit_daily_record(self, record_id, youth_id):
        """修改每日记录"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QDateEdit, QComboBox, QFormLayout
            from PyQt5.QtCore import QDate
            
            # 获取记录数据
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.id, d.record_date, d.mood, d.physical_condition, d.mental_state, d.notes,
                       y.name, y.id_card
                FROM daily_stat d
                JOIN youth y ON d.youth_id = y.id
                WHERE d.id=?
            ''', (record_id,))
            record = cursor.fetchone()
            conn.close()
            
            if not record:
                QMessageBox.warning(self, '错误', '记录不存在')
                return
            
            dialog = QDialog(self)
            dialog.setWindowTitle('修改每日记录')
            dialog.setMinimumSize(500, 300)  # 设置最小尺寸
            dialog.resize(500, 300)  # 设置初始尺寸，但允许调整
            dialog.setModal(True)
            
            # 允许窗口调整大小，只显示最大化和关闭按钮
            dialog.setWindowFlags(Qt.Dialog | 
                                 Qt.WindowMaximizeButtonHint | 
                                 Qt.WindowCloseButtonHint)
            
            layout = QVBoxLayout()
            layout.setSpacing(10)
            
            # 标题
            title = QLabel(f'修改每日记录 - {record[6]} ({record[7]})')
            title.setStyleSheet('font-size: 14px; font-weight: bold;')
            layout.addWidget(title)
            
            # 表单
            form_layout = QFormLayout()
            form_layout.setSpacing(10)
            
            # 日期
            date_input = QDateEdit()
            date_input.setCalendarPopup(True)
            if record[1]:
                date_parts = record[1].split('-')
                if len(date_parts) == 3:
                    date_input.setDate(QDate(int(date_parts[0]), int(date_parts[1]), int(date_parts[2])))
            form_layout.addRow('日期:', date_input)
            
            # 思想
            mood_input = QComboBox()
            mood_input.addItems(['正常', '异常'])
            mood_input.setCurrentText(record[2] or '正常')
            form_layout.addRow('思想:', mood_input)
            
            # 身体
            physical_input = QComboBox()
            physical_input.addItems(['正常', '异常'])
            physical_input.setCurrentText(record[3] or '正常')
            form_layout.addRow('身体:', physical_input)
            
            # 精神
            mental_input = QComboBox()
            mental_input.addItems(['正常', '异常'])
            mental_input.setCurrentText(record[4] or '正常')
            form_layout.addRow('精神:', mental_input)
            
            # 其他
            notes_input = QLineEdit()
            notes_input.setText(record[5] or '')
            form_layout.addRow('其他:', notes_input)
            
            layout.addLayout(form_layout)
            
            # 按钮
            button_layout = QHBoxLayout()
            
            save_btn = QPushButton('保存')
            save_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3498DB;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #2980B9;
                }
            """)
            
            def save_changes():
                try:
                    # 使用新的更新方法
                    self.db_manager.update_daily_stat(
                        record_id,
                        date_input.date().toString('yyyy-MM-dd'), 
                        mood_input.currentText(),
                        physical_input.currentText(),
                        mental_input.currentText(),
                        notes_input.text().strip()
                    )
                    
                    QMessageBox.information(dialog, '成功', '记录已更新')
                    self.load_daily_stats_data()
                    dialog.accept()
                except Exception as e:
                    QMessageBox.warning(dialog, '错误', f'更新失败: {str(e)}')
            
            save_btn.clicked.connect(save_changes)
            
            cancel_btn = QPushButton('取消')
            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #95A5A6;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #7F8C8D;
                }
            """)
            cancel_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(save_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.warning(self, '错误', f'修改记录时发生错误: {str(e)}')
    
    def delete_daily_record(self, record_id):
        """删除每日记录"""
        try:
            reply = QMessageBox.question(self, '确认删除', 
                                       '确定要删除这条记录吗？',
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            
            if reply == QMessageBox.StandardButton.Yes:
                # 使用新的删除方法
                success = self.db_manager.delete_daily_stat(record_id)
                
                if success:
                    QMessageBox.information(self, '成功', '记录已删除')
                    self.load_daily_stats_data()
                else:
                    QMessageBox.warning(self, '错误', '删除失败')
        except Exception as e:
            QMessageBox.warning(self, '错误', f'删除记录时发生错误: {str(e)}')
    
    def add_daily_record(self):
        """添加每日记录"""
        QMessageBox.information(self, '提示', '请先在搜索结果中选择一个青年，然后点击"查看详情"进入详情页面添加每日记录')

    def export_data(self, tab_name):
        """导出数据"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            f'导出{tab_name}', 
            f'{tab_name}_{datetime.now().strftime("%Y%m%d")}.xlsx',
            'Excel Files (*.xlsx)'
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = tab_name
            
            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            if tab_name == '基本信息':
                # 导出基本信息（新结构40字段）
                headers = [
                    '选择', '序号', '姓名', '公民身份号码', '性别', '出生日期', '民族', '政治面貌', '宗教信仰', '籍贯',
                    '文化程度', '学业情况', '学习类型', '入营时间', '应征地', '经常居住地址', '户籍所在地', '邮编',
                    '本人电话', '家庭电话', '毕业(就读)学校', '所学专业', '入学时间',
                    '初检医院', '初检结论', '初检时间', '体检结论', '体检时间', '体检不合格原因', '主检医师意见', '毕业时间',
                    '连', '排', '班', '带训班长信息', '在营状态', '离营时间', '离营原因', '操作', '删除'
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # 获取所有青年数据
                conn = self.db_manager.get_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id_card, name, gender, birth_date, nation, political_status, religion, native_place,
                           education_level, study_status, study_type, camp_entry_time, recruitment_place,
                           residence_address, household_address, postal_code, personal_phone, family_phone,
                           school, major, enrollment_time, initial_hospital, initial_conclusion, initial_time,
                           physical_conclusion, physical_time, physical_disqualification, chief_doctor_opinion,
                           graduation_time, company, platoon, squad, squad_leader, camp_status, leave_time, leave_reason, id
                    FROM youth ORDER BY id
                """)
                results = cursor.fetchall()
                conn.close()
                
                for row_idx, youth in enumerate(results, 2):
                    # 按照查询的字段顺序填充数据（共37个字段，索引0-36）
                    ws.cell(row=row_idx, column=1, value='')  # 选择
                    ws.cell(row=row_idx, column=2, value=youth[36])  # 序号 (id) - 索引36
                    ws.cell(row=row_idx, column=3, value=youth[1])   # 姓名 - 索引1
                    ws.cell(row=row_idx, column=4, value=youth[0])   # 公民身份号码 - 索引0
                    ws.cell(row=row_idx, column=5, value=youth[2])   # 性别 - 索引2
                    ws.cell(row=row_idx, column=6, value=youth[3])   # 出生日期 - 索引3
                    ws.cell(row=row_idx, column=7, value=youth[4])   # 民族 - 索引4
                    ws.cell(row=row_idx, column=8, value=youth[5])   # 政治面貌 - 索引5
                    ws.cell(row=row_idx, column=9, value=youth[6])   # 宗教信仰 - 索引6
                    ws.cell(row=row_idx, column=10, value=youth[7])  # 籍贯 - 索引7
                    ws.cell(row=row_idx, column=11, value=youth[8])  # 文化程度 - 索引8
                    ws.cell(row=row_idx, column=12, value=youth[9])  # 学业情况 - 索引9
                    ws.cell(row=row_idx, column=13, value=youth[10]) # 学习类型 - 索引10
                    ws.cell(row=row_idx, column=14, value=youth[11]) # 入营时间 - 索引11
                    ws.cell(row=row_idx, column=15, value=youth[12]) # 应征地 - 索引12
                    ws.cell(row=row_idx, column=16, value=youth[13]) # 经常居住地址 - 索引13
                    ws.cell(row=row_idx, column=17, value=youth[14]) # 户籍所在地 - 索引14
                    ws.cell(row=row_idx, column=18, value=youth[15]) # 邮编 - 索引15
                    ws.cell(row=row_idx, column=19, value=youth[16]) # 本人电话 - 索引16
                    ws.cell(row=row_idx, column=20, value=youth[17]) # 家庭电话 - 索引17
                    ws.cell(row=row_idx, column=21, value=youth[18]) # 毕业(就读)学校 - 索引18
                    ws.cell(row=row_idx, column=22, value=youth[19]) # 所学专业 - 索引19
                    ws.cell(row=row_idx, column=23, value=youth[20]) # 入学时间 - 索引20
                    ws.cell(row=row_idx, column=24, value=youth[21]) # 初检医院 - 索引21
                    ws.cell(row=row_idx, column=25, value=youth[22]) # 初检结论 - 索引22
                    ws.cell(row=row_idx, column=26, value=youth[23]) # 初检时间 - 索引23
                    ws.cell(row=row_idx, column=27, value=youth[24]) # 体检结论 - 索引24
                    ws.cell(row=row_idx, column=28, value=youth[25]) # 体检时间 - 索引25
                    ws.cell(row=row_idx, column=29, value=youth[26]) # 体检不合格原因 - 索引26
                    ws.cell(row=row_idx, column=30, value=youth[27]) # 主检医师意见 - 索引27
                    ws.cell(row=row_idx, column=31, value=youth[28]) # 毕业时间 - 索引28
                    ws.cell(row=row_idx, column=32, value=youth[29]) # 连 - 索引29
                    ws.cell(row=row_idx, column=33, value=youth[30]) # 排 - 索引30
                    ws.cell(row=row_idx, column=34, value=youth[31]) # 班 - 索引31
                    ws.cell(row=row_idx, column=35, value=youth[32]) # 带训班长信息 - 索引32
                    ws.cell(row=row_idx, column=36, value=youth[33]) # 在营状态 - 索引33
                    ws.cell(row=row_idx, column=37, value=youth[34]) # 离营时间 - 索引34
                    ws.cell(row=row_idx, column=38, value=youth[35]) # 离营原因 - 索引35
                    ws.cell(row=row_idx, column=39, value='')        # 操作
                    ws.cell(row=row_idx, column=40, value='')        # 删除
                
            elif tab_name == '异常情况统计':
                headers = ['日期', '姓名', '性别', '公民身份号码', '应征地', '连', '排', '班', '带训班长信息', '思想', '身体', '精神', '训练', '管理', '其他']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # 获取选中的数据或全部数据
                selected_records = self.get_selected_exception_statistics_data()
                
                if not selected_records:
                    # 没有选择时导出全部数据 - 使用异常统计视图
                    data = self.db_manager.get_exception_statistics_view_data()
                    results = []
                    for record in data:
                        # 转换ViewRecord对象为元组格式，按照新的导出表头顺序
                        result_tuple = (
                            record[14],  # 日期
                            record[1],   # 姓名
                            record[2],   # 性别
                            record[0],   # 公民身份号码
                            record[7],   # 应征地
                            record[3],   # 连
                            record[4],   # 排
                            record[5],   # 班
                            record[6],   # 带训班长信息
                            record[8],   # 思想是否异常
                            record[9],   # 身体是否异常
                            record[10],  # 精神是否异常
                            record[11],  # 训练是否异常
                            record[12],  # 管理是否异常
                            f"来源:{record[13]}" if record[13] else ""  # 其他（异常来源）
                        )
                        results.append(result_tuple)
                    # 注意：这里不需要conn.close()，因为get_exception_statistics_view_data()方法内部已经处理了连接
                else:
                    # 有选择时导出选中数据
                    results = selected_records
                
                for row_idx, record in enumerate(results, 2):
                    # 直接使用记录数据，因为已经是正确的格式
                    for col_idx, value in enumerate(record, 1):
                        ws.cell(row=row_idx, column=col_idx, value=str(value or ''))
            
            elif tab_name == '每日情况统计':
                headers = ['公民身份号码', '姓名', '日期', '思想', '身体', '精神', '训练', '管理', '备注']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # 获取选中的数据或全部数据
                selected_records = self.get_selected_daily_stats_data()
                
                if not selected_records:
                    # 没有选择时导出全部数据
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT y.id_card, y.name, d.record_date, d.mood, d.physical_condition, 
                               d.mental_state, d.training, d.management, d.notes
                        FROM daily_stat d
                        JOIN youth y ON d.youth_id = y.id
                        ORDER BY d.record_date DESC, d.id ASC
                    ''')
                    results = cursor.fetchall()
                    conn.close()
                else:
                    # 有选择时导出选中数据
                    results = selected_records
                
                for row_idx, record in enumerate(results, 2):
                    for col_idx, value in enumerate(record, 1):
                        ws.cell(row=row_idx, column=col_idx, value=str(value or ''))
            
            elif tab_name in ['病史调查情况', '镇街谈心谈话情况', '领导谈心谈话情况']:
                if tab_name == '镇街谈心谈话情况':
                    # 镇街谈心谈话情况特殊处理 - xlsx + 图片文件夹格式
                    import os
                    import shutil
                    
                    # 创建导出文件夹
                    export_folder = file_path.replace('.xlsx', '_导出')
                    if os.path.exists(export_folder):
                        shutil.rmtree(export_folder)
                    os.makedirs(export_folder)
                    
                    # 创建图片文件夹
                    images_folder = os.path.join(export_folder, '走访调查图片')
                    os.makedirs(images_folder)
                    
                    # 设置表头
                    headers = ['日期', '姓名', '性别', '公民身份号码', '应征地', '连', '排', '班', '带训班长信息', '思想', '精神', '走访调查情况', '导出时间']
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    # 获取选中的数据或全部数据
                    selected_records = self.get_selected_town_interview_data()
                    if not selected_records:
                        # 没有选择时导出全部数据
                        records = self.db_manager.search_town_interviews()
                    else:
                        # 有选择时导出选中数据
                        records = selected_records
                    
                    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    for row_idx, record in enumerate(records, 2):
                        # record格式: (id, youth_id_card, youth_name, gender, interview_date, 
                        #              enlistment_place, company, platoon, squad, squad_leader,
                        #              thoughts, spirit, created_at)
                        ws.cell(row=row_idx, column=1, value=record[4] or '')   # 日期
                        ws.cell(row=row_idx, column=2, value=record[2] or '')   # 姓名
                        ws.cell(row=row_idx, column=3, value=record[3] or '')   # 性别
                        ws.cell(row=row_idx, column=4, value=record[1] or '')   # 公民身份号码
                        ws.cell(row=row_idx, column=5, value=record[5] or '')   # 应征地
                        ws.cell(row=row_idx, column=6, value=record[6] or '')   # 连
                        ws.cell(row=row_idx, column=7, value=record[7] or '')   # 排
                        ws.cell(row=row_idx, column=8, value=record[8] or '')   # 班
                        ws.cell(row=row_idx, column=9, value=record[9] or '')   # 带训班长信息
                        ws.cell(row=row_idx, column=10, value=record[10] or '') # 思想
                        ws.cell(row=row_idx, column=11, value=record[11] or '') # 精神
                        ws.cell(row=row_idx, column=13, value=export_time)      # 导出时间
                        
                        # 处理图片
                        try:
                            image_data = self.db_manager.get_town_interview_image(record[0])
                            if image_data:
                                # 保存图片到文件夹
                                image_filename = f"{record[2]}_{record[1]}_走访调查.jpg"
                                image_path = os.path.join(images_folder, image_filename)
                                
                                with open(image_path, 'wb') as img_file:
                                    img_file.write(image_data)
                                
                                # 在Excel中创建超链接
                                relative_path = f"走访调查图片/{image_filename}"
                                cell = ws.cell(row=row_idx, column=12, value="查看图片")
                                cell.hyperlink = relative_path
                                cell.style = "Hyperlink"
                            else:
                                ws.cell(row=row_idx, column=12, value='无图片')
                        except Exception as e:
                            print(f"处理图片时出错: {e}")
                            ws.cell(row=row_idx, column=12, value='图片处理失败')
                    
                    # 保存Excel文件到导出文件夹
                    excel_path = os.path.join(export_folder, '镇街谈心谈话数据.xlsx')
                    wb.save(excel_path)
                    
                    # 显示成功消息
                    QMessageBox.information(self, '成功', f'数据已导出到:\n{export_folder}\n\n包含:\n- Excel文件: 镇街谈心谈话数据.xlsx\n- 图片文件夹: 走访调查图片')
                    return  # 提前返回，不执行后面的通用保存逻辑
                
                elif tab_name == '领导谈心谈话情况':
                    # 领导谈心谈话情况特殊处理 - xlsx + 图片文件夹格式
                    import os
                    import shutil
                    
                    # 创建导出文件夹
                    export_folder = file_path.replace('.xlsx', '_导出')
                    if os.path.exists(export_folder):
                        shutil.rmtree(export_folder)
                    os.makedirs(export_folder)
                    
                    # 创建图片文件夹
                    images_folder = os.path.join(export_folder, '走访调查图片')
                    os.makedirs(images_folder)
                    
                    # 设置表头
                    headers = ['日期', '姓名', '性别', '公民身份号码', '应征地', '连', '排', '班', '带训班长信息', '思想', '精神', '走访调查情况', '导出时间']
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    # 获取选中的数据或全部数据
                    selected_records = self.get_selected_leader_interview_data()
                    if not selected_records:
                        # 没有选择时导出全部数据
                        records = self.db_manager.search_leader_interviews()
                    else:
                        # 有选择时导出选中数据
                        records = selected_records
                    
                    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    for row_idx, record in enumerate(records, 2):
                        # record格式: (id, youth_id_card, youth_name, gender, interview_date, 
                        #              enlistment_place, company, platoon, squad, squad_leader,
                        #              thoughts, spirit, created_at)
                        ws.cell(row=row_idx, column=1, value=record[4] or '')   # 日期
                        ws.cell(row=row_idx, column=2, value=record[2] or '')   # 姓名
                        ws.cell(row=row_idx, column=3, value=record[3] or '')   # 性别
                        ws.cell(row=row_idx, column=4, value=record[1] or '')   # 公民身份号码
                        ws.cell(row=row_idx, column=5, value=record[5] or '')   # 应征地
                        ws.cell(row=row_idx, column=6, value=record[6] or '')   # 连
                        ws.cell(row=row_idx, column=7, value=record[7] or '')   # 排
                        ws.cell(row=row_idx, column=8, value=record[8] or '')   # 班
                        ws.cell(row=row_idx, column=9, value=record[9] or '')   # 带训班长信息
                        ws.cell(row=row_idx, column=10, value=record[10] or '') # 思想
                        ws.cell(row=row_idx, column=11, value=record[11] or '') # 精神
                        ws.cell(row=row_idx, column=13, value=export_time)      # 导出时间
                        
                        # 处理图片
                        try:
                            image_data = self.db_manager.get_leader_interview_image(record[0])
                            if image_data:
                                # 保存图片到文件夹
                                image_filename = f"{record[2]}_{record[1]}_走访调查.jpg"
                                image_path = os.path.join(images_folder, image_filename)
                                
                                with open(image_path, 'wb') as img_file:
                                    img_file.write(image_data)
                                
                                # 在Excel中创建超链接
                                relative_path = f"走访调查图片/{image_filename}"
                                cell = ws.cell(row=row_idx, column=12, value="查看图片")
                                cell.hyperlink = relative_path
                                cell.style = "Hyperlink"
                            else:
                                ws.cell(row=row_idx, column=12, value='无图片')
                        except Exception as e:
                            print(f"处理图片时出错: {e}")
                            ws.cell(row=row_idx, column=12, value='图片处理失败')
                    
                    # 保存Excel文件到导出文件夹
                    excel_path = os.path.join(export_folder, '领导谈心谈话数据.xlsx')
                    wb.save(excel_path)
                    
                    # 显示成功消息
                    QMessageBox.information(self, '成功', f'数据已导出到:\n{export_folder}\n\n包含:\n- Excel文件: 领导谈心谈话数据.xlsx\n- 图片文件夹: 走访调查图片')
                    return  # 提前返回，不执行后面的通用保存逻辑
                    
                else:
                    # 其他情况的原有处理
                    headers = ['公民身份号码', '姓名', '文件路径', '日期', '备注']
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    table_map = {
                        '病史调查情况': 'medical_history',
                        '领导谈心谈话情况': 'leader_interview'
                    }
                    table_name = table_map[tab_name]
                    date_field = 'upload_date' if tab_name == '病史调查情况' else 'interview_date'
                    
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    cursor.execute(f'''
                        SELECT y.id_card, y.name, t.file_path, t.{date_field}, t.notes
                        FROM {table_name} t
                        JOIN youth y ON t.youth_id = y.id
                        ORDER BY t.{date_field} DESC
                    ''')
                    results = cursor.fetchall()
                    conn.close()
                    
                    for row_idx, record in enumerate(results, 2):
                        for col_idx, value in enumerate(record, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 调整列宽（仅对非镇街谈心谈话情况）
            if tab_name != '镇街谈心谈话情况':
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
                
                wb.save(file_path)
                QMessageBox.information(self, '成功', f'数据已导出到:\n{file_path}')
            
        except Exception as e:
            error_msg = str(e)
            if "Permission denied" in error_msg or "PermissionError" in error_msg:
                QMessageBox.warning(self, '导出失败', 
                    f'导出失败：文件访问权限被拒绝\n\n可能的原因：\n'
                    f'1. 目标文件正在被Excel或其他程序打开，请关闭后重试\n'
                    f'2. 没有写入权限，请选择其他保存位置（如文档文件夹）\n'
                    f'3. 文件被系统锁定\n\n'
                    f'建议解决方案：\n'
                    f'• 关闭可能打开该文件的Excel程序\n'
                    f'• 选择桌面或文档文件夹作为保存位置\n'
                    f'• 以管理员身份运行程序\n\n'
                    f'详细错误信息：{error_msg}')
            else:
                QMessageBox.warning(self, '导出失败', f'导出失败: {error_msg}')

    def print_all_info(self):
        """打印全部信息"""
        QMessageBox.information(self, '提示', '请在各个模块中使用导出功能导出数据后打印')
    
    def load_town_interview_data(self):
        """加载镇街谈心谈话数据到表格"""
        self.town_interview_base.load_interview_data()
    
    def search_town_interview(self):
        """搜索镇街谈心谈话记录"""
        self.town_interview_base.search_interview()
    
    def reset_town_interview_search(self):
        """重置镇街谈心谈话搜索"""
        self.town_interview_base.reset_interview_search()
    
    def display_town_interview_records(self, records):
        """显示镇街谈心谈话记录"""
        self.town_interview_base.display_interview_records(records)
    
    def toggle_town_select_all(self, state):
        """镇街谈心谈话全选/取消全选"""
        self.town_interview_base.toggle_select_all(state)
    
    def update_town_select_all_state(self):
        """更新镇街谈心谈话全选按钮状态"""
        self.town_interview_base.update_select_all_state()
    
    def add_town_interview(self):
        """添加镇街谈心谈话记录"""
        self.town_interview_base.add_interview()
    
    def edit_town_interview_by_record(self, record_data):
        """根据记录数据编辑镇街谈心谈话记录"""
        self.town_interview_base.edit_interview_by_record(record_data)
    
    def delete_single_town_interview(self, record_data):
        """删除单个镇街谈心谈话记录"""
        self.town_interview_base.delete_single_interview(record_data)
    
    def batch_delete_town_interview(self):
        """批量删除镇街谈心谈话记录"""
        self.town_interview_base.batch_delete_interview()
    
    def view_town_interview_image(self, record_id):
        """查看镇街谈心谈话图片"""
        self.town_interview_base.view_interview_image(record_id)
    
    def get_selected_town_interview_data(self):
        """获取镇街谈心谈话中选中的数据"""
        return self.town_interview_base.get_selected_interview_data()
    
    def load_leader_interview_data(self):
        """加载领导谈心谈话数据到表格"""
        self.leader_interview_base.load_interview_data()
    
    def search_leader_interview(self):
        """搜索领导谈心谈话记录"""
        self.leader_interview_base.search_interview()
    
    def reset_leader_interview_search(self):
        """重置领导谈心谈话搜索"""
        self.leader_interview_base.reset_interview_search()
    
    def display_leader_interview_records(self, records):
        """显示领导谈心谈话记录"""
        self.leader_interview_base.display_interview_records(records)
    
    def toggle_leader_select_all(self, state):
        """领导谈心谈话全选/取消全选"""
        self.leader_interview_base.toggle_select_all(state)
    
    def update_leader_select_all_state(self):
        """更新领导谈心谈话全选按钮状态"""
        self.leader_interview_base.update_select_all_state()
    
    def add_leader_interview(self):
        """添加领导谈心谈话记录"""
        self.leader_interview_base.add_interview()
    
    def edit_leader_interview_by_record(self, record_data):
        """根据记录数据编辑领导谈心谈话记录"""
        self.leader_interview_base.edit_interview_by_record(record_data)
    
    def delete_single_leader_interview(self, record_data):
        """删除单个领导谈心谈话记录"""
        self.leader_interview_base.delete_single_interview(record_data)
    
    def batch_delete_leader_interview(self):
        """批量删除领导谈心谈话记录"""
        self.leader_interview_base.batch_delete_interview()
    
    def view_leader_interview_image(self, record_id):
        """查看领导谈心谈话图片"""
        self.leader_interview_base.view_interview_image(record_id)
    
    def get_selected_leader_interview_data(self):
        """获取领导谈心谈话中选中的数据"""
        return self.leader_interview_base.get_selected_interview_data()
    
    def get_selected_basic_info_data(self):
        """获取基本信息中选中的数据"""
        try:
            selected_data = []
            
            # 获取所有数据用于对照
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id_card, name, gender, nation, political_status,
                       school, education_level, major, study_status, study_type,
                       phone, household_address, residence_address, family_info,
                       district, street, parent_phone, personal_experience,
                       reference_person, reference_phone, id
                FROM youth ORDER BY id
            """)
            all_results = cursor.fetchall()
            conn.close()
            
            # 检查选中的行
            for row in range(self.search_table.rowCount()):
                checkbox_widget = self.search_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_results):
                            selected_data.append(all_results[row])
            
            return selected_data
            
        except Exception as e:
            print(f"获取选中基本信息数据时出错: {e}")
            return []

    def get_selected_exception_statistics_data(self):
        """获取异常情况统计中选中的数据"""
        try:
            selected_data = []
            
            # 获取所有异常统计视图数据用于对照
            all_records = self.db_manager.get_exception_statistics_view_data()
            
            # 检查选中的行
            for row in range(self.exception_statistics_table.rowCount()):
                checkbox_widget = self.exception_statistics_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_records):
                            record = all_records[row]
                            # 转换为导出格式：日期, 姓名, 性别, 公民身份号码, 连, 排, 班, 应征地, 带训班长, 思想, 身体, 精神, 训练, 管理, 其他
                            selected_tuple = (
                                record[14],  # 日期
                                record[1],   # 姓名
                                record[2],   # 性别
                                record[0],   # 公民身份号码
                                record[3],   # 连
                                record[4],   # 排
                                record[5],   # 班
                                record[7],   # 应征地
                                record[6],   # 带训班长
                                record[8],   # 思想是否异常
                                record[9],   # 身体是否异常
                                record[10],  # 精神是否异常
                                record[11],  # 训练是否异常
                                record[12],  # 管理是否异常
                                f"来源:{record[13]}" if record[13] else ""  # 其他（异常来源）
                            )
                            selected_data.append(selected_tuple)
            
            return selected_data
            
        except Exception as e:
            print(f"获取选中异常情况统计数据时出错: {e}")
            return []

    
    # ==================== 病史筛查情况模块 ====================
    
    def create_medical_screening_tab(self):
        """创建病史筛查标签页"""
        tab_widget = QWidget()
        layout = QVBoxLayout()
        
        # 搜索栏
        search_layout = QHBoxLayout()
        
        search_layout.addWidget(QLabel('搜索:'))
        
        self.medical_name_input = QLineEdit()
        self.medical_name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.medical_name_input.setPlaceholderText('姓名')
        self.medical_name_input.returnPressed.connect(self.search_medical_screening)
        search_layout.addWidget(self.medical_name_input)
        
        self.medical_id_card_input = QLineEdit()
        self.medical_id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.medical_id_card_input.setPlaceholderText('身份证号')
        self.medical_id_card_input.returnPressed.connect(self.search_medical_screening)
        search_layout.addWidget(self.medical_id_card_input)
        
        # 应征地输入框
        self.medical_recruitment_place_input = QLineEdit()
        self.medical_recruitment_place_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.medical_recruitment_place_input.setPlaceholderText('应征地')
        self.medical_recruitment_place_input.returnPressed.connect(self.search_medical_screening)
        search_layout.addWidget(self.medical_recruitment_place_input)
        
        # 连输入框
        self.medical_company_input = QLineEdit()
        self.medical_company_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.medical_company_input.setPlaceholderText('连')
        self.medical_company_input.returnPressed.connect(self.search_medical_screening)
        search_layout.addWidget(self.medical_company_input)
        
        # 排输入框
        self.medical_platoon_input = QLineEdit()
        self.medical_platoon_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.medical_platoon_input.setPlaceholderText('排')
        self.medical_platoon_input.returnPressed.connect(self.search_medical_screening)
        search_layout.addWidget(self.medical_platoon_input)
        
        # 班输入框
        self.medical_squad_input = QLineEdit()
        self.medical_squad_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.medical_squad_input.setPlaceholderText('班')
        self.medical_squad_input.returnPressed.connect(self.search_medical_screening)
        search_layout.addWidget(self.medical_squad_input)
        
        search_btn = QPushButton('搜索')
        search_btn.clicked.connect(self.search_medical_screening)
        self.setup_button_style(search_btn, 'primary')
        search_layout.addWidget(search_btn)
        
        reset_btn = QPushButton('重置')
        reset_btn.clicked.connect(self.reset_medical_screening_search)
        self.setup_button_style(reset_btn, 'normal')
        search_layout.addWidget(reset_btn)
        
        layout.addLayout(search_layout)
        
        # 全选复选框行
        select_all_layout = QHBoxLayout()
        self.medical_select_all_checkbox = QCheckBox('全选')
        self.medical_select_all_checkbox.stateChanged.connect(self.toggle_medical_select_all)
        self.medical_select_all_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #7F8C8D;
                border-radius: 2px;
                background-color: white;
            }
            QCheckBox::indicator:hover {
                border-color: #3498DB;
                background-color: #ECF0F1;
            }
            QCheckBox::indicator:checked {
                background-color: #3498DB;
                border-color: #2980B9;
            }
        """)
        select_all_layout.addWidget(self.medical_select_all_checkbox)
        select_all_layout.addStretch()
        layout.addLayout(select_all_layout)
        
        # 病史筛查表格
        self.medical_screening_table = QTableWidget()
        medical_headers = ['选择', '姓名', '性别', '公民身份号码', '应征地', '连', '排', '班', '带训班长信息', '筛查情况', '筛查日期', '备注', '身体状况', '精神状况', '操作', '删除']
        self.medical_screening_table.setColumnCount(len(medical_headers))
        self.medical_screening_table.setHorizontalHeaderLabels(medical_headers)
        self.medical_screening_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.medical_screening_table.setAlternatingRowColors(True)
        self.medical_screening_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # 应用统一的表格样式
        self.setup_table_style(self.medical_screening_table)  # 禁止直接编辑
        # 设置文本省略模式为末尾省略
        self.medical_screening_table.setTextElideMode(Qt.TextElideMode.ElideRight)
        
        # 确保行高设置生效
        self.medical_screening_table.verticalHeader().setDefaultSectionSize(40)
        
        # 连接双击事件
        self.medical_screening_table.cellDoubleClicked.connect(self.show_medical_screening_detail)
        
        # 设置表格自适应屏幕
        from PyQt5.QtWidgets import QHeaderView
        header = self.medical_screening_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)  # 允许手动调整
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 选择列
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 姓名列
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 性别列
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 身份证列
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 应征地列
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 连列
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 排列
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # 班列
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # 带训班长信息列
        header.setSectionResizeMode(9, QHeaderView.Stretch)  # 筛查情况列自动拉伸
        header.setSectionResizeMode(10, QHeaderView.ResizeToContents)  # 筛查日期列
        header.setSectionResizeMode(11, QHeaderView.ResizeToContents)  # 备注列
        header.setSectionResizeMode(12, QHeaderView.ResizeToContents)  # 身体状况列
        header.setSectionResizeMode(13, QHeaderView.ResizeToContents)  # 精神状况列
        header.setSectionResizeMode(14, QHeaderView.ResizeToContents)  # 操作列
        header.setSectionResizeMode(15, QHeaderView.ResizeToContents)  # 删除列
        header.setStretchLastSection(False)
        header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        
        layout.addWidget(self.medical_screening_table)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        import_btn = QPushButton('导入数据')
        import_btn.clicked.connect(lambda: self.import_medical_screening_data())
        self.setup_button_style(import_btn, 'normal')
        button_layout.addWidget(import_btn)
        
        export_btn = QPushButton('导出数据')
        export_btn.clicked.connect(lambda: self.export_medical_screening_data())
        self.setup_button_style(export_btn, 'normal')
        button_layout.addWidget(export_btn)
        
        add_btn = QPushButton('添加信息')
        add_btn.clicked.connect(self.add_medical_screening_record)
        self.setup_button_style(add_btn, 'normal')
        button_layout.addWidget(add_btn)
        
        batch_delete_btn = QPushButton('批量删除')
        self.setup_button_style(batch_delete_btn, 'danger')
        batch_delete_btn.clicked.connect(self.batch_delete_medical_screening)
        button_layout.addWidget(batch_delete_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        tab_widget.setLayout(layout)
        
        # 加载病史筛查数据
        self.load_medical_screening_data()
        
        return tab_widget
    
    def load_medical_screening_data(self):
        """加载病史筛查数据"""
        try:
            conn = None
            try:
                conn = self.db_manager.get_connection()
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                # 获取时间筛选条件
                time_condition, time_params = self.get_time_filter_condition("ms.screening_date")
                
                # 使用联表查询从基本信息中获取应征地、连、排、班、带训班长信息
                base_query = """
                    SELECT ms.id, ms.name, ms.gender, ms.id_card, 
                           COALESCE(y.recruitment_place, '') as recruitment_place,
                           COALESCE(y.company, '') as company,
                           COALESCE(y.platoon, '') as platoon,
                           COALESCE(y.squad, '') as squad,
                           COALESCE(y.squad_leader, '') as squad_leader,
                           ms.screening_result, ms.screening_date, 
                           ms.remark, ms.physical_status, ms.mental_status
                    FROM medical_screening ms
                    LEFT JOIN youth y ON ms.id_card = y.id_card
                """
                
                if time_condition:
                    query = base_query + f" WHERE {time_condition} ORDER BY ms.screening_date DESC, ms.id DESC"
                    cursor.execute(query, time_params)
                else:
                    query = base_query + " ORDER BY ms.screening_date DESC, ms.id DESC"
                    cursor.execute(query)
                
                results = cursor.fetchall()
                self.display_medical_screening_results(results)
            finally:
                if conn:
                    conn.close()
        except Exception as e:
            QMessageBox.warning(self, "加载错误", f"加载病史筛查数据时发生错误：{str(e)}")
            self.medical_screening_table.setRowCount(0)
    
    def search_medical_screening(self):
        """搜索病史筛查记录"""
        try:
            name = self.medical_name_input.text().strip()
            id_card = self.medical_id_card_input.text().strip()
            recruitment_place = self.medical_recruitment_place_input.text().strip()
            company = self.medical_company_input.text().strip()
            platoon = self.medical_platoon_input.text().strip()
            squad = self.medical_squad_input.text().strip()
            
            conn = None
            try:
                conn = self.db_manager.get_connection()
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                # 使用联表查询，支持新增的搜索字段
                sql = """
                    SELECT ms.id, ms.name, ms.gender, ms.id_card, 
                           COALESCE(y.recruitment_place, '') as recruitment_place,
                           COALESCE(y.company, '') as company,
                           COALESCE(y.platoon, '') as platoon,
                           COALESCE(y.squad, '') as squad,
                           COALESCE(y.squad_leader, '') as squad_leader,
                           ms.screening_result, ms.screening_date, 
                           ms.remark, ms.physical_status, ms.mental_status
                    FROM medical_screening ms
                    LEFT JOIN youth y ON ms.id_card = y.id_card
                    WHERE 1=1
                """
                params = []
                
                if name:
                    sql += " AND ms.name LIKE ?"
                    params.append(f"%{name}%")
                
                if id_card:
                    sql += " AND ms.id_card LIKE ?"
                    params.append(f"%{id_card}%")
                
                if recruitment_place:
                    sql += " AND y.recruitment_place LIKE ?"
                    params.append(f"%{recruitment_place}%")
                
                if company:
                    sql += " AND y.company LIKE ?"
                    params.append(f"%{company}%")
                
                if platoon:
                    sql += " AND y.platoon LIKE ?"
                    params.append(f"%{platoon}%")
                
                if squad:
                    sql += " AND y.squad LIKE ?"
                    params.append(f"%{squad}%")
                
                sql += " ORDER BY ms.screening_date DESC, ms.id DESC"
                
                cursor.execute(sql, params)
                results = cursor.fetchall()
                self.display_medical_screening_results(results)
                
            finally:
                if conn:
                    conn.close()
                    
        except Exception as e:
            QMessageBox.warning(self, "搜索错误", f"搜索病史筛查记录时发生错误：{str(e)}")
    
    def reset_medical_screening_search(self):
        """重置病史筛查搜索"""
        self.medical_name_input.clear()
        self.medical_id_card_input.clear()
        self.medical_recruitment_place_input.clear()
        self.medical_company_input.clear()
        self.medical_platoon_input.clear()
        self.medical_squad_input.clear()
        self.load_medical_screening_data()
    
    def display_medical_screening_results(self, results):
        """显示病史筛查结果"""
        self.medical_screening_table.setRowCount(len(results))
        
        # 保存当前结果数据供双击事件使用
        self.current_medical_screening_results = results

        for row, data in enumerate(results):
            # 检查是否为异常数据
            physical_status = str(data['physical_status'] if data['physical_status'] else '').strip()
            mental_status = str(data['mental_status'] if data['mental_status'] else '').strip()
            is_abnormal = '异常' in physical_status or '异常' in mental_status
            
            # 第一列：复选框
            checkbox = QCheckBox()
            checkbox.setStyleSheet("""
                    QCheckBox {
                        spacing: 0px;
                    }
                    QCheckBox::indicator {
                        width: 20px;
                        height: 20px;
                        border: 2px solid #7F8C8D;
                        border-radius: 2px;
                        background-color: white;
                    }
                    QCheckBox::indicator:hover {
                        border-color: #3498DB;
                        background-color: #ECF0F1;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #3498DB;
                        border-color: #2980B9;
                    }
                """)
            checkbox.setMinimumSize(24, 24)
            
            checkbox_layout = QHBoxLayout()
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.setContentsMargins(5, 5, 5, 5)
            checkbox_layout.addWidget(checkbox)
            checkbox_container = QWidget()
            checkbox_container.setLayout(checkbox_layout)
            self.medical_screening_table.setCellWidget(row, 0, checkbox_container)

            # 第二列：姓名
            name_text = str(data['name'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(name_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 1, item)

            # 第三列：性别
            gender_text = str(data['gender'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(gender_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 2, item)

            # 第四列：公民身份号码
            id_card_text = str(data['id_card'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(id_card_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 3, item)

            # 第五列：应征地
            recruitment_place_text = str(data['recruitment_place'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(recruitment_place_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 4, item)

            # 第六列：连
            company_text = str(data['company'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(company_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 5, item)

            # 第七列：排
            platoon_text = str(data['platoon'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(platoon_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 6, item)

            # 第八列：班
            squad_text = str(data['squad'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(squad_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 7, item)

            # 第九列：带训班长信息
            squad_leader_text = str(data['squad_leader'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(squad_leader_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 8, item)

            # 第十列：筛查情况
            screening_text = str(data['screening_result'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(screening_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            item.setToolTip(screening_text)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 9, item)

            # 第十一列：筛查日期
            date_text = str(data['screening_date'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(date_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 10, item)

            # 第十二列：备注
            remark_text = str(data['remark'] or '').replace('\n', ' ').replace('\r', ' ')
            item = QTableWidgetItem(remark_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 11, item)

            # 第十三列：身体状况
            item = QTableWidgetItem(physical_status)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 12, item)

            # 第十四列：精神状况
            item = QTableWidgetItem(mental_status)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            if is_abnormal:
                item.setBackground(QColor(255, 182, 193, 100))
            self.medical_screening_table.setItem(row, 13, item)

            # 第十五列：查看详情按钮
            medical_id = data['id']
            view_btn = QPushButton('查看详情')
            view_btn.setStyleSheet('background-color: #ADD8E6;')
            view_btn.clicked.connect(lambda checked, mid=medical_id: self.view_medical_screening_details(mid))
            view_btn_layout = QHBoxLayout()
            view_btn_layout.setAlignment(Qt.AlignCenter)
            view_btn_layout.setContentsMargins(5, 5, 5, 5)
            view_btn_layout.addWidget(view_btn)
            view_btn_container = QWidget()
            view_btn_container.setLayout(view_btn_layout)
            self.medical_screening_table.setCellWidget(row, 14, view_btn_container)

            # 第十六列：删除按钮
            delete_btn = QPushButton('删除')
            delete_btn.setStyleSheet('background-color: #FFB6C1;')
            delete_btn.clicked.connect(lambda checked, mid=medical_id: self.delete_medical_screening(mid))
            delete_btn_layout = QHBoxLayout()
            delete_btn_layout.setAlignment(Qt.AlignCenter)
            delete_btn_layout.setContentsMargins(5, 5, 5, 5)
            delete_btn_layout.addWidget(delete_btn)
            delete_btn_container = QWidget()
            delete_btn_container.setLayout(delete_btn_layout)
            self.medical_screening_table.setCellWidget(row, 15, delete_btn_container)
        
        # 确保所有行的行高都是40像素
        for row in range(self.medical_screening_table.rowCount()):
            self.medical_screening_table.setRowHeight(row, 40)
    
    def add_medical_screening_record(self):
        """添加病史筛查记录"""
        from ui.add_medical_screening_dialog import AddMedicalScreeningDialog
        
        dialog = AddMedicalScreeningDialog(self.db_manager, self)
        dialog.data_updated.connect(self.load_medical_screening_data)
        dialog.exec_()
    
    def show_medical_screening_detail(self, row, column):
        """显示病史筛查详情（双击事件）"""
        try:
            # 从保存的结果数据中获取记录ID
            if not hasattr(self, 'current_medical_screening_results') or row >= len(self.current_medical_screening_results):
                QMessageBox.warning(self, '错误', '无法获取记录信息')
                return
            
            data = self.current_medical_screening_results[row]
            medical_id = data['id']
            
            # 调用查看详情方法（与点击"查看详情"按钮相同）
            self.view_medical_screening_details(medical_id)
                
        except Exception as e:
            QMessageBox.warning(self, '错误', f'查看详情时发生错误: {str(e)}')
    
    def delete_medical_screening(self, medical_id):
        """删除单条病史筛查记录"""
        try:
            reply = QMessageBox.question(
                self, 
                '确认删除', 
                '确定要删除这条病史筛查记录吗？\n删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                conn = None
                try:
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    
                    cursor.execute("DELETE FROM medical_screening WHERE id = ?", (medical_id,))
                    conn.commit()
                    
                    QMessageBox.information(self, '删除成功', '病史筛查记录已删除')
                    self.load_medical_screening_data()
                    
                finally:
                    if conn:
                        conn.close()
                        
        except Exception as e:
            QMessageBox.critical(self, '删除失败', f'删除病史筛查记录时发生错误：{str(e)}')
    
    def batch_delete_medical_screening(self):
        """批量删除病史筛查记录"""
        try:
            selected_ids = []
            selected_names = []
            
            # 先从当前显示的数据中获取
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, name FROM medical_screening ORDER BY id
            """)
            all_results = cursor.fetchall()
            conn.close()
            
            for row in range(self.medical_screening_table.rowCount()):
                checkbox_widget = self.medical_screening_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_results):
                            data = all_results[row]
                            medical_id = data['id']
                            name = data['name']
                            
                            if medical_id:
                                selected_ids.append(medical_id)
                                selected_names.append(name)
            
            if not selected_ids:
                QMessageBox.information(self, '提示', '请先选择要删除的病史筛查记录')
                return
            
            names_text = '、'.join(selected_names[:5])
            if len(selected_names) > 5:
                names_text += f' 等{len(selected_names)}人'
            
            reply = QMessageBox.question(
                self,
                '确认批量删除',
                f'确定要删除以下病史筛查记录吗？\n{names_text}\n\n共{len(selected_ids)}条记录，删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                conn = None
                try:
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                    
                    # 在删除前收集需要同步的记录信息
                    records_to_sync = []
                    for medical_id in selected_ids:
                        cursor.execute("""
                            SELECT id_card, screening_date 
                            FROM medical_screening WHERE id = ?
                        """, (medical_id,))
                        record_info = cursor.fetchone()
                        if record_info:
                            records_to_sync.append(record_info)
                    
                    deleted_count = 0
                    for medical_id in selected_ids:
                        cursor.execute("DELETE FROM medical_screening WHERE id = ?", (medical_id,))
                        if cursor.rowcount > 0:
                            deleted_count += 1
                    
                    conn.commit()
                    
                    QMessageBox.information(self, '删除成功', f'成功删除 {deleted_count} 条病史筛查记录')
                    
                    # 取消全选
                    if hasattr(self, 'medical_select_all_checkbox'):
                        self.medical_select_all_checkbox.setChecked(False)
                    
                    self.load_medical_screening_data()
                    
                finally:
                    if conn:
                        conn.close()
                        
        except Exception as e:
            QMessageBox.critical(self, '批量删除失败', f'批量删除病史筛查记录时发生错误：{str(e)}')
    
    def toggle_medical_select_all(self, state):
        """全选/取消全选病史筛查"""
        is_checked = (state == Qt.CheckState.Checked.value)
        for row in range(self.medical_screening_table.rowCount()):
            checkbox_widget = self.medical_screening_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(is_checked)
    
    def view_medical_screening_details(self, medical_id):
        """查看病史筛查详情"""
        try:
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, youth_id_card, name, gender, id_card, 
                       screening_result, screening_date, remark, physical_status, mental_status
                FROM medical_screening 
                WHERE id = ?
            ''', (medical_id,))
            
            record = cursor.fetchone()
            conn.close()
            
            if record:
                from ui.medical_screening_detail_dialog import MedicalScreeningDetailDialog
                dialog = MedicalScreeningDetailDialog(self.db_manager, record, self)
                dialog.data_updated.connect(self.load_medical_screening_data)
                dialog.exec_()
            else:
                QMessageBox.warning(self, '错误', '未找到该记录')
                
        except Exception as e:
            QMessageBox.warning(self, '错误', f'查看详情时发生错误: {str(e)}')
    
    def import_medical_screening_data(self):
        """导入病史筛查数据"""
        file_path, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '', 'Excel Files (*.xlsx *.xls)')
        if not file_path:
            return
        
        count, error = self.import_service.import_medical_screening(file_path)
        if error:
            if count > 0:
                QMessageBox.warning(self, '导入完成（有错误）', error)
                self.load_medical_screening_data()
            else:
                QMessageBox.warning(self, '导入失败', error)
        else:
            QMessageBox.information(self, '导入成功', f'成功导入 {count} 条病史筛查记录')
            self.load_medical_screening_data()
            # 额外刷新所有已打开的青年详情对话框（如果有的话），确保详情页也能看到新导入的数据
            try:
                from PyQt5.QtWidgets import QApplication
                from ui.youth_detail_dialog import YouthDetailDialog

                for widget in QApplication.topLevelWidgets():
                    if isinstance(widget, YouthDetailDialog):
                        if hasattr(widget, 'load_medical_screening_data'):
                            try:
                                widget.load_medical_screening_data()
                            except Exception:
                                # 忽略单个对话框的刷新错误
                                pass
            except Exception:
                pass
    
    def export_medical_screening_data(self):
        """导出病史筛查数据 - 勾选记录时导出选中的，不勾选时导出全部"""
        try:
            # 检查是否有选中的记录
            selected_ids = []
            selected_names = []
            
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, name FROM medical_screening ORDER BY id
            """)
            all_results = cursor.fetchall()
            conn.close()
            
            for row in range(self.medical_screening_table.rowCount()):
                checkbox_widget = self.medical_screening_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        if row < len(all_results):
                            data = all_results[row]
                            medical_id = data['id']
                            name = data['name']
                            
                            if medical_id:
                                selected_ids.append(medical_id)
                                selected_names.append(name)
            
            # 根据是否有选中记录决定导出方式
            if selected_ids:
                # 有选中记录，导出选中的
                names_text = '、'.join(selected_names[:5])
                if len(selected_names) > 5:
                    names_text += f' 等{len(selected_names)}人'
                
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    '导出选中的病史筛查数据',
                    f'病史筛查数据_选中_{datetime.now().strftime("%Y%m%d")}.xlsx',
                    'Excel Files (*.xlsx)'
                )
                
                if not file_path:
                    return
                
                # 调用导出服务导出选中的记录
                success, message = self.export_service.export_selected_medical_screening_to_excel(file_path, selected_ids)
                
                if success:
                    QMessageBox.information(self, '导出成功', f'{message}\n\n导出记录：{names_text}')
                    
                    # 取消全选
                    if hasattr(self, 'medical_select_all_checkbox'):
                        self.medical_select_all_checkbox.setChecked(False)
                else:
                    QMessageBox.warning(self, '导出失败', message)
            else:
                # 没有选中记录，导出全部
                file_path, _ = QFileDialog.getSaveFileName(
                    self, 
                    '导出全部病史筛查数据', 
                    f'病史筛查数据_全部_{datetime.now().strftime("%Y%m%d")}.xlsx',
                    'Excel Files (*.xlsx)'
                )
                
                if not file_path:
                    return
                
                success, message = self.export_service.export_medical_screening_to_excel(file_path)
                if success:
                    QMessageBox.information(self, '导出成功', message)
                else:
                    QMessageBox.warning(self, '导出失败', message)
                    
        except Exception as e:
            QMessageBox.critical(self, '导出失败', f'导出病史筛查记录时发生错误：{str(e)}')
    
    def import_camp_verification_data(self):
        """导入入营点验情况数据"""
        file_path, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '', 'Excel Files (*.xlsx *.xls)')
        if not file_path:
            return
        
        try:
            from services.camp_verification_import_service import CampVerificationImportService
            
            # 创建导入服务
            import_service = CampVerificationImportService(self.db_manager)
            
            # 执行导入
            success_count, fail_count, error_messages, skipped_records = import_service.import_from_excel(file_path)
            
            # 生成消息
            total = success_count + fail_count + len(skipped_records)
            
            if total == 0:
                # 检查是否是表头问题
                if error_messages:
                    error_detail = '\n'.join(error_messages)
                    QMessageBox.critical(self, '导入失败', f'导入失败！\n\n{error_detail}')
                else:
                    QMessageBox.warning(self, '导入失败', '没有找到有效的数据行')
                return
            
            # 构建导入结果消息
            message = f'导入完成！\n\n成功：{success_count} 条\n失败：{fail_count} 条\n重复：{len(skipped_records)} 条'
            
            # 如果有被跳过（重复）的数据，添加提示和详细列表
            if len(skipped_records) > 0:
                message += '\n\n以下数据已经存在（身份证号重复）：\n'
                for idx, record in enumerate(skipped_records[:10], 1):
                    message += f'{idx}. 姓名: {record["username"]}，身份证: {record["user_id"]}\n'
                
                if len(skipped_records) > 10:
                    message += f'... 还有 {len(skipped_records) - 10} 条重复数据'
                
                QMessageBox.warning(self, '导入完成（有数据重复）', message)
            elif fail_count == 0:
                # 全部成功，没有重复
                QMessageBox.information(self, '导入完成', message)
            else:
                # 有失败数据
                error_text = '\n'.join(error_messages[:10])
                if len(error_messages) > 10:
                    error_text += f'\n... 还有 {len(error_messages) - 10} 条错误'
                
                message += f'\n\n错误详情：\n{error_text}'
                QMessageBox.warning(self, '导入完成（有错误）', message)
            
        except Exception as e:
            QMessageBox.critical(self, '导入错误', f'导入入营点验情况数据时发生错误：\n{str(e)}')
    
    def import_political_assessment_data(self):
        """导入政治考核情况数据"""
        file_path, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '', 'Excel Files (*.xlsx *.xls)')
        if not file_path:
            return
        
        try:
            # 执行导入
            imported_count, error_message = self.import_service.import_political_assessment(file_path)
            
            if error_message:
                QMessageBox.warning(self, '导入完成（有错误）', error_message)
            else:
                QMessageBox.information(self, '导入成功', f'成功导入 {imported_count} 条政治考核情况记录')
            
        except Exception as e:
            QMessageBox.critical(self, '导入错误', f'导入政治考核情况数据时发生错误：\n{str(e)}')
    
    def load_exception_statistics_data(self):
        """加载异常情况统计数据到表格"""
        # 直接调用统一的筛选方法
        self.apply_all_filters()
    
    def filter_exception_by_date_range(self):
        """根据日期范围筛选异常情况统计数据"""
        try:
            # 重置全选按钮状态
            if hasattr(self, 'exception_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操作
                self.exception_select_all_checkbox.stateChanged.disconnect()
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
            # 将下拉框选项设置为"无"
            if hasattr(self, 'exception_time_range_combo'):
                # 临时断开信号连接，避免触发下拉框变化事件
                self.exception_time_range_combo.currentTextChanged.disconnect()
                self.exception_time_range_combo.setCurrentText('无')
                # 重新连接信号
                self.exception_time_range_combo.currentTextChanged.connect(self.on_exception_time_range_changed)
            
            # 获取日期范围
            start_date = self.exception_start_date.date().toString('yyyy-MM-dd')
            end_date = self.exception_end_date.date().toString('yyyy-MM-dd')
            
            # 验证日期范围
            if start_date > end_date:
                QMessageBox.warning(self, "日期错误", "开始日期不能晚于结束日期")
                return
            
            # 获取所有搜索条件
            name = self.exception_name_input.text().strip()
            id_card = self.exception_id_card_input.text().strip()
            recruitment_place = self.exception_recruitment_place_input.text().strip()
            company = self.exception_company_input.text().strip()
            platoon = self.exception_platoon_input.text().strip()
            squad = self.exception_squad_input.text().strip()
            
            # 使用视图数据进行筛选（同时应用日期范围和搜索条件）
            view_records = self.db_manager.get_exception_statistics_view_data(
                start_date=start_date,
                end_date=end_date,
                name=name if name else None,
                id_card=id_card if id_card else None,
                recruitment_place=recruitment_place if recruitment_place else None,
                company=company if company else None,
                platoon=platoon if platoon else None,
                squad=squad if squad else None
            )
            
            # 直接使用视图数据
            self.display_exception_statistics_records(view_records)
            
        except Exception as e:
            QMessageBox.warning(self, "筛选错误", f"按日期范围筛选数据时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'exception_select_all_checkbox'):
                    self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
                if hasattr(self, 'exception_time_range_combo'):
                    self.exception_time_range_combo.currentTextChanged.connect(self.on_exception_time_range_changed)
            except:
                pass

    def on_exception_time_range_changed(self, selected_range):
        """时间范围下拉框变化处理"""
        try:
            # 重置全选按钮状态
            if hasattr(self, 'exception_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操作
                self.exception_select_all_checkbox.stateChanged.disconnect()
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
            # 如果选择的不是"无"，则清空日期筛选器
            if selected_range != '无' and hasattr(self, 'exception_start_date') and hasattr(self, 'exception_end_date'):
                from PyQt5.QtCore import QDate
                # 重置日期选择器为默认值
                self.exception_start_date.setDate(QDate.currentDate().addDays(-30))
                self.exception_end_date.setDate(QDate.currentDate())
            
            # 如果选择"无"，不进行任何筛选操作，保持当前显示
            if selected_range == '无':
                return
            
            # 统一使用搜索方法，它会同时考虑搜索条件和时间范围
            self.apply_all_filters()
            
        except Exception as e:
            QMessageBox.warning(self, "时间范围切换错误", f"切换时间范围时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'exception_select_all_checkbox'):
                    self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            except:
                pass
    
    def apply_all_filters(self):
        """应用所有筛选条件（搜索条件 + 时间范围）"""
        try:
            # 重置全选按钮状态
            if hasattr(self, 'exception_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操作
                self.exception_select_all_checkbox.stateChanged.disconnect()
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
            # 获取所有搜索条件
            name = self.exception_name_input.text().strip()
            id_card = self.exception_id_card_input.text().strip()
            recruitment_place = self.exception_recruitment_place_input.text().strip()
            company = self.exception_company_input.text().strip()
            platoon = self.exception_platoon_input.text().strip()
            squad = self.exception_squad_input.text().strip()
            
            # 获取当前选择的时间范围
            time_range = getattr(self, 'exception_time_range_combo', None)
            if time_range:
                selected_range = time_range.currentText()
            else:
                selected_range = '全部'  # 默认值
            
            # 根据时间范围计算日期条件
            start_date = None
            end_date = None
            
            # 首先检查全局时间筛选
            year = self.year_combo.currentText()
            if year != '全部':
                half_year = self.half_year_combo.currentText()
                
                if half_year == '上半年':
                    start_date = f"{year}-01-01"
                    end_date = f"{year}-06-30"
                else:  # 下半年
                    start_date = f"{year}-07-01"
                    end_date = f"{year}-12-31"
            
            # 如果有具体的时间范围选择，则覆盖全局时间筛选
            if selected_range != '全部' and selected_range != '无':
                from datetime import datetime, timedelta
                today = datetime.now()
                
                if selected_range == '当天':
                    start_date = end_date = today.strftime('%Y-%m-%d')
                elif selected_range == '三天内':
                    start_date = (today - timedelta(days=2)).strftime('%Y-%m-%d')
                    end_date = today.strftime('%Y-%m-%d')
                elif selected_range == '七天内':
                    start_date = (today - timedelta(days=6)).strftime('%Y-%m-%d')
                    end_date = today.strftime('%Y-%m-%d')
                elif selected_range == '半个月内':
                    start_date = (today - timedelta(days=14)).strftime('%Y-%m-%d')
                    end_date = today.strftime('%Y-%m-%d')
            
            # 使用视图数据进行搜索
            view_records = self.db_manager.get_exception_statistics_view_data(
                start_date=start_date,
                end_date=end_date,
                name=name if name else None,
                id_card=id_card if id_card else None,
                recruitment_place=recruitment_place if recruitment_place else None,
                company=company if company else None,
                platoon=platoon if platoon else None,
                squad=squad if squad else None
            )
            
            # 直接使用视图数据
            self.display_exception_statistics_records(view_records)
            
        except Exception as e:
            QMessageBox.warning(self, "筛选错误", f"应用筛选条件时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'exception_select_all_checkbox'):
                    self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            except:
                pass

    def search_exception_statistics(self):
        """搜索异常情况统计记录"""
        # 直接调用统一的筛选方法
        self.apply_all_filters()
    
    def reset_exception_statistics_search(self):
        """重置异常情况统计搜索"""
        try:
            # 重置全选按钮状态
            if hasattr(self, 'exception_select_all_checkbox'):
                # 临时断开信号连接，避免触发全选操作
                self.exception_select_all_checkbox.stateChanged.disconnect()
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                # 重新连接信号
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
            # 清空所有搜索框
            self.exception_name_input.clear()
            self.exception_id_card_input.clear()
            self.exception_recruitment_place_input.clear()
            self.exception_company_input.clear()
            self.exception_platoon_input.clear()
            self.exception_squad_input.clear()
            
            # 重置时间范围为默认值
            if hasattr(self, 'exception_time_range_combo'):
                self.exception_time_range_combo.setCurrentText('全部')
            
            # 重置日期筛选器
            if hasattr(self, 'exception_start_date') and hasattr(self, 'exception_end_date'):
                from PyQt5.QtCore import QDate
                self.exception_start_date.setDate(QDate.currentDate().addDays(-30))
                self.exception_end_date.setDate(QDate.currentDate())
            
            # 应用重置后的筛选条件（实际上是加载全部数据）
            self.apply_all_filters()
            
        except Exception as e:
            QMessageBox.warning(self, "重置错误", f"重置搜索时发生错误：{str(e)}")
            # 确保信号重新连接
            try:
                if hasattr(self, 'exception_select_all_checkbox'):
                    self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            except:
                pass
    
    def display_exception_statistics_records(self, records):
        """显示异常情况统计记录"""
        try:
            self.exception_statistics_table.setRowCount(len(records))
            
            for row, record in enumerate(records):
                # 选择框
                checkbox = QCheckBox()
                checkbox.setStyleSheet("""
                    QCheckBox {
                        spacing: 0px;
                    }
                    QCheckBox::indicator {
                        width: 20px;
                        height: 20px;
                        border: 2px solid #7F8C8D;
                        border-radius: 2px;
                        background-color: white;
                    }
                    QCheckBox::indicator:hover {
                        border-color: #3498DB;
                        background-color: #ECF0F1;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #3498DB;
                        border-color: #2980B9;
                    }
                """)
                checkbox.setMinimumSize(24, 24)  # 确保复选框有足够空间显示
                
                # 添加状态变化监听，当单个复选框状态改变时更新全选按钮状态
                checkbox.stateChanged.connect(self.update_exception_select_all_state)
                
                checkbox_layout = QHBoxLayout()
                checkbox_layout.setAlignment(Qt.AlignCenter)
                checkbox_layout.setContentsMargins(5, 5, 5, 5)
                checkbox_layout.addWidget(checkbox)
                checkbox_container = QWidget()
                checkbox_container.setLayout(checkbox_layout)
                self.exception_statistics_table.setCellWidget(row, 0, checkbox_container)
                
                # 从视图数据中获取字段（record是元组）
                # 视图字段顺序：公民身份号码, 姓名, 性别, 连, 排, 班, 带训班长, 应征地, 思想是否异常, 身体是否异常, 精神是否异常, 训练是否异常, 管理是否异常, 其他, 日期
                
                # 姓名
                name_item = QTableWidgetItem(str(record[1] or ''))
                name_item.setToolTip(str(record[1] or ''))  # 设置工具提示
                self.exception_statistics_table.setItem(row, 1, name_item)
                
                # 日期
                date_item = QTableWidgetItem(str(record[14] or ''))
                date_item.setToolTip(str(record[14] or ''))  # 设置工具提示
                self.exception_statistics_table.setItem(row, 2, date_item)
                
                # 公民身份号码
                id_item = QTableWidgetItem(str(record[0] or ''))
                id_item.setToolTip(str(record[0] or ''))  # 设置工具提示
                self.exception_statistics_table.setItem(row, 3, id_item)
                
                # 性别
                gender_item = QTableWidgetItem(str(record[2] or ''))
                gender_item.setToolTip(str(record[2] or ''))  # 设置工具提示
                self.exception_statistics_table.setItem(row, 4, gender_item)
                
                # 应征地
                recruitment_item = QTableWidgetItem(str(record[7] or ''))
                recruitment_item.setToolTip(str(record[7] or ''))  # 设置工具提示
                self.exception_statistics_table.setItem(row, 5, recruitment_item)
                
                # 连
                company_item = QTableWidgetItem(str(record[3] or ''))
                company_item.setToolTip(str(record[3] or ''))  # 设置工具提示
                self.exception_statistics_table.setItem(row, 6, company_item)
                
                # 排
                platoon_item = QTableWidgetItem(str(record[4] or ''))
                platoon_item.setToolTip(str(record[4] or ''))  # 设置工具提示
                self.exception_statistics_table.setItem(row, 7, platoon_item)
                
                # 班
                squad_item = QTableWidgetItem(str(record[5] or ''))
                squad_item.setToolTip(str(record[5] or ''))  # 设置工具提示
                self.exception_statistics_table.setItem(row, 8, squad_item)
                
                # 带训班长信息
                leader_item = QTableWidgetItem(str(record[6] or ''))
                leader_item.setToolTip(str(record[6] or ''))  # 设置工具提示
                self.exception_statistics_table.setItem(row, 9, leader_item)
                
                # 思想
                thought_text = record[8] if record[8] == '异常' else '正常'
                thought_item = QTableWidgetItem(thought_text)
                thought_item.setToolTip(thought_text)  # 设置工具提示
                if thought_text == '异常':
                    thought_item.setBackground(QColor(255, 204, 204))  # 浅红色
                self.exception_statistics_table.setItem(row, 10, thought_item)
                
                # 身体
                body_text = record[9] if record[9] == '异常' else '正常'
                body_item = QTableWidgetItem(body_text)
                body_item.setToolTip(body_text)  # 设置工具提示
                if body_text == '异常':
                    body_item.setBackground(QColor(255, 204, 204))  # 浅红色
                self.exception_statistics_table.setItem(row, 11, body_item)
                
                # 精神
                spirit_text = record[10] if record[10] == '异常' else '正常'
                spirit_item = QTableWidgetItem(spirit_text)
                spirit_item.setToolTip(spirit_text)  # 设置工具提示
                if spirit_text == '异常':
                    spirit_item.setBackground(QColor(255, 204, 204))  # 浅红色
                self.exception_statistics_table.setItem(row, 12, spirit_item)
                
                # 训练
                training_text = record[11] if record[11] == '异常' else '正常'
                training_item = QTableWidgetItem(training_text)
                training_item.setToolTip(training_text)  # 设置工具提示
                if training_text == '异常':
                    training_item.setBackground(QColor(255, 204, 204))  # 浅红色
                self.exception_statistics_table.setItem(row, 13, training_item)
                
                # 管理
                management_text = record[12] if record[12] == '异常' else '正常'
                management_item = QTableWidgetItem(management_text)
                management_item.setToolTip(management_text)  # 设置工具提示
                if management_text == '异常':
                    management_item.setBackground(QColor(255, 204, 204))  # 浅红色
                self.exception_statistics_table.setItem(row, 14, management_item)
                
                # 其他（异常来源）
                source_text = str(record[13] or '')
                # 在内容前面添加"来源："前缀
                if source_text.strip():
                    source_text = f"来源：{source_text}"
                source_item = QTableWidgetItem(source_text)
                # 设置文本换行和工具提示
                source_item.setToolTip(source_text)  # 设置工具提示显示完整内容
                self.exception_statistics_table.setItem(row, 15, source_item)
                
                # 查看详情按钮
                view_btn = QPushButton('查看详情')
                view_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝色
                view_btn.clicked.connect(lambda checked, user_id=record[0], name=record[1]: 
                                       self.show_exception_statistics_detail(user_id, name))
                self.exception_statistics_table.setCellWidget(row, 16, view_btn)
            
            # 数据加载完成后，调整列宽以适应内容
            self.exception_statistics_table.resizeColumnsToContents()
            
            # 确保重要列有足够的宽度
            header = self.exception_statistics_table.horizontalHeader()
            # 姓名列最小宽度
            if self.exception_statistics_table.columnWidth(1) < 120:
                self.exception_statistics_table.setColumnWidth(1, 120)
            # 公民身份号码列最小宽度
            if self.exception_statistics_table.columnWidth(3) < 180:
                self.exception_statistics_table.setColumnWidth(3, 180)
            # 异常来源列最小宽度
            if self.exception_statistics_table.columnWidth(15) < 200:
                self.exception_statistics_table.setColumnWidth(15, 200)
                
        except Exception as e:
            QMessageBox.warning(self, "显示错误", f"显示异常情况统计记录时发生错误：{str(e)}")
    
    def on_exception_statistics_double_click(self, row, column):
        """异常统计表格双击事件处理"""
        try:
            # 获取当前行的用户ID和姓名
            user_id_item = self.exception_statistics_table.item(row, 3)  # 公民身份号码列
            name_item = self.exception_statistics_table.item(row, 1)     # 姓名列
            
            if user_id_item and name_item:
                user_id = user_id_item.text()
                name = name_item.text()
                self.show_exception_statistics_detail(user_id, name)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"打开详情时发生错误：{str(e)}")
    
    def show_exception_statistics_detail(self, user_id, name):
        """显示异常统计详情弹窗"""
        try:
            from .exception_statistics_detail_dialog import ExceptionStatisticsDetailDialog
            dialog = ExceptionStatisticsDetailDialog(self.db_manager, self, user_id, name)
            dialog.exec_()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"打开异常统计详情时发生错误：{str(e)}")
    
    def toggle_exception_select_all(self, state):
        """异常情况统计全选/取消全选"""
        is_checked = (state == Qt.CheckState.Checked.value)
        
        # 临时标记，防止在全选操作时触发单个复选框的状态更新
        self._updating_exception_select_all = True
        
        for row in range(self.exception_statistics_table.rowCount()):
            checkbox_widget = self.exception_statistics_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(is_checked)
        
        # 重置标记
        self._updating_exception_select_all = False
    
    def update_exception_select_all_state(self):
        """更新异常情况统计全选按钮状态"""
        # 如果正在执行全选操作，则不更新全选按钮状态
        if getattr(self, '_updating_exception_select_all', False):
            return
        
        try:
            total_rows = self.exception_statistics_table.rowCount()
            if total_rows == 0:
                return
            
            checked_count = 0
            for row in range(total_rows):
                checkbox_widget = self.exception_statistics_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        checked_count += 1
            
            # 临时断开全选按钮的信号连接，避免触发toggle_exception_select_all
            self.exception_select_all_checkbox.stateChanged.disconnect()
            
            # 更新全选按钮状态
            if checked_count == 0:
                # 没有选中任何项
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            elif checked_count == total_rows:
                # 全部选中
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.Checked)
            else:
                # 部分选中
                self.exception_select_all_checkbox.setCheckState(Qt.CheckState.PartiallyChecked)
            
            # 重新连接信号
            self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            
        except Exception as e:
            print(f"更新异常统计全选按钮状态时出错: {e}")
            # 确保信号重新连接
            try:
                self.exception_select_all_checkbox.stateChanged.connect(self.toggle_exception_select_all)
            except:
                pass
    
    def view_exception_statistics(self, record_id):
        """查看异常情况统计记录详情"""
        try:
            # 从视图数据中获取记录详情
            all_records = self.db_manager.get_exception_statistics_view_data()
            record = None
            for r in all_records:
                # 使用身份证号和日期作为唯一标识
                if f"{r[0]}_{r[14]}" == record_id:
                    record = r
                    break
            
            if not record:
                QMessageBox.warning(self, '错误', '找不到该记录')
                return
            
            # 使用新的详情弹窗
            self.show_exception_statistics_detail(record[0], record[1])  # 身份证号, 姓名
            
        except Exception as e:
            QMessageBox.critical(self, '查看失败', f'查看异常情况统计记录时发生错误：{str(e)}')
    
    def delete_single_exception_statistics(self, record_id, name):
        """删除单个异常情况统计记录"""
        try:
            # 异常统计现在基于视图，无法直接删除
            # 需要删除对应的源数据才能从视图中移除
            QMessageBox.information(
                self,
                '操作说明',
                '异常统计数据现在基于实时视图生成。\n'
                '要移除异常记录，请到对应的数据源（如病史筛查、每日统计等）中修改或删除相关数据。'
            )
                    
        except Exception as e:
            QMessageBox.critical(self, '操作失败', f'操作异常情况统计记录时发生错误：{str(e)}')
    


    # 新的每日情况统计方法
    def create_daily_stats_tab(self):
        """创建每日情况统计标签页"""
        tab_widget = QWidget()
        layout = QVBoxLayout()
        
        # 搜索栏
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel('搜索:'))
        
        self.daily_name_input = QLineEdit()
        self.daily_name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.daily_name_input.setPlaceholderText('姓名')
        self.daily_name_input.returnPressed.connect(self.search_daily_stats)
        search_layout.addWidget(self.daily_name_input)
        
        self.daily_id_card_input = QLineEdit()
        self.daily_id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.daily_id_card_input.setPlaceholderText('公民身份号码')
        self.daily_id_card_input.returnPressed.connect(self.search_daily_stats)
        search_layout.addWidget(self.daily_id_card_input)
        
        self.daily_recruitment_place_input = QLineEdit()
        self.daily_recruitment_place_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.daily_recruitment_place_input.setPlaceholderText('应征地')
        self.daily_recruitment_place_input.returnPressed.connect(self.search_daily_stats)
        search_layout.addWidget(self.daily_recruitment_place_input)
        
        self.daily_company_input = QLineEdit()
        self.daily_company_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.daily_company_input.setPlaceholderText('连')
        self.daily_company_input.returnPressed.connect(self.search_daily_stats)
        search_layout.addWidget(self.daily_company_input)
        
        self.daily_platoon_input = QLineEdit()
        self.daily_platoon_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.daily_platoon_input.setPlaceholderText('排')
        self.daily_platoon_input.returnPressed.connect(self.search_daily_stats)
        search_layout.addWidget(self.daily_platoon_input)
        
        self.daily_squad_input = QLineEdit()
        self.daily_squad_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.daily_squad_input.setPlaceholderText('班')
        self.daily_squad_input.returnPressed.connect(self.search_daily_stats)
        search_layout.addWidget(self.daily_squad_input)
        
        search_btn = QPushButton('搜索')
        search_btn.clicked.connect(self.search_daily_stats)
        self.setup_button_style(search_btn, 'primary')
        search_layout.addWidget(search_btn)
        
        reset_btn = QPushButton('重置')
        reset_btn.clicked.connect(self.reset_daily_stats_search)
        self.setup_button_style(reset_btn, 'normal')
        search_layout.addWidget(reset_btn)
        
        layout.addLayout(search_layout)
        
        # 全选按钮行
        select_all_layout = QHBoxLayout()
        self.daily_select_all_checkbox = QCheckBox('全选')
        self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
        self.daily_select_all_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #7F8C8D;
                border-radius: 2px;
                background-color: white;
            }
            QCheckBox::indicator:hover {
                border-color: #3498DB;
                background-color: #ECF0F1;
            }
            QCheckBox::indicator:checked {
                background-color: #3498DB;
                border-color: #2980B9;
            }
        """)
        select_all_layout.addWidget(self.daily_select_all_checkbox)
        select_all_layout.addStretch()
        layout.addLayout(select_all_layout)
        
        # 每日情况统计表格
        self.daily_stats_table = QTableWidget()
        self.daily_stats_headers = [
            '选择', '公民身份号码', '姓名', '应征地', '连', '排', '班', '带训班长信息', 
            '日期', '思想', '身体', '精神', '训练', '管理', '其他', '修改', '删除'
        ]
        self.daily_stats_table.setColumnCount(len(self.daily_stats_headers))
        self.daily_stats_table.setHorizontalHeaderLabels(self.daily_stats_headers)
        self.daily_stats_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.daily_stats_table.setAlternatingRowColors(True)
        self.daily_stats_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # 应用统一的表格样式
        self.setup_table_style(self.daily_stats_table)
        
        # 设置表格列宽策略 - 采用与病史筛查情况相同的设置
        header = self.daily_stats_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)  # 允许手动调整
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 选择列
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 公民身份号码列
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 姓名列
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 应征地列
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 连列
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 排列
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 班列
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # 带训班长信息列
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # 日期列
        header.setSectionResizeMode(9, QHeaderView.Stretch)  # 思想列自动拉伸
        header.setSectionResizeMode(10, QHeaderView.ResizeToContents)  # 身体列
        header.setSectionResizeMode(11, QHeaderView.ResizeToContents)  # 精神列
        header.setSectionResizeMode(12, QHeaderView.ResizeToContents)  # 训练列
        header.setSectionResizeMode(13, QHeaderView.ResizeToContents)  # 管理列
        header.setSectionResizeMode(14, QHeaderView.ResizeToContents)  # 其他列
        header.setSectionResizeMode(15, QHeaderView.ResizeToContents)  # 修改列
        header.setSectionResizeMode(16, QHeaderView.ResizeToContents)  # 删除列
        header.setStretchLastSection(False)
        header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        
        layout.addWidget(self.daily_stats_table)
        
        # 日期筛选
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel('日期筛选:'))
        
        from datetime import datetime, date
        # 默认值：本月月初到今日
        today = date.today()
        first_day_of_month = today.replace(day=1)
        
        self.daily_start_date_input = QDateEdit()
        self.daily_start_date_input.setDate(QDate(first_day_of_month.year, first_day_of_month.month, first_day_of_month.day))
        self.daily_start_date_input.setCalendarPopup(True)
        filter_layout.addWidget(QLabel('从:'))
        filter_layout.addWidget(self.daily_start_date_input)
        
        self.daily_end_date_input = QDateEdit()
        self.daily_end_date_input.setDate(QDate(today.year, today.month, today.day))
        self.daily_end_date_input.setCalendarPopup(True)
        filter_layout.addWidget(QLabel('到:'))
        filter_layout.addWidget(self.daily_end_date_input)
        
        filter_btn = QPushButton('筛选')
        filter_btn.clicked.connect(self.filter_daily_stats)
        self.setup_button_style(filter_btn, 'primary')
        filter_layout.addWidget(filter_btn)
        
        reset_filter_btn = QPushButton('重置')
        reset_filter_btn.clicked.connect(self.reset_daily_stats_filter)
        self.setup_button_style(reset_filter_btn, 'normal')
        filter_layout.addWidget(reset_filter_btn)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        add_btn = QPushButton('添加信息')
        add_btn.clicked.connect(self.add_daily_stats_info)
        self.setup_button_style(add_btn, 'normal')
        button_layout.addWidget(add_btn)
        
        batch_modify_btn = QPushButton('批量添加信息')
        batch_modify_btn.clicked.connect(self.batch_modify_daily_stats)
        self.setup_button_style(batch_modify_btn, 'normal')
        button_layout.addWidget(batch_modify_btn)
        
        batch_delete_btn = QPushButton('批量删除')
        batch_delete_btn.clicked.connect(self.batch_delete_daily_stats)
        self.setup_button_style(batch_delete_btn, 'danger')
        button_layout.addWidget(batch_delete_btn)
        
        import_btn = QPushButton('导入数据')
        import_btn.clicked.connect(self.import_daily_stats_data)
        self.setup_button_style(import_btn, 'normal')
        button_layout.addWidget(import_btn)
        
        export_btn = QPushButton('导出数据')
        export_btn.clicked.connect(self.export_daily_stats_data)
        self.setup_button_style(export_btn, 'normal')
        button_layout.addWidget(export_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        tab_widget.setLayout(layout)
        return tab_widget

    def load_daily_stats_data(self):
        """加载每日情况统计数据（新版本）"""
        try:
            # 获取时间筛选条件
            time_condition, time_params = self.get_time_filter_condition("record_date")
            results = self.db_manager.get_all_daily_stats_with_youth_info(time_condition, time_params)
            self.display_daily_stats_data(results)
        except Exception as e:
            QMessageBox.warning(self, "加载错误", f"加载每日情况统计数据时发生错误：{str(e)}")

    def display_daily_stats_data(self, results):
        """显示每日情况统计数据"""
        try:
            self.daily_stats_table.setRowCount(len(results))
            
            for row, data in enumerate(results):
                # 数据库返回字段顺序：
                # 0:d.id, 1:y.id_card, 2:y.name, 3:y.recruitment_place, 4:y.company, 5:y.platoon, 6:y.squad, 7:y.squad_leader,
                # 8:d.record_date, 9:d.mood, 10:d.physical_condition, 11:d.mental_state, 12:d.training, 13:d.management, 14:d.notes
                
                # 检查是否有异常状态
                has_abnormal = any(str(data[i]) == '异常' for i in [9, 10, 11, 12, 13])  # mood, physical, mental, training, management
                
                # 第一列：复选框
                checkbox = QCheckBox()
                checkbox.setStyleSheet("""
                    QCheckBox {
                        spacing: 0px;
                    }
                    QCheckBox::indicator {
                        width: 20px;
                        height: 20px;
                        border: 2px solid #7F8C8D;
                        border-radius: 2px;
                        background-color: white;
                    }
                    QCheckBox::indicator:hover {
                        border-color: #3498DB;
                        background-color: #ECF0F1;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #3498DB;
                        border-color: #2980B9;
                    }
                """)
                checkbox.setMinimumSize(24, 24)
                checkbox.stateChanged.connect(self.update_daily_select_all_state)
                
                checkbox_layout = QHBoxLayout()
                checkbox_layout.setAlignment(Qt.AlignCenter)
                checkbox_layout.setContentsMargins(5, 5, 5, 5)
                checkbox_layout.addWidget(checkbox)
                checkbox_container = QWidget()
                checkbox_container.setLayout(checkbox_layout)
                self.daily_stats_table.setCellWidget(row, 0, checkbox_container)
                
                # 第二列：公民身份号码（原第三列）
                item = QTableWidgetItem(str(data[1] or ''))
                if has_abnormal:
                    from PyQt5.QtGui import QColor
                    item.setBackground(QColor(255, 182, 193, 100))  # 浅红色背景
                self.daily_stats_table.setItem(row, 1, item)
                
                # 第三列：姓名（原第四列）
                item = QTableWidgetItem(str(data[2] or ''))
                if has_abnormal:
                    from PyQt5.QtGui import QColor
                    item.setBackground(QColor(255, 182, 193, 100))  # 浅红色背景
                self.daily_stats_table.setItem(row, 2, item)
                
                # 第四列：应征地（原第五列）
                item = QTableWidgetItem(str(data[3] or ''))
                if has_abnormal:
                    from PyQt5.QtGui import QColor
                    item.setBackground(QColor(255, 182, 193, 100))  # 浅红色背景
                self.daily_stats_table.setItem(row, 3, item)
                
                # 第五列：连（原第六列）
                item = QTableWidgetItem(str(data[4] or ''))
                if has_abnormal:
                    from PyQt5.QtGui import QColor
                    item.setBackground(QColor(255, 182, 193, 100))  # 浅红色背景
                self.daily_stats_table.setItem(row, 4, item)
                
                # 第六列：排（原第七列）
                item = QTableWidgetItem(str(data[5] or ''))
                if has_abnormal:
                    from PyQt5.QtGui import QColor
                    item.setBackground(QColor(255, 182, 193, 100))  # 浅红色背景
                self.daily_stats_table.setItem(row, 5, item)
                
                # 第七列：班（原第八列）
                item = QTableWidgetItem(str(data[6] or ''))
                if has_abnormal:
                    from PyQt5.QtGui import QColor
                    item.setBackground(QColor(255, 182, 193, 100))  # 浅红色背景
                self.daily_stats_table.setItem(row, 6, item)
                
                # 第八列：带训班长信息（原第九列）
                item = QTableWidgetItem(str(data[7] or ''))
                if has_abnormal:
                    from PyQt5.QtGui import QColor
                    item.setBackground(QColor(255, 182, 193, 100))  # 浅红色背景
                self.daily_stats_table.setItem(row, 7, item)
                
                # 第九列：日期（原第十列）
                date_value = data[8]
                if date_value:
                    try:
                        from datetime import datetime
                        if isinstance(date_value, str):
                            date_obj = datetime.strptime(date_value, '%Y-%m-%d')
                            formatted_date = date_obj.strftime('%Y-%m-%d')
                        else:
                            formatted_date = str(date_value)
                        item = QTableWidgetItem(formatted_date)
                    except:
                        item = QTableWidgetItem(str(date_value or ''))
                else:
                    item = QTableWidgetItem('')
                
                if has_abnormal:
                    from PyQt5.QtGui import QColor
                    item.setBackground(QColor(255, 182, 193, 100))  # 浅红色背景
                self.daily_stats_table.setItem(row, 8, item)
                
                # 第十到十五列：状态字段（原第十一到十六列）
                status_fields = [data[9], data[10], data[11], data[12], data[13], data[14]]  # mood, physical, mental, training, management, notes
                for col, value in enumerate(status_fields, 9):
                    item = QTableWidgetItem(str(value or ''))
                    if has_abnormal:
                        from PyQt5.QtGui import QColor
                        item.setBackground(QColor(255, 182, 193, 100))  # 浅红色背景
                    self.daily_stats_table.setItem(row, col, item)
                
                # 第十六列：修改按钮（原第十七列）
                edit_btn = QPushButton('修改')
                edit_btn.setStyleSheet('background-color: #ADD8E6;')  # 淡蓝色
                edit_btn.clicked.connect(lambda checked, record_id=data[0], id_card=data[1]: self.edit_daily_stats_record(record_id, id_card))
                edit_btn_layout = QHBoxLayout()
                edit_btn_layout.setAlignment(Qt.AlignCenter)
                edit_btn_layout.setContentsMargins(5, 5, 5, 5)
                edit_btn_layout.addWidget(edit_btn)
                edit_btn_container = QWidget()
                edit_btn_container.setLayout(edit_btn_layout)
                self.daily_stats_table.setCellWidget(row, 15, edit_btn_container)
                
                # 第十七列：删除按钮（原第十八列）
                del_btn = QPushButton('删除')
                del_btn.setStyleSheet('background-color: #FFB6C1;')  # 淡红色
                del_btn.clicked.connect(lambda checked, record_id=data[0]: self.delete_daily_stats_record(record_id))
                del_btn_layout = QHBoxLayout()
                del_btn_layout.setAlignment(Qt.AlignCenter)
                del_btn_layout.setContentsMargins(5, 5, 5, 5)
                del_btn_layout.addWidget(del_btn)
                del_btn_container = QWidget()
                del_btn_container.setLayout(del_btn_layout)
                self.daily_stats_table.setCellWidget(row, 16, del_btn_container)
                
        except Exception as e:
            QMessageBox.warning(self, "显示错误", f"显示每日情况统计数据时发生错误：{str(e)}")

    def search_daily_stats(self):
        """搜索每日情况统计（新版本）"""
        try:
            name = self.daily_name_input.text().strip()
            id_card = self.daily_id_card_input.text().strip()
            recruitment_place = self.daily_recruitment_place_input.text().strip()
            company = self.daily_company_input.text().strip()
            platoon = self.daily_platoon_input.text().strip()
            squad = self.daily_squad_input.text().strip()
            
            results = self.db_manager.search_daily_stats_with_youth_info(
                name=name, id_card=id_card, recruitment_place=recruitment_place,
                company=company, platoon=platoon, squad=squad
            )
            
            self.display_daily_stats_data(results)
            
        except Exception as e:
            QMessageBox.warning(self, "搜索错误", f"搜索每日情况统计时发生错误：{str(e)}")

    def reset_daily_stats_search(self):
        """重置每日情况统计搜索（新版本）"""
        try:
            # 清空搜索框
            self.daily_name_input.clear()
            self.daily_id_card_input.clear()
            self.daily_recruitment_place_input.clear()
            self.daily_company_input.clear()
            self.daily_platoon_input.clear()
            self.daily_squad_input.clear()
            
            # 重置全选按钮
            if hasattr(self, 'daily_select_all_checkbox'):
                self.daily_select_all_checkbox.stateChanged.disconnect()
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            
            # 重新加载所有数据
            self.load_daily_stats_data()
            
        except Exception as e:
            QMessageBox.warning(self, "重置错误", f"重置搜索时发生错误：{str(e)}")

    def filter_daily_stats(self):
        """按日期筛选每日情况统计（新版本）"""
        try:
            start_date = self.daily_start_date_input.date().toString('yyyy-MM-dd')
            end_date = self.daily_end_date_input.date().toString('yyyy-MM-dd')
            
            results = self.db_manager.filter_daily_stats_by_date_range(start_date, end_date)
            self.display_daily_stats_data(results)
            
            QMessageBox.information(self, '筛选完成', f'找到 {len(results)} 条记录')
            
        except Exception as e:
            QMessageBox.warning(self, "筛选错误", f"筛选每日情况统计数据时发生错误：{str(e)}")

    def reset_daily_stats_filter(self):
        """重置每日情况统计筛选（新版本）"""
        try:
            # 重置日期范围到本月月初到今日
            from datetime import date
            today = date.today()
            first_day_of_month = today.replace(day=1)
            
            self.daily_start_date_input.setDate(QDate(first_day_of_month.year, first_day_of_month.month, first_day_of_month.day))
            self.daily_end_date_input.setDate(QDate(today.year, today.month, today.day))
            
            # 重置全选按钮
            if hasattr(self, 'daily_select_all_checkbox'):
                self.daily_select_all_checkbox.stateChanged.disconnect()
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
                self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            
            # 重新加载所有数据
            self.load_daily_stats_data()
            
        except Exception as e:
            QMessageBox.warning(self, "重置错误", f"重置筛选时发生错误：{str(e)}")

    def toggle_daily_select_all(self, state):
        """每日情况统计全选/取消全选（新版本）"""
        try:
            is_checked = (state == Qt.CheckState.Checked.value)
            
            for row in range(self.daily_stats_table.rowCount()):
                checkbox_widget = self.daily_stats_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox:
                        checkbox.setChecked(is_checked)
                        
        except Exception as e:
            QMessageBox.warning(self, "全选错误", f"全选操作时发生错误：{str(e)}")

    def update_daily_select_all_state(self):
        """更新每日情况统计全选按钮状态（新版本）"""
        try:
            if not hasattr(self, 'daily_select_all_checkbox'):
                return
            
            total_rows = self.daily_stats_table.rowCount()
            if total_rows == 0:
                return
            
            checked_count = 0
            for row in range(total_rows):
                checkbox_widget = self.daily_stats_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        checked_count += 1
            
            # 临时断开信号连接
            self.daily_select_all_checkbox.stateChanged.disconnect()
            
            if checked_count == 0:
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            elif checked_count == total_rows:
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.Checked)
            else:
                self.daily_select_all_checkbox.setCheckState(Qt.CheckState.PartiallyChecked)
            
            # 重新连接信号
            self.daily_select_all_checkbox.stateChanged.connect(self.toggle_daily_select_all)
            
        except Exception as e:
            print(f"更新全选状态时出错: {e}")
    # 每日情况统计操作方法（待实现）
    def add_daily_stats_info(self):
        """添加每日情况统计信息"""
        try:
            from ui.add_daily_record_dialog import AddDailyRecordDialog
            
            dialog = AddDailyRecordDialog(self.db_manager, self)
            
            # 连接数据更新信号
            dialog.data_updated.connect(self.load_daily_stats_data)
            
            # 显示对话框
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开添加对话框时发生错误: {str(e)}')

    def edit_daily_stats_record(self, record_id, id_card):
        """编辑每日情况统计记录"""
        try:
            from ui.edit_daily_record_dialog import EditDailyRecordDialog
            
            dialog = EditDailyRecordDialog(self.db_manager, record_id, self)
            
            # 连接数据更新信号
            dialog.data_updated.connect(self.load_daily_stats_data)
            
            # 显示对话框
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开编辑对话框时发生错误: {str(e)}')

    def delete_daily_stats_record(self, record_id):
        """删除每日情况统计记录"""
        reply = QMessageBox.question(
            self, '确认删除', 
            f'确定要删除这条记录吗？\n删除后无法恢复！',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # 删除记录
                conn = self.db_manager.get_connection()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM daily_stat WHERE id = ?", (record_id,))
                conn.commit()
                conn.close()
                
                QMessageBox.information(self, '删除成功', '记录已删除')
                self.load_daily_stats_data()  # 重新加载数据
                
            except Exception as e:
                QMessageBox.critical(self, '删除失败', f'删除记录时发生错误：{str(e)}')

    def batch_modify_daily_stats(self):
        """批量添加每日情况统计信息"""
        try:
            from ui.batch_add_daily_record_dialog import BatchAddDailyRecordDialog
            
            dialog = BatchAddDailyRecordDialog(self.db_manager, self)
            
            # 连接数据更新信号
            dialog.data_updated.connect(self.load_daily_stats_data)
            
            # 显示对话框
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开批量添加对话框时发生错误: {str(e)}')
    
    def get_selected_youth_for_daily_stats(self):
        """获取每日情况统计中选中的青年"""
        selected_youth = []
        
        try:
            # 从每日情况统计表格中获取选中的青年
            if hasattr(self, 'daily_stats_table'):
                for row in range(self.daily_stats_table.rowCount()):
                    # 获取复选框
                    checkbox_widget = self.daily_stats_table.cellWidget(row, 0)
                    if checkbox_widget:
                        checkbox = checkbox_widget.findChild(QCheckBox)
                        if checkbox and checkbox.isChecked():
                            # 获取青年信息
                            id_card_item = self.daily_stats_table.item(row, 1)  # 公民身份号码
                            name_item = self.daily_stats_table.item(row, 2)     # 姓名
                            
                            if id_card_item and name_item:
                                id_card = id_card_item.text()
                                name = name_item.text()
                                
                                # 获取youth_id
                                conn = self.db_manager.get_connection()
                                cursor = conn.cursor()
                                cursor.execute('SELECT id FROM youth WHERE id_card = ?', (id_card,))
                                result = cursor.fetchone()
                                conn.close()
                                
                                if result:
                                    youth_info = {
                                        'youth_id': result[0],
                                        'id_card': id_card,
                                        'name': name
                                    }
                                    
                                    # 避免重复添加同一个青年
                                    if not any(y['youth_id'] == youth_info['youth_id'] for y in selected_youth):
                                        selected_youth.append(youth_info)
            
            return selected_youth
            
        except Exception as e:
            print(f"获取选中青年时出错: {e}")
            return []

    def get_selected_daily_stats_records(self):
        """获取每日情况统计中选中的记录数据"""
        selected_records = []
        
        try:
            # 从每日情况统计表格中获取选中的记录
            if hasattr(self, 'daily_stats_table'):
                for row in range(self.daily_stats_table.rowCount()):
                    # 获取复选框
                    checkbox_widget = self.daily_stats_table.cellWidget(row, 0)
                    if checkbox_widget:
                        checkbox = checkbox_widget.findChild(QCheckBox)
                        if checkbox and checkbox.isChecked():
                            # 获取记录数据
                            record_data = []
                            # 按照导出需要的列顺序获取数据：公民身份号码, 姓名, 日期, 思想, 身体, 精神, 训练, 管理, 其他
                            for col in [1, 2, 8, 9, 10, 11, 12, 13, 14]:  # 对应表格中的列索引
                                item = self.daily_stats_table.item(row, col)
                                if item:
                                    record_data.append(item.text())
                                else:
                                    record_data.append('')
                            
                            if len(record_data) >= 9:  # 确保有足够的数据
                                selected_records.append(tuple(record_data))
            
            return selected_records
            
        except Exception as e:
            print(f"获取选中记录时出错: {e}")
            return []

    def batch_delete_daily_stats(self):
        """批量删除每日情况统计（新版本）"""
        try:
            # 获取选中的记录ID
            selected_record_ids = []
            selected_names = []
            
            for row in range(self.daily_stats_table.rowCount()):
                checkbox_widget = self.daily_stats_table.cellWidget(row, 0)
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        # 从display_daily_stats_data方法中我们知道data[0]是记录ID
                        # 我们需要从表格数据中获取记录ID
                        # 由于表格显示的是处理后的数据，我们需要重新查询获取ID
                        id_card_item = self.daily_stats_table.item(row, 1)  # 公民身份号码
                        name_item = self.daily_stats_table.item(row, 2)     # 姓名
                        date_item = self.daily_stats_table.item(row, 8)     # 日期
                        
                        if id_card_item and name_item and date_item:
                            id_card = id_card_item.text()
                            name = name_item.text()
                            date = date_item.text()
                            
                            # 查询获取记录ID
                            conn = self.db_manager.get_connection()
                            cursor = conn.cursor()
                            cursor.execute('''
                                SELECT d.id FROM daily_stat d
                                JOIN youth y ON d.youth_id = y.id
                                WHERE y.id_card = ? AND d.record_date = ?
                            ''', (id_card, date))
                            result = cursor.fetchone()
                            conn.close()
                            
                            if result:
                                selected_record_ids.append(result[0])
                                selected_names.append(f"{name}({date})")
            
            if not selected_record_ids:
                QMessageBox.warning(self, '提示', '请先选择要删除的记录')
                return
            
            # 显示要删除的记录
            names_text = '\n'.join(selected_names[:10])
            if len(selected_names) > 10:
                names_text += f'\n... 等共 {len(selected_names)} 条记录'
            
            reply = QMessageBox.question(
                self, '确认删除', 
                f'确定要删除以下记录吗？\n\n{names_text}\n\n删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # 执行批量删除
                success_count = 0
                fail_count = 0
                
                for record_id in selected_record_ids:
                    try:
                        conn = self.db_manager.get_connection()
                        cursor = conn.cursor()
                        cursor.execute('DELETE FROM daily_stat WHERE id = ?', (record_id,))
                        conn.commit()
                        conn.close()
                        success_count += 1
                    except Exception as e:
                        print(f"删除记录 {record_id} 时出错: {e}")
                        fail_count += 1
                
                # 刷新数据
                self.load_daily_stats_data()
                
                # 显示结果
                if fail_count == 0:
                    QMessageBox.information(self, '删除成功', f'成功删除 {success_count} 条记录')
                else:
                    QMessageBox.warning(self, '部分删除失败', 
                                      f'成功删除 {success_count} 条记录\n失败 {fail_count} 条记录')
                
        except Exception as e:
            QMessageBox.critical(self, '批量删除失败', f'批量删除时发生错误：{str(e)}')

    def import_daily_stats_data(self):
        """导入每日情况统计数据"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, '选择导入文件', '', 
                'Excel Files (*.xlsx *.xls);;All Files (*)'
            )
            
            if not file_path:
                return
            
            try:
                import openpyxl
            except ImportError:
                QMessageBox.critical(self, '错误', '需要安装openpyxl库才能导入Excel文件\n请运行: pip install openpyxl')
                return
            
            try:
                # 读取Excel文件
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # 获取表头
                headers = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value).strip())
                    else:
                        break
                
                # 检查必需的列
                required_columns = ['公民身份号码', '姓名', '日期', '思想', '身体', '精神', '训练', '管理', '其他']
                missing_columns = [col for col in required_columns if col not in headers]
                
                if missing_columns:
                    QMessageBox.warning(self, '格式错误', 
                                      f'Excel文件缺少以下必需列：\n{", ".join(missing_columns)}')
                    return
                
                # 获取列索引
                col_indices = {}
                for i, header in enumerate(headers):
                    if header in required_columns:
                        col_indices[header] = i + 1
                
                # 处理数据
                success_count = 0
                fail_count = 0
                error_messages = []
                duplicate_records = []
                
                # 首先检查所有重复记录
                for row_num in range(2, ws.max_row + 1):
                    try:
                        id_card = str(ws.cell(row=row_num, column=col_indices['公民身份号码']).value or '').strip()
                        date_value = ws.cell(row=row_num, column=col_indices['日期']).value
                        
                        if not id_card or not date_value:
                            continue
                        
                        # 处理日期
                        if isinstance(date_value, datetime):
                            date = date_value.strftime('%Y-%m-%d')
                        else:
                            date = str(date_value).strip()
                            # 验证日期格式
                            try:
                                date_obj = datetime.strptime(date, '%Y-%m-%d')
                                date = date_obj.strftime('%Y-%m-%d')
                            except:
                                continue
                        
                        # 检查是否已存在记录
                        conn = self.db_manager.get_connection()
                        cursor = conn.cursor()
                        cursor.execute('''
                            SELECT d.id, y.name FROM daily_stat d
                            JOIN youth y ON d.youth_id = y.id
                            WHERE y.id_card = ? AND d.record_date = ?
                        ''', (id_card, date))
                        existing_record = cursor.fetchone()
                        conn.close()
                        
                        if existing_record:
                            name = str(ws.cell(row=row_num, column=col_indices['姓名']).value or '').strip()
                            duplicate_records.append(f"{name}({date})")
                    
                    except Exception as e:
                        continue
                
                # 如果有重复记录，询问是否覆盖
                overwrite = False
                if duplicate_records:
                    duplicate_text = '\n'.join(duplicate_records[:10])
                    if len(duplicate_records) > 10:
                        duplicate_text += f'\n... 等共 {len(duplicate_records)} 条记录'
                    
                    reply = QMessageBox.question(
                        self, '发现重复记录', 
                        f'以下记录在数据库中已存在：\n\n{duplicate_text}\n\n是否要覆盖现有记录？',
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
                    )
                    
                    if reply == QMessageBox.StandardButton.Cancel:
                        return
                    elif reply == QMessageBox.StandardButton.Yes:
                        overwrite = True
                
                # 第一步：收集和验证所有数据
                valid_records = []
                for row_num in range(2, ws.max_row + 1):
                    try:
                        id_card = str(ws.cell(row=row_num, column=col_indices['公民身份号码']).value or '').strip()
                        name = str(ws.cell(row=row_num, column=col_indices['姓名']).value or '').strip()
                        date_value = ws.cell(row=row_num, column=col_indices['日期']).value
                        mood = str(ws.cell(row=row_num, column=col_indices['思想']).value or '').strip()
                        physical = str(ws.cell(row=row_num, column=col_indices['身体']).value or '').strip()
                        mental = str(ws.cell(row=row_num, column=col_indices['精神']).value or '').strip()
                        training = str(ws.cell(row=row_num, column=col_indices['训练']).value or '').strip()
                        management = str(ws.cell(row=row_num, column=col_indices['管理']).value or '').strip()
                        notes_value = ws.cell(row=row_num, column=col_indices['其他']).value
                        notes = str(notes_value).strip() if notes_value else ''
                        
                        # 验证必填字段
                        if not id_card or not name or not date_value:
                            error_messages.append(f'第{row_num}行：公民身份号码、姓名、日期不能为空')
                            fail_count += 1
                            continue
                        
                        # 异常数据验证：如果任一状态为异常，其他列不能为空
                        is_abnormal = any(status == '异常' for status in [mood, physical, mental, training, management])
                        if is_abnormal and not notes:
                            error_messages.append(f'第{row_num}行：当思想、身体、精神、训练或管理任一项为异常时，"其他"列不能为空')
                            fail_count += 1
                            continue
                        
                        # 处理日期
                        if isinstance(date_value, datetime):
                            date = date_value.strftime('%Y-%m-%d')
                        else:
                            date = str(date_value).strip()
                            # 验证日期格式
                            try:
                                date_obj = datetime.strptime(date, '%Y-%m-%d')
                                date = date_obj.strftime('%Y-%m-%d')
                            except:
                                error_messages.append(f'第{row_num}行：日期格式错误，应为YYYY-MM-DD')
                                fail_count += 1
                                continue
                        
                        valid_records.append({
                            'row_num': row_num,
                            'id_card': id_card,
                            'name': name,
                            'date': date,
                            'mood': mood,
                            'physical': physical,
                            'mental': mental,
                            'training': training,
                            'management': management,
                            'notes': notes
                        })
                        
                    except Exception as e:
                        error_messages.append(f'第{row_num}行：数据处理错误 - {str(e)}')
                        fail_count += 1
                        continue
                
                if not valid_records:
                    QMessageBox.warning(self, '导入失败', '没有有效的数据可以导入')
                    return
                
                # 第二步：批量查询青年ID和重复记录（一次数据库连接）
                conn = self.db_manager.get_connection()
                cursor = conn.cursor()
                
                try:
                    # 批量查询青年ID
                    id_cards = [record['id_card'] for record in valid_records]
                    placeholders = ','.join(['?' for _ in id_cards])
                    cursor.execute(f'SELECT id_card, id FROM youth WHERE id_card IN ({placeholders})', id_cards)
                    youth_id_map = {row[0]: row[1] for row in cursor.fetchall()}
                    
                    # 过滤出存在的青年记录
                    existing_records = []
                    for record in valid_records:
                        if record['id_card'] in youth_id_map:
                            record['youth_id'] = youth_id_map[record['id_card']]
                            existing_records.append(record)
                        else:
                            error_messages.append(f'第{record["row_num"]}行：找不到身份证号为 {record["id_card"]} 的青年')
                            fail_count += 1
                    
                    if not existing_records:
                        conn.close()
                        QMessageBox.warning(self, '导入失败', '没有找到对应的青年信息')
                        return
                    
                    # 批量检查重复记录 - 修复：不能使用executemany执行SELECT
                    existing_daily_records = set()
                    for record in existing_records:
                        cursor.execute('SELECT youth_id, record_date FROM daily_stat WHERE youth_id = ? AND record_date = ?', 
                                     (record['youth_id'], record['date']))
                        result = cursor.fetchone()
                        if result:
                            existing_daily_records.add((result[0], result[1]))
                    
                    # 分离新记录和重复记录
                    new_records = []
                    duplicate_records = []
                    
                    for record in existing_records:
                        key = (record['youth_id'], record['date'])
                        if key in existing_daily_records:
                            duplicate_records.append(record)
                        else:
                            new_records.append(record)
                    
                    # 第三步：处理重复记录询问
                    if duplicate_records and not overwrite:
                        duplicate_names = [f"{record['name']}({record['date']})" for record in duplicate_records[:10]]
                        duplicate_text = '\n'.join(duplicate_names)
                        if len(duplicate_records) > 10:
                            duplicate_text += f'\n... 等共 {len(duplicate_records)} 条记录'
                        
                        reply = QMessageBox.question(
                            self, '发现重复记录', 
                            f'以下记录在数据库中已存在：\n\n{duplicate_text}\n\n是否要覆盖现有记录？',
                            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
                        )
                        
                        if reply == QMessageBox.StandardButton.Cancel:
                            conn.close()
                            return
                        elif reply == QMessageBox.StandardButton.Yes:
                            overwrite = True
                    
                    # 第四步：批量插入新记录
                    if new_records:
                        insert_data = [
                            (record['youth_id'], record['date'], record['mood'], record['physical'], 
                             record['mental'], record['training'], record['management'], record['notes'])
                            for record in new_records
                        ]
                        
                        cursor.executemany('''
                            INSERT INTO daily_stat (youth_id, record_date, mood, physical_condition, 
                                                  mental_state, training, management, notes)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', insert_data)
                        success_count += len(new_records)
                    
                    # 第五步：批量更新重复记录（如果选择覆盖）
                    if duplicate_records and overwrite:
                        update_data = [
                            (record['mood'], record['physical'], record['mental'], record['training'], 
                             record['management'], record['notes'], record['youth_id'], record['date'])
                            for record in duplicate_records
                        ]
                        
                        cursor.executemany('''
                            UPDATE daily_stat 
                            SET mood = ?, physical_condition = ?, mental_state = ?, 
                                training = ?, management = ?, notes = ?
                            WHERE youth_id = ? AND record_date = ?
                        ''', update_data)
                        success_count += len(duplicate_records)
                    
                    # 提交事务
                    conn.commit()
                    
                except Exception as e:
                    conn.rollback()
                    error_messages.append(f'数据库操作失败：{str(e)}')
                    fail_count += len(valid_records)
                finally:
                    conn.close()
                
                # 刷新数据
                if success_count > 0:
                    self.load_daily_stats_data()
                
                # 显示结果
                result_msg = f'导入完成！\n成功：{success_count} 条\n失败：{fail_count} 条'
                if error_messages:
                    result_msg += f'\n\n错误详情（前10条）：\n' + '\n'.join(error_messages[:10])
                    if len(error_messages) > 10:
                        result_msg += f'\n... 还有 {len(error_messages) - 10} 条错误'
                
                if fail_count == 0:
                    QMessageBox.information(self, '导入成功', result_msg)
                else:
                    QMessageBox.warning(self, '导入完成', result_msg)
                
            except Exception as e:
                QMessageBox.critical(self, '导入失败', f'读取Excel文件时发生错误：{str(e)}')
                
        except Exception as e:
            QMessageBox.critical(self, '导入失败', f'导入数据时发生错误：{str(e)}')

    def export_daily_stats_data(self):
        """导出每日情况统计数据"""
        try:
            # 首先检查是否有选中的数据
            selected_records = self.get_selected_daily_stats_records()
            
            if selected_records:
                # 有选中数据，询问用户是否只导出选中数据
                reply = QMessageBox.question(
                    self, '导出选择', 
                    f'检测到您已选择了 {len(selected_records)} 条记录。\n\n'
                    f'是否只导出选中的记录？\n\n'
                    f'选择"是"：导出选中的 {len(selected_records)} 条记录\n'
                    f'选择"否"：导出所有记录',
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
                    QMessageBox.StandardButton.Yes
                )
                
                if reply == QMessageBox.StandardButton.Cancel:
                    return
                
                export_selected_only = (reply == QMessageBox.StandardButton.Yes)
            else:
                # 没有选中数据，默认导出所有数据
                export_selected_only = False
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, '保存导出文件', 
                f'每日情况统计_{datetime.now().strftime("%Y%m%d")}.xlsx',
                'Excel Files (*.xlsx);;All Files (*)'
            )
            
            if not file_path:
                return
            
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                QMessageBox.critical(self, '错误', '需要安装openpyxl库才能导出Excel文件\n请运行: pip install openpyxl')
                return
            
            try:
                # 根据选择获取数据
                if export_selected_only and selected_records:
                    # 导出选中的数据
                    results = selected_records
                    export_count = len(selected_records)
                    export_type = "选中"
                else:
                    # 导出所有数据
                    all_results = self.db_manager.get_all_daily_stats_with_youth_info()
                    # 转换为导出格式：公民身份号码, 姓名, 日期, 思想, 身体, 精神, 训练, 管理, 其他
                    results = []
                    for record in all_results:
                        # record格式: (d.id, y.id_card, y.name, y.recruitment_place, y.company, y.platoon, y.squad, y.squad_leader,
                        #              d.record_date, d.mood, d.physical_condition, d.mental_state, d.training, d.management, d.notes)
                        export_record = (
                            record[1] or '',  # 公民身份号码
                            record[2] or '',  # 姓名
                            record[8] or '',  # 日期
                            record[9] or '',  # 思想
                            record[10] or '', # 身体
                            record[11] or '', # 精神
                            record[12] or '', # 训练
                            record[13] or '', # 管理
                            record[14] or ''  # 其他
                        )
                        results.append(export_record)
                    export_count = len(results)
                    export_type = "全部"
                
                if not results:
                    QMessageBox.warning(self, '提示', '没有数据可以导出')
                    return
                
                # 创建工作簿
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = '每日情况统计'
                
                # 设置表头样式
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                # 设置表头
                headers = ['公民身份号码', '姓名', '日期', '思想', '身体', '精神', '训练', '管理', '其他']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # 填充数据
                for row_idx, record in enumerate(results, 2):
                    for col_idx, value in enumerate(record, 1):
                        ws.cell(row=row_idx, column=col_idx, value=str(value or ''))
                
                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # 保存文件
                wb.save(file_path)
                
                QMessageBox.information(self, '导出成功', 
                                      f'成功导出 {export_count} 条{export_type}记录到：\n{file_path}')
                
            except Exception as e:
                QMessageBox.critical(self, '导出失败', f'导出Excel文件时发生错误：{str(e)}')
                
        except Exception as e:
            QMessageBox.critical(self, '导出失败', f'导出数据时发生错误：{str(e)}')
    
    def import_town_interview_files(self):
        """导入镇街谈心谈话文件"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QDateEdit, QPushButton, QFileDialog, QMessageBox
            from PyQt5.QtCore import QDate, Qt
            import os
            from datetime import datetime
            
            # 创建导入对话框
            dialog = QDialog(self)
            dialog.setWindowTitle('导入镇街谈心谈话文件')
            dialog.setModal(True)
            dialog.resize(400, 200)
            
            layout = QVBoxLayout()
            
            # 日期选择
            date_layout = QHBoxLayout()
            date_layout.addWidget(QLabel('选择日期:'))
            date_edit = QDateEdit()
            date_edit.setDate(QDate.currentDate())
            date_edit.setCalendarPopup(True)
            date_layout.addWidget(date_edit)
            layout.addLayout(date_layout)
            
            # 文件选择
            file_layout = QHBoxLayout()
            file_layout.addWidget(QLabel('选择文件:'))
            file_btn = QPushButton('选择文件')
            selected_files = []
            
            def select_files():
                files, _ = QFileDialog.getOpenFileNames(
                    dialog, '选择文件', '',
                    '支持的文件 (*.pdf *.jpg *.jpeg *.png *.bmp *.gif *.tiff);;PDF文件 (*.pdf);;图片文件 (*.jpg *.jpeg *.png *.bmp *.gif *.tiff);;所有文件 (*.*)'
                )
                selected_files.clear()
                selected_files.extend(files)
                file_btn.setText(f'已选择 {len(files)} 个文件')
            
            file_btn.clicked.connect(select_files)
            file_layout.addWidget(file_btn)
            layout.addLayout(file_layout)
            
            # 按钮
            button_layout = QHBoxLayout()
            import_btn = QPushButton('开始导入')
            cancel_btn = QPushButton('取消')
            
            def start_import():
                if not selected_files:
                    QMessageBox.warning(dialog, '警告', '请先选择文件')
                    return
                
                interview_date = date_edit.date().toString('yyyy-MM-dd')
                self.process_interview_files(selected_files, interview_date, 'town')
                dialog.accept()
            
            import_btn.clicked.connect(start_import)
            cancel_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(import_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导入文件时发生错误: {str(e)}')
    
    def import_leader_interview_files(self):
        """导入领导谈心谈话文件"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QDateEdit, QPushButton, QFileDialog, QMessageBox
            from PyQt5.QtCore import QDate, Qt
            import os
            from datetime import datetime
            
            # 创建导入对话框
            dialog = QDialog(self)
            dialog.setWindowTitle('导入领导谈心谈话文件')
            dialog.setModal(True)
            dialog.resize(400, 200)
            
            layout = QVBoxLayout()
            
            # 日期选择
            date_layout = QHBoxLayout()
            date_layout.addWidget(QLabel('选择日期:'))
            date_edit = QDateEdit()
            date_edit.setDate(QDate.currentDate())
            date_edit.setCalendarPopup(True)
            date_layout.addWidget(date_edit)
            layout.addLayout(date_layout)
            
            # 文件选择
            file_layout = QHBoxLayout()
            file_layout.addWidget(QLabel('选择文件:'))
            file_btn = QPushButton('选择文件')
            selected_files = []
            
            def select_files():
                files, _ = QFileDialog.getOpenFileNames(
                    dialog, '选择文件', '',
                    '支持的文件 (*.pdf *.jpg *.jpeg *.png *.bmp *.gif *.tiff);;PDF文件 (*.pdf);;图片文件 (*.jpg *.jpeg *.png *.bmp *.gif *.tiff);;所有文件 (*.*)'
                )
                selected_files.clear()
                selected_files.extend(files)
                file_btn.setText(f'已选择 {len(files)} 个文件')
            
            file_btn.clicked.connect(select_files)
            file_layout.addWidget(file_btn)
            layout.addLayout(file_layout)
            
            # 按钮
            button_layout = QHBoxLayout()
            import_btn = QPushButton('开始导入')
            cancel_btn = QPushButton('取消')
            
            def start_import():
                if not selected_files:
                    QMessageBox.warning(dialog, '警告', '请先选择文件')
                    return
                
                interview_date = date_edit.date().toString('yyyy-MM-dd')
                self.process_interview_files(selected_files, interview_date, 'leader')
                dialog.accept()
            
            import_btn.clicked.connect(start_import)
            cancel_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(import_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导入文件时发生错误: {str(e)}')
    
    def process_interview_files(self, file_paths, interview_date, interview_type):
        """处理导入的谈心谈话文件"""
        try:
            import os
            from PyQt5.QtWidgets import QMessageBox
            
            success_count = 0
            failed_files = []
            valid_files = []
            
            # 验证所有文件并提取身份证号
            for file_path in file_paths:
                try:
                    # 从文件名提取身份证号
                    filename = os.path.basename(file_path)
                    name_without_ext = os.path.splitext(filename)[0]
                    
                    # 取文件名最后18位字符作为身份证号
                    if len(name_without_ext) < 18:
                        failed_files.append(f'{filename}: 文件名长度不足18位，无法提取身份证号')
                        continue
                    
                    id_card = name_without_ext[-18:]  # 取最后18位字符作为身份证号
                    
                    # 读取文件数据
                    file_ext = os.path.splitext(filename)[1].lower()
                    if file_ext == '.pdf':
                        # 转换PDF为JPG
                        image_data = self.convert_pdf_to_jpg(file_path)
                        if not image_data:
                            failed_files.append(f'{filename}: PDF转换失败')
                            continue
                    else:
                        # 读取图片文件
                        with open(file_path, 'rb') as f:
                            image_data = f.read()
                    
                    valid_files.append({
                        'filename': filename,
                        'id_card': id_card,
                        'image_data': image_data
                    })
                    
                except Exception as e:
                    failed_files.append(f'{filename}: 文件处理错误 - {str(e)}')
            
            if not valid_files:
                QMessageBox.warning(self, '导入失败', '没有有效的文件可以导入')
                return
            
            # 批量验证身份证号是否存在
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            
            try:
                # 批量查询青年信息
                id_cards = [file_info['id_card'] for file_info in valid_files]
                placeholders = ','.join(['?' for _ in id_cards])
                cursor.execute(f'SELECT id_card, name, gender FROM youth WHERE id_card IN ({placeholders})', id_cards)
                youth_info_map = {row[0]: {'name': row[1], 'gender': row[2]} for row in cursor.fetchall()}
                
                # 过滤出存在的青年记录
                valid_records = []
                for file_info in valid_files:
                    id_card = file_info['id_card']
                    if id_card in youth_info_map:
                        youth_info = youth_info_map[id_card]
                        valid_records.append({
                            'filename': file_info['filename'],
                            'id_card': id_card,
                            'name': youth_info['name'],
                            'gender': youth_info['gender'],
                            'image_data': file_info['image_data']
                        })
                    else:
                        failed_files.append(f'{file_info["filename"]}: 身份证号 {id_card} 在基本信息表中不存在')
                
                if not valid_records:
                    conn.close()
                    QMessageBox.warning(self, '导入失败', '没有找到对应的青年信息')
                    return
                
                # 检查是否有需要覆盖的数据
                existing_records = []
                new_records = []
                
                for record in valid_records:
                    # 检查是否已存在记录
                    if interview_type == 'town':
                        cursor.execute('''
                            SELECT id, youth_name FROM town_interview 
                            WHERE youth_id_card = ? AND interview_date = ?
                        ''', (record['id_card'], interview_date))
                    else:  # leader
                        cursor.execute('''
                            SELECT id, youth_name FROM leader_interview 
                            WHERE youth_id_card = ? AND interview_date = ?
                        ''', (record['id_card'], interview_date))
                    
                    existing_record = cursor.fetchone()
                    if existing_record:
                        existing_records.append({
                            'record': record,
                            'existing_name': existing_record[1]
                        })
                    else:
                        new_records.append(record)
                
                # 如果有需要覆盖的数据，弹出确认框
                if existing_records:
                    conn.close()
                    
                    # 构建确认消息
                    confirm_msg = f"检测到以下 {len(existing_records)} 条记录已存在，将被覆盖：\n\n"
                    for i, item in enumerate(existing_records[:5]):  # 最多显示5条
                        confirm_msg += f"{i+1}. {item['record']['name']} ({item['record']['id_card'][-4:]})\n"
                    
                    if len(existing_records) > 5:
                        confirm_msg += f"... 还有 {len(existing_records) - 5} 条记录\n"
                    
                    confirm_msg += f"\n新增记录：{len(new_records)} 条"
                    confirm_msg += f"\n日期：{interview_date}"
                    confirm_msg += "\n\n是否继续导入并覆盖现有数据？"
                    
                    reply = QMessageBox.question(
                        self, '确认覆盖数据', confirm_msg,
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No
                    )
                    
                    if reply != QMessageBox.StandardButton.Yes:
                        return
                    
                    # 重新连接数据库继续处理
                    conn = self.db_manager.get_connection()
                    cursor = conn.cursor()
                
                # 批量处理数据
                if interview_type == 'town':
                    # 处理镇街谈心谈话记录
                    for record in valid_records:
                        # 检查是否已存在记录
                        cursor.execute('''
                            SELECT id FROM town_interview 
                            WHERE youth_id_card = ? AND interview_date = ?
                        ''', (record['id_card'], interview_date))
                        existing_record = cursor.fetchone()
                        
                        if existing_record:
                            # 更新现有记录
                            cursor.execute('''
                                UPDATE town_interview 
                                SET youth_name = ?, gender = ?, visit_survey_image = ?, 
                                    thoughts = ?, spirit = ?
                                WHERE youth_id_card = ? AND interview_date = ?
                            ''', (record['name'], record['gender'], record['image_data'], 
                                  '正常', '正常', record['id_card'], interview_date))
                        else:
                            # 插入新记录
                            cursor.execute('''
                                INSERT INTO town_interview (youth_id_card, youth_name, gender, interview_date, 
                                                           visit_survey_image, thoughts, spirit)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            ''', (record['id_card'], record['name'], record['gender'], interview_date, 
                                  record['image_data'], '正常', '正常'))
                        success_count += 1
                else:  # leader
                    # 处理领导谈心谈话记录
                    for record in valid_records:
                        # 检查是否已存在记录
                        cursor.execute('''
                            SELECT id FROM leader_interview 
                            WHERE youth_id_card = ? AND interview_date = ?
                        ''', (record['id_card'], interview_date))
                        existing_record = cursor.fetchone()
                        
                        if existing_record:
                            # 更新现有记录
                            cursor.execute('''
                                UPDATE leader_interview 
                                SET youth_name = ?, gender = ?, visit_survey_image = ?, 
                                    thoughts = ?, spirit = ?
                                WHERE youth_id_card = ? AND interview_date = ?
                            ''', (record['name'], record['gender'], record['image_data'], 
                                  '正常', '正常', record['id_card'], interview_date))
                        else:
                            # 插入新记录
                            cursor.execute('''
                                INSERT INTO leader_interview (youth_id_card, youth_name, gender, interview_date, 
                                                             visit_survey_image, thoughts, spirit)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            ''', (record['id_card'], record['name'], record['gender'], interview_date, 
                                  record['image_data'], '正常', '正常'))
                        success_count += 1
                
                conn.commit()
                
            except Exception as e:
                conn.rollback()
                failed_files.append(f'数据库操作失败: {str(e)}')
            finally:
                conn.close()
            
            # 刷新数据
            if interview_type == 'town':
                self.load_town_interview_data()
            else:
                self.load_leader_interview_data()
            
            # 显示结果
            result_msg = f'导入完成！\n成功导入: {success_count} 个文件'
            if failed_files:
                result_msg += f'\n失败: {len(failed_files)} 个文件\n\n失败详情:\n' + '\n'.join(failed_files[:10])
                if len(failed_files) > 10:
                    result_msg += f'\n... 还有 {len(failed_files) - 10} 个失败文件'
            
            QMessageBox.information(self, '导入结果', result_msg)
            
        except Exception as e:
            QMessageBox.critical(self, '错误', f'处理文件时发生错误: {str(e)}')
    
    def convert_pdf_to_jpg(self, pdf_path):
        """将PDF文件转换为JPG图片"""
        try:
            # 尝试导入pdf2image库
            try:
                from pdf2image import convert_from_path
            except ImportError:
                return None
            
            # 转换PDF的第一页为图片
            images = convert_from_path(pdf_path, first_page=1, last_page=1, dpi=200)
            
            if not images:
                return None
            
            # 将PIL图片转换为字节数据
            import io
            img_byte_arr = io.BytesIO()
            images[0].save(img_byte_arr, format='JPEG', quality=85)
            img_byte_arr = img_byte_arr.getvalue()
            
            return img_byte_arr
            
        except Exception as e:
            print(f"PDF转换错误: {e}")
            return None
    def resizeEvent(self, event):
        """窗口大小改变事件处理"""
        super().resizeEvent(event)
        
        # 延迟调整表格列宽，确保窗口大小改变完成后再调整
        from PyQt5.QtCore import QTimer
        QTimer.singleShot(200, self._adjust_all_table_columns)
    
    def _adjust_all_table_columns(self):
        """调整所有表格的列宽"""
        try:
            # 获取当前显示的标签页索引
            current_index = self.tabs.currentIndex()
            
            # 根据当前标签页调整对应表格的列宽
            # 每日情况统计表格列宽调整已移除
                    
        except Exception as e:
            print(f"调整表格列宽时出错: {e}")

    # ==================== 体检情况统计表相关方法 ====================
    
    def create_physical_examination_tab(self):
        """创建体检情况统计表标签页"""
        tab_widget = QWidget()
        layout = QVBoxLayout()
        
        # 搜索栏
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel('搜索:'))
        
        self.physical_name_input = QLineEdit()
        self.physical_name_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.physical_name_input.setPlaceholderText('姓名')
        self.physical_name_input.returnPressed.connect(self.search_physical_examination)
        search_layout.addWidget(self.physical_name_input)
        
        self.physical_id_card_input = QLineEdit()
        self.physical_id_card_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.physical_id_card_input.setPlaceholderText('公民身份号码')
        self.physical_id_card_input.returnPressed.connect(self.search_physical_examination)
        search_layout.addWidget(self.physical_id_card_input)
        
        self.physical_recruitment_place_input = QLineEdit()
        self.physical_recruitment_place_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.physical_recruitment_place_input.setPlaceholderText('应征地')
        self.physical_recruitment_place_input.returnPressed.connect(self.search_physical_examination)
        search_layout.addWidget(self.physical_recruitment_place_input)
        
        self.physical_company_input = QLineEdit()
        self.physical_company_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.physical_company_input.setPlaceholderText('连')
        self.physical_company_input.returnPressed.connect(self.search_physical_examination)
        search_layout.addWidget(self.physical_company_input)
        
        self.physical_platoon_input = QLineEdit()
        self.physical_platoon_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.physical_platoon_input.setPlaceholderText('排')
        self.physical_platoon_input.returnPressed.connect(self.search_physical_examination)
        search_layout.addWidget(self.physical_platoon_input)
        
        self.physical_squad_input = QLineEdit()
        self.physical_squad_input.setStyleSheet("""
            QLineEdit {
                font-size: 11pt;
                padding: 8px;
                min-height: 30px;
            }
        """)
        self.physical_squad_input.setPlaceholderText('班')
        self.physical_squad_input.returnPressed.connect(self.search_physical_examination)
        search_layout.addWidget(self.physical_squad_input)
        
        search_btn = QPushButton('搜索')
        search_btn.clicked.connect(self.search_physical_examination)
        self.setup_button_style(search_btn, 'primary')
        search_layout.addWidget(search_btn)
        
        reset_btn = QPushButton('重置')
        reset_btn.clicked.connect(self.reset_physical_examination_search)
        self.setup_button_style(reset_btn, 'normal')
        search_layout.addWidget(reset_btn)
        
        layout.addLayout(search_layout)
        
        # 全选复选框行
        select_all_layout = QHBoxLayout()
        self.physical_select_all_checkbox = QCheckBox('全选')
        self.physical_select_all_checkbox.stateChanged.connect(self.toggle_physical_select_all)
        self.physical_select_all_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #7F8C8D;
                border-radius: 2px;
                background-color: white;
            }
            QCheckBox::indicator:hover {
                border-color: #3498DB;
                background-color: #ECF0F1;
            }
            QCheckBox::indicator:checked {
                background-color: #3498DB;
                border-color: #2980B9;
            }
        """)
        select_all_layout.addWidget(self.physical_select_all_checkbox)
        select_all_layout.addStretch()
        layout.addLayout(select_all_layout)
        
        # 体检情况统计表格
        self.physical_examination_table = QTableWidget()
        physical_headers = ['选择', '姓名', '性别', '公民身份号码', '应征地', '连', '排', '班', '带训班长信息',
                           '检查阶段', '阳性特征/边缘问题', '日期', '身体状况',
                           '士兵职业基本适应性检测机检类型', '跟踪处置意见', '处置意见落实情况', '查看详情', '删除']
        self.physical_examination_table.setColumnCount(len(physical_headers))
        self.physical_examination_table.setHorizontalHeaderLabels(physical_headers)
        self.physical_examination_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.physical_examination_table.setAlternatingRowColors(True)
        self.physical_examination_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # 隐藏垂直表头（行号）
        self.physical_examination_table.verticalHeader().setVisible(False)
        
        # 连接双击事件
        self.physical_examination_table.cellDoubleClicked.connect(self.on_physical_examination_double_click)
        
        # 应用统一的表格样式
        self.setup_table_style(self.physical_examination_table)
        
        # 设置表格自适应屏幕
        header = self.physical_examination_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 选择
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 姓名
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 性别
        header.setSectionResizeMode(3, QHeaderView.Interactive)  # 公民身份号码 - 可手动调整宽度
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 应征地
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # 连
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # 排
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # 班
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # 带训班长信息
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # 检查阶段
        header.setSectionResizeMode(10, QHeaderView.ResizeToContents)  # 阳性特征/边缘问题
        header.setSectionResizeMode(11, QHeaderView.ResizeToContents)  # 日期
        header.setSectionResizeMode(12, QHeaderView.ResizeToContents)  # 身体状况
        header.setSectionResizeMode(13, QHeaderView.ResizeToContents)  # 士兵职业基本适应性检测机检类型
        header.setSectionResizeMode(14, QHeaderView.ResizeToContents)  # 跟踪处置意见
        header.setSectionResizeMode(15, QHeaderView.ResizeToContents)  # 处置意见落实情况
        header.setSectionResizeMode(16, QHeaderView.ResizeToContents)  # 查看详情
        header.setSectionResizeMode(17, QHeaderView.ResizeToContents)  # 删除
        header.setStretchLastSection(False)
        header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        
        # 设置公民身份号码列的宽度，确保18位身份证号完整显示
        self.physical_examination_table.setColumnWidth(3, 250)
        
        # 设置应征地、连、排、班、带训班长信息列的宽度
        self.physical_examination_table.setColumnWidth(4, 80)   # 应征地
        self.physical_examination_table.setColumnWidth(5, 50)   # 连
        self.physical_examination_table.setColumnWidth(6, 50)   # 排
        self.physical_examination_table.setColumnWidth(7, 50)   # 班
        self.physical_examination_table.setColumnWidth(8, 120)  # 带训班长信息
        
        layout.addWidget(self.physical_examination_table)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        import_btn = QPushButton('导入数据')
        import_btn.clicked.connect(lambda: self.import_physical_examination_data())
        self.setup_button_style(import_btn, 'normal')
        button_layout.addWidget(import_btn)
        
        export_btn = QPushButton('导出数据')
        export_btn.clicked.connect(lambda: self.export_physical_examination_data())
        self.setup_button_style(export_btn, 'normal')
        button_layout.addWidget(export_btn)
        
        add_btn = QPushButton('添加信息')
        add_btn.clicked.connect(self.add_physical_examination_record)
        self.setup_button_style(add_btn, 'normal')
        button_layout.addWidget(add_btn)
        
        batch_delete_btn = QPushButton('批量删除')
        self.setup_button_style(batch_delete_btn, 'danger')
        batch_delete_btn.clicked.connect(self.batch_delete_physical_examination)
        button_layout.addWidget(batch_delete_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        tab_widget.setLayout(layout)
        
        # 加载体检情况数据
        # self.load_physical_examination_data()  # 延迟加载，在switch_tab中调用
        
        return tab_widget

    def load_physical_examination_data(self):
        """加载体检情况数据"""
        try:
            conn = None
            try:
                conn = self.db_manager.get_connection()
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                # 获取时间筛选条件
                time_condition, time_params = self.get_time_filter_condition("pe.district_date")
                
                base_query = """
                    SELECT pe.id, pe.youth_id_card, pe.name, pe.gender,
                           pe.district_exam, pe.district_positive, pe.district_date,
                           pe.city_exam, pe.city_positive, pe.city_date,
                           pe.special_exam, pe.special_positive, pe.special_date,
                           pe.body_status, pe.psychological_test_type, pe.tracking_opinion,
                           pe.implementation_status,
                           COALESCE(y.recruitment_place, '') as recruitment_place,
                           COALESCE(y.company, '') as company,
                           COALESCE(y.platoon, '') as platoon,
                           COALESCE(y.squad, '') as squad,
                           COALESCE(y.squad_leader, '') as squad_leader
                    FROM physical_examination pe
                    LEFT JOIN youth y ON pe.youth_id_card = y.id_card
                """
                
                if time_condition:
                    query = base_query + f" WHERE {time_condition} ORDER BY pe.id"
                    cursor.execute(query, time_params)
                else:
                    query = base_query + " ORDER BY pe.id"
                    cursor.execute(query)
                
                results = cursor.fetchall()
                self.display_physical_examination_results(results)
            finally:
                if conn:
                    conn.close()
        except Exception as e:
            QMessageBox.warning(self, "加载错误", f"加载体检情况数据时发生错误：{str(e)}")
            self.physical_examination_table.setRowCount(0)
    
    def search_physical_examination(self):
        """搜索体检情况记录"""
        try:
            name = self.physical_name_input.text().strip()
            id_card = self.physical_id_card_input.text().strip()
            recruitment_place = self.physical_recruitment_place_input.text().strip()
            company = self.physical_company_input.text().strip()
            platoon = self.physical_platoon_input.text().strip()
            squad = self.physical_squad_input.text().strip()
            
            conn = None
            try:
                conn = self.db_manager.get_connection()
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                sql = """
                    SELECT pe.id, pe.youth_id_card, pe.name, pe.gender,
                           pe.district_exam, pe.district_positive, pe.district_date,
                           pe.city_exam, pe.city_positive, pe.city_date,
                           pe.special_exam, pe.special_positive, pe.special_date,
                           pe.body_status, pe.psychological_test_type, pe.tracking_opinion,
                           pe.implementation_status,
                           COALESCE(y.recruitment_place, '') as recruitment_place,
                           COALESCE(y.company, '') as company,
                           COALESCE(y.platoon, '') as platoon,
                           COALESCE(y.squad, '') as squad,
                           COALESCE(y.squad_leader, '') as squad_leader
                    FROM physical_examination pe
                    LEFT JOIN youth y ON pe.youth_id_card = y.id_card
                    WHERE 1=1
                """
                params = []
                
                if name:
                    sql += " AND pe.name LIKE ?"
                    params.append(f"%{name}%")
                
                if id_card:
                    sql += " AND pe.youth_id_card LIKE ?"
                    params.append(f"%{id_card}%")
                
                if recruitment_place:
                    sql += " AND y.recruitment_place LIKE ?"
                    params.append(f"%{recruitment_place}%")
                
                if company:
                    sql += " AND y.company LIKE ?"
                    params.append(f"%{company}%")
                
                if platoon:
                    sql += " AND y.platoon LIKE ?"
                    params.append(f"%{platoon}%")
                
                if squad:
                    sql += " AND y.squad LIKE ?"
                    params.append(f"%{squad}%")
                
                sql += " ORDER BY pe.id"
                
                cursor.execute(sql, params)
                results = cursor.fetchall()
                self.display_physical_examination_results(results)
                
            finally:
                if conn:
                    conn.close()
                    
        except Exception as e:
            QMessageBox.warning(self, "搜索错误", f"搜索体检情况记录时发生错误：{str(e)}")
    
    def reset_physical_examination_search(self):
        """重置体检情况搜索"""
        self.physical_name_input.clear()
        self.physical_id_card_input.clear()
        self.physical_recruitment_place_input.clear()
        self.physical_company_input.clear()
        self.physical_platoon_input.clear()
        self.physical_squad_input.clear()
        self.load_physical_examination_data()

    def display_physical_examination_results(self, results):
        """显示体检情况结果 - 每条记录占3行（区体检、市体检、专项复查各一行）"""
        self.physical_examination_table.setRowCount(len(results) * 3)

        for idx, data in enumerate(results):
            base_row = idx * 3
            
            checkbox = QCheckBox()
            checkbox.setStyleSheet("""
                QCheckBox {
                    spacing: 0px;
                }
                QCheckBox::indicator {
                    width: 20px;
                    height: 20px;
                    border: 2px solid #7F8C8D;
                    border-radius: 2px;
                    background-color: white;
                }
                QCheckBox::indicator:hover {
                    border-color: #3498DB;
                    background-color: #ECF0F1;
                }
                QCheckBox::indicator:checked {
                    background-color: #3498DB;
                    border-color: #2980B9;
                }
            """)
            checkbox.setMinimumSize(24, 24)
            
            checkbox_layout = QHBoxLayout()
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.setContentsMargins(5, 5, 5, 5)
            checkbox_layout.addWidget(checkbox)
            checkbox_container = QWidget()
            checkbox_container.setLayout(checkbox_layout)
            self.physical_examination_table.setCellWidget(base_row, 0, checkbox_container)
            self.physical_examination_table.setSpan(base_row, 0, 3, 1)
            
            item = QTableWidgetItem(str(data['name'] or ''))
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 1, item)
            self.physical_examination_table.setSpan(base_row, 1, 3, 1)
            
            item = QTableWidgetItem(str(data['gender'] or ''))
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 2, item)
            self.physical_examination_table.setSpan(base_row, 2, 3, 1)
            
            item = QTableWidgetItem(str(data['youth_id_card'] or ''))
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 3, item)
            self.physical_examination_table.setSpan(base_row, 3, 3, 1)
            
            item = QTableWidgetItem(str(data['recruitment_place'] or ''))
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 4, item)
            self.physical_examination_table.setSpan(base_row, 4, 3, 1)
            
            item = QTableWidgetItem(str(data['company'] or ''))
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 5, item)
            self.physical_examination_table.setSpan(base_row, 5, 3, 1)
            
            item = QTableWidgetItem(str(data['platoon'] or ''))
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 6, item)
            self.physical_examination_table.setSpan(base_row, 6, 3, 1)
            
            item = QTableWidgetItem(str(data['squad'] or ''))
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 7, item)
            self.physical_examination_table.setSpan(base_row, 7, 3, 1)
            
            item = QTableWidgetItem(str(data['squad_leader'] or ''))
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 8, item)
            self.physical_examination_table.setSpan(base_row, 8, 3, 1)
            
            item = QTableWidgetItem('区体检')
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 9, item)
            
            positive_text = str(data['district_positive'] or '')
            item = QTableWidgetItem(positive_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            item.setToolTip(positive_text)
            self.physical_examination_table.setItem(base_row, 10, item)
            
            district_date = str(data['district_date'] or '')
            item = QTableWidgetItem(district_date)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 11, item)
            
            body_status = str(data['body_status'] or '')
            item = QTableWidgetItem(body_status)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 12, item)
            self.physical_examination_table.setSpan(base_row, 12, 3, 1)
            
            item = QTableWidgetItem(str(data['psychological_test_type'] or ''))
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row, 13, item)
            self.physical_examination_table.setSpan(base_row, 13, 3, 1)
            
            tracking_text = str(data['tracking_opinion'] or '')
            item = QTableWidgetItem(tracking_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            item.setToolTip(tracking_text)
            self.physical_examination_table.setItem(base_row, 14, item)
            self.physical_examination_table.setSpan(base_row, 14, 3, 1)
            
            implementation_text = str(data['implementation_status'] or '')
            item = QTableWidgetItem(implementation_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            item.setToolTip(implementation_text)
            self.physical_examination_table.setItem(base_row, 15, item)
            self.physical_examination_table.setSpan(base_row, 15, 3, 1)
            
            item = QTableWidgetItem('市体检')
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row + 1, 9, item)
            
            positive_text = str(data['city_positive'] or '')
            item = QTableWidgetItem(positive_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            item.setToolTip(positive_text)
            self.physical_examination_table.setItem(base_row + 1, 10, item)
            
            city_date = str(data['city_date'] or '')
            item = QTableWidgetItem(city_date)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row + 1, 11, item)
            
            item = QTableWidgetItem('专项复查')
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row + 2, 9, item)
            
            positive_text = str(data['special_positive'] or '')
            item = QTableWidgetItem(positive_text)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignLeft | Qt.AlignVCenter)
            item.setToolTip(positive_text)
            self.physical_examination_table.setItem(base_row + 2, 10, item)
            
            special_date = str(data['special_date'] or '')
            item = QTableWidgetItem(special_date)
            item.setData(Qt.ItemDataRole.TextAlignmentRole, Qt.AlignCenter | Qt.AlignVCenter)
            self.physical_examination_table.setItem(base_row + 2, 11, item)
            
            physical_id = data['id']
            view_btn = QPushButton('查看详情')
            view_btn.setStyleSheet('background-color: #ADD8E6;')
            view_btn.clicked.connect(lambda checked, pid=physical_id: self.view_physical_examination_details(pid))
            view_btn_layout = QHBoxLayout()
            view_btn_layout.setAlignment(Qt.AlignCenter)
            view_btn_layout.setContentsMargins(5, 5, 5, 5)
            view_btn_layout.addWidget(view_btn)
            view_btn_container = QWidget()
            view_btn_container.setLayout(view_btn_layout)
            self.physical_examination_table.setCellWidget(base_row, 16, view_btn_container)
            self.physical_examination_table.setSpan(base_row, 16, 3, 1)

            delete_btn = QPushButton('删除')
            delete_btn.setStyleSheet('background-color: #FFB6C1;')
            delete_btn.clicked.connect(lambda checked, pid=physical_id: self.delete_physical_examination(pid))
            delete_btn_layout = QHBoxLayout()
            delete_btn_layout.setAlignment(Qt.AlignCenter)
            delete_btn_layout.setContentsMargins(5, 5, 5, 5)
            delete_btn_layout.addWidget(delete_btn)
            delete_btn_container = QWidget()
            delete_btn_container.setLayout(delete_btn_layout)
            self.physical_examination_table.setCellWidget(base_row, 17, delete_btn_container)
            self.physical_examination_table.setSpan(base_row, 17, 3, 1)
        
        for row in range(self.physical_examination_table.rowCount()):
            self.physical_examination_table.setRowHeight(row, 40)
    
    def add_physical_examination_record(self):
        """添加体检情况记录"""
        from ui.physical_examination_detail_dialog import PhysicalExaminationDetailDialog
        
        dialog = PhysicalExaminationDetailDialog(self.db_manager, self)
        dialog.data_updated.connect(self.on_physical_examination_data_updated)
        dialog.exec_()
    
    def on_physical_examination_data_updated(self):
        """体检数据更新时的回调"""
        self.load_physical_examination_data()
    
    def on_physical_examination_double_click(self, row, column):
        """双击体检情况表格行时查看详情"""
        try:
            record_index = row // 3
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM physical_examination ORDER BY id")
            all_results = cursor.fetchall()
            conn.close()
            
            if record_index < len(all_results):
                physical_id = all_results[record_index]['id']
                self.view_physical_examination_details(physical_id)
        except Exception as e:
            QMessageBox.warning(self, '错误', f'打开详情时发生错误: {str(e)}')
    
    def view_physical_examination_details(self, physical_id):
        """查看体检情况详情"""
        try:
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, youth_id_card, name, gender,
                       district_exam, district_positive, district_date,
                       city_exam, city_positive, city_date,
                       special_exam, special_positive, special_date,
                       body_status, psychological_test_type, tracking_opinion,
                       implementation_status
                FROM physical_examination 
                WHERE id = ?
            ''', (physical_id,))
            
            record = cursor.fetchone()
            conn.close()
            
            if record:
                from ui.physical_examination_detail_dialog import PhysicalExaminationDetailDialog
                dialog = PhysicalExaminationDetailDialog(self.db_manager, self, record_data=record)
                dialog.data_updated.connect(self.on_physical_examination_data_updated)
                dialog.exec_()
            else:
                QMessageBox.warning(self, '错误', '未找到该记录')
        except Exception as e:
            QMessageBox.warning(self, '错误', f'查看详情时发生错误: {str(e)}')
    
    def delete_physical_examination(self, physical_id):
        """删除单条体检情况记录"""
        try:
            reply = QMessageBox.question(
                self, '确认删除', '确定要删除这条体检情况记录吗？\n删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                conn = self.db_manager.get_connection()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM physical_examination WHERE id = ?", (physical_id,))
                conn.commit()
                conn.close()
                
                QMessageBox.information(self, '删除成功', '体检情况记录已删除')
                self.on_physical_examination_data_updated()
        except Exception as e:
            QMessageBox.critical(self, '删除失败', f'删除体检情况记录时发生错误：{str(e)}')
    
    def batch_delete_physical_examination(self):
        """批量删除体检情况记录"""
        try:
            selected_ids = []
            selected_names = []
            
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT id, name FROM physical_examination ORDER BY id")
            all_results = cursor.fetchall()
            conn.close()
            
            for idx in range(len(all_results)):
                base_row = idx * 3
                if base_row < self.physical_examination_table.rowCount():
                    checkbox_widget = self.physical_examination_table.cellWidget(base_row, 0)
                    if checkbox_widget:
                        checkbox = checkbox_widget.findChild(QCheckBox)
                        if checkbox and checkbox.isChecked():
                            data = all_results[idx]
                            selected_ids.append(data['id'])
                            selected_names.append(data['name'])
            
            if not selected_ids:
                QMessageBox.information(self, '提示', '请先选择要删除的体检情况记录')
                return
            
            names_text = '、'.join(selected_names[:5])
            if len(selected_names) > 5:
                names_text += f' 等{len(selected_names)}人'
            
            reply = QMessageBox.question(
                self, '确认批量删除',
                f'确定要删除以下体检情况记录吗？\n{names_text}\n\n共{len(selected_ids)}条记录，删除后无法恢复！',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                conn = self.db_manager.get_connection()
                cursor = conn.cursor()
                
                deleted_count = 0
                for physical_id in selected_ids:
                    cursor.execute("DELETE FROM physical_examination WHERE id = ?", (physical_id,))
                    if cursor.rowcount > 0:
                        deleted_count += 1
                
                conn.commit()
                conn.close()
                
                QMessageBox.information(self, '删除成功', f'成功删除 {deleted_count} 条体检情况记录')
                
                if hasattr(self, 'physical_select_all_checkbox'):
                    self.physical_select_all_checkbox.setChecked(False)
                
                self.on_physical_examination_data_updated()
        except Exception as e:
            QMessageBox.critical(self, '批量删除失败', f'批量删除体检情况记录时发生错误：{str(e)}')
    
    def toggle_physical_select_all(self, state):
        """全选/取消全选体检情况"""
        is_checked = (state == Qt.CheckState.Checked.value)
        for row in range(0, self.physical_examination_table.rowCount(), 3):
            checkbox_widget = self.physical_examination_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(is_checked)
    
    def import_physical_examination_data(self):
        """导入体检情况数据（支持合并单元格格式）"""
        file_path, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '', 'Excel Files (*.xlsx *.xls)')
        if not file_path:
            return
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            
            count = 0
            updated_count = 0
            errors = []
            
            is_merged_format = False
            # 检查第1行是否有"检查阶段"列（表头在第1行）
            for cell in ws[1]:
                if cell.value == '检查阶段':
                    is_merged_format = True
                    break
            
            # 如果第1行没有，检查第2行（可能有标题行）
            if not is_merged_format:
                for cell in ws[2]:
                    if cell.value == '检查阶段':
                        is_merged_format = True
                        break
            
            if is_merged_format:
                # 确定数据起始行
                data_start_row = 2 if ws[1][4].value == '检查阶段' else 3
                
                row_idx = data_start_row
                while row_idx <= ws.max_row:
                    try:
                        id_card_cell = ws[f'D{row_idx}']
                        name_cell = ws[f'B{row_idx}']
                        gender_cell = ws[f'C{row_idx}']
                        
                        youth_id_card = str(id_card_cell.value).strip() if id_card_cell.value else ''
                        name = str(name_cell.value).strip() if name_cell.value else ''
                        gender = str(gender_cell.value).strip() if gender_cell.value else ''
                        
                        if not youth_id_card:
                            row_idx += 3
                            continue
                        
                        district_positive = str(ws[f'F{row_idx}'].value).strip() if ws[f'F{row_idx}'].value else ''
                        district_date = str(ws[f'G{row_idx}'].value).strip() if ws[f'G{row_idx}'].value else ''
                        city_positive = str(ws[f'F{row_idx + 1}'].value).strip() if ws[f'F{row_idx + 1}'].value else ''
                        city_date = str(ws[f'G{row_idx + 1}'].value).strip() if ws[f'G{row_idx + 1}'].value else ''
                        special_positive = str(ws[f'F{row_idx + 2}'].value).strip() if ws[f'F{row_idx + 2}'].value else ''
                        special_date = str(ws[f'G{row_idx + 2}'].value).strip() if ws[f'G{row_idx + 2}'].value else ''
                        
                        # 身体状况和检测类型从第一行读取（如果三行都有值，取第一行）
                        body_status_1 = str(ws[f'H{row_idx}'].value).strip() if ws[f'H{row_idx}'].value else ''
                        body_status_2 = str(ws[f'H{row_idx + 1}'].value).strip() if ws[f'H{row_idx + 1}'].value else ''
                        body_status_3 = str(ws[f'H{row_idx + 2}'].value).strip() if ws[f'H{row_idx + 2}'].value else ''
                        body_status = body_status_1 or body_status_2 or body_status_3
                        
                        test_type_1 = str(ws[f'I{row_idx}'].value).strip() if ws[f'I{row_idx}'].value else ''
                        test_type_2 = str(ws[f'I{row_idx + 1}'].value).strip() if ws[f'I{row_idx + 1}'].value else ''
                        test_type_3 = str(ws[f'I{row_idx + 2}'].value).strip() if ws[f'I{row_idx + 2}'].value else ''
                        psychological_test_type = test_type_1 or test_type_2 or test_type_3
                        
                        tracking_opinion_cell = ws[f'J{row_idx}']
                        implementation_status_cell = ws[f'K{row_idx}']
                        
                        tracking_opinion = str(tracking_opinion_cell.value).strip() if tracking_opinion_cell.value else ''
                        implementation_status = str(implementation_status_cell.value).strip() if implementation_status_cell.value else ''
                        
                        district_exam = ''
                        city_exam = ''
                        special_exam = ''
                        
                        cursor.execute("SELECT id_card FROM youth WHERE id_card = ?", (youth_id_card,))
                        if not cursor.fetchone():
                            errors.append(f"第{row_idx}行：身份证号 {youth_id_card} 不存在于基本信息中")
                            row_idx += 3
                            continue
                        
                        cursor.execute("""
                            SELECT id FROM physical_examination 
                            WHERE youth_id_card = ? 
                            AND district_date = ? 
                            AND city_date = ? 
                            AND special_date = ?
                        """, (youth_id_card, district_date, city_date, special_date))
                        existing_record = cursor.fetchone()
                        
                        if existing_record:
                            cursor.execute("""
                                UPDATE physical_examination 
                                SET name = ?, gender = ?,
                                    district_exam = ?, district_positive = ?,
                                    city_exam = ?, city_positive = ?,
                                    special_exam = ?, special_positive = ?,
                                    body_status = ?, psychological_test_type = ?,
                                    tracking_opinion = ?, implementation_status = ?
                                WHERE id = ?
                            """, (name, gender,
                                  district_exam, district_positive,
                                  city_exam, city_positive,
                                  special_exam, special_positive,
                                  body_status, psychological_test_type,
                                  tracking_opinion, implementation_status,
                                  existing_record[0]))
                            updated_count += 1
                        else:
                            cursor.execute("""
                                INSERT INTO physical_examination 
                                (youth_id_card, name, gender, 
                                 district_exam, district_positive, district_date,
                                 city_exam, city_positive, city_date,
                                 special_exam, special_positive, special_date,
                                 body_status, psychological_test_type, tracking_opinion, implementation_status)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (youth_id_card, name, gender, 
                                  district_exam, district_positive, district_date,
                                  city_exam, city_positive, city_date,
                                  special_exam, special_positive, special_date,
                                  body_status, psychological_test_type, tracking_opinion, implementation_status))
                        
                        count += 1
                        row_idx += 3
                        
                    except Exception as e:
                        errors.append(f"第{row_idx}行：{str(e)}")
                        row_idx += 3
            else:
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        if not row[0]:
                            continue
                        
                        youth_id_card = str(row[0]).strip()
                        name = str(row[1]).strip() if row[1] else ''
                        gender = str(row[2]).strip() if row[2] else ''
                        district_exam = str(row[3]).strip() if row[3] else ''
                        district_positive = str(row[4]).strip() if row[4] else ''
                        district_date = str(row[5]).strip() if row[5] else ''
                        city_exam = str(row[6]).strip() if row[6] else ''
                        city_positive = str(row[7]).strip() if row[7] else ''
                        city_date = str(row[8]).strip() if row[8] else ''
                        special_exam = str(row[9]).strip() if row[9] else ''
                        special_positive = str(row[10]).strip() if row[10] else ''
                        special_date = str(row[11]).strip() if row[11] else ''
                        body_status = str(row[12]).strip() if row[12] else ''
                        psychological_test_type = str(row[13]).strip() if row[13] else ''
                        tracking_opinion = str(row[14]).strip() if row[14] else ''
                        implementation_status = str(row[15]).strip() if row[15] else ''
                        
                        cursor.execute("SELECT id_card FROM youth WHERE id_card = ?", (youth_id_card,))
                        if not cursor.fetchone():
                            errors.append(f"第{row_idx}行：身份证号 {youth_id_card} 不存在于基本信息中")
                            continue
                        
                        cursor.execute("""
                            SELECT id FROM physical_examination 
                            WHERE youth_id_card = ? 
                            AND district_date = ? 
                            AND city_date = ? 
                            AND special_date = ?
                        """, (youth_id_card, district_date, city_date, special_date))
                        existing_record = cursor.fetchone()
                        
                        if existing_record:
                            cursor.execute("""
                                UPDATE physical_examination 
                                SET name = ?, gender = ?,
                                    district_exam = ?, district_positive = ?,
                                    city_exam = ?, city_positive = ?,
                                    special_exam = ?, special_positive = ?,
                                    body_status = ?, psychological_test_type = ?,
                                    tracking_opinion = ?, implementation_status = ?
                                WHERE id = ?
                            """, (name, gender,
                                  district_exam, district_positive,
                                  city_exam, city_positive,
                                  special_exam, special_positive,
                                  body_status, psychological_test_type,
                                  tracking_opinion, implementation_status,
                                  existing_record[0]))
                            updated_count += 1
                        else:
                            cursor.execute("""
                                INSERT INTO physical_examination 
                                (youth_id_card, name, gender, 
                                 district_exam, district_positive, district_date,
                                 city_exam, city_positive, city_date,
                                 special_exam, special_positive, special_date,
                                 body_status, psychological_test_type, tracking_opinion, implementation_status)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (youth_id_card, name, gender, 
                                  district_exam, district_positive, district_date,
                                  city_exam, city_positive, city_date,
                                  special_exam, special_positive, special_date,
                                  body_status, psychological_test_type, tracking_opinion, implementation_status))
                        
                        count += 1
                        
                    except Exception as e:
                        errors.append(f"第{row_idx}行：{str(e)}")
            
            conn.commit()
            conn.close()
            
            new_count = count - updated_count
            result_parts = []
            if new_count > 0:
                result_parts.append(f"新增 {new_count} 条记录")
            if updated_count > 0:
                result_parts.append(f"覆盖更新 {updated_count} 条记录（日期相同）")
            result_msg = "、".join(result_parts) if result_parts else "未导入任何记录"
            
            if errors:
                error_msg = f"成功处理 {count} 条记录（{result_msg}）\n\n以下记录导入失败：\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    error_msg += f"\n... 还有 {len(errors) - 10} 条错误"
                QMessageBox.warning(self, '导入完成（有错误）', error_msg)
            else:
                QMessageBox.information(self, '导入成功', f'成功处理 {count} 条记录\n{result_msg}')
            
            self.load_physical_examination_data()
            
        except Exception as e:
            QMessageBox.critical(self, '导入失败', f'导入体检情况数据时发生错误：{str(e)}')
    
    def export_physical_examination_data(self):
        """导出体检情况数据（合并单元格格式）"""
        try:
            selected_ids = []
            
            conn = self.db_manager.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM physical_examination ORDER BY id")
            all_results = cursor.fetchall()
            conn.close()
            
            for idx in range(len(all_results)):
                base_row = idx * 3
                if base_row < self.physical_examination_table.rowCount():
                    checkbox_widget = self.physical_examination_table.cellWidget(base_row, 0)
                    if checkbox_widget:
                        checkbox = checkbox_widget.findChild(QCheckBox)
                        if checkbox and checkbox.isChecked():
                            selected_ids.append(all_results[idx]['id'])
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, '导出体检情况统计表',
                f'体检情况统计表_{datetime.now().strftime("%Y%m%d")}.xlsx',
                'Excel Files (*.xlsx)'
            )
            
            if not file_path:
                return
            
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '体检情况统计表'
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 25
            ws.column_dimensions['K'].width = 20
            
            ws.merge_cells('A1:K1')
            title_cell = ws['A1']
            title_cell.value = '体检情况统计表'
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = center_alignment
            
            headers = [
                ('A2', '序号'), ('B2', '姓名'), ('C2', '性别'), ('D2', '公民身份号码'),
                ('E2', '检查阶段'), ('F2', '阳性特征/边缘问题'), ('G2', '日期'),
                ('H2', '身体状况'), ('I2', '士兵职业基本适应性检测机检类型'),
                ('J2', '跟踪处置意见'), ('K2', '处置意见落实情况')
            ]
            
            for cell_ref, header_text in headers:
                cell = ws[cell_ref]
                cell.value = header_text
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            
            if selected_ids:
                placeholders = ','.join('?' * len(selected_ids))
                cursor.execute(f"""
                    SELECT youth_id_card, name, gender, 
                           district_exam, district_positive, district_date,
                           city_exam, city_positive, city_date,
                           special_exam, special_positive, special_date,
                           body_status, psychological_test_type, tracking_opinion, implementation_status
                    FROM physical_examination
                    WHERE id IN ({placeholders})
                    ORDER BY id
                """, selected_ids)
            else:
                cursor.execute("""
                    SELECT youth_id_card, name, gender, 
                           district_exam, district_positive, district_date,
                           city_exam, city_positive, city_date,
                           special_exam, special_positive, special_date,
                           body_status, psychological_test_type, tracking_opinion, implementation_status
                    FROM physical_examination
                    ORDER BY id
                """)
            
            results = cursor.fetchall()
            conn.close()
            
            current_row = 3
            for idx, record in enumerate(results, 1):
                start_row = current_row
                end_row = current_row + 2
                
                ws.merge_cells(f'A{start_row}:A{end_row}')
                for r in range(start_row, end_row + 1):
                    ws[f'A{r}'].border = border
                cell = ws[f'A{start_row}']
                cell.value = idx
                cell.alignment = center_alignment
                
                ws.merge_cells(f'B{start_row}:B{end_row}')
                for r in range(start_row, end_row + 1):
                    ws[f'B{r}'].border = border
                cell = ws[f'B{start_row}']
                cell.value = record[1]
                cell.alignment = center_alignment
                
                ws.merge_cells(f'C{start_row}:C{end_row}')
                for r in range(start_row, end_row + 1):
                    ws[f'C{r}'].border = border
                cell = ws[f'C{start_row}']
                cell.value = record[2]
                cell.alignment = center_alignment
                
                ws.merge_cells(f'D{start_row}:D{end_row}')
                for r in range(start_row, end_row + 1):
                    ws[f'D{r}'].border = border
                cell = ws[f'D{start_row}']
                cell.value = record[0]
                cell.alignment = center_alignment
                
                ws[f'E{start_row}'].value = '区体检'
                ws[f'E{start_row}'].alignment = center_alignment
                ws[f'E{start_row}'].border = border
                ws[f'F{start_row}'].value = record[4] or ''
                ws[f'F{start_row}'].alignment = left_alignment
                ws[f'F{start_row}'].border = border
                ws[f'G{start_row}'].value = record[5] or ''
                ws[f'G{start_row}'].alignment = center_alignment
                ws[f'G{start_row}'].border = border
                
                ws[f'E{start_row + 1}'].value = '市体检'
                ws[f'E{start_row + 1}'].alignment = center_alignment
                ws[f'E{start_row + 1}'].border = border
                ws[f'F{start_row + 1}'].value = record[7] or ''
                ws[f'F{start_row + 1}'].alignment = left_alignment
                ws[f'F{start_row + 1}'].border = border
                ws[f'G{start_row + 1}'].value = record[8] or ''
                ws[f'G{start_row + 1}'].alignment = center_alignment
                ws[f'G{start_row + 1}'].border = border
                
                ws[f'E{start_row + 2}'].value = '专项复查'
                ws[f'E{start_row + 2}'].alignment = center_alignment
                ws[f'E{start_row + 2}'].border = border
                ws[f'F{start_row + 2}'].value = record[10] or ''
                ws[f'F{start_row + 2}'].alignment = left_alignment
                ws[f'F{start_row + 2}'].border = border
                ws[f'G{start_row + 2}'].value = record[11] or ''
                ws[f'G{start_row + 2}'].alignment = center_alignment
                ws[f'G{start_row + 2}'].border = border
                
                # 身体状况 - 每行都显示
                ws[f'H{start_row}'].value = record[12] or ''
                ws[f'H{start_row}'].alignment = center_alignment
                ws[f'H{start_row}'].border = border
                ws[f'H{start_row + 1}'].value = record[12] or ''
                ws[f'H{start_row + 1}'].alignment = center_alignment
                ws[f'H{start_row + 1}'].border = border
                ws[f'H{start_row + 2}'].value = record[12] or ''
                ws[f'H{start_row + 2}'].alignment = center_alignment
                ws[f'H{start_row + 2}'].border = border
                
                # 士兵职业基本适应性检测机检类型 - 每行都显示
                ws[f'I{start_row}'].value = record[13] or ''
                ws[f'I{start_row}'].alignment = center_alignment
                ws[f'I{start_row}'].border = border
                ws[f'I{start_row + 1}'].value = record[13] or ''
                ws[f'I{start_row + 1}'].alignment = center_alignment
                ws[f'I{start_row + 1}'].border = border
                ws[f'I{start_row + 2}'].value = record[13] or ''
                ws[f'I{start_row + 2}'].alignment = center_alignment
                ws[f'I{start_row + 2}'].border = border
                
                ws.merge_cells(f'J{start_row}:J{end_row}')
                for r in range(start_row, end_row + 1):
                    ws[f'J{r}'].border = border
                cell = ws[f'J{start_row}']
                cell.value = record[14] or ''
                cell.alignment = left_alignment
                
                ws.merge_cells(f'K{start_row}:K{end_row}')
                for r in range(start_row, end_row + 1):
                    ws[f'K{r}'].border = border
                cell = ws[f'K{start_row}']
                cell.value = record[15] or ''
                cell.alignment = left_alignment
                
                ws.row_dimensions[start_row].height = 25
                ws.row_dimensions[start_row + 1].height = 25
                ws.row_dimensions[start_row + 2].height = 25
                
                current_row += 3
            
            wb.save(file_path)
            
            if selected_ids:
                QMessageBox.information(self, '导出成功', f'成功导出 {len(results)} 条选中的体检情况记录')
                if hasattr(self, 'physical_select_all_checkbox'):
                    self.physical_select_all_checkbox.setChecked(False)
            else:
                QMessageBox.information(self, '导出成功', f'成功导出 {len(results)} 条体检情况记录')
                
        except Exception as e:
            QMessageBox.critical(self, '导出失败', f'导出体检情况数据时发生错误：{str(e)}')