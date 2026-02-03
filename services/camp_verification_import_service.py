"""
Excel 导入服务 - 用于导入入营点验情况数据
"""
import openpyxl
from openpyxl import load_workbook
import os


class CampVerificationImportService:
    """入营点验情况导入服务"""
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
    
    def import_from_excel(self, file_path):
        """
        从 Excel 文件导入入营点验情况数据
        
        Excel 文件格式：
        - 姓名 → username
        - 携带物品 → item
        - 用途 → usage
        - 处置措施 → Disposal
        - 公民身份号码 → user_id
        - 日期 → data
        
        去重逻辑：检查身份证号是否已存在于数据库，若存在则询问用户选择跳过或覆盖
        
        返回：(成功数量, 失败数量, 错误信息列表, 跳过的数据列表)
        """
        success_count = 0
        fail_count = 0
        error_messages = []
        skipped_records = []  # 记录被跳过的数据
        
        try:
            # 加载 Excel 文件
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            # 获取表头 - 改进容错性
            headers = {}
            for col_idx, cell in enumerate(worksheet[1], 1):
                if cell.value:
                    # 去除前后空格和特殊字符
                    header_text = str(cell.value).strip()
                    headers[header_text] = col_idx
            
            # 验证必需的列是否存在
            required_cols = ['姓名', '携带物品', '用途', '处置措施', '公民身份号码', '日期']
            missing_cols = [col for col in required_cols if col not in headers]
            
            if missing_cols:
                # 详细错误信息：显示实际找到的列名
                actual_cols = list(headers.keys())
                error_msg = f'缺少必需的列: {", ".join(missing_cols)}\n实际找到的列: {", ".join(actual_cols)}'
                return 0, 0, [error_msg], []
            
            # 预先检查所有数据中是否有重复的身份证号
            duplicate_user_ids = self._check_for_duplicates(worksheet, headers)
            user_choice = None  # 用户选择：None=未选择, 'skip'=跳过, 'overwrite'=覆盖
            
            if duplicate_user_ids:
                # 有重复数据，询问用户处理方式
                user_choice = self._ask_user_for_duplicate_handling(duplicate_user_ids)
                if user_choice is None:
                    # 用户取消操作
                    return 0, 0, ['用户取消了导入操作'], []
            
            # 遍历数据行
            for row_idx in range(2, worksheet.max_row + 1):
                try:
                    # 提取数据
                    name_col = headers.get('姓名')
                    item_col = headers.get('携带物品')
                    usage_col = headers.get('用途')
                    disposal_col = headers.get('处置措施')
                    user_id_col = headers.get('公民身份号码')
                    date_col = headers.get('日期')
                    
                    username = self._get_cell_value(worksheet, row_idx, name_col)
                    item = self._get_cell_value(worksheet, row_idx, item_col)
                    usage = self._get_cell_value(worksheet, row_idx, usage_col)
                    Disposal = self._get_cell_value(worksheet, row_idx, disposal_col)
                    user_id = self._get_cell_value(worksheet, row_idx, user_id_col)
                    data = self._get_cell_value(worksheet, row_idx, date_col)
                    
                    # 验证必填项
                    if not username:
                        fail_count += 1
                        error_messages.append(f'第 {row_idx} 行: 姓名为空')
                        continue
                    
                    if not user_id:
                        fail_count += 1
                        error_messages.append(f'第 {row_idx} 行: 公民身份号码为空')
                        continue
                    
                    if not item:
                        fail_count += 1
                        error_messages.append(f'第 {row_idx} 行: 携带物品为空')
                        continue
                    
                    # 转换日期格式 - 日期不能为空或格式无效
                    if not data or not str(data).strip():
                        fail_count += 1
                        error_messages.append(f'第 {row_idx} 行: 日期为空')
                        continue
                    
                    date_str = self._format_date(data)
                    if not date_str:
                        fail_count += 1
                        error_messages.append(f'第 {row_idx} 行: 日期格式无效')
                        continue
                    
                    # 处理重复数据
                    user_id_str = str(user_id).strip()
                    if self._user_id_exists(user_id_str):
                        if user_choice == 'skip':
                            # 跳过重复数据
                            skipped_records.append({
                                'row': row_idx,
                                'username': username,
                                'user_id': user_id_str,
                                'item': item
                            })
                            continue
                        elif user_choice == 'overwrite':
                            # 覆盖现有数据 - 先删除旧记录
                            self._delete_existing_record(user_id_str)
                    
                    # 添加到数据库
                    try:
                        record_id = self.db_manager.add_camp_verification(
                            username=username,
                            user_id=user_id_str,
                            item=str(item).strip(),
                            usage=str(usage).strip() if usage else '',
                            Disposal=str(Disposal).strip() if Disposal else '',
                            data=date_str
                        )
                        
                        if record_id:
                            success_count += 1
                        else:
                            fail_count += 1
                            error_messages.append(f'第 {row_idx} 行: 添加数据库失败')
                    except Exception as db_error:
                        fail_count += 1
                        error_messages.append(f'第 {row_idx} 行: 数据库错误 - {str(db_error)}')
                
                except Exception as row_error:
                    fail_count += 1
                    error_messages.append(f'第 {row_idx} 行: 处理行数据出错 - {str(row_error)}')
            
            workbook.close()
            
        except Exception as e:
            error_messages.insert(0, f'读取 Excel 文件出错: {str(e)}')
            return 0, 0, error_messages, []
        
        return success_count, fail_count, error_messages, skipped_records
    
    def _get_cell_value(self, worksheet, row_idx, col_idx):
        """获取单元格的值，正确处理日期类型"""
        if col_idx is None:
            return None
        
        try:
            from datetime import datetime, date as date_class
            
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            
            if value is None:
                return None
            
            # 如果是 datetime 或 date 对象，格式化为字符串
            if isinstance(value, datetime):
                return value.strftime('%Y/%m/%d')
            elif isinstance(value, date_class):
                return value.strftime('%Y/%m/%d')
            
            # 否则转换为字符串并去空格
            return str(value).strip()
        except:
            return None
    
    def _user_id_exists(self, user_id):
        """
        检查身份证号是否已存在于数据库
        
        参数：user_id (str) - 身份证号
        返回：bool - 如果存在返回 True，不存在返回 False
        """
        try:
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute(
                'SELECT COUNT(*) FROM camp_verification WHERE user_id = ?',
                (user_id,)
            )
            result = cursor.fetchone()
            conn.close()
            
            # 如果查询结果的第一列大于0，说明存在相同的身份证号
            return result[0] > 0 if result else False
        except Exception as e:
            print(f"检查身份证号时出错: {e}")
            return False
    
    def _check_for_duplicates(self, worksheet, headers):
        """
        预先检查Excel中的身份证号是否在数据库中已存在
        
        参数：
            worksheet - Excel工作表
            headers - 列名映射字典
        返回：list - 存在重复的身份证号列表
        """
        duplicate_user_ids = []
        user_id_col = headers.get('公民身份号码')
        
        if not user_id_col:
            return duplicate_user_ids
        
        # 遍历数据行，检查身份证号
        for row_idx in range(2, worksheet.max_row + 1):
            try:
                user_id = self._get_cell_value(worksheet, row_idx, user_id_col)
                if user_id:
                    user_id_str = str(user_id).strip()
                    if user_id_str and self._user_id_exists(user_id_str):
                        if user_id_str not in duplicate_user_ids:
                            duplicate_user_ids.append(user_id_str)
                        # 优化：找到一个重复就可以停止后续检查（如用户要求）
                        if len(duplicate_user_ids) >= 1:
                            break
            except:
                continue
        
        return duplicate_user_ids
    
    def _ask_user_for_duplicate_handling(self, duplicate_user_ids):
        """
        询问用户如何处理重复数据
        
        参数：duplicate_user_ids - 重复的身份证号列表
        返回：str - 'skip'=跳过, 'overwrite'=覆盖, None=取消
        """
        try:
            from PyQt5.QtWidgets import QMessageBox, QPushButton
            
            # 创建消息框
            msg_box = QMessageBox()
            msg_box.setWindowTitle('发现重复数据')
            
            # 构建提示信息
            if len(duplicate_user_ids) == 1:
                msg_text = f'发现重复的身份证号：{duplicate_user_ids[0]}\n\n请选择处理方式：'
            else:
                msg_text = f'发现 {len(duplicate_user_ids)} 个重复的身份证号\n\n请选择处理方式：'
            
            msg_box.setText(msg_text)
            msg_box.setIcon(QMessageBox.Icon.Question)
            
            # 添加自定义按钮
            skip_button = msg_box.addButton('跳过重复数据', QMessageBox.ButtonRole.AcceptRole)
            overwrite_button = msg_box.addButton('覆盖现有数据', QMessageBox.ButtonRole.AcceptRole)
            cancel_button = msg_box.addButton('取消导入', QMessageBox.ButtonRole.RejectRole)
            
            # 设置按钮样式
            skip_button.setStyleSheet("""
                QPushButton {
                    background-color: #3498DB;
                    color: white;
                    border: none;
                    border-radius: 3px;
                    padding: 8px 16px;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background-color: #2980B9;
                }
            """)
            
            overwrite_button.setStyleSheet("""
                QPushButton {
                    background-color: #E67E22;
                    color: white;
                    border: none;
                    border-radius: 3px;
                    padding: 8px 16px;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background-color: #D35400;
                }
            """)
            
            cancel_button.setStyleSheet("""
                QPushButton {
                    background-color: #95A5A6;
                    color: white;
                    border: none;
                    border-radius: 3px;
                    padding: 8px 16px;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background-color: #7F8C8D;
                }
            """)
            
            # 显示对话框并获取用户选择
            msg_box.exec_()
            clicked_button = msg_box.clickedButton()
            
            if clicked_button == skip_button:
                return 'skip'
            elif clicked_button == overwrite_button:
                return 'overwrite'
            else:
                return None  # 取消
                
        except Exception as e:
            print(f"显示重复数据对话框时出错: {e}")
            return None
    
    def _delete_existing_record(self, user_id):
        """
        删除指定身份证号的现有记录
        
        参数：user_id (str) - 身份证号
        """
        try:
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('DELETE FROM camp_verification WHERE user_id = ?', (user_id,))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"删除现有记录时出错: {e}")
    
    def _format_date(self, date_value):
        """
        将 Excel 中的日期转换为 YYYY-MM-DD 格式
        
        支持的日期格式：
        - datetime 对象（直接转换）
        - Excel 日期序列号（如 45000，自动转换）
        - "2026-01-25" 或 "2025-8-8" 格式
        - "2026/01/25" 或 "2025/8/8" 格式（支持无补零）
        - "01/25/2026" 格式
        - "25/01/2026" 格式
        - "2026年01月25日" 或 "2026年1月8日" 格式
        """
        if not date_value:
            return None
        
        try:
            from datetime import datetime, timedelta, date as date_class
            
            # 情况1：如果是 datetime 或 date 对象，直接转换
            if isinstance(date_value, datetime):
                return date_value.strftime('%Y-%m-%d')
            elif isinstance(date_value, date_class):
                return date_value.strftime('%Y-%m-%d')
            
            # 情况2：如果是数字（可能是 Excel 的日期序列号或 float）
            if isinstance(date_value, (int, float)):
                try:
                    # Excel 日期序列号起点是 1900-01-01
                    # 注意：Excel 有一个 bug，1900-02-29 不存在但被计算为有效日期
                    excel_date = int(date_value)
                    if excel_date > 0:
                        # 从 1899-12-30 (Excel 中的 0) 到目标日期
                        date_obj = datetime(1899, 12, 30) + timedelta(days=excel_date)
                        return date_obj.strftime('%Y-%m-%d')
                except:
                    pass
            
            # 情况3：字符串处理
            date_str = str(date_value).strip()
            
            # 预处理：将中文格式转换为标准格式
            if '年' in date_str and '月' in date_str and '日' in date_str:
                # 处理 "2026年01月25日" 或 "2026年1月8日"
                date_str = date_str.replace('年', '-').replace('月', '-').replace('日', '')
                # 此时格式应该是 "2026-01-25" 或 "2026-1-8"
            
            # 标准化日期格式：补零
            # 目的：将 "2025/8/8" 转换为 "2025/08/08"
            parts = date_str.replace('-', '/').split('/')
            if len(parts) == 3:
                try:
                    # 尝试判断日期格式（YYYY/MM/DD 或 DD/MM/YYYY 等）
                    # 如果第一部分 > 31，则第一部分必然是年份
                    if int(parts[0]) > 31:
                        # 格式应该是 YYYY/MM/DD 或 YYYY-MM-DD
                        year = int(parts[0])
                        month = int(parts[1])
                        day = int(parts[2])
                    elif int(parts[2]) > 31:
                        # 格式应该是 MM/DD/YYYY 或 DD/MM/YYYY，第三部分是年份
                        year = int(parts[2])
                        month = int(parts[0])
                        day = int(parts[1])
                    else:
                        # 无法确定，尝试标准格式
                        raise ValueError("无法确定日期格式")
                    
                    # 验证月份和日期
                    if 1 <= month <= 12 and 1 <= day <= 31:
                        date_obj = datetime(year, month, day)
                        return date_obj.strftime('%Y-%m-%d')
                except:
                    pass
            
            # 尝试各种标准格式
            formats = [
                '%Y-%m-%d',      # 2026-01-25
                '%Y/%m/%d',      # 2026/01/25
                '%m/%d/%Y',      # 01/25/2026
                '%d/%m/%Y',      # 25/01/2026
                '%Y年%m月%d日',  # 2026年01月25日
            ]
            
            for fmt in formats:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.strftime('%Y-%m-%d')
                except:
                    continue
            
            # 如果所有方法都失败，返回 None
            return None
            
        except Exception as e:
            print(f"日期处理错误: {e}")
            return None


class GeneralImportService:
    """通用导入服务 - 支持多种数据类型的导入"""
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
        self.camp_verification_service = CampVerificationImportService(db_manager)
    
    def import_camp_verification_from_excel(self, file_path):
        """
        导入入营点验情况数据
        
        参数：file_path (str) - Excel 文件路径
        返回：(成功数量, 失败数量, 错误信息列表)
        """
        return self.camp_verification_service.import_from_excel(file_path)
