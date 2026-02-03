"""
数据导入服务
"""
import openpyxl
from datetime import datetime
from utils.validators import validate_youth_data


class ImportService:
    def __init__(self, db_manager):
        self.db_manager = db_manager
    
    def import_youth_from_excel(self, file_path):
        """从Excel导入青年基本信息（新结构40字段）- 优化版本"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            imported_count = 0
            skipped_count = 0
            error_rows = []
            duplicate_data = []  # 存储重复的数据
            valid_data = []  # 存储验证通过的数据
            
            # 第一步：收集所有数据并验证
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[2]:  # 跳过空行（检查姓名列，现在是第3列）
                    continue
                
                # 安全地获取单元格值，处理可能的None值和数字类型
                def safe_get_cell_value(row, index):
                    if len(row) > index and row[index] is not None:
                        return str(row[index]).strip()
                    return ''
                
                youth_data = (
                    safe_get_cell_value(row, 3),   # id_card 公民身份号码 (第4列) - 第一个参数
                    safe_get_cell_value(row, 2),   # name 姓名 (第3列) - 第二个参数
                    safe_get_cell_value(row, 4),   # gender 性别 (第5列) - 第三个参数
                    safe_get_cell_value(row, 5),   # birth_date 出生日期 (第6列)
                    safe_get_cell_value(row, 6),   # nation 民族 (第7列)
                    safe_get_cell_value(row, 7),   # political_status 政治面貌 (第8列)
                    safe_get_cell_value(row, 8),   # religion 宗教信仰 (第9列)
                    safe_get_cell_value(row, 9),   # native_place 籍贯 (第10列)
                    safe_get_cell_value(row, 10),  # education_level 文化程度 (第11列)
                    safe_get_cell_value(row, 11),  # study_status 学业情况 (第12列)
                    safe_get_cell_value(row, 12),  # study_type 学习类型 (第13列)
                    safe_get_cell_value(row, 13),  # camp_entry_time 入营时间 (第14列)
                    safe_get_cell_value(row, 14),  # recruitment_place 应征地 (第15列)
                    safe_get_cell_value(row, 15),  # residence_address 经常居住地址 (第16列)
                    safe_get_cell_value(row, 16),  # household_address 户籍所在地 (第17列)
                    safe_get_cell_value(row, 17),  # postal_code 邮编 (第18列)
                    safe_get_cell_value(row, 18),  # personal_phone 本人电话 (第19列)
                    safe_get_cell_value(row, 19),  # family_phone 家庭电话 (第20列)
                    safe_get_cell_value(row, 20),  # school 毕业(就读)学校 (第21列)
                    safe_get_cell_value(row, 21),  # major 所学专业 (第22列)
                    safe_get_cell_value(row, 22),  # enrollment_time 入学时间 (第23列)
                    safe_get_cell_value(row, 23),  # initial_hospital 初检医院 (第24列)
                    safe_get_cell_value(row, 24),  # initial_conclusion 初检结论 (第25列)
                    safe_get_cell_value(row, 25),  # initial_time 初检时间 (第26列)
                    safe_get_cell_value(row, 26),  # physical_conclusion 体检结论 (第27列)
                    safe_get_cell_value(row, 27),  # physical_time 体检时间 (第28列)
                    safe_get_cell_value(row, 28),  # physical_disqualification 体检不合格原因 (第29列)
                    safe_get_cell_value(row, 29),  # chief_doctor_opinion 主检医师意见 (第30列)
                    safe_get_cell_value(row, 30),  # graduation_time 毕业时间 (第31列)
                    safe_get_cell_value(row, 31),  # company 连 (第32列)
                    safe_get_cell_value(row, 32),  # platoon 排 (第33列)
                    safe_get_cell_value(row, 33),  # squad 班 (第34列)
                    safe_get_cell_value(row, 34),  # squad_leader 带训班长信息 (第35列)
                    safe_get_cell_value(row, 35),  # camp_status 在营状态 (第36列)
                    safe_get_cell_value(row, 36),  # leave_time 离营时间 (第37列)
                    safe_get_cell_value(row, 37)   # leave_reason 离营原因 (第38列)
                )
                
                # 验证数据
                is_valid, error_msg = validate_youth_data(youth_data)
                if not is_valid:
                    error_rows.append(f"第{row_num}行: {error_msg}")
                    continue
                
                valid_data.append((row_num, youth_data))
            
            if not valid_data:
                return 0, "没有有效的数据可以导入"
            
            # 第二步：批量检查重复数据（一次数据库查询）
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            
            # 构建批量查询的身份证号列表
            id_cards = [data[1][0] for data in valid_data if data[1][0]]  # 提取所有身份证号
            
            if id_cards:
                # 使用IN查询批量检查重复
                placeholders = ','.join(['?' for _ in id_cards])
                cursor.execute(f'SELECT id_card, name FROM youth WHERE id_card IN ({placeholders})', id_cards)
                existing_youth_dict = {row[0]: row[1] for row in cursor.fetchall()}
            else:
                existing_youth_dict = {}
            
            # 分离新数据和重复数据
            new_data = []
            for row_num, youth_data in valid_data:
                id_card = youth_data[0]
                if id_card in existing_youth_dict:
                    duplicate_data.append({
                        'row_num': row_num,
                        'youth_data': youth_data,
                        'existing_name': existing_youth_dict[id_card]
                    })
                else:
                    new_data.append(youth_data)
            
            # 第三步：批量插入新数据（一次事务）
            if new_data:
                try:
                    cursor.executemany('''
                        INSERT INTO youth (
                            id_card, name, gender, birth_date, nation, political_status, religion, native_place,
                            education_level, study_status, study_type, camp_entry_time, recruitment_place,
                            residence_address, household_address, postal_code, personal_phone, family_phone,
                            school, major, enrollment_time, initial_hospital, initial_conclusion, initial_time,
                            physical_conclusion, physical_time, physical_disqualification, chief_doctor_opinion,
                            graduation_time, company, platoon, squad, squad_leader, camp_status, leave_time, leave_reason
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', new_data)
                    conn.commit()
                    imported_count = len(new_data)
                except Exception as e:
                    conn.rollback()
                    error_rows.append(f"批量插入失败: {str(e)}")
            
            # 第四步：处理重复数据
            if duplicate_data:
                overwrite_choice = self._ask_overwrite_duplicates(duplicate_data)
                
                if overwrite_choice == 'overwrite_all':
                    # 批量更新重复数据
                    update_data = []
                    for dup in duplicate_data:
                        youth_data = dup['youth_data']
                        # 构建更新数据（除了id_card，其他字段都要更新）
                        update_data.append(youth_data[1:] + (youth_data[0],))  # 将id_card移到最后作为WHERE条件
                    
                    try:
                        cursor.executemany('''
                            UPDATE youth SET
                                name=?, gender=?, birth_date=?, nation=?, political_status=?, religion=?, native_place=?,
                                education_level=?, study_status=?, study_type=?, camp_entry_time=?, recruitment_place=?,
                                residence_address=?, household_address=?, postal_code=?, personal_phone=?, family_phone=?,
                                school=?, major=?, enrollment_time=?, initial_hospital=?, initial_conclusion=?, initial_time=?,
                                physical_conclusion=?, physical_time=?, physical_disqualification=?, chief_doctor_opinion=?,
                                graduation_time=?, company=?, platoon=?, squad=?, squad_leader=?, camp_status=?, leave_time=?, leave_reason=?
                            WHERE id_card=?
                        ''', update_data)
                        conn.commit()
                        imported_count += len(duplicate_data)
                    except Exception as e:
                        conn.rollback()
                        error_rows.append(f"批量更新失败: {str(e)}")
                elif overwrite_choice == 'skip_all':
                    # 跳过所有重复数据
                    skipped_count += len(duplicate_data)
                    for dup in duplicate_data:
                        error_rows.append(f"第{dup['row_num']}行: 身份证号 {dup['youth_data'][0]} 已存在，跳过导入")
                else:
                    # 用户取消操作
                    skipped_count += len(duplicate_data)
                    for dup in duplicate_data:
                        error_rows.append(f"第{dup['row_num']}行: 身份证号 {dup['youth_data'][0]} 已存在，用户选择跳过")
            
            conn.close()
            
            if error_rows:
                error_summary = f"成功导入 {imported_count} 条记录"
                if skipped_count > 0:
                    error_summary += f"，跳过 {skipped_count} 条已存在的记录"
                error_summary += f"\n\n以下 {len(error_rows)} 条记录导入失败或跳过：\n" + "\n".join(error_rows[:10])
                if len(error_rows) > 10:
                    error_summary += f"\n... 还有 {len(error_rows) - 10} 条错误"
                return imported_count, error_summary
            
            return imported_count, None
        except Exception as e:
            return 0, f"文件读取失败: {str(e)}"
    
    def _ask_overwrite_duplicates(self, duplicate_data):
        """询问用户是否覆盖重复数据"""
        from PyQt5.QtWidgets import QMessageBox, QPushButton
        
        # 构建重复数据的详细信息
        duplicate_info = []
        for i, dup in enumerate(duplicate_data[:5], 1):  # 最多显示5条
            youth_data = dup['youth_data']
            existing_name = dup['existing_name']
            duplicate_info.append(f"{i}. 姓名: {youth_data[1]}, 身份证: {youth_data[0]}")
        
        if len(duplicate_data) > 5:
            duplicate_info.append(f"... 还有 {len(duplicate_data) - 5} 条重复记录")
        
        msg_box = QMessageBox()
        msg_box.setWindowTitle("发现重复数据")
        msg_box.setText(f"发现 {len(duplicate_data)} 条重复的身份证号记录：\n\n请选择处理方式：")
        msg_box.setDetailedText("\n".join(duplicate_info))
        msg_box.setIcon(QMessageBox.Icon.Question)
        
        # 添加自定义按钮
        overwrite_btn = QPushButton("覆盖所有重复数据")
        skip_btn = QPushButton("跳过所有重复数据")
        cancel_btn = QPushButton("取消导入")
        
        msg_box.addButton(overwrite_btn, QMessageBox.ButtonRole.AcceptRole)
        msg_box.addButton(skip_btn, QMessageBox.ButtonRole.RejectRole)
        msg_box.addButton(cancel_btn, QMessageBox.ButtonRole.DestructiveRole)
        
        msg_box.setDefaultButton(skip_btn)
        
        result = msg_box.exec_()
        clicked_button = msg_box.clickedButton()
        
        if clicked_button == overwrite_btn:
            return 'overwrite_all'
        elif clicked_button == skip_btn:
            return 'skip_all'
        else:
            return 'cancel'
    
    def import_abnormal_stats(self, file_path):
        """导入异常情况统计"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            imported_count = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                # 根据身份证号查找youth_id
                cursor.execute("SELECT id FROM youth WHERE id_card=?", (row[0],))
                result = cursor.fetchone()
                if not result:
                    continue
                
                youth_id = result[0]
                cursor.execute('''
                    INSERT INTO abnormal_stat (youth_id, abnormal_type, description,
                                              record_date, handler, status)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (youth_id, row[1], row[2], row[3], row[4], row[5]))
                imported_count += 1
            
            conn.commit()
            conn.close()
            return imported_count, None
        except Exception as e:
            return 0, str(e)
    
    def import_health_screening(self, file_path):
        """导入健康筛查数据"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            imported_count = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                cursor.execute("SELECT id FROM youth WHERE id_card=?", (row[0],))
                result = cursor.fetchone()
                if not result:
                    continue
                
                youth_id = result[0]
                cursor.execute('''
                    INSERT INTO health_screening (youth_id, screening_type, result,
                                                 screening_date, follow_up)
                    VALUES (?, ?, ?, ?, ?)
                ''', (youth_id, row[1], row[2], row[3], row[4]))
                imported_count += 1
            
            conn.commit()
            conn.close()
            return imported_count, None
        except Exception as e:
            return 0, str(e)
    
    def import_medical_screening(self, file_path):
        """导入病史筛查数据"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # 读取表头，建立列名到索引的映射
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            header_map = {header: idx for idx, header in enumerate(headers)}
            
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            imported_count = 0
            error_rows = []
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                
                try:
                    # 根据表头名称获取数据
                    def get_value(header_name):
                        idx = header_map.get(header_name, -1)
                        if idx >= 0 and idx < len(row) and row[idx]:
                            return str(row[idx]).strip()
                        return ''
                    
                    name = get_value('姓名')
                    gender = get_value('性别')
                    id_card = get_value('公民身份号码')
                    screening_result = get_value('筛查情况')
                    screening_date = get_value('筛查日期')
                    remark = get_value('备注')
                    physical_status = get_value('身体状况')
                    mental_status = get_value('精神状况')
                    
                    # 验证必填字段
                    if not name or not id_card:
                        error_rows.append(f"第{row_num}行: 姓名和身份证号不能为空")
                        continue
                    
                    # 验证1：身份证号码格式（18位）
                    if len(id_card) != 18:
                        error_rows.append(f"第{row_num}行 ({name}): 公民身份号码错误，必须是18位")
                        continue
                    
                    # 验证2：检查此人是否在基本信息库中存在
                    cursor.execute('SELECT COUNT(*) FROM youth WHERE id_card = ?', (id_card,))
                    exists_in_youth = cursor.fetchone()[0] > 0
                    
                    if not exists_in_youth:
                        error_rows.append(f"第{row_num}行 ({name}, {id_card}): 此人在基本信息库中不存在")
                        continue
                    
                    # 验证3：检查数据是否重复（同一个人在同一天的筛查记录）
                    cursor.execute('''
                        SELECT COUNT(*) FROM medical_screening 
                        WHERE id_card = ? AND screening_date = ?
                    ''', (id_card, screening_date))
                    is_duplicate = cursor.fetchone()[0] > 0
                    
                    if is_duplicate:
                        error_rows.append(f"第{row_num}行 ({name}, {id_card}): 数据重复，该人员在 {screening_date} 已有筛查记录")
                        continue
                    
                    # 从基本信息中获取应征地、连、排、班
                    cursor.execute("""
                        SELECT recruitment_place, company, platoon, squad 
                        FROM youth WHERE id_card = ?
                    """, (id_card,))
                    youth_info = cursor.fetchone()
                    
                    recruitment_place = youth_info[0] if youth_info and youth_info[0] else ''
                    company = youth_info[1] if youth_info and youth_info[1] else ''
                    platoon = youth_info[2] if youth_info and youth_info[2] else ''
                    squad = youth_info[3] if youth_info and youth_info[3] else ''
                    
                    # 插入病史筛查记录
                    cursor.execute('''
                        INSERT INTO medical_screening (youth_id_card, name, gender, id_card,
                                                     screening_result, screening_date, remark,
                                                     physical_status, mental_status,
                                                     recruitment_place, company, platoon, squad)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (id_card, name, gender, id_card, screening_result, screening_date, remark,
                          physical_status, mental_status, recruitment_place, company, platoon, squad))
                    
                    imported_count += 1
                    
                except Exception as e:
                    error_rows.append(f"第{row_num}行: {str(e)}")
                    continue
            
            conn.commit()
            
            # 导入完成后，同步所有新导入的病史筛查记录到异常情况统计表
            # 重新查询刚插入的记录，并进行异常情况同步
            cursor.execute('''
                SELECT id_card, name, gender, screening_date, physical_status, mental_status
                FROM medical_screening
                ORDER BY id DESC
                LIMIT ?
            ''', (imported_count,))
            
            newly_imported_records = cursor.fetchall()
            conn.close()
            
            if error_rows:
                error_summary = f"成功导入 {imported_count} 条记录\n\n以下 {len(error_rows)} 条记录导入失败：\n" + "\n".join(error_rows[:10])
                if len(error_rows) > 10:
                    error_summary += f"\n... 还有 {len(error_rows) - 10} 条错误"
                return imported_count, error_summary
            
            return imported_count, None
        except Exception as e:
            return 0, str(e)
    
    def import_daily_stats(self, file_path):
        """导入每日情况统计数据"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            imported_count = 0
            updated_count = 0
            skipped_count = 0
            error_rows = []
            duplicate_records = []  # 存储重复的记录
            
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            
            # 第一遍：收集所有数据并检查重复
            valid_data = []
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):  # 跳过完全空行
                    continue
                
                try:
                    # Excel列顺序：公民身份号码, 姓名, 日期, 思想, 身体, 精神, 训练, 管理, 其他
                    def safe_get_cell_value(row, index):
                        if len(row) > index and row[index] is not None:
                            return str(row[index]).strip()
                        return ''
                    
                    id_card = safe_get_cell_value(row, 0)  # 公民身份号码
                    name = safe_get_cell_value(row, 1)     # 姓名
                    record_date = safe_get_cell_value(row, 2)  # 日期
                    mood = safe_get_cell_value(row, 3)     # 思想
                    physical_condition = safe_get_cell_value(row, 4)  # 身体
                    mental_state = safe_get_cell_value(row, 5)  # 精神
                    training = safe_get_cell_value(row, 6)  # 训练
                    management = safe_get_cell_value(row, 7)  # 管理
                    notes = safe_get_cell_value(row, 8)    # 其他
                    
                    # 验证必填字段
                    if not id_card or not name:
                        error_rows.append(f"第{row_num}行: 公民身份号码和姓名不能为空")
                        continue
                    
                    # 验证身份证号码格式（18位）
                    if len(id_card) != 18:
                        error_rows.append(f"第{row_num}行 ({name}): 公民身份号码错误，必须是18位")
                        continue
                    
                    # 数据校验：检查此人是否在基本信息库中存在
                    cursor.execute('SELECT id FROM youth WHERE id_card = ?', (id_card,))
                    youth_result = cursor.fetchone()
                    
                    if not youth_result:
                        error_rows.append(f"第{row_num}行 ({name}, {id_card}): 此人在基本信息库中不存在，无法导入")
                        continue
                    
                    youth_id = youth_result[0]
                    
                    # 日期处理：如果为空则使用系统当前日期
                    if not record_date:
                        from datetime import datetime
                        record_date = datetime.now().strftime('%Y-%m-%d')
                    
                    # 设置默认值
                    if not mood:
                        mood = '正常'
                    if not physical_condition:
                        physical_condition = '正常'
                    if not mental_state:
                        mental_state = '正常'
                    if not training:
                        training = '正常'
                    if not management:
                        management = '正常'
                    
                    # 检查是否已存在相同日期的记录
                    cursor.execute('''
                        SELECT id FROM daily_stat 
                        WHERE youth_id = ? AND record_date = ?
                    ''', (youth_id, record_date))
                    
                    existing_record = cursor.fetchone()
                    if existing_record:
                        duplicate_records.append({
                            'row_num': row_num,
                            'name': name,
                            'youth_id': youth_id,
                            'record_date': record_date,
                            'data': (mood, physical_condition, mental_state, training, management, notes, youth_id, record_date)
                        })
                    else:
                        valid_data.append({
                            'row_num': row_num,
                            'data': (youth_id, record_date, mood, physical_condition, mental_state, training, management, notes)
                        })
                    
                except Exception as e:
                    error_rows.append(f"第{row_num}行: {str(e)}")
                    continue
            
            # 第二步：询问用户如何处理重复数据
            overwrite_choice = None
            if duplicate_records:
                overwrite_choice = self._ask_overwrite_duplicates_daily(duplicate_records)
                
                if overwrite_choice == 'cancel':
                    conn.close()
                    return 0, '用户取消导入操作'
            
            # 第三步：插入新数据
            for item in valid_data:
                try:
                    cursor.execute('''
                        INSERT INTO daily_stat (youth_id, record_date, mood, physical_condition, 
                                              mental_state, training, management, notes)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', item['data'])
                    imported_count += 1
                except Exception as e:
                    error_rows.append(f"第{item['row_num']}行: 插入失败 - {str(e)}")
            
            # 第四步：处理重复数据
            if duplicate_records:
                if overwrite_choice == 'overwrite_all':
                    # 逐条更新重复数据
                    for dup in duplicate_records:
                        try:
                            cursor.execute('''
                                UPDATE daily_stat SET
                                    mood=?, physical_condition=?, mental_state=?, training=?, management=?, notes=?
                                WHERE youth_id=? AND record_date=?
                            ''', dup['data'])
                            updated_count += 1
                        except Exception as e:
                            error_rows.append(f"第{dup['row_num']}行: 更新失败 - {str(e)}")
                elif overwrite_choice == 'skip_all':
                    skipped_count = len(duplicate_records)
                    for dup in duplicate_records:
                        error_rows.append(f"第{dup['row_num']}行 ({dup['name']}): 该人员在 {dup['record_date']} 已有记录，跳过导入")
            
            conn.commit()
            conn.close()
            
            # 构建结果消息
            result_parts = []
            if imported_count > 0:
                result_parts.append(f"成功导入 {imported_count} 条新记录")
            if updated_count > 0:
                result_parts.append(f"覆盖更新 {updated_count} 条记录")
            if skipped_count > 0:
                result_parts.append(f"跳过 {skipped_count} 条重复记录")
            
            result_message = "、".join(result_parts) if result_parts else "没有导入任何记录"
            
            if error_rows:
                error_summary = f"{result_message}\n\n以下 {len(error_rows)} 条记录导入失败或跳过：\n" + "\n".join(error_rows[:10])
                if len(error_rows) > 10:
                    error_summary += f"\n... 还有 {len(error_rows) - 10} 条错误"
                return imported_count + updated_count, error_summary
            
            return imported_count + updated_count, None if (imported_count + updated_count) > 0 else "没有导入任何记录"
        except Exception as e:
            return 0, f"文件读取失败: {str(e)}"
    
    def _ask_overwrite_duplicates_daily(self, duplicate_records):
        """询问用户是否覆盖重复的每日情况统计记录"""
        from PyQt5.QtWidgets import QMessageBox, QPushButton
        
        # 构建重复数据的详细信息
        duplicate_info = []
        for i, dup in enumerate(duplicate_records[:5], 1):
            duplicate_info.append(f"{i}. 姓名: {dup['name']}, 日期: {dup['record_date']}")
        
        if len(duplicate_records) > 5:
            duplicate_info.append(f"... 还有 {len(duplicate_records) - 5} 条重复记录")
        
        msg_box = QMessageBox()
        msg_box.setWindowTitle("发现重复数据")
        msg_box.setText(f"发现 {len(duplicate_records)} 条重复记录（同一人员在同一日期已有记录）：\n\n请选择处理方式：")
        msg_box.setDetailedText("\n".join(duplicate_info))
        msg_box.setIcon(QMessageBox.Icon.Question)
        
        # 添加自定义按钮
        overwrite_btn = QPushButton("覆盖所有重复数据")
        skip_btn = QPushButton("跳过所有重复数据")
        cancel_btn = QPushButton("取消导入")
        
        msg_box.addButton(overwrite_btn, QMessageBox.ButtonRole.AcceptRole)
        msg_box.addButton(skip_btn, QMessageBox.ButtonRole.RejectRole)
        msg_box.addButton(cancel_btn, QMessageBox.ButtonRole.DestructiveRole)
        
        msg_box.setDefaultButton(skip_btn)
        
        result = msg_box.exec_()
        clicked_button = msg_box.clickedButton()
        
        if clicked_button == overwrite_btn:
            return 'overwrite_all'
        elif clicked_button == skip_btn:
            return 'skip_all'
        else:
            return 'cancel'

    def save_scanned_document(self, youth_id, module_type, file_path, notes=''):
        """保存扫描文档"""
        conn = self.db_manager.get_connection()
        cursor = conn.cursor()
        
        table_map = {
            'medical': 'medical_history',
            'town': 'town_interview',
            'leader': 'leader_interview'
        }
        
        table_name = table_map.get(module_type)
        if not table_name:
            return False
        
        date_field = 'upload_date' if module_type == 'medical' else 'interview_date'
        current_date = datetime.now().strftime('%Y-%m-%d')
        
        cursor.execute(f'''
            INSERT INTO {table_name} (youth_id, file_path, {date_field}, notes)
            VALUES (?, ?, ?, ?)
        ''', (youth_id, file_path, current_date, notes))
        
        conn.commit()
        conn.close()
        return True

    def import_political_assessment(self, file_path):
        """导入政治考核情况统计数据"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            imported_count = 0
            updated_count = 0
            skipped_count = 0
            error_rows = []
            duplicate_records = []  # 存储重复的记录
            
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            
            # 第一遍：检查所有数据，找出重复记录
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                
                try:
                    def safe_get_cell_value(row, index):
                        if len(row) > index and row[index] is not None:
                            return str(row[index]).strip()
                        return ''
                    
                    name = safe_get_cell_value(row, 1)
                    id_card = safe_get_cell_value(row, 3)
                    assessment_date = safe_get_cell_value(row, 8)
                    
                    if not name or not id_card:
                        continue
                    
                    # 日期处理
                    if not assessment_date:
                        from datetime import datetime
                        assessment_date = datetime.now().strftime('%Y-%m-%d')
                    
                    # 检查是否已存在
                    existing_id = self.db_manager.check_political_assessment_exists(id_card, name, assessment_date)
                    if existing_id:
                        duplicate_records.append({
                            'row_num': row_num,
                            'name': name,
                            'id_card': id_card,
                            'date': assessment_date,
                            'row_data': row
                        })
                except Exception as e:
                    continue
            
            # 如果有重复记录，询问用户如何处理
            overwrite_choice = None
            if duplicate_records:
                overwrite_choice = self._ask_overwrite_duplicates_political(duplicate_records)
                
                if overwrite_choice == 'cancel':
                    conn.close()
                    return 0, '用户取消导入操作'
            
            # 第二遍：执行导入
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                
                try:
                    def safe_get_cell_value(row, index):
                        if len(row) > index and row[index] is not None:
                            return str(row[index]).strip()
                        return ''
                    
                    name = safe_get_cell_value(row, 1)
                    gender = safe_get_cell_value(row, 2)
                    id_card = safe_get_cell_value(row, 3)
                    family_member_info = safe_get_cell_value(row, 4)
                    visit_survey = safe_get_cell_value(row, 5)
                    political_assessment = safe_get_cell_value(row, 6)
                    key_attention = safe_get_cell_value(row, 7)
                    assessment_date = safe_get_cell_value(row, 8)
                    thoughts = safe_get_cell_value(row, 9)
                    spirit = safe_get_cell_value(row, 10)
                    
                    if not name or not id_card:
                        error_rows.append(f"第{row_num}行: 姓名和公民身份号码不能为空")
                        continue
                    
                    if len(id_card) != 18:
                        error_rows.append(f"第{row_num}行 ({name}): 公民身份号码错误，必须是18位")
                        continue
                    
                    cursor.execute('SELECT COUNT(*) FROM youth WHERE id_card = ?', (id_card,))
                    exists_in_youth = cursor.fetchone()[0] > 0
                    
                    if not exists_in_youth:
                        error_rows.append(f"第{row_num}行 ({name}, {id_card}): 此人在基本信息库中不存在")
                        continue
                    
                    if not assessment_date:
                        from datetime import datetime
                        assessment_date = datetime.now().strftime('%Y-%m-%d')
                    
                    # 检查是否是重复记录
                    existing_id = self.db_manager.check_political_assessment_exists(id_card, name, assessment_date)
                    
                    if existing_id:
                        # 根据用户选择处理重复记录
                        if overwrite_choice == 'overwrite_all':
                            # 覆盖更新
                            success = self.db_manager.update_political_assessment_by_unique_key(
                                id_card, name, assessment_date, family_member_info, visit_survey,
                                political_assessment, key_attention, thoughts, spirit
                            )
                            if success:
                                updated_count += 1
                        elif overwrite_choice == 'skip_all':
                            # 跳过
                            skipped_count += 1
                            continue
                    else:
                        # 插入新记录
                        cursor.execute('''
                            INSERT INTO political_assessment (youth_id_card, name, gender, id_card,
                                                            family_member_info, visit_survey, political_assessment,
                                                            key_attention, assessment_date, thoughts, spirit)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (id_card, name, gender, id_card, family_member_info, visit_survey,
                              political_assessment, key_attention, assessment_date, thoughts, spirit))
                        
                        imported_count += 1
                    
                except Exception as e:
                    error_rows.append(f"第{row_num}行: {str(e)}")
                    continue
            
            conn.commit()
            conn.close()
            
            # 构建结果消息
            result_parts = []
            if imported_count > 0:
                result_parts.append(f"成功导入 {imported_count} 条新记录")
            if updated_count > 0:
                result_parts.append(f"覆盖更新 {updated_count} 条记录")
            if skipped_count > 0:
                result_parts.append(f"跳过 {skipped_count} 条重复记录")
            
            result_message = "、".join(result_parts) if result_parts else "没有导入任何记录"
            
            if error_rows:
                error_summary = f"{result_message}\n\n以下 {len(error_rows)} 条记录导入失败：\n" + "\n".join(error_rows[:10])
                if len(error_rows) > 10:
                    error_summary += f"\n... 还有 {len(error_rows) - 10} 条错误"
                return imported_count + updated_count, error_summary
            
            return imported_count + updated_count, None if (imported_count + updated_count) > 0 else "没有导入任何记录"
        except Exception as e:
            return 0, f"文件读取失败: {str(e)}"
    
    def _ask_overwrite_duplicates_political(self, duplicate_records):
        """询问用户是否覆盖重复的政治考核情况记录"""
        from PyQt5.QtWidgets import QMessageBox, QPushButton
        
        # 构建重复数据的详细信息
        duplicate_info = []
        for i, dup in enumerate(duplicate_records[:5], 1):
            duplicate_info.append(f"{i}. 姓名: {dup['name']}, 身份证: {dup['id_card']}, 日期: {dup['date']}")
        
        if len(duplicate_records) > 5:
            duplicate_info.append(f"... 还有 {len(duplicate_records) - 5} 条重复记录")
        
        msg_box = QMessageBox()
        msg_box.setWindowTitle("发现重复数据")
        msg_box.setText(f"发现 {len(duplicate_records)} 条重复记录（身份证号、姓名和日期相同）：\n\n请选择处理方式：")
        msg_box.setDetailedText("\n".join(duplicate_info))
        msg_box.setIcon(QMessageBox.Icon.Question)
        
        # 添加自定义按钮
        overwrite_btn = QPushButton("覆盖所有重复数据")
        skip_btn = QPushButton("跳过所有重复数据")
        cancel_btn = QPushButton("取消导入")
        
        msg_box.addButton(overwrite_btn, QMessageBox.ButtonRole.AcceptRole)
        msg_box.addButton(skip_btn, QMessageBox.ButtonRole.RejectRole)
        msg_box.addButton(cancel_btn, QMessageBox.ButtonRole.DestructiveRole)
        
        msg_box.setDefaultButton(skip_btn)
        
        result = msg_box.exec_()
        clicked_button = msg_box.clickedButton()
        
        if clicked_button == overwrite_btn:
            return 'overwrite_all'
        elif clicked_button == skip_btn:
            return 'skip_all'
        else:
            return 'cancel'
