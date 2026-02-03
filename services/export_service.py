"""
导出和打印服务
"""
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
import os


class ExportService:
    def __init__(self, db_manager):
        self.db_manager = db_manager
    
    def export_youth_to_pdf(self, youth_id_card, output_path):
        """导出青年完整信息为PDF - 根据现有的6个模块+2个谈心谈话模块进行导出"""
        try:
            # 注册中文字体
            font_path = "C:\\Windows\\Fonts\\simhei.ttf"  # 黑体
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('SimHei', font_path))
                font_name = 'SimHei'
            else:
                font_name = 'Helvetica'
            
            # 获取青年基本信息
            youth = self.db_manager.get_youth_by_id_card(youth_id_card)
            if not youth:
                return False, "未找到青年信息"
            
            # 创建输出文件夹
            output_dir = os.path.dirname(output_path)
            pdf_filename = os.path.basename(output_path)
            images_dir = os.path.join(output_dir, f"{youth.name}_谈心谈话图片")
            os.makedirs(images_dir, exist_ok=True)
            
            # 创建PDF
            c = canvas.Canvas(output_path, pagesize=A4)
            width, height = A4
            
            # 标题
            c.setFont(font_name, 18)
            c.drawCentredString(width/2, height - 2*cm, f"青年详细信息 - {youth.name}")
            
            y_position = height - 3.5*cm
            
            # ========== 模块一：基本信息 ==========
            c.setFont(font_name, 14)
            c.drawString(2*cm, y_position, "一、基本信息")
            y_position -= 0.8*cm
            
            c.setFont(font_name, 10)
            basic_fields = [
                ("姓名", youth.name),
                ("性别", youth.gender),
                ("公民身份号码", youth.id_card),
                ("出生日期", getattr(youth, 'birth_date', '')),
                ("民族", youth.nation),
                ("政治面貌", youth.political_status),
                ("宗教信仰", getattr(youth, 'religion', '')),
                ("籍贯", getattr(youth, 'native_place', '')),
                ("文化程度", youth.education_level),
                ("学业情况", youth.study_status),
                ("学习类型", youth.study_type),
                ("入营时间", getattr(youth, 'camp_entry_time', '')),
                ("应征地", getattr(youth, 'recruitment_place', '')),
                ("经常居住地址", youth.residence_address),
                ("户籍所在地", youth.household_address),
                ("邮编", getattr(youth, 'postal_code', '')),
                ("本人电话", getattr(youth, 'personal_phone', '')),
                ("家庭电话", getattr(youth, 'family_phone', '')),
                ("毕业(就读)学校", youth.school),
                ("所学专业", youth.major),
                ("入学时间", getattr(youth, 'enrollment_time', '')),
                ("初检医院", getattr(youth, 'initial_hospital', '')),
                ("初检结论", getattr(youth, 'initial_conclusion', '')),
                ("初检时间", getattr(youth, 'initial_time', '')),
                ("体检结论", getattr(youth, 'physical_conclusion', '')),
                ("体检时间", getattr(youth, 'physical_time', '')),
                ("体检不合格原因", getattr(youth, 'physical_disqualification', '')),
                ("主检医师意见", getattr(youth, 'chief_doctor_opinion', '')),
                ("毕业时间", getattr(youth, 'graduation_time', '')),
                ("连", youth.company),
                ("排", youth.platoon),
                ("班", youth.squad),
                ("带训班长信息", youth.squad_leader),
                ("在营状态", youth.camp_status),
                ("离营时间", youth.leave_time),
                ("离营原因", getattr(youth, 'leave_reason', ''))
            ]
            
            for label, value in basic_fields:
                if y_position < 3*cm:
                    c.showPage()
                    y_position = height - 2*cm
                    c.setFont(font_name, 10)
                
                text = f"{label}: {value or '无'}"
                # 处理长文本换行
                if len(text) > 60:
                    lines = [text[i:i+60] for i in range(0, len(text), 60)]
                    for line in lines:
                        c.drawString(2.5*cm, y_position, line)
                        y_position -= 0.5*cm
                else:
                    c.drawString(2.5*cm, y_position, text)
                    y_position -= 0.5*cm
            
            # ========== 模块二：病史筛查情况 ==========
            if y_position < 5*cm:
                c.showPage()
                y_position = height - 2*cm
            
            y_position -= 0.5*cm
            c.setFont(font_name, 14)
            c.drawString(2*cm, y_position, "二、病史筛查情况")
            y_position -= 0.8*cm
            
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT screening_result, screening_date, physical_status, mental_status, remark
                FROM medical_screening 
                WHERE id_card = ?
                ORDER BY screening_date DESC
            ''', (youth_id_card,))
            medical_records = cursor.fetchall()
            
            c.setFont(font_name, 10)
            if medical_records:
                for idx, record in enumerate(medical_records, 1):
                    if y_position < 3*cm:
                        c.showPage()
                        y_position = height - 2*cm
                        c.setFont(font_name, 10)
                    
                    c.drawString(2.5*cm, y_position, f"记录{idx}:")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"筛查日期: {record[1] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"筛查情况: {record[0] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"身体状况: {record[2] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"精神状况: {record[3] or '无'}")
                    y_position -= 0.5*cm
                    if record[4]:  # 备注
                        c.drawString(3*cm, y_position, f"备注: {record[4]}")
                        y_position -= 0.5*cm
                    y_position -= 0.2*cm
            else:
                c.drawString(2.5*cm, y_position, "暂无病史筛查记录")
                y_position -= 0.7*cm
            
            # ========== 模块三：体检情况统计表 ==========
            if y_position < 5*cm:
                c.showPage()
                y_position = height - 2*cm
            
            y_position -= 0.5*cm
            c.setFont(font_name, 14)
            c.drawString(2*cm, y_position, "三、体检情况统计表")
            y_position -= 0.8*cm
            
            cursor.execute('''
                SELECT district_exam, district_positive, district_date, 
                       city_exam, city_positive, city_date,
                       special_exam, special_positive, special_date,
                       body_status, psychological_test_type, tracking_opinion, 
                       implementation_status, company, platoon, squad, recruitment_place
                FROM physical_examination 
                WHERE youth_id_card = ?
                ORDER BY id DESC
            ''', (youth_id_card,))
            physical_records = cursor.fetchall()
            
            c.setFont(font_name, 10)
            if physical_records:
                for idx, record in enumerate(physical_records, 1):
                    if y_position < 3*cm:
                        c.showPage()
                        y_position = height - 2*cm
                        c.setFont(font_name, 10)
                    
                    c.drawString(2.5*cm, y_position, f"记录{idx}:")
                    y_position -= 0.5*cm
                    
                    # 区级检查
                    if record[0] or record[1] or record[2]:
                        c.drawString(3*cm, y_position, f"区级检查: {record[0] or '无'}")
                        y_position -= 0.4*cm
                        c.drawString(3*cm, y_position, f"  阳性特征/边缘问题: {record[1] or '无'}")
                        y_position -= 0.4*cm
                        c.drawString(3*cm, y_position, f"  日期: {record[2] or '无'}")
                        y_position -= 0.4*cm
                    
                    # 市级检查
                    if record[3] or record[4] or record[5]:
                        c.drawString(3*cm, y_position, f"市级检查: {record[3] or '无'}")
                        y_position -= 0.4*cm
                        c.drawString(3*cm, y_position, f"  阳性特征/边缘问题: {record[4] or '无'}")
                        y_position -= 0.4*cm
                        c.drawString(3*cm, y_position, f"  日期: {record[5] or '无'}")
                        y_position -= 0.4*cm
                    
                    # 专项检查
                    if record[6] or record[7] or record[8]:
                        c.drawString(3*cm, y_position, f"专项检查: {record[6] or '无'}")
                        y_position -= 0.4*cm
                        c.drawString(3*cm, y_position, f"  阳性特征/边缘问题: {record[7] or '无'}")
                        y_position -= 0.4*cm
                        c.drawString(3*cm, y_position, f"  日期: {record[8] or '无'}")
                        y_position -= 0.4*cm
                    
                    c.drawString(3*cm, y_position, f"身体状况: {record[9] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"心理测试类型: {record[10] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"跟踪处置意见: {record[11] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"处置意见落实情况: {record[12] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"连排班: {record[13] or '无'}-{record[14] or '无'}-{record[15] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"应征地: {record[16] or '无'}")
                    y_position -= 0.7*cm
            else:
                c.drawString(2.5*cm, y_position, "暂无体检情况记录")
                y_position -= 0.7*cm
            
            # ========== 模块四：每日情况统计 ==========
            if y_position < 5*cm:
                c.showPage()
                y_position = height - 2*cm
            
            y_position -= 0.5*cm
            c.setFont(font_name, 14)
            c.drawString(2*cm, y_position, "四、每日情况统计")
            y_position -= 0.8*cm
            
            cursor.execute('''
                SELECT record_date, mood, physical_condition, mental_state, training, management, notes
                FROM daily_stat 
                WHERE youth_id = (SELECT id FROM youth WHERE id_card = ?)
                ORDER BY record_date DESC
                LIMIT 20
            ''', (youth_id_card,))
            daily_records = cursor.fetchall()
            
            c.setFont(font_name, 10)
            if daily_records:
                for idx, record in enumerate(daily_records, 1):
                    if y_position < 3*cm:
                        c.showPage()
                        y_position = height - 2*cm
                        c.setFont(font_name, 10)
                    
                    c.drawString(2.5*cm, y_position, f"记录{idx}:")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"日期: {record[0] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"思想: {record[1] or '无'} | 身体: {record[2] or '无'} | 精神: {record[3] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"训练: {record[4] or '无'} | 管理: {record[5] or '无'}")
                    y_position -= 0.5*cm
                    if record[6]:  # 备注
                        c.drawString(3*cm, y_position, f"备注: {record[6][:50]}")
                        y_position -= 0.5*cm
                    y_position -= 0.2*cm
            else:
                c.drawString(2.5*cm, y_position, "暂无每日情况记录")
                y_position -= 0.7*cm
            
            # ========== 模块五：政治考核情况统计 ==========
            if y_position < 5*cm:
                c.showPage()
                y_position = height - 2*cm
            
            y_position -= 0.5*cm
            c.setFont(font_name, 14)
            c.drawString(2*cm, y_position, "五、政治考核情况统计")
            y_position -= 0.8*cm
            
            cursor.execute('''
                SELECT assessment_date, family_member_info, visit_survey, political_assessment, 
                       key_attention, thoughts, spirit
                FROM political_assessment 
                WHERE youth_id_card = ?
                ORDER BY assessment_date DESC
            ''', (youth_id_card,))
            political_records = cursor.fetchall()
            
            c.setFont(font_name, 10)
            if political_records:
                for idx, record in enumerate(political_records, 1):
                    if y_position < 3*cm:
                        c.showPage()
                        y_position = height - 2*cm
                        c.setFont(font_name, 10)
                    
                    c.drawString(2.5*cm, y_position, f"记录{idx}:")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"考核日期: {record[0] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"家庭成员信息: {record[1] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"走访调查: {record[2] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"政治考核: {record[3] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"重点关注: {record[4] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"思想: {record[5] or '无'}")
                    y_position -= 0.4*cm
                    c.drawString(3*cm, y_position, f"精神: {record[6] or '无'}")
                    y_position -= 0.7*cm
            else:
                c.drawString(2.5*cm, y_position, "暂无政治考核记录")
                y_position -= 0.7*cm
            
            # ========== 模块六：入营点验情况 ==========
            if y_position < 5*cm:
                c.showPage()
                y_position = height - 2*cm
            
            y_position -= 0.5*cm
            c.setFont(font_name, 14)
            c.drawString(2*cm, y_position, "六、入营点验情况")
            y_position -= 0.8*cm
            
            cursor.execute('''
                SELECT data, item, usage, Disposal
                FROM camp_verification 
                WHERE user_id = ?
                ORDER BY data DESC
            ''', (youth_id_card,))
            camp_records = cursor.fetchall()
            
            c.setFont(font_name, 10)
            if camp_records:
                for idx, record in enumerate(camp_records, 1):
                    if y_position < 3*cm:
                        c.showPage()
                        y_position = height - 2*cm
                        c.setFont(font_name, 10)
                    
                    c.drawString(2.5*cm, y_position, f"记录{idx}:")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"日期: {record[0] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"项目: {record[1] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"用途: {record[2] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"处理: {record[3] or '无'}")
                    y_position -= 0.7*cm
            else:
                c.drawString(2.5*cm, y_position, "暂无入营点验记录")
                y_position -= 0.7*cm
            
            # ========== 模块七：镇街谈心谈话情况 ==========
            if y_position < 5*cm:
                c.showPage()
                y_position = height - 2*cm
            
            y_position -= 0.5*cm
            c.setFont(font_name, 14)
            c.drawString(2*cm, y_position, "七、镇街谈心谈话情况")
            y_position -= 0.8*cm
            
            cursor.execute('''
                SELECT id, interview_date, thoughts, spirit, visit_survey_image
                FROM town_interview 
                WHERE youth_id_card = ?
                ORDER BY interview_date DESC
            ''', (youth_id_card,))
            town_records = cursor.fetchall()
            
            c.setFont(font_name, 10)
            if town_records:
                for idx, record in enumerate(town_records, 1):
                    if y_position < 3*cm:
                        c.showPage()
                        y_position = height - 2*cm
                        c.setFont(font_name, 10)
                    
                    c.drawString(2.5*cm, y_position, f"记录{idx}:")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"日期: {record[1] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"思想: {record[2] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"精神: {record[3] or '无'}")
                    y_position -= 0.5*cm
                    
                    # 处理图片
                    if record[4]:  # visit_survey_image
                        try:
                            image_filename = f"镇街谈心谈话_{youth.name}_{record[0]}.jpg"
                            image_path = os.path.join(images_dir, image_filename)
                            
                            # 保存图片到文件（作为备份）
                            with open(image_path, 'wb') as img_file:
                                img_file.write(record[4])
                            
                            # 在PDF中直接插入图片
                            c.drawString(3*cm, y_position, "图片:")
                            y_position -= 0.5*cm
                            
                            # 检查页面空间，如果空间不足则换页
                            if y_position < 8*cm:  # 为图片预留足够空间
                                c.showPage()
                                y_position = height - 2*cm
                                c.setFont(font_name, 10)
                            
                            try:
                                # 将图片插入PDF
                                from reportlab.lib.utils import ImageReader
                                from io import BytesIO
                                
                                # 从BLOB数据创建图片对象
                                img_data = BytesIO(record[4])
                                img = ImageReader(img_data)
                                
                                # 获取图片尺寸并计算合适的显示尺寸
                                img_width, img_height = img.getSize()
                                
                                # 设置最大显示尺寸（适配PDF宽度，保持宽高比）
                                page_width = 21*cm  # A4页面宽度
                                left_margin = 2*cm  # 左边距
                                right_margin = 2*cm  # 右边距
                                available_width = page_width - left_margin - right_margin  # 可用宽度约17cm
                                
                                max_width = available_width - 1*cm  # 预留1cm边距，实际约16cm
                                max_height = 8*cm  # 适当增加最大高度
                                
                                # 计算缩放比例（适配PDF宽度，保持宽高比）
                                width_ratio = max_width / img_width
                                height_ratio = max_height / img_height
                                scale_ratio = min(width_ratio, height_ratio)  # 移除1.0限制，允许放大和缩小
                                
                                display_width = img_width * scale_ratio
                                display_height = img_height * scale_ratio
                                
                                # 计算居中位置
                                # 页面中心位置
                                page_center_x = page_width / 2
                                # 图片左边位置（居中）
                                img_x = page_center_x - (display_width / 2)
                                
                                # 在PDF中绘制居中的图片
                                c.drawImage(img, img_x, y_position - display_height, 
                                          width=display_width, height=display_height)
                                
                                y_position -= (display_height + 0.5*cm)
                                
                            except Exception as img_error:
                                c.drawString(3*cm, y_position, f"图片显示失败: {str(img_error)}")
                                c.setFillColorRGB(0.5, 0.5, 0.5)
                                c.drawString(3*cm, y_position - 0.4*cm, f"图片已保存到: {image_filename}")
                                c.setFillColorRGB(0, 0, 0)
                                y_position -= 1*cm
                                
                        except Exception as e:
                            c.drawString(3*cm, y_position, f"图片处理失败: {str(e)}")
                            y_position -= 0.5*cm
                    else:
                        c.drawString(3*cm, y_position, "图片: 无")
                        y_position -= 0.5*cm
                    
                    y_position -= 0.2*cm
            else:
                c.drawString(2.5*cm, y_position, "暂无镇街谈心谈话记录")
                y_position -= 0.7*cm
            
            # ========== 模块八：领导谈心谈话情况 ==========
            if y_position < 5*cm:
                c.showPage()
                y_position = height - 2*cm
            
            y_position -= 0.5*cm
            c.setFont(font_name, 14)
            c.drawString(2*cm, y_position, "八、领导谈心谈话情况")
            y_position -= 0.8*cm
            
            cursor.execute('''
                SELECT id, interview_date, thoughts, spirit, visit_survey_image
                FROM leader_interview 
                WHERE youth_id_card = ?
                ORDER BY interview_date DESC
            ''', (youth_id_card,))
            leader_records = cursor.fetchall()
            
            c.setFont(font_name, 10)
            if leader_records:
                for idx, record in enumerate(leader_records, 1):
                    if y_position < 3*cm:
                        c.showPage()
                        y_position = height - 2*cm
                        c.setFont(font_name, 10)
                    
                    c.drawString(2.5*cm, y_position, f"记录{idx}:")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"日期: {record[1] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"思想: {record[2] or '无'}")
                    y_position -= 0.5*cm
                    c.drawString(3*cm, y_position, f"精神: {record[3] or '无'}")
                    y_position -= 0.5*cm
                    
                    # 处理图片
                    if record[4]:  # visit_survey_image
                        try:
                            image_filename = f"领导谈心谈话_{youth.name}_{record[0]}.jpg"
                            image_path = os.path.join(images_dir, image_filename)
                            
                            # 保存图片到文件（作为备份）
                            with open(image_path, 'wb') as img_file:
                                img_file.write(record[4])
                            
                            # 在PDF中直接插入图片
                            c.drawString(3*cm, y_position, "图片:")
                            y_position -= 0.5*cm
                            
                            # 检查页面空间，如果空间不足则换页
                            if y_position < 8*cm:  # 为图片预留足够空间
                                c.showPage()
                                y_position = height - 2*cm
                                c.setFont(font_name, 10)
                            
                            try:
                                # 将图片插入PDF
                                from reportlab.lib.utils import ImageReader
                                from io import BytesIO
                                
                                # 从BLOB数据创建图片对象
                                img_data = BytesIO(record[4])
                                img = ImageReader(img_data)
                                
                                # 获取图片尺寸并计算合适的显示尺寸
                                img_width, img_height = img.getSize()
                                
                                # 设置最大显示尺寸（适配PDF宽度，保持宽高比）
                                page_width = 21*cm  # A4页面宽度
                                left_margin = 2*cm  # 左边距
                                right_margin = 2*cm  # 右边距
                                available_width = page_width - left_margin - right_margin  # 可用宽度约17cm
                                
                                max_width = available_width - 1*cm  # 预留1cm边距，实际约16cm
                                max_height = 8*cm  # 适当增加最大高度
                                
                                # 计算缩放比例（适配PDF宽度，保持宽高比）
                                width_ratio = max_width / img_width
                                height_ratio = max_height / img_height
                                scale_ratio = min(width_ratio, height_ratio)  # 移除1.0限制，允许放大和缩小
                                
                                display_width = img_width * scale_ratio
                                display_height = img_height * scale_ratio
                                
                                # 计算居中位置
                                # 页面中心位置
                                page_center_x = page_width / 2
                                # 图片左边位置（居中）
                                img_x = page_center_x - (display_width / 2)
                                
                                # 在PDF中绘制居中的图片
                                c.drawImage(img, img_x, y_position - display_height, 
                                          width=display_width, height=display_height)
                                
                                y_position -= (display_height + 0.5*cm)
                                
                            except Exception as img_error:
                                c.drawString(3*cm, y_position, f"图片显示失败: {str(img_error)}")
                                c.setFillColorRGB(0.5, 0.5, 0.5)
                                c.drawString(3*cm, y_position - 0.4*cm, f"图片已保存到: {image_filename}")
                                c.setFillColorRGB(0, 0, 0)
                                y_position -= 1*cm
                                
                        except Exception as e:
                            c.drawString(3*cm, y_position, f"图片处理失败: {str(e)}")
                            y_position -= 0.5*cm
                    else:
                        c.drawString(3*cm, y_position, "图片: 无")
                        y_position -= 0.5*cm
                    
                    y_position -= 0.2*cm
            else:
                c.drawString(2.5*cm, y_position, "暂无领导谈心谈话记录")
                y_position -= 0.7*cm
            
            conn.close()
            
            c.save()
            
            # 检查是否有图片被导出
            image_count = 0
            if os.path.exists(images_dir):
                image_count = len([f for f in os.listdir(images_dir) if f.endswith('.jpg')])
            
            success_msg = f"成功导出 {youth.name} 的完整信息到PDF"
            if image_count > 0:
                success_msg += f"\n同时导出了 {image_count} 张谈心谈话图片到文件夹: {os.path.basename(images_dir)}"
            
            return True, success_msg
        
        except Exception as e:
            return False, f"导出失败: {str(e)}"
    
    def generate_daily_report(self, youth_id, output_path):
        """生成每日情况报告"""
        try:
            youth = self.db_manager.get_youth_by_id(youth_id)
            daily_stats = self.db_manager.get_daily_stats_for_chart(youth_id, 30)
            
            c = canvas.Canvas(output_path, pagesize=A4)
            width, height = A4
            
            c.setFont("Helvetica-Bold", 16)
            c.drawString(2*cm, height - 2*cm, f"Daily Report - {youth[1]}")
            
            y_position = height - 4*cm
            c.setFont("Helvetica", 10)
            
            for stat in daily_stats:
                record_text = f"{stat[0]}: Mood-{stat[1]}, Physical-{stat[2]}, Training-{stat[3]}"
                c.drawString(2*cm, y_position, record_text)
                y_position -= 0.7*cm
                
                if y_position < 2*cm:
                    break
            
            c.save()
            return True, "报告生成成功"
        
        except Exception as e:
            return False, f"生成失败: {str(e)}"
    
    def export_visit_surveys_with_images(self, records, export_dir):
        """导出走访调查情况数据和图片"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from datetime import datetime
            import os
            
            # 创建图片文件夹
            images_folder = os.path.join(export_dir, '走访调查图片')
            os.makedirs(images_folder, exist_ok=True)
            
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '走访调查情况'
            
            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # 设置表头
            headers = ['日期', '姓名', '性别', '公民身份号码', '思想', '精神', '走访调查情况', '导出时间']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            for row_idx, record in enumerate(records, 2):
                # record格式: (id, youth_id_card, youth_name, gender, survey_date, thoughts, spirit, created_at)
                ws.cell(row=row_idx, column=1, value=record[4] or '')  # 日期
                ws.cell(row=row_idx, column=2, value=record[2] or '')  # 姓名
                ws.cell(row=row_idx, column=3, value=record[3] or '')  # 性别
                ws.cell(row=row_idx, column=4, value=record[1] or '')  # 公民身份号码
                ws.cell(row=row_idx, column=5, value=record[5] or '')  # 思想
                ws.cell(row=row_idx, column=6, value=record[6] or '')  # 精神
                ws.cell(row=row_idx, column=8, value=export_time)  # 导出时间
                
                # 处理图片
                try:
                    image_data = self.db_manager.get_visit_survey_image(record[0])
                    if image_data:
                        # 保存图片到文件夹
                        image_filename = f"{record[2]}_{record[1]}_走访调查.jpg"
                        image_path = os.path.join(images_folder, image_filename)
                        
                        with open(image_path, 'wb') as img_file:
                            img_file.write(image_data)
                        
                        # 在Excel中创建超链接
                        relative_path = f"走访调查图片/{image_filename}"
                        cell = ws.cell(row=row_idx, column=7, value="查看图片")
                        cell.hyperlink = relative_path
                        cell.style = "Hyperlink"
                    else:
                        ws.cell(row=row_idx, column=7, value='无图片')
                except Exception as e:
                    print(f"处理图片时出错: {e}")
                    ws.cell(row=row_idx, column=7, value='图片处理失败')
            
            # 保存Excel文件到导出文件夹
            excel_path = os.path.join(export_dir, '走访调查情况数据.xlsx')
            wb.save(excel_path)
            
            return True, f'数据已导出到:\n{export_dir}\n\n包含:\n- Excel文件: 走访调查情况数据.xlsx\n- 图片文件夹: 走访调查图片'
            
        except Exception as e:
            return False, f'导出失败: {str(e)}'

    
    def export_medical_screening_to_excel(self, output_path):
        """导出病史筛查数据到Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "病史筛查数据"
            
            # 设置表头
            headers = ['序号', '姓名', '性别', '公民身份号码', '筛查情况', '筛查日期', '备注', '身体状况', '精神状况']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, name='Microsoft YaHei')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            
            # 获取病史筛查数据
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT name, gender, id_card, screening_result, screening_date, remark,
                       physical_status, mental_status
                FROM medical_screening 
                ORDER BY id
            """)
            results = cursor.fetchall()
            conn.close()
            
            # 填充数据
            for row_num, data in enumerate(results, 2):
                ws.cell(row=row_num, column=1, value=row_num - 1)  # 序号
                for col, value in enumerate(data, 2):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.font = Font(name='Microsoft YaHei')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 调整列宽
            column_widths = [8, 12, 8, 20, 30, 12, 15, 12, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # 保存文件
            wb.save(output_path)
            return True, f"成功导出 {len(results)} 条病史筛查记录"
            
        except Exception as e:
            return False, f"导出失败: {str(e)}"
    
    def export_selected_medical_screening_to_excel(self, output_path, selected_ids):
        """批量导出选中的病史筛查数据到Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            if not selected_ids:
                return False, "没有选中任何记录"
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "病史筛查数据"
            
            # 设置表头
            headers = ['序号', '姓名', '性别', '公民身份号码', '筛查情况', '筛查日期', '备注', '身体状况', '精神状况']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, name='Microsoft YaHei')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            
            # 获取选中的病史筛查数据
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            
            # 构建 SQL 查询，使用 IN 子句
            placeholders = ','.join('?' * len(selected_ids))
            query = f"""
                SELECT name, gender, id_card, screening_result, screening_date, remark,
                       physical_status, mental_status
                FROM medical_screening 
                WHERE id IN ({placeholders})
                ORDER BY id
            """
            cursor.execute(query, selected_ids)
            results = cursor.fetchall()
            conn.close()
            
            # 填充数据
            for row_num, data in enumerate(results, 2):
                ws.cell(row=row_num, column=1, value=row_num - 1)  # 序号
                for col, value in enumerate(data, 2):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.font = Font(name='Microsoft YaHei')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 调整列宽
            column_widths = [8, 12, 8, 20, 30, 12, 15, 12, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # 保存文件
            wb.save(output_path)
            return True, f"成功导出 {len(results)} 条病史筛查记录"
            
        except Exception as e:
            return False, f"导出失败: {str(e)}"
